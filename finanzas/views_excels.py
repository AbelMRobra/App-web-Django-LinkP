import numpy as np
import string
import datetime
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from django.views.generic.base import TemplateView 
from django.http import HttpResponse
from .models import CuentaCorriente, Cuota, Pago
from proyectos.models import Proyectos
from presupuestos.models import Constantes, Registrodeconstantes
from ventas.funciones.f_pricing import unidades_calculo_m2
from finanzas.funciones import funciones_ctacte

class ExcelCuentasCorrientes(TemplateView):

    def get(self, request, id_proyecto, *args, **kwargs):

        proyecto = Proyectos.objects.get(id = id_proyecto)

        con_cuentas_corrientes = CuentaCorriente.objects.filter(venta__proyecto = proyecto).order_by("venta__unidad__orden")

        # Comenzamos a crear las distintas pestañas

        wb = Workbook()

        contador = 3

        ws = wb.active
        ws.title = "Glosario"

        sheets_names = []

        valor_constante_hormigon = Constantes.objects.get(nombre = "Hº VIVIENDA").valor

        for cuenta in con_cuentas_corrientes:

            cuotas_iterar = Cuota.objects.filter(cuenta_corriente = cuenta).order_by("fecha")

            total_pesos = sum(np.array(cuotas_iterar.values_list("precio", flat=True))*np.array(cuotas_iterar.values_list("constante__valor", flat=True)))
            total_cuentas = round(total_pesos/valor_constante_hormigon, 2)


            name_sheet = f'{cuenta.venta.unidad.piso_unidad}-{cuenta.venta.unidad.nombre_unidad}, {cuenta.venta.comprador}'.replace("-", " ").replace("º", "").replace("/", " ").replace(" ", "_").replace(",", "_")
            ws = wb.create_sheet(name_sheet[0:20])
            sheets_names.append(name_sheet[0:20])
            ws.sheet_view.showGridLines = False

            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 20
            ws.column_dimensions['M'].width = 20
            ws.column_dimensions['N'].width = 30
            ws.column_dimensions['O'].width = 20
            ws.column_dimensions['P'].width = 20
            ws.column_dimensions['Q'].width = 20

            ws["A1"].hyperlink = Hyperlink(display="Glosario", ref="A2", location="'Glosario'!A1")

            ws["A2"] = "NOMBRE COMPRADOR"
            ws["A3"] = "DIRECCIÓN"
            ws["A4"] = "TELEFONO FIJO"
            ws["A5"] = "TELEFONO CELULAR"
            ws["A6"] = "E-MAIL"

            ws["A2"].font = Font(bold = True)
            ws["A3"].font = Font(bold = True)
            ws["A4"].font = Font(bold = True)
            ws["A5"].font = Font(bold = True)
            ws["A6"].font = Font(bold = True)

            ws["F2"] = f'Proyecto: {cuenta.venta.proyecto.nombre}'
            ws["F2"].font = Font(bold = True)

            ws["C2"] = cuenta.venta.comprador

            if cuenta.direccion:
                ws["C6"] = cuenta.direccion
            else:
                ws["C3"] = "Dirección no asginada **"
                mensaje = 1

            if cuenta.telefono_fijo:
                ws["C6"] = cuenta.telefono_fijo
            else:
                ws["C4"] = "Telefono fijo no asignado **"
                mensaje = 1
            
            if cuenta.telefono_celular:
                ws["C6"] = cuenta.telefono_celular
            else:
                ws["C5"] = "Celular no asignado **"
                mensaje = 1
            
            if cuenta.venta.email:
                ws["C6"] = cuenta.venta.email
            else: 
                ws["C6"] = "Email no asignado **"
                mensaje = 1

            if mensaje:
                ws["A7"] = "** Para asginar estos datos ingrese a LinkP"

            ws.row_dimensions[7].height = 24
            # Cabeza de la operación real

            ws["B9"] = "OPERACIÓN REAL"
            ws["B9"].alignment = Alignment(horizontal = "center")
            ws["B9"].fill =  PatternFill("solid", fgColor= "CFC9D6")
            ws["B9"].font = Font(bold = True)
            ws["B9"].border = thin_border
            ws.merge_cells("B9:I9")


            ws["B10"] = "ASIGNACIÓN"
            ws["C10"] = "UNIDAD"
            ws["D10"] = "NUMERO"
            ws["E10"] = "SUPERFICIE POR UNIDAD EN M2"
            ws["G10"] = "PRECIO M2"
            ws["H10"] = "PRECIO TOTAL"
            ws["I10"] = "PRECIO M3H"

            ws["B10"].font = Font(bold = True, color="EAE3F2")
            ws["C10"].font = Font(bold = True, color="EAE3F2")
            ws["D10"].font = Font(bold = True, color="EAE3F2")
            ws["E10"].font = Font(bold = True, color="EAE3F2")
            ws["G10"].font = Font(bold = True, color="EAE3F2")
            ws["H10"].font = Font(bold = True, color="EAE3F2")
            ws["I10"].font = Font(bold = True, color="EAE3F2")

            ws["B10"].fill = PatternFill("solid", fgColor= "625E66")
            ws["C10"].fill = PatternFill("solid", fgColor= "625E66")
            ws["D10"].fill = PatternFill("solid", fgColor= "625E66")
            ws["E10"].fill = PatternFill("solid", fgColor= "625E66")
            ws["G10"].fill = PatternFill("solid", fgColor= "625E66")
            ws["H10"].fill = PatternFill("solid", fgColor= "625E66")
            ws["I10"].fill = PatternFill("solid", fgColor= "625E66")

            ws["B10"].border = thin_border
            ws["C10"].border = thin_border
            ws["D10"].border = thin_border
            ws["E10"].border = thin_border
            ws["G10"].border = thin_border
            ws["H10"].border = thin_border
            ws["I10"].border = thin_border

            ws.merge_cells("E10:F10")

            ws["B11"] = cuenta.venta.unidad.asig
            ws["C11"] = f'{cuenta.venta.unidad.piso_unidad}-{cuenta.venta.unidad.nombre_unidad}'
            ws["D11"] = cuenta.venta.unidad.orden
            ws["E11"] = cuenta.venta.unidad.tipologia
            ws["F11"] = unidades_calculo_m2(cuenta.venta.unidad.id)

            try:
                ws["G11"] = cuenta.venta.precio_venta/unidades_calculo_m2(cuenta.venta.unidad.id)
                ws["G11"].number_format = '"$"#,##0.00_-'
            except:
                ws["G11"] = "Error"

            ws["H11"] = cuenta.venta.precio_venta
            ws["H11"].number_format = '"$"#,##0.00_-'
            ws["I11"] = cuenta.venta.precio_venta_hormigon
            ws["I11"].number_format = '#,##0.00_-"M3"'

            ws["D11"].alignment = Alignment(horizontal = "center")
            ws["E11"].alignment = Alignment(horizontal = "center")
            ws["F11"].alignment = Alignment(horizontal = "center")

            ws["I11"].fill =  PatternFill("solid", fgColor= "FAEB9A")
            ws["H11"].fill =  PatternFill("solid", fgColor= "FAEB9A")
            
            ws["B11"].border = thin_border
            ws["C11"].border = thin_border
            ws["D11"].border = thin_border
            ws["E11"].border = thin_border
            ws["F11"].border = thin_border
            ws["G11"].border = thin_border
            ws["H11"].border = thin_border
            ws["I11"].border = thin_border


            # Cabeza de la operación boleto

            ws["B13"] = "OPERACIÓN BOLETO"
            ws["B13"].alignment = Alignment(horizontal = "center")
            ws["B13"].fill =  PatternFill("solid", fgColor= "CFC9D6")
            ws["B13"].font = Font(bold = True)
            ws["B13"].border = thin_border
            ws.merge_cells("B13:I13")


            ws["B14"] = "ASIGNACIÓN"
            ws["C14"] = "UNIDAD"
            ws["D14"] = "NUMERO"
            ws["E14"] = "SUPERFICIE POR UNIDAD EN M2"
            ws["G14"] = "PRECIO M2"
            ws["H14"] = "PRECIO TOTAL"
            ws["I14"] = "PRECIO M3"

            ws["B14"].font = Font(bold = True, color="EAE3F2")
            ws["C14"].font = Font(bold = True, color="EAE3F2")
            ws["D14"].font = Font(bold = True, color="EAE3F2")
            ws["E14"].font = Font(bold = True, color="EAE3F2")
            ws["G14"].font = Font(bold = True, color="EAE3F2")
            ws["H14"].font = Font(bold = True, color="EAE3F2")
            ws["I14"].font = Font(bold = True, color="EAE3F2")

            ws["B14"].fill = PatternFill("solid", fgColor= "625E66")
            ws["C14"].fill = PatternFill("solid", fgColor= "625E66")
            ws["D14"].fill = PatternFill("solid", fgColor= "625E66")
            ws["E14"].fill = PatternFill("solid", fgColor= "625E66")
            ws["G14"].fill = PatternFill("solid", fgColor= "625E66")
            ws["H14"].fill = PatternFill("solid", fgColor= "625E66")
            ws["I14"].fill = PatternFill("solid", fgColor= "625E66")

            ws["B14"].border = thin_border
            ws["C14"].border = thin_border
            ws["D14"].border = thin_border
            ws["E14"].border = thin_border
            ws["F14"].border = thin_border
            ws["G14"].border = thin_border
            ws["H14"].border = thin_border
            ws["I14"].border = thin_border

            ws.merge_cells("E14:F14")


            ws["B15"] = cuenta.venta.unidad.asig
            ws["C15"] = f'{cuenta.venta.unidad.piso_unidad}-{cuenta.venta.unidad.nombre_unidad}'
            ws["D15"] = cuenta.venta.unidad.orden
            ws["E15"] = cuenta.venta.unidad.tipologia
            ws["F15"] = unidades_calculo_m2(cuenta.venta.unidad.id)

            ws["D15"].alignment = Alignment(horizontal = "center")
            ws["E15"].alignment = Alignment(horizontal = "center")
            ws["F15"].alignment = Alignment(horizontal = "center")

            ws["I15"].fill =  PatternFill("solid", fgColor= "FAEB9A")
            ws["H15"].fill =  PatternFill("solid", fgColor= "FAEB9A")

            ws["B15"].border = thin_border
            ws["C15"].border = thin_border
            ws["D15"].border = thin_border
            ws["E15"].border = thin_border
            ws["F15"].border = thin_border
            ws["G15"].border = thin_border
            ws["H15"].border = thin_border
            ws["I15"].border = thin_border

            ws["B17"] = "* Toda la información detalla es resultado de los datos subidos en LinkP"
            ws.row_dimensions[17].height = 24

            ws["B18"] = "INCREMENTO SOBRE SALDO/CUENTAS"
            ws["B18"].alignment = Alignment(horizontal = "center")
            ws["B18"].fill =  PatternFill("solid", fgColor= "CFC9D6")
            ws["B18"].font = Font(bold = True)
            ws["B18"].border = thin_border
            ws.merge_cells("B18:Q18")

            ws.row_dimensions[20].height = 36

            ws["B19"] = "AÑO"
            ws["B19"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["B19"].font = Font(bold = True, color="EAE3F2")
            ws["B19"].fill = PatternFill("solid", fgColor= "625E66")
            ws["B19"].border = thin_border
            ws.merge_cells("B19:B20")

            ws["C19"] = "MES PAGO"
            ws["C19"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["C19"].font = Font(bold = True, color="EAE3F2")
            ws["C19"].fill = PatternFill("solid", fgColor= "625E66")
            ws["C19"].border = thin_border
            ws.merge_cells("C19:C20")

            ws["D19"] = "CONCEPTO"
            ws["D19"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["D19"].font = Font(bold = True, color="EAE3F2")
            ws["D19"].fill = PatternFill("solid", fgColor= "625E66")
            ws["D19"].border = thin_border
            ws.merge_cells("D19:D20")

            ws["E19"] = "CUOTAS"
            ws["E19"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["E19"].font = Font(bold = True, color="EAE3F2")
            ws["E19"].fill = PatternFill("solid", fgColor= "625E66")
            ws["E19"].border = thin_border

            ws["E20"] = "PARCIALES"
            ws["E20"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["E20"].font = Font(bold = True, color="EAE3F2")
            ws["E20"].fill = PatternFill("solid", fgColor= "625E66")
            ws["E20"].border = thin_border

            ws["F19"] = f"{cuotas_iterar.count()}"
            ws["F19"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["F19"].font = Font(bold = True, color="EAE3F2")
            ws["F19"].fill = PatternFill("solid", fgColor= "625E66")
            ws["F19"].border = thin_border

            ws["F20"] = round(total_cuentas,2)
            ws["F20"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["F20"].font = Font(bold = True, color="EAE3F2")
            ws["F20"].fill = PatternFill("solid", fgColor= "625E66")
            ws["F20"].border = thin_border

            ws["G19"] = "AUMENTO SALDO % MENSUAL"
            ws["G19"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["G19"].font = Font(bold = True, color="EAE3F2")
            ws["G19"].fill = PatternFill("solid", fgColor= "625E66")
            ws["G19"].border = thin_border
            ws.merge_cells("G19:L19")

            ws["G20"] = """M3 de hormigón
             de deuda"""
            ws["G20"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["G20"].font = Font(bold = True, color="EAE3F2")
            ws["G20"].fill = PatternFill("solid", fgColor= "625E66")
            ws["G20"].border = thin_border

            ws["H20"] = """Valor del M3 de
            hormigon"""
            ws["H20"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["H20"].font = Font(bold = True, color="EAE3F2")
            ws["H20"].fill = PatternFill("solid", fgColor= "625E66")
            ws["H20"].border = thin_border

            ws["I20"] = """Aumento
            porcentual"""
            ws["I20"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["I20"].font = Font(bold = True, color="EAE3F2")
            ws["I20"].fill = PatternFill("solid", fgColor= "625E66")
            ws["I20"].border = thin_border

            ws["J20"] = """Aumento
            numero"""
            ws["J20"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["J20"].font = Font(bold = True, color="EAE3F2")
            ws["J20"].fill = PatternFill("solid", fgColor= "625E66")
            ws["J20"].border = thin_border

            ws["K20"] = "Cuota"
            ws["K20"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["K20"].font = Font(bold = True, color="EAE3F2")
            ws["K20"].fill = PatternFill("solid", fgColor= "625E66")
            ws["K20"].border = thin_border

            ws["L20"] = "Saldo"
            ws["L20"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["L20"].font = Font(bold = True, color="EAE3F2")
            ws["L20"].fill = PatternFill("solid", fgColor= "625E66")
            ws["L20"].border = thin_border

            ws["M19"] = """FECHA DE
            COBRO"""
            ws["M19"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["M19"].font = Font(bold = True, color="EAE3F2")
            ws["M19"].fill = PatternFill("solid", fgColor= "625E66")
            ws["M19"].border = thin_border
            ws.merge_cells("M19:M20")

            ws["N19"] = """FORMA DE 
            PAGO"""
            ws["N19"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["N19"].font = Font(bold = True, color="EAE3F2")
            ws["N19"].fill = PatternFill("solid", fgColor= "625E66")
            ws["N19"].border = thin_border
            ws.merge_cells("N19:N20")

            ws["O19"] = "MONEDA"
            ws["O19"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["O19"].font = Font(bold = True, color="EAE3F2")
            ws["O19"].fill = PatternFill("solid", fgColor= "625E66")
            ws["O19"].border = thin_border
            ws.merge_cells("O19:O20")

            ws["P19"] = "MONTO"
            ws["P19"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["P19"].font = Font(bold = True, color="EAE3F2")
            ws["P19"].fill = PatternFill("solid", fgColor= "625E66")
            ws["P19"].border = thin_border
            ws.merge_cells("P19:P20")

            ws["Q19"] = "PAGADA"
            ws["Q19"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["Q19"].font = Font(bold = True, color="EAE3F2")
            ws["Q19"].fill = PatternFill("solid", fgColor= "625E66")
            ws["Q19"].border = thin_border
            ws.merge_cells("Q19:Q20")

            valor_inicial = 21
            year_pivote = 0
            celda_pivote = 0
            celda_pivote_2 = 0
            parciales = 1
            valor_anterior = 0
            total_cuentas_acumulado = total_cuentas

            month_names = {
                "1":"Enero",
                "2":"Febrero",
                "3":"Marzo",
                "4":"Abril",
                "5":"Mayo",
                "6":"Junio",
                "7":"Julio",
                "8":"Agosto",
                "9":"Septiembre",
                "10":"Octubre",
                "11":"Noviembre",
                "12":"Diciembre",
            }

            for cuota in cuotas_iterar:

                con_pago = Pago.objects.filter(cuota = cuota)
                pago_pesos = sum(np.array(con_pago.values_list("pago", flat=True))*np.array(con_pago.values_list("cuota__constante__valor", flat=True)))
                pago = round(pago_pesos/valor_constante_hormigon, 2)
                total_cuentas_acumulado = total_cuentas_acumulado - pago

                if year_pivote == 0:
                    year_pivote = cuota.fecha.year
                    celda_pivote = "B"+str(valor_inicial)

                if cuota.fecha.year != year_pivote:

                    ws.merge_cells(str(celda_pivote)+":"+str(celda_pivote_2))
                    year_pivote = cuota.fecha.year
                    celda_pivote = "B"+str(valor_inicial)

                ws["B"+str(valor_inicial)] = cuota.fecha.year
                ws["B"+str(valor_inicial)].alignment = Alignment(horizontal = "center", vertical="center")
                ws["B"+str(valor_inicial)].font = Font(bold = True, color="EAE3F2")
                ws["B"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "625E66")
                ws["B"+str(valor_inicial)].border = thin_border
                
                ws["C"+str(valor_inicial)] = month_names[str(cuota.fecha.month)]
                ws["D"+str(valor_inicial)] = cuota.concepto
                ws["E"+str(valor_inicial)] = parciales
                ws["F"+str(valor_inicial)] = round(total_cuentas_acumulado, 2)
                ws["G"+str(valor_inicial)] = pago

                if pago != 0:
                    pago_pesos = sum(np.array(con_pago.values_list("pago_pesos", flat=True)))
                    valor_hormigon = pago_pesos/pago
                    
                else:

                    facha_busqueda = datetime.date(cuota.fecha.year, cuota.fecha.month, 1)

                    consulta = Registrodeconstantes.objects.filter(fecha = facha_busqueda, constante__nombre = "Hº VIVIENDA")

                    if consulta.count() > 0:
                        
                        valor_hormigon = consulta[0].valor
                        
                        ws["H"+str(valor_inicial)] = valor_hormigon
                        ws["H"+str(valor_inicial)].number_format = '"$"#,##0.00_-'


                    elif cuota.fecha < datetime.date.today():

                        valor_hormigon = 0

                        ws["H"+str(valor_inicial)] = valor_hormigon

                    else:

                        valor_hormigon = valor_constante_hormigon

                        ws["H"+str(valor_inicial)] = valor_hormigon
                        ws["H"+str(valor_inicial)].number_format = '"$"#,##0.00_-'
                
                
                if valor_anterior == 0:

                    ws["I"+str(valor_inicial)] = "-"

                else:
                    
                    ws["I"+str(valor_inicial)] = (valor_hormigon/valor_anterior-1)*100
                    ws["I"+str(valor_inicial)].number_format = '#,##0.00_-"%"'
                
                ws["J"+str(valor_inicial)] = valor_hormigon - valor_anterior
                ws["J"+str(valor_inicial)].number_format = '"$"#,##0.00_-'

                ws["K"+str(valor_inicial)] = valor_hormigon*pago
                ws["K"+str(valor_inicial)].number_format = '#,##0.00_-"M3"'

                ws["L"+str(valor_inicial)] = (total_cuentas_acumulado - pago)*valor_hormigon
                ws["L"+str(valor_inicial)].number_format = '#,##0.00_-"M3"'

                fechas_pago = ""
                metodo_pago = ""

                for consulta in con_pago:

                    fechas_pago += "/"+str(consulta.fecha)
                    metodo_pago += "/"+str(consulta.metodo)

                ws["M"+str(valor_inicial)] = fechas_pago
                ws["N"+str(valor_inicial)] = metodo_pago

                ws["O"+str(valor_inicial)] = cuota.constante.nombre
                ws["P"+str(valor_inicial)] = cuota.precio
                ws["Q"+str(valor_inicial)] = round(float(cuota.pago_moneda_dura()))
                
                valor_anterior = valor_hormigon

                ws["C"+str(valor_inicial)].border = thin_border
                ws["D"+str(valor_inicial)].border = thin_border
                ws["E"+str(valor_inicial)].border = thin_border
                ws["F"+str(valor_inicial)].border = thin_border
                ws["G"+str(valor_inicial)].border = thin_border
                ws["H"+str(valor_inicial)].border = thin_border
                ws["I"+str(valor_inicial)].border = thin_border
                ws["J"+str(valor_inicial)].border = thin_border
                ws["K"+str(valor_inicial)].border = thin_border
                ws["L"+str(valor_inicial)].border = thin_border
                ws["M"+str(valor_inicial)].border = thin_border
                ws["N"+str(valor_inicial)].border = thin_border
                ws["O"+str(valor_inicial)].border = thin_border
                ws["P"+str(valor_inicial)].border = thin_border
                ws["Q"+str(valor_inicial)].border = thin_border

                if pago:

                    ws["C"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["D"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["E"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["F"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["G"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["H"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["I"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["J"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["K"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["L"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["M"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["N"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["O"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["P"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")
                    ws["Q"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F5F584")

                if cuota.precio == 0:

                    ws["C"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["D"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["E"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["F"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["G"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["H"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["I"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["J"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["K"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["L"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["M"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["N"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["O"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["P"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")
                    ws["Q"+str(valor_inicial)].fill = PatternFill("solid", fgColor= "F7C382")

                celda_pivote_2 = "B"+str(valor_inicial)
                valor_inicial += 1
                parciales += 1

            ws.merge_cells(str(celda_pivote)+":"+str(celda_pivote_2))

        paginas = ["INGRESOS TOTALES", "INGRESOS TOTALES M3H", "INGRESOS TOTALES PROYECTO", "INGRESOS TOTALES LINK"]

        for pagina in paginas:

            ws = wb.create_sheet(pagina)
            sheets_names.append(pagina)
            ws.sheet_view.showGridLines = False   

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20

            ws["A1"].hyperlink = Hyperlink(display="Glosario", ref="A2", location="'Glosario'!A1")
            ws["A4"] = Constantes.objects.get(id = 7).valor
            ws["A4"].number_format = '#,##0.00_-"M3"'
            ws["A4"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["A4"].font = Font(bold = True, color="EAE3F2")
            ws["A4"].fill = PatternFill("solid", fgColor= "625E66")
            ws["A4"].border = thin_border


            ws["A5"] = "Fecha:"
            ws["A5"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["A5"].font = Font(bold = True, color="EAE3F2")
            ws["A5"].fill = PatternFill("solid", fgColor= "625E66")
            ws["A5"].border = thin_border


            ws["B5"] = str(datetime.date.today())
            ws["B5"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["B5"].font = Font(bold = True)
            ws["B5"].border = thin_border

            ws["A8"] = "PRECIO INICIAL"
            ws["A8"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["A8"].font = Font(bold = True, color="EAE3F2")
            ws["A8"].fill = PatternFill("solid", fgColor= "625E66")
            ws["A8"].border = thin_border

            ws["A9"] = "PAGADO"
            ws["A9"].alignment = Alignment(horizontal = "center", vertical="center")
            ws["A9"].font = Font(bold = True, color="EAE3F2")
            ws["A9"].fill = PatternFill("solid", fgColor= "625E66")
            ws["A9"].border = thin_border
            
            ws.merge_cells("A8:B8")
            ws.merge_cells("A9:B9")

            fechas = funciones_ctacte.fechas_flujo_excel(proyecto.id)

            contador = 10
            year_pivote = 0
            celda_pivote = 0
            celda_pivote_2 = 0
            parciales = 1
            valor_anterior = 0
            total_cuentas_acumulado = total_cuentas

            for fecha in fechas:
                
                ws["A"+str(contador)] = fecha.year
                ws["A"+str(contador)].alignment = Alignment(horizontal = "center", vertical="center")
                ws["A"+str(contador)].font = Font(bold = True, color="EAE3F2")
                ws["A"+str(contador)].fill = PatternFill("solid", fgColor= "625E66")
                ws["A"+str(contador)].border = thin_border



                ws["B"+str(contador)] = month_names[str(fecha.month)]
                ws["B"+str(contador)].alignment = Alignment(horizontal = "center", vertical="center")
                ws["B"+str(contador)].font = Font(bold = True)
                ws["B"+str(contador)].border = thin_border

                celda_pivote_2 = "A"+str(contador - 1)

                if year_pivote == 0:
                        year_pivote = fecha.year
                        celda_pivote = "A"+str(contador)

                if fecha.year != year_pivote:

                    ws.merge_cells(str(celda_pivote)+":"+str(celda_pivote_2))
                    year_pivote = fecha.year
                    celda_pivote = "A"+str(contador)

                contador += 1

            ws.merge_cells(str(celda_pivote)+":A"+str(contador-1))
        
        cuotas = Cuota.objects.filter(cuenta_corriente__venta__proyecto = proyecto)

        for pagina in paginas:

            chr_contador = 99
            chr_contador_2 = 0

            ws = wb[pagina]

            if pagina == "INGRESOS TOTALES PROYECTO":

                consulta = con_cuentas_corrientes.filter(venta__unidad__asig = "PROYECTO")

            elif pagina == "INGRESOS TOTALES LINK":

                consulta = con_cuentas_corrientes.filter(venta__unidad__asig = "HON. LINK")

            else:

                consulta = con_cuentas_corrientes

            for cuenta in consulta:
            
                if chr_contador_2 == 0:

                    columna = str(chr(chr_contador))

                    if chr_contador == 122:
                        chr_contador = 97
                        chr_contador_2 = 97
                    else:
                        chr_contador += 1

                else:

                    columna = str(chr(chr_contador_2))+str(chr(chr_contador))

                ws.column_dimensions[columna ].width = 20

                if pagina != "INGRESOS TOTALES M3H":

                    ws[columna +"5"] = f'{cuenta.venta.unidad.piso_unidad}-{cuenta.venta.unidad.nombre_unidad}'
                    ws[columna +"6"] = f'{cuenta.venta.comprador}'
                    ws[columna +"7"] = f'{cuenta.venta.unidad.asig}'

                    ws[columna +"5"].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna +"5"].font = Font(bold = True, color="EAE3F2")
                    ws[columna +"5"].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna +"5"].border = thin_border


                    if cuenta.estado == "activo":
                        ws[columna +"6"].font = Font(bold = True, color="0FF728")
                    else:
                        ws[columna +"6"].font = Font(bold = True, color="F41C0E")

                    ws[columna +"6"].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna +"6"].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna +"6"].border = thin_border

                    ws[columna +"7"].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna +"7"].font = Font(bold = True, color="EAE3F2")
                    ws[columna +"7"].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna +"7"].border = thin_border

                    ws[columna +"8"] = cuenta.venta.precio_venta
                    ws[columna +"8"].number_format = '"$"#,##0.00_-'
                    ws[columna +"8"].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna +"8"].font = Font(bold = True, color="EAE3F2")
                    ws[columna +"8"].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna +"8"].border = thin_border

                    ws[columna +"9"] = float(cuenta.pagado_pesos_cuenta())
                    ws[columna +"9"].number_format = '"$"#,##0.00_-'
                    ws[columna +"9"].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna +"9"].font = Font(bold = True, color="EAE3F2")
                    ws[columna +"9"].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna +"9"].border = thin_border

                    ws[columna + str(contador)] = float(cuenta.saldo_pesos())
                    ws[columna + str(contador + 1)] = float(cuenta.estado_pesos())

                    ws[columna + str(contador)].number_format = '"$"#,##0.00_-'
                    ws[columna + str(contador)].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna + str(contador)].font = Font(bold = True)
                    ws[columna + str(contador)].border = thin_border

                    ws[columna + str(contador + 1)].number_format = '"$"#,##0.00_-'
                    ws[columna + str(contador + 1)].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna + str(contador + 1)].font = Font(bold = True, color="EAE3F2")
                    ws[columna + str(contador + 1)].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna + str(contador + 1)].border = thin_border
                    
                else:

                    ws[columna +"5"] = f'{cuenta.venta.unidad.piso_unidad}-{cuenta.venta.unidad.nombre_unidad}'
                    ws[columna +"6"] = f'{cuenta.venta.comprador}'
                    ws[columna +"7"] = f'{cuenta.venta.unidad.asig}'

                    ws[columna +"5"].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna +"5"].font = Font(bold = True, color="EAE3F2")
                    ws[columna +"5"].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna +"5"].border = thin_border

                    if cuenta.estado == "activo":
                        ws[columna +"6"].font = Font(bold = True, color="0FF728")
                    else:
                        ws[columna +"6"].font = Font(bold = True, color="F41C0E")

                    ws[columna +"6"].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna +"6"].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna +"6"].border = thin_border

                    ws[columna +"7"].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna +"7"].font = Font(bold = True, color="EAE3F2")
                    ws[columna +"7"].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna +"7"].border = thin_border

                    if cuenta.venta.precio_venta_hormigon:

                        ws[columna +"8"] = cuenta.venta.precio_venta_hormigon
                        ws[columna +"8"].number_format = '#,##0.00_-"M3"'

                    else:

                        ws[columna +"8"] = 0

                    ws[columna +"8"].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna +"8"].font = Font(bold = True, color="EAE3F2")
                    ws[columna +"8"].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna +"8"].border = thin_border

                    ws[columna +"9"] = float(cuenta.pagado_m3h_cuenta())
                    ws[columna +"9"].number_format = '#,##0.00_-"M3"'
                    ws[columna +"9"].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna +"9"].font = Font(bold = True, color="EAE3F2")
                    ws[columna +"9"].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna +"9"].border = thin_border

                    ws[columna + str(contador)] = float(cuenta.saldo_m3h())
                    ws[columna + str(contador)].number_format = '#,##0.00_-"M3"'


                    ws[columna + str(contador + 1)] = float(cuenta.estado_m3h())
                    ws[columna + str(contador + 1)].number_format = '#,##0.00_-"M3"'

                    
                    ws[columna + str(contador)].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna + str(contador)].font = Font(bold = True)
                    ws[columna + str(contador)].border = thin_border

                    
                    ws[columna + str(contador + 1)].alignment = Alignment(horizontal = "center", vertical="center")
                    ws[columna + str(contador + 1)].font = Font(bold = True, color="EAE3F2")
                    ws[columna + str(contador + 1)].fill = PatternFill("solid", fgColor= "625E66")
                    ws[columna + str(contador + 1)].border = thin_border

                row_fechas = 10

                if pagina != "INGRESOS TOTALES M3H":
                
                    for fecha in fechas:

                        filter_cuotas = cuotas.filter(cuenta_corriente = cuenta, fecha__year = fecha.year, fecha__month = fecha.month)

                        pagado = 0
                        saldo = 0
                        
                        for cuota in filter_cuotas:

                            pagado += float(cuota.pago_pesos())
                            saldo += float(cuota.saldo_pesos())

                        ws[(columna+str(row_fechas))] = (pagado + saldo)

                        ws[(columna+str(row_fechas))].number_format = '"$"#,##0.00_-'
                        ws[(columna+str(row_fechas))].alignment = Alignment(horizontal = "center", vertical="center")
                        ws[(columna+str(row_fechas))].border = thin_border

                        if filter_cuotas.count() == 0:

                            ws[(columna+str(row_fechas))].fill = PatternFill("solid", fgColor= "CDCABC")

                        elif saldo == 0:

                            ws[(columna+str(row_fechas))].fill = PatternFill("solid", fgColor= "C0EC8B")

                        elif pagado > 0:

                            ws[(columna+str(row_fechas))].fill = PatternFill("solid", fgColor= "ECE18B")

                        else:

                            ws[(columna+str(row_fechas))].fill = PatternFill("solid", fgColor= "E4F6F4")

                        row_fechas += 1
                
                else:

                      for fecha in fechas:

                        filter_cuotas = cuotas.filter(cuenta_corriente = cuenta, fecha__year = fecha.year, fecha__month = fecha.month)

                        pagado = 0
                        saldo = 0
                        
                        for cuota in filter_cuotas:

                            if cuota.precio == 0:

                                cuota.pagada == "SI"
                                cuota.save()

                            else:

                                if float(cuota.saldo_moneda_dura())/cuota.precio > 0.03:

                                    cuota.pagada == "NO"
                                    cuota.save()

                            pagado += float(cuota.pago_m3h())
                            saldo += float(cuota.saldo_m3h())

                        ws[(columna+str(row_fechas))] = (pagado + saldo)

                        ws[(columna+str(row_fechas))].number_format = '#,##0.00_-"M3"'
                        ws[(columna+str(row_fechas))].alignment = Alignment(horizontal = "center", vertical="center")
                        ws[(columna+str(row_fechas))].border = thin_border

                        
                        if filter_cuotas.count() == 0:

                            ws[(columna+str(row_fechas))].fill = PatternFill("solid", fgColor= "CDCABC")

                        elif saldo == 0:

                            ws[(columna+str(row_fechas))].fill = PatternFill("solid", fgColor= "C0EC8B")

                        elif pagado > 0:

                            ws[(columna+str(row_fechas))].fill = PatternFill("solid", fgColor= "ECE18B")

                        else:

                            ws[(columna+str(row_fechas))].fill = PatternFill("solid", fgColor= "E4F6F4")

                        row_fechas += 1
                
            columna_final = columna

            row_fechas_total  = 10

            if chr_contador_2 == 0:

                columna = str(chr(chr_contador))

                if chr_contador == 122:
                    chr_contador = 97
                    chr_contador_2 = 97
                else:
                    chr_contador += 1

            else:

                columna = str(chr(chr_contador_2))+str(chr(chr_contador))
            
            ws.column_dimensions[columna ].width = 25

            for fecha in fechas:

                ws[str(columna) + str(row_fechas_total)] = f"=SUM(C{str(row_fechas_total)}:{str(columna_final) + str(row_fechas_total)})"
                row_fechas_total += 1

                ws[str(columna) + str(row_fechas_total)].alignment = Alignment(horizontal = "center", vertical="center")
                ws[str(columna) + str(row_fechas_total)].font = Font(bold = True, color="EAE3F2")
                ws[str(columna) + str(row_fechas_total)].fill = PatternFill("solid", fgColor= "625E66")
                ws[str(columna) + str(row_fechas_total)].border = thin_border

            ws["A"+str(contador)] = "A PAGAR"
            ws["A"+str(contador + 1)] = "TOTAL"

            ws["A"+str(contador)].alignment = Alignment(horizontal = "center", vertical="center")
            ws["A"+str(contador)].font = Font(bold = True, color="EAE3F2")
            ws["A"+str(contador)].fill = PatternFill("solid", fgColor= "625E66")
            ws["A"+str(contador)].border = thin_border
            
            ws.merge_cells("A"+str(contador)+":"+"B"+str(contador))

            ws["A"+str(contador + 1)].alignment = Alignment(horizontal = "center", vertical="center")
            ws["A"+str(contador + 1)].font = Font(bold = True, color="EAE3F2")
            ws["A"+str(contador + 1)].fill = PatternFill("solid", fgColor= "625E66")
            ws["A"+str(contador + 1)].border = thin_border
            
            ws.merge_cells("A"+str(contador + 1)+":"+"B"+str(contador + 1))

        # Aqui creamos la parte del glosario

        ws = wb["Glosario"]
        ws["A2"] = "Glosario del Excel"
        ws["A2"].font = Font(bold = True)

        ws.sheet_view.showGridLines = False  
        ws.column_dimensions['A'].width = 100

        contador = 4

        for name in sheets_names:
            
            to_location = "'{0}'!{1}".format(name, 'A1')
            at_cell = "A"+str(contador)
            ws[at_cell].hyperlink = Hyperlink(display=name, ref=at_cell, location=to_location)
            contador += 1

        #Establecer el nombre del archivo
        nombre_archivo = "CuentasCorrientes{}.xls".format(proyecto.nombre).replace(" ", "_").replace(",", "_")
        
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo).replace(',', '_')
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
