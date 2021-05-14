import os
import datetime
import pandas as pd
import numpy as np
from io import  BytesIO
from xhtml2pdf import pisa
from django.shortcuts import render
from django.shortcuts import redirect
from django.http import HttpResponse
from django.template import Context
from django.views.generic import View
from django.template.loader import get_template
from django.contrib.staticfiles import finders
from django.views.generic.base import TemplateView  
from django.conf import settings
from presupuestos.models import Proyectos, Presupuestos, Constantes, Modelopresupuesto, Registrodeconstantes
from .models import Almacenero, CuentaCorriente, Cuota, Pago, RegistroAlmacenero, ArchivosAdmFin, Arqueo, RetirodeSocios, MovimientoAdmin, Honorarios
from proyectos.models import Unidades, Proyectos
from ventas.models import Pricing, VentasRealizadas
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from rrhh.models import datosusuario, RegistroContable


# Create your views here.


# Vista para imprimir reportes de las cuentas corrientes

class PdfPrueba(View):

    def link_callback(self, uri, rel):
            """
            Convert HTML URIs to absolute system paths so xhtml2pdf can access those
            resources
            """
            sUrl = settings.STATIC_URL        # Typically /static/
            sRoot = settings.STATIC_ROOT      # Typically /home/userX/project_static/
            mUrl = settings.MEDIA_URL         # Typically /media/
            mRoot = settings.MEDIA_ROOT       # Typically /home/userX/project_static/media/

            if uri.startswith(mUrl):
                path = os.path.join(mRoot, uri.replace(mUrl, ""))
            elif uri.startswith(sUrl):
                path = os.path.join(sRoot, uri.replace(sUrl, ""))
            else:
                return uri

            # make sure that file exists
            if not os.path.isfile(path):
                raise Exception(
                        'media URI must start with %s or %s' % (sUrl, mUrl)
                )
            return path

    def get(self, request, id_cuenta, *args, **kwargs):

        #Creamos la información

        ctacte = CuentaCorriente.objects.get(id = id_cuenta)

        cuotas = Cuota.objects.filter(cuenta_corriente = ctacte)

        nombre_conceptos = []

        datos = []

        for cuota in cuotas:

            nombre_conceptos.append(cuota.concepto)

        nombre_conceptos = set(nombre_conceptos)

        saldo_total_pesos = 0

        for nombre in nombre_conceptos:

            moneda = 0
            total_moneda = 0
            cuotas_t = 0
            total_pagado = 0

            for cuota in cuotas:

                if nombre == cuota.concepto:

                    moneda = cuota.constante
                    total_moneda = total_moneda + cuota.precio

                    cuotas_t = (cuotas_t + 1)

                    pagos = Pago.objects.filter(cuota = cuota)

                    for pago in pagos:

                        total_pagado = total_pagado + pago.pago

            saldo_moneda = total_moneda - total_pagado
            saldo_pesos = saldo_moneda*moneda.valor
            saldo_total_pesos = saldo_total_pesos + saldo_pesos
            
            if total_moneda == 0:

                avance = 0
            else:

                avance = total_pagado/total_moneda

            datos.append((nombre, moneda, total_moneda, cuotas_t, total_pagado, saldo_moneda, saldo_pesos, avance))

        datos = sorted(datos, key=lambda datos: datos[7], reverse=True)

        #Aqui creo la curva de ingresos por mes

        fechas = []

        for cuota in cuotas:

            fecha_nueva = date(cuota.fecha.year, cuota.fecha.month, 1 )
            fechas.append(fecha_nueva)

        fechas = list(set(fechas))

        fechas.sort()

        datos_cuotas = []

        deuda_md = 0
        pago_md = 0

        for fecha in fechas:

            deuda_md = 0
            pago_md = 0

            hoy = datetime.date.today()

            if fecha < hoy:

                basura = 1

            else:

                if fecha.month == 12:

                    año = fecha.year + 1

                    fecha_final = date(año, 1, fecha.day)

                else:

                    mes = fecha.month + 1

                    fecha_final = date(fecha.year, mes, fecha.day)

                fecha_final = fecha_final + timedelta(days = -1)

                cuotas = Cuota.objects.filter(fecha__range = (fecha, fecha_final), cuenta_corriente  = ctacte)

                for cuota in cuotas:

                    pagos = Pago.objects.filter(cuota = cuota)

                    for pago in pagos:

                        pago_md = pago_md + pago.pago*pago.cuota.constante.valor 

                    deuda_md = deuda_md + cuota.precio*cuota.constante.valor 

                saldo_md = deuda_md - pago_md

                datos_cuotas.append((fecha, saldo_md))

        
        # Aqui vemos el pago actual

        hoy = datetime.date.today()
        fecha_1 = datetime.date(hoy.year, hoy.month, 1)
        fecha_2 = datetime.date(hoy.year, hoy.month, 28)
        vencimiento = datetime.date(hoy.year, hoy.month, 10)

        venc_cuotas = []

        cuotas_vencimiento = Cuota.objects.filter(cuenta_corriente = ctacte, fecha__range = (fecha_1, fecha_2))

        total_mes = 0
        
        for c in cuotas_vencimiento:
            pesos = c.precio*c.constante.valor
            total_mes = total_mes + pesos
            venc_cuotas.append((c, pesos))

        datos_vencimiento = [venc_cuotas, vencimiento]


        # Aqui llamamos y armamos el PDF
      
        template = get_template('reportepdf.html')
        contexto = {'ctacte':ctacte, 
        'datos':datos, 
        'datos_cuotas':datos_cuotas,
        'datos_vencimiento':datos_vencimiento, 
        'saldo_total_pesos':saldo_total_pesos,
        'total_mes':total_mes,
        'fecha':datetime.date.today(),
        'logo':'{}{}'.format(settings.STATIC_URL, 'img/link.png'),
        'cabecera':'{}{}'.format(settings.STATIC_URL, 'img/fondo.png'),
        'fondo':'{}{}'.format(settings.STATIC_URL, 'img/fondo.jpg')}
        html = template.render(contexto)
        response = HttpResponse(content_type = "application/pdf")
        
        #response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'
        
        pisaStatus = pisa.CreatePDF(html, dest=response, link_callback=self.link_callback)
        
        if pisaStatus.err:
            
            return HttpResponse("Hay un error")

        return response

def mandarmail(request, id_cuenta):

    datos = CuentaCorriente.objects.get(id = id_cuenta)

    mensaje = 0


    if request.method == 'POST':

        venta = VentasRealizadas.objects.get(id = datos.venta.id)

        venta.email = request.POST["email"]

        venta.save()

        hoy = datetime.date.today()

        h = Constantes.objects.get(nombre = "Hº VIVIENDA")

        try:

            # Establecemos conexion con el servidor smtp de gmail
            mailServer = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
            mailServer.ehlo()
            mailServer.starttls()
            mailServer.ehlo()
            mailServer.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

            # Construimos el mensaje simple
            
            mensaje = MIMEMultipart()
            mensaje.attach(MIMEText("""
            
            Estimado cliente {},

            Adjunto estado de cuenta con el valor de la cuota de {}-{}.

            El valor del hormigón de este mes es de ${}.

            Le recordamos que los horarios de tesoreria para realizar pagos son:

            Lunes a Viernes de 08:30 a 12:30

            En caso de realizar pagos por Depósito o Transferencias bancarias, solicitamos enviar el comprobante al siguiente mail: 
            
            mjp@linkinversiones.com.ar
            
            
            Saludos cordiales

            
            """.format(venta.comprador, str(hoy.month), str(hoy.year), str(h.valor)), 'plain'))
            mensaje['From']=settings.EMAIL_HOST_USER
            mensaje['To']=venta.email
            mensaje['Subject']="ESTADO DE CUENTA {}-{}".format(str(hoy.month), str(hoy.year))

            # Esta es la parte para adjuntar (prueba)

            adjunto_MIME = MIMEBase('application', "octet-stream")
            pdf_adjunto = PdfPrueba()
            pdf_adjunto = pdf_adjunto.get(request = request, id_cuenta = id_cuenta).content
            adjunto_MIME.set_payload(pdf_adjunto)
            encoders.encode_base64(adjunto_MIME)
            adjunto_MIME.add_header('Content-Disposition', 'attachment; filename="EstadoDeCuenta{}.{}.pdf"'.format(str(hoy.month), str(hoy.year)))
            mensaje.attach(adjunto_MIME)

            # Envio del mensaje

            mailServer.sendmail(settings.EMAIL_HOST_USER,
                            venta.email,
                            mensaje.as_string())


            datos = CuentaCorriente.objects.get(id = id_cuenta)

            mensaje = "ok"
        except:
            mensaje = "error"

    return render(request, 'mandarmail.html', {"datos":datos, "mensaje":mensaje})
     
def resumencredinv(request):

    busqueda = 1
    datos_almacenados = ArchivosAdmFin.objects.filter(resumen_credito_inv__isnull = False)
    datos = 0
    fecha = 0

    fechas = []

    for dato in datos_almacenados:
        if dato.resumen_credito_inv: 
            fechas.append((dato.fecha, str(dato.fecha)))

    fechas = list(set(fechas))

    fechas.sort( reverse=True)

    if request.method == 'POST':

        #Trae los datos elegidos
        datos_elegidos = request.POST.items()

        for dato in datos_elegidos:

            if dato[0] == "fecha":
                datos = ArchivosAdmFin.objects.get(fecha = dato[1])
                busqueda = 0
                fecha = dato[1]


    datos = {"fechas":fechas,
        "busqueda":busqueda,
        "datos":datos,
        "fecha":fecha}

    return render(request, 'resumencredinv.html', {"datos":datos})

def eliminar_pago(request, id_pago):

    pago = Pago.objects.get(id = id_pago)

    if request.method == 'POST':

        pago.delete()

        return redirect('Pagos', id_cuota = pago.cuota.id)

    return render(request, 'eliminar_pago.html', {"pago":pago})

def editar_cuota(request, id_cuota):

    cuota = Cuota.objects.get(id = id_cuota)

    if request.method == 'POST':

        datos = request.POST.items()

        for i in datos:

            if  'fecha' in i[0] and i[1] != "":

                cuota.fecha = str(i[1])

                cuota.save()

            if  'concepto' in i[0] and i[1] != "":

                cuota.concepto = str(i[1])

                cuota.save()

            if  'precio' in i[0] and i[1] != "":

                cuota.precio = float(i[1])

                cuota.save()

            if  'tipo_venta' in i[0]:

                if i[1] == "HORM":

                    cuota.constante = Constantes.objects.get(nombre = "Hº VIVIENDA")
                    
                    cuota.save()

                if i[1] == "USD":

                    cuota.constante = Constantes.objects.get(nombre = "USD")

                    cuota.save()

        return redirect('Cuenta corriente venta', id_cliente = cuota.cuenta_corriente.id)

    return render(request, 'editar_cuota.html', {"cuota":cuota})

def agregar_cuota(request, id_cuenta):

    cuenta = CuentaCorriente.objects.get(id = id_cuenta)

    if request.method == 'POST':

        datos = request.POST.items()

        fecha = request.POST['fecha1']

        concepto = request.POST['concepto1']

        precio = request.POST['precio1']

        if request.POST['tipo_venta'] == "HORM":
            constante = Constantes.objects.get(nombre = "Hº VIVIENDA"),

        if request.POST['tipo_venta'] == "USD":
                constante = Constantes.objects.get(nombre = "USD"),

        precio_pesos = float(precio)/constante[0].valor


        for i in range(int(request.POST['cantidad'])):

            c = Cuota(

                cuenta_corriente = cuenta,
                fecha = fecha,
                precio = float(precio),
                constante = constante[0],
                precio_pesos = precio_pesos,                        
                concepto = concepto,
                )

            c.save()

            fecha_objeto = datetime.datetime.strptime(str(fecha), '%Y-%m-%d')

            fecha_dia = fecha_objeto.day

            if fecha_objeto.month == 12:

                fecha_mes = 1

                fecha_ano = (fecha_objeto.year + 1)

            else:

                fecha_mes = (fecha_objeto.month + 1)

                fecha_ano = fecha_objeto.year

            hoy = date.today()

            fecha = hoy.replace(fecha_ano, fecha_mes, fecha_dia)

        return redirect('Cuenta corriente venta', id_cliente = cuenta.id)

    return render(request, 'agregar_cuota.html', {"cuenta":cuenta})

def deudores(request, id_proyecto):

    # ---> Lista de proyectos

    proyectos = Proyectos.objects.all()

    list_proyectos = []

    for p in proyectos:

        cantidad = len(CuentaCorriente.objects.filter(venta__proyecto__id = p.id))

        if cantidad > 0:

            list_proyectos.append((p, cantidad))

    if id_proyecto == "0":

        mensaje = "Todos"

        ctas_ctes = CuentaCorriente.objects.all()

    else:

        mensaje = Proyectos.objects.get(id = id_proyecto)

        ctas_ctes = CuentaCorriente.objects.filter(venta__proyecto__id = id_proyecto)

    fecha_hoy = datetime.date.today()

    h = Constantes.objects.get(id = 7).valor
    usd = Constantes.objects.get(id = 1).valor

    datos = []

    deuda_total = 0

    for c in ctas_ctes:

        cuotas_anteriores_h = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__lte = fecha_hoy, constante__id = 7, cuenta_corriente = c)))*h
        cuotas_anteriores_usd = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__lte = fecha_hoy, constante__id = 1, cuenta_corriente = c)))*usd
        pagos_h = sum(np.array(Pago.objects.values_list('pago').filter(fecha__lte = fecha_hoy, cuota__constante__id = 7, cuota__cuenta_corriente = c)))*h
        pagos_usd = sum(np.array(Pago.objects.values_list('pago').filter(fecha__lte = fecha_hoy, cuota__constante__id = 1, cuota__cuenta_corriente = c)))*usd
        
        cuotas = (cuotas_anteriores_h + cuotas_anteriores_usd)/h
        pagos = (pagos_h + pagos_usd)/h
        deuda = cuotas - pagos
        deuda_pesos = deuda*h
        deuda_total = deuda_total + deuda_pesos

        datos.append((c, cuotas, pagos, deuda, deuda_pesos))

    # Aqui empieza el filtro

    if request.method == 'POST':

        datos_viejos = datos

        datos = []   

        palabra_buscar = request.POST["palabra"]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                lista_palabra = palabra.split()

                buscar = (str(i[0].venta.proyecto.nombre)+str(i[0].venta.comprador))

                contador = 0

                for palabra in lista_palabra:

                    contador2 = 0

                    if palabra.lower() in buscar.lower():
  
                        contador += 1

                if contador == len(lista_palabra):

                    datos.append(i)


    # Aqui termina el filtro

    # Ordenamos los datos

    datos = sorted(datos, key=lambda tup: tup[3], reverse=True)


    return render(request, 'deudores.html', {'datos':datos, 'mensaje':mensaje, 'list_proyectos':list_proyectos, 'deuda_total': deuda_total})

def pagos(request, id_cuota):

    cuota = Cuota.objects.get(id = id_cuota)

    datos = Pago.objects.filter(cuota = cuota)

    datos_total = []

    try:

        for d in datos:

            cotizacion = d.pago_pesos/d.pago

            datos_total.append((d, cotizacion))

    except:

        datos_total = 0

    return render(request, 'pagos.html', {'datos':datos_total, 'cuota':cuota})

def editar_pagos(request, id_pago):

    pago = Pago.objects.get(id = id_pago)

    cotizacion = pago.pago_pesos/pago.pago

    if request.method == 'POST':

        datos_crear = request.POST.items()

        pagado = 0

        for i in datos_crear:

            if  'fecha' in i[0] and i[1] != "":

                pago.fecha = str(i[1])
                pago.save()

            if  'documento1' in i[0] and i[1] != "":

                pago.documento_1 = i[1]
                pago.save()

            if  'documento2' in i[0] and i[1] != "":

                pago.documento_2 = i[1]
                pago.save()

            if  'precio1' in i[0] and i[1] != "":

                cotizacion = i[1]

                precio1 = float(pagado)/float(cotizacion)
                pago.pago = precio1
                pago.save()

            if  'precio2' in i[0] and i[1] != "":

                pago.pago_pesos = i[1]
                pagado = i[1]

                pago.save()


        return redirect('Pagos', id_cuota = pago.cuota.id)

    return render(request, 'editar_pagos.html', {'pago':pago, 'cotizacion':cotizacion})

def pagos(request, id_cuota):

    cuota = Cuota.objects.get(id = id_cuota)

    datos = Pago.objects.filter(cuota = cuota)

    datos_total = []

    if len(datos)>0:

        for d in datos:

            if d.pago == 0:
                
                cotizacion = 0

            else:

                cotizacion = d.pago_pesos/d.pago

            datos_total.append((d, cotizacion))

    else:

        datos_total = 0

    return render(request, 'pagos.html', {'datos':datos_total, 'cuota':cuota})

def agregar_pagos(request, id_cuota):

    cuota = Cuota.objects.get(id = id_cuota)

    if request.method == 'POST':

        datos_crear = request.POST.items()

        pagado = float(request.POST['precio2'])

        fecha = request.POST['fecha']

        documento1 = request.POST['documento1']

        documento2 = request.POST['documento2']

        precio1 = float(pagado)/float(request.POST['precio1'])

        c = Pago(

            cuota = cuota,
            fecha = fecha,
            pago = precio1,
            pago_pesos = float(pagado),                       
            documento_1 = documento1,
            documento_2 = documento2,
            )

        c.save()


        return redirect('Pagos', id_cuota = cuota.id)

    return render(request, 'agregar_pagos.html', {'cuota':cuota})

def eliminar_cuota(request, id_cuota):

    cuota = Cuota.objects.get(id = id_cuota)

    if request.method == 'POST':

        cuota.delete()

        return redirect('Cuenta corriente venta', id_cliente = cuota.cuenta_corriente.id)

    return render(request, 'eliminar_cuota.html', {"cuota":cuota})

def crearcuenta(request, id_proyecto):

    proyecto = Proyectos.objects.get(id = id_proyecto)

    datos = VentasRealizadas.objects.filter(proyecto = id_proyecto)

    if request.method == 'POST':

        datos_crear = request.POST.items()

        for i in datos_crear:

            if i[0] == 'ventas':

                b = CuentaCorriente(
                    venta = VentasRealizadas.objects.get(id=i[1]),

                    )

                b.save()

            if  'fecha' in i[0]:

                fecha = i[1]

            if  'concepto' in i[0]:

                concepto = i[1]

            if  'precio' in i[0]:

                precio = i[1]

            if  'cuotas' in i[0]:

                cuotas = i[1]

                constante = Constantes.objects.get(nombre = "Hº VIVIENDA")

                precio_pesos = float(precio)/constante.valor

                for i in range(int(cuotas)):

                    c = Cuota(

                        cuenta_corriente = b,
                        fecha = fecha,
                        precio = float(precio),
                        constante = constante,
                        precio_pesos = precio_pesos,                        
                        concepto = concepto,
                        )

                    c.save()

                    fecha_objeto = datetime.datetime.strptime(str(fecha), '%Y-%m-%d')

                    fecha_dia = fecha_objeto.day

                    if fecha_objeto.month == 12:

                        fecha_mes = 1

                        fecha_ano = (fecha_objeto.year + 1)

                    else:

                        fecha_mes = (fecha_objeto.month + 1)

                        fecha_ano = fecha_objeto.year

                    hoy = date.today()

                    fecha = hoy.replace(fecha_ano, fecha_mes, fecha_dia)

        return redirect('Cuenta corriente venta', b.id)


    return render(request, 'crearcuenta.html', {"datos":datos, "proyecto":proyecto})

def ctacteproyecto(request, id_proyecto):

    proyecto = Proyectos.objects.get(id = id_proyecto)

    datos = CuentaCorriente.objects.filter(venta__proyecto = proyecto)

    return render(request, 'ctacteproyecto.html', {"proyecto":proyecto, "datos":datos})

def principalfinanzas(request):

    return render(request, 'principalfinanzas.html')

def EliminarCuentaCorriente(request, id_cuenta):

    datos = CuentaCorriente.objects.get(id = id_cuenta)

    cuotas = len(Cuota.objects.filter(cuenta_corriente = datos))
    pagos = len(Pago.objects.filter(cuota__cuenta_corriente = datos))

    otros_datos = [cuotas, pagos]

    if request.method == 'POST':

        datos.delete()

        return redirect('Cuenta corriente proyecto', id_proyecto = datos.venta.proyecto.id)

    return render(request, 'eliminar_cuenta.html', {"datos":datos, "otros_datos":otros_datos})

def totalcuentacte(request, id_proyecto):

    proyectos = Proyectos.objects.all()

    datos_primeros = []

    # Listado de los proyectos que tienen cuenta corrientes

    listado = []

    for proyecto in proyectos:

        if len(Cuota.objects.filter(cuenta_corriente__venta__proyecto = proyecto)) > 0:

            listado.append(proyecto)

    proy = 0

    cantidad_cuentas = len(CuentaCorriente.objects.all())


    #Esto es para cuando se selecciona un proyecto en especifico

    if id_proyecto != "0":

        proy = Proyectos.objects.get(id = id_proyecto)

        cantidad_cuentas = len(CuentaCorriente.objects.filter(venta__proyecto__id = id_proyecto))
    
    #Establecemos un rango para hacer el cash de ingreso
    
    fecha_inicial_hoy = datetime.date.today()

    fecha_inicial_2 = datetime.date(fecha_inicial_hoy.year, fecha_inicial_hoy.month, 1)

    fechas = []

    contador = 0
    contador_year = 1

    # El cash en template es un rango de 26 dias

    for f in range(26):

        if (fecha_inicial_2.month + contador) == 13:
            
            year = fecha_inicial_2.year + contador_year
            
            fecha_cargar = date(year, 1, 1)

            fechas.append(fecha_cargar)
            
            contador_year += 1

            contador = - (12 - contador)

        else:

            mes = fecha_inicial_2.month + contador

            year = fecha_inicial_2.year + contador_year - 1

            fecha_cargar = date(year, mes, 1)

            fechas.append(fecha_cargar)

        contador += 1

    #Aqui hacemos los totalizadores generales

    if id_proyecto != "0":

        cuotas_anteriores_h = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__lt = fecha_inicial_hoy, constante__id = 7, cuenta_corriente__venta__proyecto__id = id_proyecto)))*Constantes.objects.get(id = 7).valor
        cuotas_anteriores_usd = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__lt = fecha_inicial_hoy, constante__id = 1, cuenta_corriente__venta__proyecto__id = id_proyecto)))*Constantes.objects.get(id = 1).valor
        pagos_anteriores_h = sum(np.array(Pago.objects.values_list('pago').filter(fecha__lt = fecha_inicial_hoy, cuota__constante__id = 7, cuota__cuenta_corriente__venta__proyecto__id = id_proyecto)))*Constantes.objects.get(id = 7).valor
        pagos_anteriores_usd = sum(np.array(Pago.objects.values_list('pago').filter(fecha__lt = fecha_inicial_hoy, cuota__constante__id = 1, cuota__cuenta_corriente__venta__proyecto__id = id_proyecto)))*Constantes.objects.get(id = 1).valor
        cuotas_posteriores_h = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__gt = fecha_inicial_hoy, constante__id = 7, cuenta_corriente__venta__proyecto__id = id_proyecto)))*Constantes.objects.get(id = 7).valor
        cuotas_posteriores_usd = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__gt = fecha_inicial_hoy, constante__id = 1, cuenta_corriente__venta__proyecto__id = id_proyecto)))*Constantes.objects.get(id = 1).valor


    else:

        
        # Traemos las cuotas y pagos anteriores en hormigon y en dolares (por si constante para hacerlo en pesos), luego dividimos en el valor del Hº para tener todo en 1 moneda

        cuotas_anteriores_h = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__lt = fecha_inicial_hoy, constante__id = 7)))*Constantes.objects.get(id = 7).valor
        cuotas_anteriores_usd = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__lt = fecha_inicial_hoy, constante__id = 1)))*Constantes.objects.get(id = 1).valor
        pagos_anteriores_h = sum(np.array(Pago.objects.values_list('pago').filter(fecha__lt = fecha_inicial_hoy, cuota__constante__id = 7)))*Constantes.objects.get(id = 7).valor
        pagos_anteriores_usd = sum(np.array(Pago.objects.values_list('pago').filter(fecha__lt = fecha_inicial_hoy, cuota__constante__id = 1)))*Constantes.objects.get(id = 1).valor
        cuotas_posteriores_h = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__gt = fecha_inicial_hoy, constante__id = 7)))*Constantes.objects.get(id = 7).valor
        cuotas_posteriores_usd = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__gt = fecha_inicial_hoy, constante__id = 1)))*Constantes.objects.get(id = 1).valor

    total_original = (cuotas_anteriores_h + cuotas_anteriores_usd)
    total_cobrado = (pagos_anteriores_h, pagos_anteriores_usd)
    total_pendiente = total_original - total_cobrado
    total_acobrar= (cuotas_posteriores_h + cuotas_posteriores_usd)
    
    h = Constantes.objects.get(nombre = "Hº VIVIENDA")
    total = total_cobrado[0]/h.valor + total_pendiente[0]/h.valor + total_acobrar[0]/h.valor

    otros_datos = [total_cobrado[0]/h.valor, total_pendiente[0]/h.valor, total_acobrar[0]/h.valor, total] 
    

    #Aqui buscamos agrupar proyecto - sumatorias de cuotas y pagos - mes
    
    datos_segundos = []

    total_fecha = []

    fecha_inicial = 0

    for f in fechas:

        total = 0
        total_link = 0

        datos_terceros = []

        if fecha_inicial == 0:

                fecha_inicial = datetime.date(fecha_inicial_hoy.year, fecha_inicial_hoy.month, 1)

        else:

            #Aqui calculamos el saldo de cuotas totales - Uso el id_proyecto para filtrar

            if id_proyecto != "0":

                cuotas_h = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__range = (fecha_inicial, f), constante__id = 7, cuenta_corriente__venta__proyecto__id = id_proyecto)))*Constantes.objects.get(id = 7).valor
                cuotas_usd = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__range = (fecha_inicial, f), constante__id = 1, cuenta_corriente__venta__proyecto__id = id_proyecto)))*Constantes.objects.get(id = 1).valor
                cuotas_h_link = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__range = (fecha_inicial, f), constante__id = 7, cuenta_corriente__venta__proyecto__id = id_proyecto, cuenta_corriente__venta__unidad__asig = "HON. LINK")))*Constantes.objects.get(id = 7).valor
                cuotas_usd_link = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__range = (fecha_inicial, f), constante__id = 1, cuenta_corriente__venta__proyecto__id = id_proyecto, cuenta_corriente__venta__unidad__asig = "HON. LINK")))*Constantes.objects.get(id = 1).valor
                
                pagos_h = sum(np.array(Pago.objects.values_list('pago').filter(fecha__range = (fecha_inicial, f), cuota__constante__id = 7, cuota__cuenta_corriente__venta__proyecto__id = id_proyecto)))*Constantes.objects.get(id = 7).valor
                pagos_usd = sum(np.array(Pago.objects.values_list('pago').filter(fecha__range = (fecha_inicial, f), cuota__constante__id = 1, cuota__cuenta_corriente__venta__proyecto__id = id_proyecto)))*Constantes.objects.get(id = 1).valor
                pagos_h_link = sum(np.array(Pago.objects.values_list('pago').filter(fecha__range = (fecha_inicial, f), cuota__constante__id = 7, cuota__cuenta_corriente__venta__proyecto__id = id_proyecto, cuota__cuenta_corriente__venta__unidad__asig = "HON. LINK")))*Constantes.objects.get(id = 7).valor
                pagos_usd_link = sum(np.array(Pago.objects.values_list('pago').filter(fecha__range = (fecha_inicial, f), cuota__constante__id = 1, cuota__cuenta_corriente__venta__proyecto__id = id_proyecto , cuota__cuenta_corriente__venta__unidad__asig = "HON. LINK")))*Constantes.objects.get(id = 1).valor



            else:

                cuotas_h = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__range = (fecha_inicial, f), constante__id = 7)))*Constantes.objects.get(id = 7).valor
                cuotas_usd = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__range = (fecha_inicial, f), constante__id = 1)))*Constantes.objects.get(id = 1).valor
                cuotas_h_link = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__range = (fecha_inicial, f), constante__id = 7, cuenta_corriente__venta__unidad__asig = "HON. LINK")))*Constantes.objects.get(id = 7).valor
                cuotas_usd_link = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__range = (fecha_inicial, f), constante__id = 1, cuenta_corriente__venta__unidad__asig = "HON. LINK")))*Constantes.objects.get(id = 1).valor
                
                pagos_h = sum(np.array(Pago.objects.values_list('pago').filter(fecha__range = (fecha_inicial, f), cuota__constante__id = 7)))*Constantes.objects.get(id = 7).valor
                pagos_usd = sum(np.array(Pago.objects.values_list('pago').filter(fecha__range = (fecha_inicial, f), cuota__constante__id = 1)))*Constantes.objects.get(id = 1).valor
                pagos_h_link = sum(np.array(Pago.objects.values_list('pago').filter(fecha__range = (fecha_inicial, f), cuota__constante__id = 7, cuota__cuenta_corriente__venta__unidad__asig = "HON. LINK")))*Constantes.objects.get(id = 7).valor
                pagos_usd_link = sum(np.array(Pago.objects.values_list('pago').filter(fecha__range = (fecha_inicial, f), cuota__constante__id = 1, cuota__cuenta_corriente__venta__unidad__asig = "HON. LINK")))*Constantes.objects.get(id = 1).valor

            total_cuotas = cuotas_h + cuotas_usd
            total_pagado = pagos_h  + pagos_usd                 
            saldo = total_cuotas-total_pagado
            total = total + saldo
            

            total_cuotas_link = cuotas_h_link + cuotas_usd_link
            total_pagado_link = pagos_h_link + pagos_usd_link
            saldo_link = total_cuotas_link-total_pagado_link
            total_link = total_link + saldo_link
           
            datos_terceros.append((fecha_inicial, saldo, saldo_link))

            fecha_inicial = f

            horm = Constantes.objects.get(nombre = "Hº VIVIENDA")
            
            total_horm = total/horm.valor

            total_proy = total - total_link

            total_horm_link = total_link/horm.valor

            total_horm_proy = total_proy/horm.valor

            datos_segundos.append((datos_terceros, total, total_horm, total_link, total_horm_link, total_proy, total_horm_proy))
            
    return render(request, 'totalcuentas.html', {"fechas":fechas, "datos":datos_segundos, "datos_primero":datos_primeros, "total_fechas":total_fecha, "listado":listado, "proy":proy, "cantidad_cuentas":cantidad_cuentas, "otros_datos":otros_datos})

def resumenctacte(request, id_cliente):

    ctacte = CuentaCorriente.objects.get(id = id_cliente)

    cuotas = Cuota.objects.filter(cuenta_corriente = ctacte)

    nombre_conceptos = []

    datos = []

    for cuota in cuotas:

        nombre_conceptos.append(cuota.concepto)

    nombre_conceptos = set(nombre_conceptos)

    saldo_total_pesos = 0

    for nombre in nombre_conceptos:

        moneda = 0
        total_moneda = 0
        cuotas_t = 0
        total_pagado = 0

        for cuota in cuotas:

            if nombre == cuota.concepto:

                moneda = cuota.constante
                total_moneda = total_moneda + cuota.precio

                cuotas_t = (cuotas_t + 1)

                pagos = Pago.objects.filter(cuota = cuota)

                for pago in pagos:

                    total_pagado = total_pagado + pago.pago

        saldo_moneda = total_moneda - total_pagado
        saldo_pesos = saldo_moneda*moneda.valor
        saldo_total_pesos = saldo_total_pesos + saldo_pesos
        
        if total_moneda == 0:
            avance = 0
        else:
            avance = total_pagado/total_moneda

        datos.append((nombre, moneda, total_moneda, cuotas_t, total_pagado, saldo_moneda, saldo_pesos, avance))

    datos = sorted(datos, key=lambda datos: datos[7], reverse=True)

    #Aqui creo la curva de ingresos por mes

    fechas = []

    for cuota in cuotas:

        fecha_nueva = date(cuota.fecha.year, cuota.fecha.month, 1 )
        fechas.append(fecha_nueva)

    fechas = list(set(fechas))

    fechas.sort()

    datos_cuotas = []

    deuda_md = 0
    pago_md = 0

    for fecha in fechas:

        deuda_md = 0
        pago_md = 0

        hoy = datetime.date.today()

        if fecha < hoy:

            basura = 1

        else:

            if fecha.month == 12:

                año = fecha.year + 1

                fecha_final = date(año, 1, fecha.day)

            else:

                mes = fecha.month + 1

                fecha_final = date(fecha.year, mes, fecha.day)

            fecha_final = fecha_final + timedelta(days = -1)

            cuotas = Cuota.objects.filter(fecha__range = (fecha, fecha_final), cuenta_corriente  = ctacte)

            for cuota in cuotas:

                pagos = Pago.objects.filter(cuota = cuota)

                for pago in pagos:

                    pago_md = pago_md + pago.pago*pago.cuota.constante.valor 

                deuda_md = deuda_md + cuota.precio*cuota.constante.valor 

            saldo_md = deuda_md - pago_md

            datos_cuotas.append((fecha, saldo_md))
        

    return render(request, 'resumencta.html', {"ctacte":ctacte, "datos":datos, "datos_cuotas":datos_cuotas, "saldo_total_pesos":saldo_total_pesos})

def ctactecliente(request, id_cliente):

    ctacte = CuentaCorriente.objects.get(id = id_cliente)

    cuotas = Cuota.objects.filter(cuenta_corriente = ctacte)

    pagos = Pago.objects.all()

    datos_cuenta = []

    for cuota in cuotas:

        pago_cuota = sum(np.array(Pago.objects.filter(cuota = cuota).values_list("pago")))
        pago_pesos = sum(np.array(Pago.objects.filter(cuota = cuota).values_list("pago_pesos")))
        saldo_cuota = cuota.precio - pago_cuota
        saldo_pesos = saldo_cuota*cuota.constante.valor
        pagos_realizados = Pago.objects.filter(cuota = cuota)
        saldo_cuota = cuota.precio - pago_cuota

        if pago_cuota == 0:
            cotizacion = 0
        else:
            cotizacion = pago_pesos/pago_cuota

        if cuota.precio != 0 and cuota.pagada == "NO":
            if abs(saldo_cuota*cuota.constante.valor) < 50:
                cuota.precio = pago_cuota
                cuota.pagada = "SI"
                cuota.save()
                saldo_cuota = 0
                saldo_pesos = 0

        datos_cuenta.append((cuota, pago_cuota, saldo_cuota, saldo_pesos, pagos_realizados, cotizacion))

    datos_cuenta = sorted(datos_cuenta, key=lambda datos: datos[0].fecha)

    return render(request, 'ctacte.html', {"ctacte":ctacte, "datos_cuenta":datos_cuenta})

def estructura_boleto(request, id_cliente):

    ctacte = CuentaCorriente.objects.get(id = id_cliente)

    cuotas = Cuota.objects.filter(cuenta_corriente = ctacte)

    operacion_real = 0
    operacion_boleto = 0
    pagado = 0
    saldo = 0

    datos_cuenta = []

    for cuota in cuotas:

        valor_real = cuota.precio*cuota.constante.valor -sum(np.array(Pago.objects.filter(cuota = cuota).values_list("pago")))*cuota.constante.valor + sum(np.array(Pago.objects.filter(cuota = cuota).values_list("pago_pesos")))
        operacion_real = operacion_real + valor_real

        if cuota.boleto  == "BOLETO":

            valor_boleto  = cuota.precio*cuota.constante.valor*cuota.porc_boleto
            
            pago_cuota = sum(np.array(Pago.objects.filter(cuota = cuota).values_list("pago")))*cuota.porc_boleto
            pago_pesos = sum(np.array(Pago.objects.filter(cuota = cuota).values_list("pago_pesos")))*cuota.porc_boleto
            
            operacion_boleto = operacion_boleto + valor_boleto - pago_cuota*cuota.constante.valor + pago_pesos
            
            pagado = pagado + pago_pesos
            saldo_cuota = cuota.precio*cuota.porc_boleto - pago_cuota
            saldo_pesos = saldo_cuota*cuota.constante.valor
            pagos_realizados = Pago.objects.filter(cuota = cuota)
            saldo_cuota = cuota.precio - pago_cuota
            if pago_cuota == 0:
                cotizacion = 0
            else:
                cotizacion = pago_pesos/pago_cuota
            datos_cuenta.append((cuota, pago_pesos, saldo_cuota, saldo_pesos, pagos_realizados, cotizacion, valor_boleto ))

    datos_cuenta = sorted(datos_cuenta, key=lambda datos: datos[0].fecha)

    saldo = operacion_boleto - pagado

    datos_totales = [operacion_real, operacion_boleto, pagado, saldo]

    return render(request, 'ctacte_boleto.html', {"ctacte":ctacte, "datos_cuenta":datos_cuenta, "datos_totales":datos_totales})

def boleto(request, id_cuenta, id_cuota):

    ctacte = CuentaCorriente.objects.get(id = id_cuenta)

    nombre_cuotas = list(set(Cuota.objects.filter(cuenta_corriente = ctacte).values_list("concepto")))

    cuota = Cuota.objects.get(id = id_cuota)

    if request.method == 'POST':

        if request.POST["cuota_edit"] != "":

            cuota.boleto = "BOLETO"
            cuota.porc_boleto = float(request.POST["cuota_edit"])
            cuota.save()

        for n in nombre_cuotas:

            if request.POST[n[0]] != "":

                cuotas = Cuota.objects.filter(concepto = n[0], cuenta_corriente = ctacte)

                for c in cuotas:

                    c.boleto = "BOLETO"
                    c.porc_boleto = float(request.POST[n[0]])
                    c.save()

        return redirect('Cuenta corriente venta', ctacte.id)


    return render(request, 'boleto.html', {"ctacte":ctacte, "nombre_cuotas":nombre_cuotas})

def panelctacote(request):

    datos_ventas = VentasRealizadas.objects.all()

    datos = []

    for dato in datos_ventas:

        datos.append((dato.proyecto.id, dato.proyecto.nombre))

    datos = set(datos)

    if request.method == 'POST':

        proyecto_elegido = request.POST.items()

        for i in proyecto_elegido:

            if i[0] == 'proyecto':

                return redirect('Cuenta corriente proyecto', id_proyecto=i[1])

    return render(request, 'panelctacte.html', {"datos":datos})

def consultapagos(request, id_proyecto):

    # ---> Lista de proyectos

    proyectos = Proyectos.objects.all()

    list_proyectos = []

    for p in proyectos:

        cantidad = len(Pago.objects.filter(cuota__cuenta_corriente__venta__proyecto__id = p.id))

        if cantidad > 0:

            list_proyectos.append((p, cantidad))

    if id_proyecto == "0":

        mensaje = "Todos"

        datos_viejo = Pago.objects.order_by("-fecha")

    else:

        mensaje = Proyectos.objects.get(id = id_proyecto).nombre

        datos_viejo = Pago.objects.filter(cuota__cuenta_corriente__venta__proyecto__id = id_proyecto).order_by("-fecha")


    datos = []

    for d in datos_viejo:

        cotizacion = 0

        if d.pago != 0:

            cotizacion = d.pago_pesos/d.pago

        datos_subir = (d, cotizacion)

        datos.append(datos_subir)

    return render(request, 'pagos_total.html', {"datos":datos, "list_proyectos":list_proyectos, "mensaje":mensaje})

def honorarios(request):

    proyectos = Proyectos.objects.all()

    datos_totales = []

    for p in proyectos:

        if len(Pricing.objects.filter(unidad__proyecto = p)) > 0:

            if len(Unidades.objects.filter(proyecto = p, asig = "HON. LINK")) > 0:

                datos_unidades = Unidades.objects.filter(proyecto = p, estado = "DISPONIBLE", asig = "HON. LINK")
                
                m2_totales = 0

                sumatoria_contado = 0

                cochera = 0

                departamento = 0

                for d in datos_unidades:

                    if d.sup_equiv > 0:

                        m2 = d.sup_equiv

                    else:

                        m2 = d.sup_propia + d.sup_balcon + d.sup_comun + d.sup_patio

                    param_uni = Pricing.objects.get(unidad = d)
                    
                    desde = d.proyecto.desde

                    departamento += 1

                    if d.tipo == "COCHERA":
                        desde = d.proyecto.desde*d.proyecto.descuento_cochera
                        cochera += 1
                        departamento -= 1

                    if param_uni.frente == "SI":
                        desde = desde*d.proyecto.recargo_frente

                    if param_uni.piso_intermedio == "SI":
                        desde =desde*d.proyecto.recargo_piso_intermedio

                    if param_uni.cocina_separada == "SI":
                        desde = desde*d.proyecto.recargo_cocina_separada

                    if param_uni.local == "SI":
                        desde = desde*d.proyecto.recargo_local

                    if param_uni.menor_45_m2 == "SI":
                        desde = desde*d.proyecto.recargo_menor_45

                    if param_uni.menor_50_m2 == "SI":
                        desde = desde*d.proyecto.recargo_menor_50

                    if param_uni.otros == "SI":
                        desde = desde*d.proyecto.recargo_otros 

                    #Aqui calculamos el contado/financiado
                    
                    contado = desde*m2 

                    sumatoria_contado = sumatoria_contado + contado
                    
                    m2_totales = m2_totales + m2

                if m2_totales > 0:

                    precio_promedio_contado = sumatoria_contado/m2_totales

                else:

                    precio_promedio_contado = 0

                h = Constantes.objects.get(nombre = "Hº VIVIENDA")

                sumatoria_contado = sumatoria_contado/h.valor

                # Empezamos cuenta corriente (calculamos el Hº en deuda y pendiente)


                fecha_inicial_hoy = datetime.date.today()

                cuotas_anteriores = Cuota.objects.filter(fecha__lt = fecha_inicial_hoy, cuenta_corriente__venta__proyecto = p, cuenta_corriente__venta__unidad__asig = "HON. LINK")
                pagos_anteriores = Pago.objects.filter(fecha__lt = fecha_inicial_hoy, cuota__cuenta_corriente__venta__proyecto = p, cuota__cuenta_corriente__venta__unidad__asig = "HON. LINK")
                cuotas_posteriores= Cuota.objects.filter(fecha__gt = fecha_inicial_hoy, cuenta_corriente__venta__proyecto = p, cuenta_corriente__venta__unidad__asig = "HON. LINK")
                pagos_posteriores = Pago.objects.filter(fecha__gt = fecha_inicial_hoy, cuota__cuenta_corriente__venta__proyecto = p, cuota__cuenta_corriente__venta__unidad__asig = "HON. LINK")


                total_anterior = 0
                total_pagado_anterior = 0
                total_cobrar = 0
                
                for c in cuotas_anteriores:

                    total_anterior = total_anterior + c.precio*Constantes.objects.get(id = c.constante.id).valor

                for f in pagos_anteriores:

                    total_pagado_anterior = total_pagado_anterior + f.pago*Constantes.objects.get(id = f.cuota.constante.id).valor

                for d in cuotas_posteriores:

                    total_cobrar = total_cobrar + d.precio*Constantes.objects.get(id = d.constante.id).valor

                total_deuda = (total_anterior - total_pagado_anterior)/h.valor

                total_cobrar = total_cobrar/h.valor

                total_m3 = sumatoria_contado + total_cobrar + total_deuda

                datos = (p, cochera, departamento, sumatoria_contado, m2_totales, precio_promedio_contado, total_deuda, total_cobrar, total_m3)

                datos_totales.append(datos)


        honorarios = Honorarios.objects.order_by("-fecha")

        subtotal_1 = honorarios[0].cuotas + honorarios[0].ventas
        ingresos = subtotal_1 + honorarios[0].creditos
        comision = honorarios[0].comision_venta*honorarios[0].ventas
        subtotal_2 = honorarios[0].estructura_gio + honorarios[0].aportes + honorarios[0].socios + comision
        costos = subtotal_2 + honorarios[0].deudas
        beneficio = ingresos - costos + honorarios[0].caja_actual
        porc_beneficio = beneficio/ingresos*100
        beneficio_2 = beneficio - honorarios[0].retiro_socios
        if beneficio == 0:
            porc_costo = 0

        else:
            porc_costo = costos/ingresos*100
        datos_honorarios = [subtotal_1, ingresos, comision, subtotal_2, costos, beneficio, porc_beneficio, porc_costo, beneficio_2]
        

    return render(request, 'honorarios.html', {"datos_totales":datos_totales, "honorarios":honorarios, "datos_honorarios":datos_honorarios})

def modhonorarios(request):

    honorarios = Honorarios.objects.order_by("-fecha")

    if request.method == 'POST':

        # Cuotas

        if request.POST['cuotas'] != "":
            cuotas = request.POST['cuotas']
        else:
            cuotas = honorarios[0].cuotas

        # Ventas

        if request.POST['ventas'] != "":
            ventas = request.POST['ventas']
        else:
            ventas = honorarios[0].ventas

        # Estructura de gastos y GIO

        if request.POST['estructura_gio'] != "":
            estructura_gio = request.POST['estructura_gio']
        else:
            estructura_gio = honorarios[0].estructura_gio

        # Aportes

        if request.POST['aportes'] != "":
            aportes = request.POST['aportes']
        else:
            aportes = honorarios[0].aportes

        # Socios

        if request.POST['socios'] != "":
            socios = request.POST['socios']
        else:
            socios = honorarios[0].socios

        # comision_venta

        if request.POST['comision_venta'] != "":
            comision_venta = request.POST['comision_venta']
        else:
            comision_venta = honorarios[0].comision_venta

        # deudas

        if request.POST['deudas'] != "":
            deudas = request.POST['deudas']
        else:
            deudas = honorarios[0].deudas

        # retiro_socios

        if request.POST['retiro_socios'] != "":
            retiro_socios = request.POST['retiro_socios']
        else:
            retiro_socios = honorarios[0].retiro_socios

        # creditos

        if request.POST['creditos'] != "":
            creditos = request.POST['creditos']
        else:
            creditos = honorarios[0].creditos

        # caja actual

        if request.POST['caja_actual'] != "":
            caja_actual = request.POST['caja_actual']
        else:
            caja_actual = honorarios[0].caja_actual

        b = Honorarios(

            cuotas = cuotas,
            ventas = ventas,
            estructura_gio = estructura_gio,
            aportes = aportes,
            socios = socios,
            comision_venta = comision_venta,
            deudas = deudas,
            retiro_socios = retiro_socios,
            creditos = creditos,
            caja_actual = caja_actual,

        )

        b.save()

        return redirect('Honorarios')


    return render(request, 'modificar_hono.html', {"honorarios":honorarios})

def ingresounidades(request, estado, proyecto):

    if request.method == 'POST':

        proyecto_elegido = request.POST.items()

        for i in proyecto_elegido:

            if i[0] == 'estado':

                unidad = Unidades.objects.get(id = int(request.POST['nombre']))

                unidad.estado = i[1]

                unidad.save()

            if i[0] == 'asig':

                unidad = Unidades.objects.get(id = int(request.POST['nombre']))

                unidad.asig = i[1]

                unidad.save()

            if i[0] == 'comisionsi':

                unidad = Unidades.objects.get(id = int(request.POST['nombre']))
                unidad.estado_comision = "SI"
                unidad.save()

            if i[0] == 'comisionno':

                unidad = Unidades.objects.get(id = int(request.POST['nombre']))
                unidad.estado_comision = "NO"
                unidad.save()

            if i[0] == 'iibbsi':

                unidad = Unidades.objects.get(id = int(request.POST['nombre']))
                unidad.estado_iibb = "SI"
                unidad.save()

            if i[0] == 'iibbno':

                unidad = Unidades.objects.get(id = int(request.POST['nombre']))
                unidad.estado_iibb = "NO"
                unidad.save()



    estado_marcado = estado


    listado = []

    datos_proyectos = datos = Unidades.objects.all()
    for d in datos_proyectos:

        listado.append(d.proyecto)

    listado = set(list(listado))

    if proyecto == "0":
        proyecto_marcado = "Proyecto"
        datos = Unidades.objects.all().order_by("orden")
    else:
        proyecto_marcado = Proyectos.objects.get(id = int(proyecto)).nombre
        datos = Unidades.objects.filter(proyecto__id = int(proyecto)).order_by("orden")
    
    # -----------> Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()
        
        for i in palabra_buscar:

            datos_viejos = datos

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

                datos = []

            if str(palabra_buscar) == "":

                datos = datos_viejos

            else:
            
                for i in datos_viejos:

                    palabra =(str(palabra_buscar))

                    lista_palabra = palabra.split()

                    buscar = (str(i.asig)+str(i.nombre_unidad)+str(i.estado))

                    contador = 0

                    for palabra in lista_palabra:

                        contador2 = 0

                        if palabra.lower() in buscar.lower():
    
                            contador += 1

                    if contador == len(lista_palabra):

                        datos.append(i)

    # -----------> Aqui termina el filtro

    return render(request, 'ingresounidades.html',{'datos':datos, 'estado':estado_marcado, 'proyecto_marcado':proyecto_marcado, 'listado':listado})

def indicelink(request, id_moneda, id_time):

    id_time = id_time
    id_moneda = id_moneda

    if id_moneda == "0":
        moneda = 1
    if id_moneda == "1":
        moneda = Constantes.objects.get(id = 7).valor
    if id_moneda == "2":
        moneda = Constantes.objects.get(id = 2).valor

    datos = Almacenero.objects.all()
    datos_completos = []
    datos_finales = []
    saldo_caja_total = 0
    pendiente_gastar_total = 0
    ingresos_total = 0
    descuento_total = 0

    for dato in datos:

        presupuesto = "NO"

        pricing = "NO"

        almacenero = dato

        presupuesto = Presupuestos.objects.get(proyecto = dato.proyecto)
        
        # Calculamos los 2 valores de cuenta corriente (Pendiente y Cobrado)

        if almacenero.auto_cta == "SI":

            cuota = Cuota.objects.filter(cuenta_corriente__venta__proyecto = dato.proyecto)

            total_pesos = 0
            total_pagado_pesos = 0
            total_pagado_historico = 0
            
            if len(cuota) > 0:

                for c in cuota:

                    moneda_cuota = c.constante
                    total_pesos = total_pesos + c.precio*moneda_cuota.valor
                    pagos = Pago.objects.filter(cuota = c)

                    for p in pagos:
                        total_pagado_pesos = total_pagado_pesos + p.pago*moneda_cuota.valor
                        total_pagado_historico = total_pagado_historico + p.pago_pesos
            
            saldo_pesos = total_pesos - total_pagado_pesos
            almacenero.cuotas_a_cobrar = saldo_pesos
            almacenero.cuotas_cobradas = total_pagado_historico
            almacenero.save()

        # Calculo el resto de las cosas

        if id_time == "1":

            #===================================================
            # Calculamos la logica de los meses
            #===================================================

            ahora = datetime.date.today()

            if dato.proyecto.fecha_i > ahora:
                meses_costo = dato.proyecto.fecha_f.month - dato.proyecto.fecha_i.month
                meses_costo = int(meses_costo/2 + dato.proyecto.fecha_i.month - ahora.month)
            else:
                if dato.proyecto.fecha_f > ahora:
                    meses_costo = int((dato.proyecto.fecha_f.month - ahora.month)/2)
                else:
                    meses_costo = 0   

            meses_ingreso = int((dato.proyecto.fecha_f.month - ahora.month)/2)

            if meses_costo:
                array_costo = np.zeros((meses_costo + 1), dtype = int)
            else:
                array_costo =np.zeros(1, dtype = int)

            if meses_ingreso > 0:
                array_ingreso = np.zeros((meses_ingreso+1), dtype = int)
            else:
                array_ingreso = np.zeros(1, dtype = int)

        retiro_socios = sum(np.array(RetirodeSocios.objects.values_list('monto_pesos').filter(proyecto = dato.proyecto)))
        saldo_caja = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados - retiro_socios + almacenero.tenencia
        saldo_caja_total = saldo_caja_total + saldo_caja


        pend_gast = almacenero.pendiente_admin + almacenero.pendiente_comision + presupuesto.saldo_mat + presupuesto.saldo_mo + presupuesto.imprevisto + presupuesto.credito + presupuesto.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos
        
        if id_time == "1":
            array_costo = np.append(array_costo, [pend_gast])
            pend_gast = np.npv(rate=(dato.proyecto.tasa_f/100), values=array_costo)
     
        pendiente_gastar_total = pendiente_gastar_total + pend_gast
        prest_cobrar = almacenero.prestamos_proyecto + almacenero.prestamos_otros 

        cuenta_corriente = almacenero.cuotas_a_cobrar
        total_ingresos = prest_cobrar  + almacenero.ingreso_ventas + almacenero.financiacion + almacenero.inmuebles

        if id_time == "0":
            total_ingresos = total_ingresos + cuenta_corriente
            
        if id_time == "1" and almacenero.auto_cta == "NO":
            total_ingresos = total_ingresos + cuenta_corriente
            array_ingreso_1 = np.append(array_ingreso, [total_ingresos])
            total_ingresos = np.npv(rate=(dato.proyecto.tasa_f/100), values=array_ingreso_1)
      
        if id_time == "1" and almacenero.auto_cta == "SI":


            #===================================================
            # Armamos el flujo de cuentas corriente
            #===================================================

            #===================================================
            # Armamos las fechas necesarias
            #===================================================

            #Establecemos un rango para hacer el cash de ingreso
        
            fecha_inicial_hoy = datetime.date.today()

            fecha_inicial_2 = datetime.date(fecha_inicial_hoy.year, fecha_inicial_hoy.month, 1)

            fechas = []

            #Aqui buscamos la ultima fecha

            fecha_ultima = Cuota.objects.filter(cuenta_corriente__venta__proyecto = dato.proyecto, cuenta_corriente__venta__unidad__asig = "PROYECTO").order_by("-fecha")

            contador = 0
            
            contador_year = 1

            fecha_cargar = fecha_inicial_2

            while fecha_cargar < fecha_ultima[0].fecha:

                if (fecha_inicial_2.month + contador) == 13:
                    
                    year = fecha_inicial_2.year + contador_year
                    
                    fecha_cargar = date(year, 1, fecha_inicial_2.day)

                    fechas.append(fecha_cargar)
                    
                    contador_year += 1

                    contador = - (12 - contador)

                else:

                    mes = fecha_inicial_2.month + contador

                    year = fecha_inicial_2.year + contador_year - 1

                    fecha_cargar = date(year, mes, fecha_inicial_2.day)

                    fechas.append(fecha_cargar)

                contador += 1
            
            #===================================================
            # Armamos el cash
            #===================================================
            
            array_cuotas = [0]

            fecha_inicial = 0

            for f in fechas:

                total = 0
                
                if fecha_inicial == 0:

                        fecha_inicial = fecha_inicial_hoy
                else:

                    cuotas = Cuota.objects.filter(fecha__range = (fecha_inicial, f), cuenta_corriente__venta__proyecto = dato.proyecto, cuenta_corriente__venta__unidad__asig = "PROYECTO")
                        
                    pagos = Pago.objects.filter(fecha__range = (fecha_inicial, f), cuota__cuenta_corriente__venta__proyecto = dato.proyecto, cuota__cuenta_corriente__venta__unidad__asig = "PROYECTO")

                    total_cuotas_proyecto = 0
                    total_pagado_proyecto = 0

                    if len(cuotas)>0:

                        for c in cuotas:

                            if c.cuenta_corriente.venta.unidad.asig == "PROYECTO":

                                total_cuotas_proyecto = total_cuotas_proyecto + c.precio*c.constante.valor 

                    if len(pagos)>0:

                        for p in pagos:

                            if p.cuota.cuenta_corriente.venta.unidad.asig == "PROYECTO":

                                total_pagado_proyecto = total_pagado_proyecto + p.pago*p.cuota.constante.valor 

                    saldo = total_cuotas_proyecto - total_pagado_proyecto

                    #Aqui calculamos el saldo de cuotas de LINK
                    
                    array_cuotas.append(saldo)

                    fecha_inicial = f



            cuenta_corriente = np.npv(rate=(dato.proyecto.tasa_f/100), values=array_cuotas)
            array_ingreso_1 = np.append(array_ingreso, [total_ingresos])
            total_ingresos = np.npv(rate=(dato.proyecto.tasa_f/100), values=array_ingreso_1) + cuenta_corriente


        ingresos_total = ingresos_total + total_ingresos
        margen = total_ingresos - pend_gast + saldo_caja
        descuento = almacenero.ingreso_ventas*0.06

        if id_time == "1":
            array_ingreso_2 = np.append(array_ingreso, [descuento])
            descuento = np.npv(rate=(dato.proyecto.tasa_f/100), values=array_ingreso_2)

        descuento_total = descuento_total + descuento
        margen_2 = margen - descuento 
               
        datos_completos.append((dato, saldo_caja/moneda, pend_gast/moneda, total_ingresos/moneda, margen/moneda, descuento/moneda, margen_2/moneda))

    
    # -----------------> Aqui calculo los totalizadores

    margen1_total = ingresos_total - pendiente_gastar_total + saldo_caja_total
    margen2_total = margen1_total - descuento_total


    # -----------------> Aqui calculo la parte de honorarios

    honorarios = Honorarios.objects.order_by("-fecha")
    caja_actual = honorarios[0].caja_actual
    subtotal_1 = honorarios[0].cuotas + honorarios[0].ventas
    ingresos = subtotal_1 + honorarios[0].creditos
    comision = honorarios[0].comision_venta*honorarios[0].ventas
    subtotal_2 = honorarios[0].estructura_gio + honorarios[0].aportes + honorarios[0].socios + comision
    costos = subtotal_2  + honorarios[0].deudas
    honorario = ingresos - costos + honorarios[0].caja_actual
    honorarios2 = honorario

    # -----------------> Aqui calculo los totalizadores con los honorarios

    caja_total = caja_actual + saldo_caja_total
    costos_totales = pendiente_gastar_total + costos
    ingresos_totales = ingresos_total + ingresos
    margen1_completo = margen1_total + honorario
    margen2_completo = margen1_completo - descuento_total

    # -----------------> Información para graficos

    retiro_honorarios = 0
    honorarios_beneficio2 = 0
    honorarios_beneficio1 = 0

    datos_finales.append((saldo_caja_total , pendiente_gastar_total, ingresos_total, descuento_total, margen1_total, margen2_total))

    datos_finales_2 = [honorario, ingresos, costos, honorarios2, ingresos_totales, costos_totales, margen1_completo , margen2_completo, retiro_honorarios, honorarios_beneficio2, honorarios_beneficio1, caja_actual, caja_total]

    # -----------------> Bucle para moneda

    datos_finales = np.array(datos_finales)/moneda
    datos_finales_2 = np.array(datos_finales_2)/moneda

    # -----------------> Esta es la parte del historico

    datos_historicos = RegistroAlmacenero.objects.order_by("fecha")

    fechas = []

    for d in datos_historicos:

        if not d.fecha in fechas:

            fechas.append(d.fecha)

    datos_registro = []

    for fecha in fechas:

        datos = RegistroAlmacenero.objects.filter(fecha = fecha)

        saldo_caja_total = 0
        pendiente_gastar_total = 0
        ingresos_total = 0
        descuento_total = 0
        honorario = 0


        for dato in datos:

            almacenero = dato

            #Calculo el resto de las cosas

            retiro_socios = almacenero.retiro_socios
            saldo_caja = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados - retiro_socios + almacenero.tenencia
            
            pend_gast = almacenero.pendiente_admin + almacenero.pendiente_comision + almacenero.saldo_mat + almacenero.saldo_mo + almacenero.imprevisto + almacenero.credito + almacenero.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos
            
            prest_cobrar = almacenero.prestamos_proyecto + almacenero.prestamos_otros
            total_ingresos = prest_cobrar + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + almacenero.financiacion + almacenero.inmuebles
            
            margen = total_ingresos - pend_gast + saldo_caja
            descuento = almacenero.ingreso_ventas*0.06
            
            margen_2 = margen - descuento
            honorario = almacenero.honorarios

            pendiente_gastar_total = pendiente_gastar_total + pend_gast
            saldo_caja_total = saldo_caja_total + saldo_caja
            descuento_total = descuento_total + descuento
            ingresos_total = ingresos_total + total_ingresos

            # Me falta calcular la parte de honorarios


        margen1 = ingresos_total - pendiente_gastar_total + honorario + saldo_caja_total
        margen2 = margen1 - descuento_total

        if id_moneda == "0":

            datos_registro.append((margen1, margen2))

        if id_moneda == "1":

            fecha_valor = datetime.date(fecha.year, fecha.month, 1)
            
            valor = Registrodeconstantes.objects.get(fecha=fecha_valor, constante__id = 7)

            datos_registro.append((margen1/valor.valor, margen2/valor.valor))

        if id_moneda == "2":

            fecha_valor = datetime.date(fecha.year, fecha.month, 1)
            
            valor = Registrodeconstantes.objects.get(fecha=fecha_valor, constante__id = 2)

            datos_registro.append((margen1/valor.valor, margen2/valor.valor))
  

    return render(request, 'indicelink.html', {"id_time":id_time, "id_moneda":id_moneda, "datos_completos":datos_completos, 'datos_finales':datos_finales, "datos_registro":datos_registro, "fechas":fechas, "datos_finales_2":datos_finales_2})

def registro_almacenero(request):

    datos = []
    var_especifica = []
    datos_finales = []
    mensaje = 0

    if request.method == 'POST':

        fecha_1 = request.POST['fechafinal']
        fecha_2 = request.POST['fechainicial']

        # --> Prueba de existencia

        datos_fecha1 = RegistroAlmacenero.objects.filter(fecha = fecha_1)
        datos_fecha2 = RegistroAlmacenero.objects.filter(fecha = fecha_2)

        if (len(datos_fecha1) == 0) or (len(datos_fecha2) == 0):
            mensaje = (1, "Algunas de las fechas ingresadas no tiene registros")

        # --> Prueba de cambio

        elif (len(datos_fecha1) != len(datos_fecha2)):

            proyectos_fecha1 = RegistroAlmacenero.objects.filter(fecha = fecha_1).values_list('proyecto', flat = True).distinct()
            proyectos_fecha2 = RegistroAlmacenero.objects.filter(fecha = fecha_2).values_list('proyecto', flat = True).distinct()

            proyectos_agregados = []

            for p in proyectos_fecha1:
                if p not in proyectos_fecha2:
                    proyectos_agregados.append(p)

            mensaje = (3, "Hubo variación con la cantidad de proyectos")

            almacenero_1 = datos_fecha1

            saldo_caja_total_1 = 0
            pendiente_gastar_total_1 = 0
            ingresos_total_1 = 0
            descuento_total_1 = 0
            honorario_1 = 0

            for dato in almacenero_1:

                almacenero = dato

                # Calculo el resto de las cosas

                retiro_socios_1 = almacenero.retiro_socios
                saldo_caja_1 = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados - retiro_socios_1 + almacenero.tenencia
                pend_gast_1 = almacenero.pendiente_admin + almacenero.pendiente_comision + almacenero.saldo_mat + almacenero.saldo_mo + almacenero.imprevisto + almacenero.credito + almacenero.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos                
                prest_cobrar_1 = almacenero.prestamos_proyecto + almacenero.prestamos_otros
                total_ingresos_1 = prest_cobrar_1 + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + almacenero.financiacion + almacenero.inmuebles
                margen_1 = total_ingresos_1 - pend_gast_1 + saldo_caja_1
                descuento_1 = almacenero.ingreso_ventas*0.06
                
                try:
                    # Variación especifica

                    aux = RegistroAlmacenero.objects.get(fecha = fecha_2, proyecto = almacenero.proyecto)

                    retiro_socios_aux = aux.retiro_socios
                    saldo_caja_aux = aux.cuotas_cobradas - aux.gastos_fecha - aux.Prestamos_dados - retiro_socios_aux + aux.tenencia
                    pend_gast_aux = aux.pendiente_admin + aux.pendiente_comision + aux.saldo_mat + aux.saldo_mo + aux.imprevisto + aux.credito + aux.fdr - aux.pendiente_adelantos + aux.pendiente_iva_ventas + aux.pendiente_iibb_tem +aux.cheques_emitidos                
                    prest_cobrar_aux = aux.prestamos_proyecto + aux.prestamos_otros
                    total_ingresos_aux = prest_cobrar_aux + aux.cuotas_a_cobrar + aux.ingreso_ventas + aux.financiacion + aux.inmuebles
                    margen_aux = total_ingresos_aux - pend_gast_aux + saldo_caja_aux
                    descuento_aux = aux.ingreso_ventas

                    var_especifica.append((almacenero.proyecto, [(margen_aux - margen_1), (total_ingresos_aux - total_ingresos_1), (pend_gast_aux - pend_gast_1), (saldo_caja_aux - saldo_caja_1), (descuento_aux - descuento_1)]))

                except:
                    var_especifica.append((almacenero.proyecto, [(margen_1), (total_ingresos_1), (pend_gast_1), (saldo_caja_1), (descuento_1)]))

                # Suma a los totalizadores

                margen_2_1 = margen_1 - descuento_1
                pendiente_gastar_total_1 = pendiente_gastar_total_1 + pend_gast_1
                saldo_caja_total_1 = saldo_caja_total_1 + saldo_caja_1
                descuento_total_1 = descuento_total_1 + descuento_1
                ingresos_total_1 = ingresos_total_1 + total_ingresos_1


                # Honorarios
                
                honorario_1 = almacenero.honorarios


            margen1_1 = ingresos_total_1 - pendiente_gastar_total_1 + honorario_1 + saldo_caja_total_1
            margen2_1 = margen1_1 - descuento_total_1


            almacenero_2 = datos_fecha2

            saldo_caja_total_2 = 0
            pendiente_gastar_total_2 = 0
            ingresos_total_2 = 0
            descuento_total_2 = 0
            honorario_2 = 0

            for dato in almacenero_2:

                almacenero = dato

                #Calculo el resto de las cosas

                retiro_socios_2 = almacenero.retiro_socios
                saldo_caja_2 = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados - retiro_socios_2 + almacenero.tenencia
                
                pend_gast_2 = almacenero.pendiente_admin + almacenero.pendiente_comision + almacenero.saldo_mat + almacenero.saldo_mo + almacenero.imprevisto + almacenero.credito + almacenero.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos
                
                prest_cobrar_2 = almacenero.prestamos_proyecto + almacenero.prestamos_otros
                total_ingresos_2 = prest_cobrar_2 + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + almacenero.financiacion + almacenero.inmuebles
                
                margen_2 = total_ingresos_2 - pend_gast_2 + saldo_caja_2
                descuento_2 = almacenero.ingreso_ventas*0.06
                
                margen_2_2 = margen_2 - descuento_2
                honorario_2 = almacenero.honorarios

                pendiente_gastar_total_2 = pendiente_gastar_total_2 + pend_gast_2
                saldo_caja_total_2 = saldo_caja_total_2 + saldo_caja_2
                descuento_total_2 = descuento_total_2 + descuento_2
                ingresos_total_2 = ingresos_total_2 + total_ingresos_2

                # Me falta calcular la parte de honorarios

            margen1_2 = ingresos_total_2 - pendiente_gastar_total_2 + honorario_2 + saldo_caja_total_2
            margen2_2 = margen1_2 - descuento_total_2

            mensaje = (2, "Solicitud procesada correctamente!, Sin variación de proyectos")

            datos.append((round(margen2_1, 2), round(margen2_2, 2), round((margen2_2/margen2_1 -1 )*100, 10), round((margen2_2 - margen2_1), 10)))
            datos.append((fecha_1, fecha_2))
            datos.append((round((ingresos_total_2/ingresos_total_1 -1 )*100, 10), round((descuento_total_2/descuento_total_1 -1 )*100), round((saldo_caja_total_2/saldo_caja_total_1 -1 )*100, 10), round((pendiente_gastar_total_2/pendiente_gastar_total_1 -1 )*100, 10)))


        elif (len(datos_fecha1) == len(datos_fecha2)):

            almacenero_1 = datos_fecha1

            saldo_caja_total_1 = 0
            pendiente_gastar_total_1 = 0
            ingresos_total_1 = 0
            descuento_total_1 = 0
            honorario_1 = 0

            for dato in almacenero_1:

                almacenero = dato

                # Calculo el resto de las cosas

                retiro_socios_1 = almacenero.retiro_socios
                saldo_caja_1 = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados - retiro_socios_1 + almacenero.tenencia
                pend_gast_1 = almacenero.pendiente_admin + almacenero.pendiente_comision + almacenero.saldo_mat + almacenero.saldo_mo + almacenero.imprevisto + almacenero.credito + almacenero.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos                
                prest_cobrar_1 = almacenero.prestamos_proyecto + almacenero.prestamos_otros
                total_ingresos_1 = prest_cobrar_1 + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + almacenero.financiacion + almacenero.inmuebles
                margen_1 = total_ingresos_1 - pend_gast_1 + saldo_caja_1
                descuento_1 = almacenero.ingreso_ventas*0.06
                

                # Variación especifica

                aux = RegistroAlmacenero.objects.get(fecha = fecha_2, proyecto = almacenero.proyecto)

                retiro_socios_aux = aux.retiro_socios
                saldo_caja_aux = aux.cuotas_cobradas - aux.gastos_fecha - aux.Prestamos_dados - retiro_socios_aux + aux.tenencia
                pend_gast_aux = aux.pendiente_admin + aux.pendiente_comision + aux.saldo_mat + aux.saldo_mo + aux.imprevisto + aux.credito + aux.fdr - aux.pendiente_adelantos + aux.pendiente_iva_ventas + aux.pendiente_iibb_tem +aux.cheques_emitidos                
                prest_cobrar_aux = aux.prestamos_proyecto + aux.prestamos_otros
                total_ingresos_aux = prest_cobrar_aux + aux.cuotas_a_cobrar + aux.ingreso_ventas + aux.financiacion + aux.inmuebles
                margen_aux = total_ingresos_aux - pend_gast_aux + saldo_caja_aux
                descuento_aux = aux.ingreso_ventas

                # Variación ultra especifica

                list_gasto = [(aux.pendiente_admin - almacenero.pendiente_admin), (aux.pendiente_admin - almacenero.pendiente_admin), (aux.pendiente_comision - almacenero.pendiente_comision), (aux.saldo_mat - almacenero.saldo_mat), (aux.saldo_mo - almacenero.saldo_mo), (aux.imprevisto - almacenero.imprevisto), (aux.credito - almacenero.credito), (aux.fdr - almacenero.fdr), (aux.pendiente_adelantos - almacenero.pendiente_adelantos), (aux.pendiente_iva_ventas - almacenero.pendiente_iva_ventas), (aux.pendiente_iibb_tem - almacenero.pendiente_iibb_tem), (aux.cheques_emitidos - almacenero.cheques_emitidos), ]
                list_caja = [(aux.cuotas_cobradas - almacenero.cuotas_cobradas), (aux.gastos_fecha - almacenero.gastos_fecha), (aux.Prestamos_dados - almacenero.Prestamos_dados), (retiro_socios_aux - retiro_socios_1), (aux.tenencia - almacenero.tenencia)]
                list_ingresos = [(prest_cobrar_aux - prest_cobrar_1), (aux.cuotas_a_cobrar - almacenero.cuotas_a_cobrar), (aux.ingreso_ventas - almacenero.ingreso_ventas), (aux.financiacion - almacenero.financiacion), (aux.inmuebles - almacenero.inmuebles)]
                
                var_especifica.append((almacenero.proyecto, [(margen_aux - margen_1), (total_ingresos_aux - total_ingresos_1), (pend_gast_aux - pend_gast_1), (saldo_caja_aux - saldo_caja_1), (descuento_aux - descuento_1), list_gasto, list_caja, list_ingresos]))
                
                # Suma a los totalizadores

                margen_2_1 = margen_1 - descuento_1
                pendiente_gastar_total_1 = pendiente_gastar_total_1 + pend_gast_1
                saldo_caja_total_1 = saldo_caja_total_1 + saldo_caja_1
                descuento_total_1 = descuento_total_1 + descuento_1
                ingresos_total_1 = ingresos_total_1 + total_ingresos_1


                # Honorarios
                
                honorario_1 = almacenero.honorarios


            margen1_1 = ingresos_total_1 - pendiente_gastar_total_1 + honorario_1 + saldo_caja_total_1
            margen2_1 = margen1_1 - descuento_total_1


            almacenero_2 = datos_fecha2

            saldo_caja_total_2 = 0
            pendiente_gastar_total_2 = 0
            ingresos_total_2 = 0
            descuento_total_2 = 0
            honorario_2 = 0

            for dato in almacenero_2:

                almacenero = dato

                #Calculo el resto de las cosas

                retiro_socios_2 = almacenero.retiro_socios
                saldo_caja_2 = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados - retiro_socios_2 + almacenero.tenencia
                
                pend_gast_2 = almacenero.pendiente_admin + almacenero.pendiente_comision + almacenero.saldo_mat + almacenero.saldo_mo + almacenero.imprevisto + almacenero.credito + almacenero.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos
                
                prest_cobrar_2 = almacenero.prestamos_proyecto + almacenero.prestamos_otros
                total_ingresos_2 = prest_cobrar_2 + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + almacenero.financiacion + almacenero.inmuebles
                
                margen_2 = total_ingresos_2 - pend_gast_2 + saldo_caja_2
                descuento_2 = almacenero.ingreso_ventas*0.06
                
                margen_2_2 = margen_2 - descuento_2
                honorario_2 = almacenero.honorarios

                pendiente_gastar_total_2 = pendiente_gastar_total_2 + pend_gast_2
                saldo_caja_total_2 = saldo_caja_total_2 + saldo_caja_2
                descuento_total_2 = descuento_total_2 + descuento_2
                ingresos_total_2 = ingresos_total_2 + total_ingresos_2

                # Me falta calcular la parte de honorarios

            margen1_2 = ingresos_total_2 - pendiente_gastar_total_2 + honorario_2 + saldo_caja_total_2
            margen2_2 = margen1_2 - descuento_total_2

            mensaje = (2, "Solicitud procesada correctamente!, Sin variación de proyectos")

            datos.append((round(margen2_1, 2), round(margen2_2, 2), round((margen2_2/margen2_1 -1 )*100, 2), round((margen2_2 - margen2_1), 2)))
            datos.append((fecha_1, fecha_2))
            datos.append((round((ingresos_total_2/ingresos_total_1 -1 )*100, 3), round((descuento_total_2/descuento_total_1 -1 )*100, 3), round((saldo_caja_total_2/saldo_caja_total_1 -1 )*100, 3), round((pendiente_gastar_total_2/pendiente_gastar_total_1 -1 )*100, 3), round((honorario_2/honorario_1 -1 )*100, 3), round(honorario_2 - honorario_1), round((honorario_2 - honorario_1) , 2)/round((margen2_2 - margen2_1), 2)*100))
            datos_finales = [(ingresos_total_2 - ingresos_total_1), (pendiente_gastar_total_2 - pendiente_gastar_total_1), (saldo_caja_total_2 - saldo_caja_total_1), ((margen1_2 - honorario_2) - (margen1_1 - honorario_1)), (descuento_total_1 - descuento_total_2)]
    
            for v in var_especifica:
                v[1].append((v[1][0]/((margen1_2 - honorario_2) - (margen1_1 - honorario_1))*100))

    return render(request, 'historicoalmacenero.html', {"datos":datos, "mensaje":mensaje, "var_especifica":var_especifica, "datos_finales":datos_finales })

def estudioindice(request, fecha1, fecha2):

    fecha_1 = datetime.date(year = int(fecha1[0:4]), month=int(fecha1[4:6]), day=int(fecha1[6:8]))
    fecha_2= datetime.date(year = int(fecha2[0:4]), month=int(fecha2[4:6]), day=int(fecha2[6:8]))

    registro_2 = RegistroAlmacenero.objects.filter(fecha= fecha_2)
    registro_1 = RegistroAlmacenero.objects.filter(fecha= fecha_1)

    proyectos_agregados = []
    proyectos_eliminados = []

    ingresos_diferencia = []

    costos_diferencia = []

    for r in registro_2:
        try:
            r_aux = RegistroAlmacenero.objects.filter(fecha= fecha_1, proyecto = r.proyecto)

            if abs((r.ingreso_ventas-r_aux.ingreso_ventas)/r_aux.ingreso_ventas) > 0.03:
                ingresos_diferencia.append((r.proyecto, (r.ingreso_ventas/r_aux.ingreso_ventas - 1)))
        except:
            proyectos_agregados.append(r.proyecto)

    print(ingresos_diferencia)

    return render(request, 'estudio_indice.html')

def indicelinkajustado(request):

    datos = Almacenero.objects.all()
    datos_completos = []
    datos_finales = []
    saldo_caja_total = 0
    pendiente_gastar_total = 0
    ingresos_total = 0
    descuento_total = 0

    for dato in datos:

        presupuesto = "NO"
        pricing = "NO"
        almacenero = dato
        presupuesto = Presupuestos.objects.get(proyecto = dato.proyecto)

        # Aqui calculo el IVA sobre compras

        iva_compras = (presupuesto.imprevisto + presupuesto.saldo_mat + presupuesto.saldo_mo + presupuesto.credito + presupuesto.fdr)*0.07875

        almacenero.pendiente_iva_ventas = iva_compras

        almacenero.save()
        

        # Calculo el resto de las cosas

        retiro_socios = sum(np.array(RetirodeSocios.objects.values_list('monto_pesos').filter(proyecto = dato.proyecto)))
        saldo_caja = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados - retiro_socios + almacenero.tenencia
        saldo_caja_total = saldo_caja_total + saldo_caja
        pend_gast = almacenero.pendiente_admin + almacenero.pendiente_comision + presupuesto.saldo_mat + presupuesto.saldo_mo + presupuesto.imprevisto + presupuesto.credito + presupuesto.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos
        pendiente_gastar_total = pendiente_gastar_total + pend_gast
        prest_cobrar = almacenero.prestamos_proyecto + almacenero.prestamos_otros 
        total_ingresos = prest_cobrar + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + almacenero.financiacion 
        ingresos_total = ingresos_total + total_ingresos
        margen = total_ingresos - pend_gast + saldo_caja
        descuento = almacenero.ingreso_ventas*0.06
        descuento_total = descuento_total + descuento
        margen_2 = margen - descuento 
               
        datos_completos.append((dato, saldo_caja, pend_gast, total_ingresos, margen, descuento, margen_2))

    # -----------------> Aqui calculo los totalizadores

    margen1_total = ingresos_total - pendiente_gastar_total + saldo_caja_total
    margen2_total = margen1_total - descuento_total


    # -----------------> Aqui calculo la parte de honorarios

    honorarios = Honorarios.objects.order_by("-fecha")
    caja_actual = honorarios[0].caja_actual
    subtotal_1 = honorarios[0].cuotas + honorarios[0].ventas
    ingresos = subtotal_1 + honorarios[0].creditos
    comision = honorarios[0].comision_venta*honorarios[0].ventas
    subtotal_2 = honorarios[0].estructura_gio + honorarios[0].aportes + honorarios[0].socios + comision
    costos = subtotal_2  + honorarios[0].deudas
    honorario = ingresos - costos + honorarios[0].caja_actual
    honorarios2 = honorario

    # -----------------> Aqui calculo los totalizadores con los honorarios

    caja_total = caja_actual + saldo_caja_total
    costos_totales = pendiente_gastar_total + costos
    ingresos_totales = ingresos_total + ingresos
    margen1_completo = margen1_total + honorario
    margen2_completo = margen1_completo - descuento_total

    # -----------------> Información para graficos

    retiro_honorarios = 0
    honorarios_beneficio2 = 0
    honorarios_beneficio1 = 0

    datos_finales.append((saldo_caja_total , pendiente_gastar_total, ingresos_total, descuento_total, margen1_total, margen2_total))

    datos_finales_2 = [honorario, ingresos, costos, honorarios2, ingresos_totales, costos_totales, margen1_completo , margen2_completo, retiro_honorarios, honorarios_beneficio2, honorarios_beneficio1, caja_actual, caja_total]


    # -----------------> Esta es la parte del historico

    datos_historicos = RegistroAlmacenero.objects.order_by("fecha")

    fechas = []

    for d in datos_historicos:

        if not d.fecha in fechas:

            fechas.append(d.fecha)

    datos_registro = []

    for fecha in fechas:

        datos = RegistroAlmacenero.objects.filter(fecha = fecha)

        saldo_caja_total = 0
        pendiente_gastar_total = 0
        ingresos_total = 0
        descuento_total = 0
        honorario = 0


        for dato in datos:

            almacenero = dato

            #Calculo el resto de las cosas

            retiro_socios = almacenero.retiro_socios
            saldo_caja = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados - retiro_socios + almacenero.tenencia
            
            pend_gast = almacenero.pendiente_admin + almacenero.pendiente_comision + almacenero.saldo_mat + almacenero.saldo_mo + almacenero.imprevisto + almacenero.credito + almacenero.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos
            
            prest_cobrar = almacenero.prestamos_proyecto + almacenero.prestamos_otros
            total_ingresos = prest_cobrar + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + almacenero.financiacion
            
            margen = total_ingresos - pend_gast + saldo_caja
            descuento = almacenero.ingreso_ventas*0.06
            
            margen_2 = margen - descuento
            honorario = almacenero.honorarios

            pendiente_gastar_total = pendiente_gastar_total + pend_gast
            saldo_caja_total = saldo_caja_total + saldo_caja
            descuento_total = descuento_total + descuento
            ingresos_total = ingresos_total + total_ingresos

            # Me falta calcular la parte de honorarios


        margen1 = ingresos_total - pendiente_gastar_total + honorario + saldo_caja_total
        margen2 = margen1 - descuento_total

        datos_registro.append((margen1, margen2))

    return render(request, 'indicelink.html', {"datos_completos":datos_completos, 'datos_finales':datos_finales, "datos_registro":datos_registro, "fechas":fechas, "datos_finales_2":datos_finales_2})

def consolidado(request):

    datos = Almacenero.objects.all()

    datos_completos = []
    datos_finales = []

    costo_total = 0
    ingresos_total = 0
    descuento_total = 0
    retiro_totales = 0
    tenencia_totales = 0
    financiacion_totales = 0


    for dato in datos:

        presupuesto = "NO"

        pricing = "NO"

        almacenero = dato

        presupuesto = Presupuestos.objects.get(proyecto = dato.proyecto)
        dato_presupuesto = presupuesto

        # Calculo el resto de las cosas

        pend_gast = almacenero.pendiente_admin + almacenero.pendiente_comision + presupuesto.saldo_mat + presupuesto.saldo_mo + presupuesto.imprevisto + presupuesto.credito + presupuesto.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem
        prest_cobrar = almacenero.prestamos_proyecto + almacenero.prestamos_otros
        retiro_socios = sum(np.array(RetirodeSocios.objects.values_list('monto_pesos').filter(proyecto = dato.proyecto)))  
        total_costo = almacenero.cheques_emitidos + almacenero.gastos_fecha + pend_gast + almacenero.Prestamos_dados      
        
        
        costo_total = costo_total + total_costo

        descuento = almacenero.ingreso_ventas*0.06
        descuento_total = descuento_total + descuento
        
        total_ingresos = prest_cobrar + almacenero.cuotas_cobradas + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas
        
        ingresos_total = ingresos_total + total_ingresos

        saldo_caja = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados
        saldo_proyecto = total_ingresos - total_costo - descuento
        tenencia_totales = tenencia_totales + almacenero.tenencia
        financiacion_totales = financiacion_totales + almacenero.financiacion
        rentabilidad = (saldo_proyecto/total_costo)*100


        total_ingresos_pesimista = saldo_proyecto  - retiro_socios
        saldo_proyecto_pesimista = total_ingresos_pesimista
        rentabilidad_pesimista = (saldo_proyecto_pesimista/total_costo)*100
        retiro_totales = retiro_totales + retiro_socios 

        try:

            modelo = Modelopresupuesto.objects.filter(proyecto = dato.proyecto)

            presupuesto = len(modelo)

        except:

            pass

        try:

            pricing = Pricing.objects.filter(unidad__proyecto = dato.proyecto)

            pricing = len(pricing)

        except:

            pass

        # -----------------> Aqui empieza para el precio promedio contado

        if len(Unidades.objects.filter(proyecto = dato.proyecto, estado = "DISPONIBLE")) > 0:

            datos_unidades = Unidades.objects.filter(proyecto = dato.proyecto, estado = "DISPONIBLE")

            m2_totales = 0

            sumatoria_contado = 0

            for d in datos_unidades:

                if d.sup_equiv > 0:

                    m2 = d.sup_equiv

                else:

                    m2 = d.sup_propia + d.sup_balcon + d.sup_comun + d.sup_patio
            
                try:

                    param_uni = Pricing.objects.get(unidad = d)
                    
                    desde = d.proyecto.desde

                    if d.tipo == "COCHERA":
                        desde = d.proyecto.desde*d.proyecto.descuento_cochera

                    if param_uni.frente == "SI":
                        desde = desde*d.proyecto.recargo_frente

                    if param_uni.piso_intermedio == "SI":
                        desde =desde*d.proyecto.recargo_piso_intermedio

                    if param_uni.cocina_separada == "SI":
                        desde = desde*d.proyecto.recargo_cocina_separada

                    if param_uni.local == "SI":
                        desde = desde*d.proyecto.recargo_local

                    if param_uni.menor_45_m2 == "SI":
                        desde = desde*d.proyecto.recargo_menor_45

                    if param_uni.menor_50_m2 == "SI":
                        desde = desde*d.proyecto.recargo_menor_50

                    if param_uni.otros == "SI":
                        desde = desde*d.proyecto.recargo_otros 

                    #Aqui calculamos el contado/financiado
                    
                    contado = desde*m2 

                    sumatoria_contado = sumatoria_contado + contado
                    m2_totales = m2_totales + m2

                except:

                    basura = 1


            if m2_totales == 0:

                precio_promedio_contado = 0

            else:

                if "2UO" in dato.proyecto.nombre:

                    precio_promedio_contado = 98029

                else:

                    precio_promedio_contado = sumatoria_contado/m2_totales

        else:

            if "#300" in dato.proyecto.nombre:

                precio_promedio_contado = 133979

            else:

                precio_promedio_contado = 0

    # -----------------> Aqui termina para el precio promedio contado

        datos_completos.append((dato, total_costo, total_ingresos, saldo_proyecto, rentabilidad, presupuesto, pricing, saldo_proyecto_pesimista, rentabilidad_pesimista, precio_promedio_contado, retiro_socios, descuento, dato_presupuesto))



    # -----------------> Aqui calculo los totalizadores

    beneficio_total = ingresos_total - costo_total - descuento_total
    beneficio_total_pesimista = beneficio_total - retiro_totales
    beneficio_retiros = beneficio_total - descuento_total
    rendimiento_total = beneficio_total/costo_total*100
    rendimiento_total_pesimista = beneficio_total_pesimista/costo_total*100

    datos_finales.append((ingresos_total, costo_total, beneficio_total, rendimiento_total, descuento_total, rendimiento_total_pesimista, beneficio_total_pesimista, retiro_totales, beneficio_retiros, tenencia_totales, financiacion_totales))

    # -----------------> Información para graficos

    beneficio_total = ingresos_total - costo_total - descuento_total - retiro_totales + tenencia_totales + financiacion_totales
    beneficio_descuento = ingresos_total - costo_total - retiro_totales + tenencia_totales + financiacion_totales
    retiros_completo = ingresos_total - costo_total + tenencia_totales + financiacion_totales

    datos_finales_2 = [ingresos_total, costo_total, beneficio_total, beneficio_descuento, retiros_completo]

    # -----------------> Esta es la parte del historico

    datos_historicos = RegistroAlmacenero.objects.order_by("fecha")

    fechas = []

    for d in datos_historicos:

        if not d.fecha in fechas:

            fechas.append(d.fecha)

    datos_registro = []

    for fecha in fechas:

        datos = RegistroAlmacenero.objects.filter(fecha = fecha)

        datos_finales_registro = []

        costo_total = 0
        ingresos_total = 0
        descuento_total = 0
        retiro_totales = 0
        tenencia_totales = 0
        financiacion_totales = 0

        for dato in datos:

            almacenero = dato

            #Calculo el resto de las cosas

            pend_gast = almacenero.pendiente_admin + almacenero.pendiente_comision + almacenero.saldo_mat + almacenero.saldo_mo + almacenero.imprevisto + almacenero.credito + almacenero.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem
            prest_cobrar = almacenero.prestamos_proyecto + almacenero.prestamos_otros
            total_costo = almacenero.cheques_emitidos + almacenero.gastos_fecha + pend_gast + almacenero.Prestamos_dados                
            
            costo_total = costo_total + total_costo

            descuento = almacenero.ingreso_ventas*0.06
            descuento_total = descuento_total + descuento
            
            total_ingresos = prest_cobrar + almacenero.cuotas_cobradas + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas
            
            ingresos_total = ingresos_total + total_ingresos
            retiro_totales = retiro_totales + dato.retiro_socios

            saldo_caja = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados
            saldo_proyecto = total_ingresos - total_costo
            tenencia_totales = tenencia_totales + almacenero.tenencia
            financiacion_totales = financiacion_totales + almacenero.financiacion
            

            total_ingresos_pesimista = total_ingresos - descuento
            saldo_proyecto_pesimista = total_ingresos_pesimista - total_costo
            


        beneficio_total = ingresos_total - costo_total - descuento_total - retiro_totales + tenencia_totales + financiacion_totales
        beneficio_descuento = ingresos_total - costo_total - retiro_totales + tenencia_totales + financiacion_totales
        retiros_completo = ingresos_total - costo_total + tenencia_totales + financiacion_totales


        datos_finales_registro.append((ingresos_total, costo_total, retiros_completo, beneficio_total, beneficio_descuento))

        datos_registro.append(datos_finales_registro)


    return render(request, 'consolidado.html', {"datos_completos":datos_completos, 'datos_finales':datos_finales, "datos_registro":datos_registro, "fechas":fechas, "datos_finales_2":datos_finales_2})

def indicelinkmoneda(request, id_moneda):

    moneda = Constantes.objects.get(id = id_moneda)

    datos = Almacenero.objects.all()

    datos_completos = []
    datos_finales = []

    saldo_caja_total = 0
    pendiente_gastar_total = 0
    ingresos_total = 0
    descuento_total = 0

    for dato in datos:

        presupuesto = "NO"

        pricing = "NO"

        almacenero = dato

        presupuesto = Presupuestos.objects.get(proyecto = dato.proyecto)

        # Aqui calculo el IVA sobre compras

        iva_compras = (presupuesto.imprevisto + presupuesto.saldo_mat + presupuesto.saldo_mo + presupuesto.credito + presupuesto.fdr)*0.07875

        almacenero.pendiente_iva_ventas = iva_compras

        almacenero.save()

        # Calculo el resto de las cosas

        retiro_socios = sum(np.array(RetirodeSocios.objects.values_list('monto_pesos').filter(proyecto = dato.proyecto)))
        saldo_caja = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados - retiro_socios + almacenero.tenencia
        saldo_caja_total = saldo_caja_total + saldo_caja
        pend_gast = almacenero.pendiente_admin + almacenero.pendiente_comision + presupuesto.saldo_mat + presupuesto.saldo_mo + presupuesto.imprevisto + presupuesto.credito + presupuesto.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos
        pendiente_gastar_total = pendiente_gastar_total + pend_gast
        prest_cobrar = almacenero.prestamos_proyecto + almacenero.prestamos_otros
        total_ingresos = prest_cobrar + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + almacenero.financiacion
        ingresos_total = ingresos_total + total_ingresos
        margen = total_ingresos - pend_gast + saldo_caja
        descuento = almacenero.ingreso_ventas*0.06
        descuento_total = descuento_total + descuento
        margen_2 = margen - descuento 
               
        datos_completos.append((dato, saldo_caja/moneda.valor, pend_gast/moneda.valor, total_ingresos/moneda.valor, margen/moneda.valor, descuento/moneda.valor, margen_2/moneda.valor))

    # -----------------> Aqui calculo los totalizadores

    margen1_total = ingresos_total - pendiente_gastar_total + saldo_caja_total
    margen2_total = margen1_total - descuento_total


    # -----------------> Aqui calculo la parte de honorarios

    honorarios = Honorarios.objects.order_by("-fecha")
    caja_actual = honorarios[0].caja_actual
    subtotal_1 = honorarios[0].cuotas + honorarios[0].ventas
    ingresos = subtotal_1 + honorarios[0].creditos
    comision = honorarios[0].comision_venta*honorarios[0].ventas
    subtotal_2 = honorarios[0].estructura_gio + honorarios[0].aportes + honorarios[0].socios + comision
    costos = subtotal_2  + honorarios[0].deudas
    honorario = ingresos - costos + honorarios[0].caja_actual
    honorarios2 = honorario

    # -----------------> Aqui calculo los totalizadores con los honorarios

    caja_total = caja_actual + saldo_caja_total
    costos_totales = pendiente_gastar_total + costos
    ingresos_totales = ingresos_total + ingresos
    margen1_completo = margen1_total + honorario
    margen2_completo = margen1_completo - descuento_total

    # -----------------> Información para graficos

    retiro_honorarios = 0
    honorarios_beneficio2 = 0
    honorarios_beneficio1 = 0

    datos_finales.append((saldo_caja_total , pendiente_gastar_total, ingresos_total, descuento_total, margen1_total, margen2_total))

    datos_finales_2 = [honorario, ingresos, costos, honorarios2, ingresos_totales, costos_totales, margen1_completo , margen2_completo, retiro_honorarios, honorarios_beneficio2, honorarios_beneficio1, caja_actual, caja_total]

    # -----------------> Bucle para moneda

    datos_finales = np.array(datos_finales)/moneda.valor
    datos_finales_2 = np.array(datos_finales_2)/moneda.valor

    # -----------------> Esta es la parte del historico

    datos_historicos = RegistroAlmacenero.objects.order_by("fecha")

    fechas = []

    for d in datos_historicos:

        if not d.fecha in fechas:

            fechas.append(d.fecha)

    datos_registro = []

    for fecha in fechas:

        datos = RegistroAlmacenero.objects.filter(fecha = fecha)

        saldo_caja_total = 0
        pendiente_gastar_total = 0
        ingresos_total = 0
        descuento_total = 0
        honorario = 0


        for dato in datos:

            almacenero = dato

            #Calculo el resto de las cosas

            retiro_socios = almacenero.retiro_socios
            saldo_caja = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados - retiro_socios + almacenero.tenencia
            
            pend_gast = almacenero.pendiente_admin + almacenero.pendiente_comision + almacenero.saldo_mat + almacenero.saldo_mo + almacenero.imprevisto + almacenero.credito + almacenero.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos
            
            prest_cobrar = almacenero.prestamos_proyecto + almacenero.prestamos_otros
            total_ingresos = prest_cobrar + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + almacenero.financiacion
            
            margen = total_ingresos - pend_gast + saldo_caja
            descuento = almacenero.ingreso_ventas*0.06
            
            margen_2 = margen - descuento
            honorario = almacenero.honorarios

            pendiente_gastar_total = pendiente_gastar_total + pend_gast
            saldo_caja_total = saldo_caja_total + saldo_caja
            descuento_total = descuento_total + descuento
            ingresos_total = ingresos_total + total_ingresos

            # Me falta calcular la parte de honorarios


        margen1 = ingresos_total - pendiente_gastar_total + honorario + saldo_caja_total
        margen2 = margen1 - descuento_total

        # Valor para dividir de la moneda

        #try:

        fecha_valor = datetime.date(fecha.year, fecha.month, 1)
        
        valor = Registrodeconstantes.objects.get(fecha=fecha_valor, constante__id = id_moneda)

        datos_registro.append((margen1/valor.valor, margen2/valor.valor))

        #except:

            #datos_registro.append((0, 0))

    return render(request, 'indicelinkmoneda.html', {"datos_completos":datos_completos, 'datos_finales':datos_finales, "datos_registro":datos_registro, "fechas":fechas, "datos_finales_2":datos_finales_2, "moneda":moneda})

def almacenero(request):

    datos = Almacenero.objects.all()

    proyectos = []

    for dato in datos:
        proyectos.append((dato.proyecto.id, dato.proyecto.nombre))

    proyectos = list(set(proyectos))

    usd_blue = Constantes.objects.get(nombre = "USD_BLUE")

    datos = 0

    mensaje = 0

    lista = 0

    if request.method == 'POST':

        try:

            #Trae el proyecto elegido

            proyecto_elegido = request.POST.items()

            #Crea los datos del pricing

            datos = []

            for i in proyecto_elegido:

                if i[0] == "proyecto" and i[1] != "":

                    proyecto = Proyectos.objects.get(id = i[1])
                    presupuesto = Presupuestos.objects.get(proyecto = proyecto)
                    almacenero = Almacenero.objects.get(proyecto = proyecto)

                    #Aqui calculo el IVA sobre compras

                    iva_compras = (presupuesto.imprevisto + presupuesto.saldo_mat + presupuesto.saldo_mo + presupuesto.credito + presupuesto.fdr)*0.07875

                    almacenero.pendiente_iva_ventas = iva_compras

                    almacenero.save()

                    #Calculo el resto de las cosas
                    
                    pend_gast = almacenero.pendiente_admin + almacenero.pendiente_comision + presupuesto.saldo_mat + presupuesto.saldo_mo + presupuesto.imprevisto + presupuesto.credito + presupuesto.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem
                    prest_cobrar = almacenero.prestamos_proyecto + almacenero.prestamos_otros
                    retiro_socios = sum(np.array(RetirodeSocios.objects.values_list('monto_pesos').filter(proyecto = proyecto)))
                    total_costo = almacenero.cheques_emitidos + almacenero.gastos_fecha + pend_gast + almacenero.Prestamos_dados
                    
                    descuento = almacenero.ingreso_ventas*0.06 
                    
                    total_ingresos = prest_cobrar + almacenero.cuotas_cobradas + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + almacenero.tenencia + almacenero.financiacion + almacenero.inmuebles
                    saldo_caja = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados + almacenero.tenencia
                    saldo_proyecto = total_ingresos - total_costo
                    rentabilidad = (saldo_proyecto/total_costo)*100

                    total_ingresos_pesimista = total_ingresos - descuento
                    saldo_proyecto_pesimista = total_ingresos_pesimista - total_costo  - retiro_socios - almacenero.unidades_socios
                    rentabilidad_pesimista = (saldo_proyecto_pesimista/total_costo)*100

                    #Cargo todo a datos

                    datos.append(proyecto)
                    datos.append(presupuesto)
                    datos.append(almacenero)

                    datos.append((pend_gast, prest_cobrar, total_costo, total_ingresos, rentabilidad, saldo_caja, saldo_proyecto, descuento, saldo_proyecto_pesimista, rentabilidad_pesimista, retiro_socios))

                    lista = RegistroAlmacenero.objects.filter(proyecto = proyecto)

            proyectos = 0

        except:

            mensaje = 1

    if request.method == 'GET':

        datos = request.GET.items()

        for i in datos:

            if i[0] != "csrfmiddlewaretoken":

                print(i)

                return redirect('Historico almacenero', id_proyecto = int(i[0]), fecha = int(i[1]))


    datos = {"proyectos":proyectos,
    'datos':datos,
    "usd":usd_blue,
    "mensaje":mensaje,
    'lista':lista}

    return render(request, 'almacenero.html', {"datos":datos} )

def movimientoadmin(request):

    if request.method == 'POST':

        req = request.POST.items()

        for r in req:

            if r[0] == "APROBADA" and r[1] != "":

                b = MovimientoAdmin.objects.get(id = r[1])

                b.estado = "APROBADA"

                b.save()

            if r[0] == "RECHAZADA" and r[1] != "":

                b = MovimientoAdmin.objects.get(id = r[1])

                b.estado = "RECHAZADA"

                b.save()

    datos = MovimientoAdmin.objects.order_by("-fecha")

    return render(request, 'movimientoadmin.html', {'datos':datos})

def subirmovimiento(request):

    if request.method == 'POST':

        b = MovimientoAdmin(
            fecha = request.POST['fecha'],
            archivo = request.FILES['archivo'],
            comentario = request.POST['comentario'],
            )

        b.save()

        return redirect ('Movimiento administración')

    return render(request, 'crearmovimiento.html')

def borrarmovimiento(request, id_mov):

    datos = MovimientoAdmin.objects.get(id = id_mov)

    if request.method == 'POST':

        datos.delete()

        return redirect ('Movimiento administración')


    return render(request, 'borrarmovimiento.html', {'datos':datos})

def retirodesocios(request):

    if request.method == 'POST':

        b = RetirodeSocios(
            proyecto = Proyectos.objects.get(nombre = request.POST['proyecto']),
            fecha = request.POST['fecha'],
            monto_pesos = request.POST['pesos'],
            comentario = request.POST['comentario']
        )

        b.save()

    datos_retiro = RetirodeSocios.objects.order_by("-fecha")

    proyectos = Proyectos.objects.all()

    total_pesos = 0
    total_h = 0

    datos = []

    for d in datos_retiro:

        fecha = datetime.date(d.fecha.year, d.fecha.month, 1)
        registro = Registrodeconstantes.objects.get(fecha = fecha, constante__nombre='Hº VIVIENDA')
        monto_h = d.monto_pesos/registro.valor

        total_pesos = total_pesos + d.monto_pesos
        total_h = total_h + monto_h

        datos.append((d, monto_h))

    return render(request, 'retirodesocios.html', {'datos':datos, 'total_pesos':total_pesos, 'total_h':total_h, 'proyectos':proyectos})

def arqueo_diario(request, id_arqueo):

    data_cruda = Arqueo.objects.get(id = id_arqueo)

    data = data_cruda

    data_frame = pd.read_excel(data.arqueo)

    lista_proyecto = list(data_frame['PROYECTO'])

    datos = []

    array_pesos = np.array(data_frame['EFECTIVO'])

    pesos = sum(array_pesos)

    array_cheques = np.array(data_frame['CHEQUES'])

    cheques = sum(array_cheques)

    array_usd = np.array(data_frame['USD'])

    usd = sum(array_usd)

    array_euro = np.array(data_frame['EUROS'])

    euro = sum(array_euro)

    cambio_usd = data_frame['CAMBIO USD'][0]
    cambio_euro = data_frame['CAMBIO EURO'][0]

    pesos_usd = usd*Constantes.objects.get(nombre = "USD_BLUE").valor
    pesos_euros = euro*Constantes.objects.get(nombre = "EURO_BLUE").valor

    pesos_pesos = pesos - pesos_usd - pesos_euros

    porcentaje_usd = pesos_usd/pesos*100
    porcentaje_euros = pesos_euros/pesos*100
    porcentaje_pesos = 100 - porcentaje_euros - porcentaje_usd

    bancos = 0

    consolidados = 0

    consolidado_actual = 0

    moneda_extranjera_actual = 0
  
    datos_grafico = [porcentaje_usd, porcentaje_euros, porcentaje_pesos]

    numero = 0

    for i in lista_proyecto:

        try:

            proyecto = Proyectos.objects.get(id = data_frame.loc[numero, 'ID LINK-P'])

        except:

            proyecto = 0

        nombre_columnas = data_frame.columns

        banco = 0

        list_bank_proj_info = []

        for n in nombre_columnas:

            if "BANCO" in n:
                banco = banco + data_frame.loc[numero, n]
                bancos = bancos + data_frame.loc[numero, n]
                if data_frame.loc[numero, n] != 0:
                    list_bank_proj_info.append((n, data_frame.loc[numero, n]))



        consolidado = data_frame.loc[numero, 'EFECTIVO'] + data_frame.loc[numero, 'CHEQUES'] + data_frame.loc[numero, 'MONEDA EXTRANJERA'] + banco

        consolidados = consolidados + consolidado

        moneda_extranjera_actual = moneda_extranjera_actual + data_frame.loc[numero, 'USD']*cambio_usd + data_frame.loc[numero, 'EUROS']*cambio_euro

        consolidado_actual = consolidado_actual + consolidado - data_frame.loc[numero, 'MONEDA EXTRANJERA'] + data_frame.loc[numero, 'USD']*cambio_usd + data_frame.loc[numero, 'EUROS']*cambio_euro

        
        try:
            # ----> Armemos lo de los cheques

            nombre_proyecto = data_frame.loc[numero, 'PROYECTO']

            cheque_numero = list(data_frame[data_frame["CHEQUE PROYECTO"] == nombre_proyecto]["CHEQUE NUMERO"])
            cheque_cliente = list(data_frame[data_frame["CHEQUE PROYECTO"] == nombre_proyecto]["CHEQUE CLIENTE"])
            cheque_emitido = list(data_frame[data_frame["CHEQUE PROYECTO"] == nombre_proyecto]["CHEQUE EMITIDO"])
            cheque_vencimiento = list(data_frame[data_frame["CHEQUE PROYECTO"] == nombre_proyecto]["CHEQUE VENCIMIENTO"])
            cheque_monto = list(data_frame[data_frame["CHEQUE PROYECTO"] == nombre_proyecto]["CHEQUE MONTO"])
            cheque_tipo = list(data_frame[data_frame["CHEQUE PROYECTO"] == nombre_proyecto]["CHEQUE TIPO"])

            info_cheque = []

            cont_info = 0

            for i in cheque_numero:
                info_cheque.append(("Nº "+str(cheque_numero[cont_info]), cheque_cliente[cont_info], cheque_emitido[cont_info], cheque_vencimiento[cont_info], cheque_monto[cont_info], cheque_tipo[cont_info]))
                cont_info += 1
        except:
            info_cheque = []

        datos.append((proyecto, data_frame.loc[numero, 'PROYECTO'], data_frame.loc[numero, 'EFECTIVO'], data_frame.loc[numero, 'USD'], data_frame.loc[numero, 'EUROS'], data_frame.loc[numero, 'CHEQUES'], data_frame.loc[numero, 'MONEDA EXTRANJERA'], banco, consolidado, list_bank_proj_info, info_cheque))

        numero += 1

    otros_datos = [usd, euro, pesos, cheques, bancos, consolidados, consolidado_actual, moneda_extranjera_actual]


    grafico = []

    n = data_cruda

    frame = pd.read_excel(n.arqueo)

    array_extranjera = np.array(frame['MONEDA EXTRANJERA'])

    extranjera = sum(array_extranjera)

    array_efectivo = np.array(frame['EFECTIVO'])

    efectivo = sum(array_efectivo)

    #Aqui sumamos los bancos

    banco = 0

    for m in nombre_columnas:

        if "BANCO" in m:

            array_banco = np.array(frame[m])

            banco = banco +  sum(array_banco)

    array_cheque = np.array(frame['CHEQUES'])

    cheque = sum(array_cheque)

    grafico.append((n.fecha, extranjera, efectivo, banco, cheque))

    grafico = sorted(grafico, key=lambda tup: tup[0])

    return render(request, 'arqueo.html', {'datos':datos, 'data_cruda':data_cruda, 'otros_datos':otros_datos, 'grafico':grafico, 'cambio_usd':cambio_usd, 'cambio_euro':cambio_euro})

def arqueos(request):

    if request.method == 'POST':
        datos_p = request.POST.items()
        for d in datos_p:
            if d[0] == "fecha":
                b = Arqueo(
                    fecha = request.POST['fecha'],
                    arqueo = request.FILES['adjunto']
                )

                b.save()

                #try:

                destino = [str(datosusuario.objects.get(identificacion = "GB").email),
                str(datosusuario.objects.get(identificacion = "AP").email),
                str(datosusuario.objects.get(identificacion = "ALM").email),
                str(datosusuario.objects.get(identificacion = "PL").email),
                str(datosusuario.objects.get(identificacion = "SP").email),
                str(datosusuario.objects.get(identificacion = "AR").email),
                str(datosusuario.objects.get(identificacion = "CS").email),
                str(datosusuario.objects.get(identificacion = "CA").email)]

                for d in destino:

                    # Establecemos conexion con el servidor smtp de gmail
                    mailServer = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
                    mailServer.ehlo()
                    mailServer.starttls()
                    mailServer.ehlo()
                    mailServer.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

                    # Construimos el mensaje simple
                    
                    mensaje = MIMEText("""
                    
Buenas!,

{} acaba de cargar un arqueo

Fecha del mismo: {}

Entra a Link-P para chequearlo www.linkp.online/finanzas/arqueos/

Gracias!

Saludos!
                    """.format(request.user.username, request.POST['fecha']))
                    mensaje['From']=settings.EMAIL_HOST_USER
                    mensaje['To']= d
                    mensaje['Subject']="Se cargo un arqueo con fecha {}".format(request.POST['fecha'])


                    # Envio del mensaje

                    mailServer.sendmail(settings.EMAIL_HOST_USER,
                                    d,
                                    mensaje.as_string())


            if d[0] == "delete":
                arqueo = Arqueo.objects.get(id = int(request.POST['delete']))
                arqueo.delete()

        

    data = Arqueo.objects.all().order_by('-fecha')

    return render(request, 'arqueos.html', {'data':data})

def cuentacte_resumen(request):

    data_project = []
    data_project_b = []

    list_p = Cuota.objects.values_list("cuenta_corriente__venta__proyecto__id", flat = True)
    list_p = list(set(list_p))
    cuotas_anteriores_t = 0
    pagos_anteriores_t = 0
    deuda_t = 0
    cuotas_posterior_t = 0
    pagos_posterior_t = 0
    pagos_pesos_t = 0
    adelantos_t = 0

    for project in list_p:

        proyecto = Proyectos.objects.get(id = int(project))

        today = datetime.date.today()

        cuotas_anteriores = sum(np.array(Cuota.objects.values_list('precio', flat=True).filter(fecha__lt = today, cuenta_corriente__venta__proyecto__id = project))*np.array(Cuota.objects.values_list('constante__valor', flat=True).filter(fecha__lt = today, cuenta_corriente__venta__proyecto__id = project)))
        pagos_anteriores = sum(np.array(Pago.objects.values_list('pago', flat=True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente__venta__proyecto__id = project))*np.array(Pago.objects.values_list('cuota__constante__valor', flat=True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente__venta__proyecto__id = project)))
        deuda = cuotas_anteriores - pagos_anteriores

        cuotas_posterior = sum(np.array(Cuota.objects.values_list('precio', flat=True).filter(fecha__gte = today, cuenta_corriente__venta__proyecto__id = project))*np.array(Cuota.objects.values_list('constante__valor', flat=True).filter(fecha__gte = today, cuenta_corriente__venta__proyecto__id = project)))
        pagos_posterior= sum(np.array(Pago.objects.values_list('pago', flat=True).filter(cuota__fecha__gte = today, cuota__cuenta_corriente__venta__proyecto__id = project))*np.array(Pago.objects.values_list('cuota__constante__valor', flat=True).filter(cuota__fecha__gte = today, cuota__cuenta_corriente__venta__proyecto__id = project)))
        pagos_pesos= sum(np.array(Pago.objects.values_list('pago_pesos', flat=True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente__venta__proyecto__id = project)))
        adelantos = cuotas_posterior - pagos_posterior

        array_pesos = np.array([cuotas_anteriores, pagos_anteriores, deuda, cuotas_posterior, pagos_posterior, adelantos])
        array_h = array_pesos/Constantes.objects.get(id = 7).valor

        cuotas_anteriores_t += cuotas_anteriores
        pagos_anteriores_t += pagos_anteriores
        deuda_t += deuda
        cuotas_posterior_t += cuotas_posterior
        pagos_posterior_t += pagos_posterior
        pagos_pesos_t += pagos_pesos
        adelantos_t += adelantos

        data_project.append((proyecto, array_pesos, array_h, pagos_pesos))

    array_total = [cuotas_anteriores_t, pagos_pesos_t, deuda_t, cuotas_posterior_t, pagos_posterior_t, adelantos_t]

    for project in list_p:

        proyecto = Proyectos.objects.get(id = int(project))

        today = datetime.date.today()

        cuotas_anteriores = sum(np.array(Cuota.objects.values_list('precio', flat=True).filter(fecha__lt = today, cuenta_corriente__venta__proyecto__id = project).exclude(porc_boleto = None))*np.array(Cuota.objects.values_list('constante__valor', flat=True).filter(fecha__lt = today, cuenta_corriente__venta__proyecto__id = project).exclude(porc_boleto = None))*np.array(Cuota.objects.values_list('porc_boleto', flat=True).filter(fecha__lt = today, cuenta_corriente__venta__proyecto__id = project).exclude(porc_boleto = None)))
        pagos_anteriores = sum(np.array(Pago.objects.values_list('pago', flat=True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente__venta__proyecto__id = project).exclude(cuota__porc_boleto = None))*np.array(Pago.objects.values_list('cuota__constante__valor', flat=True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente__venta__proyecto__id = project).exclude(cuota__porc_boleto = None))*np.array(Pago.objects.values_list('cuota__porc_boleto', flat=True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente__venta__proyecto__id = project).exclude(cuota__porc_boleto = None)))
        deuda = cuotas_anteriores - pagos_anteriores

        cuotas_posterior = sum(np.array(Cuota.objects.values_list('precio', flat=True).filter(fecha__gte = today, cuenta_corriente__venta__proyecto__id = project).exclude(porc_boleto = None))*np.array(Cuota.objects.values_list('constante__valor', flat=True).filter(fecha__gte = today, cuenta_corriente__venta__proyecto__id = project).exclude(porc_boleto = None))*np.array(Cuota.objects.values_list('porc_boleto', flat=True).filter(fecha__gte = today, cuenta_corriente__venta__proyecto__id = project).exclude(porc_boleto = None)))
        pagos_posterior= sum(np.array(Pago.objects.values_list('pago', flat=True).filter(cuota__fecha__gte = today, cuota__cuenta_corriente__venta__proyecto__id = project).exclude(cuota__porc_boleto = None))*np.array(Pago.objects.values_list('cuota__constante__valor', flat=True).filter(cuota__fecha__gte = today, cuota__cuenta_corriente__venta__proyecto__id = project).exclude(cuota__porc_boleto = None))*np.array(Pago.objects.values_list('cuota__porc_boleto', flat=True).filter(cuota__fecha__gte = today, cuota__cuenta_corriente__venta__proyecto__id = project).exclude(cuota__porc_boleto = None)))
        pagos_pesos= sum(np.array(Pago.objects.values_list('pago_pesos', flat=True).filter(cuota__fecha__gte = today, cuota__cuenta_corriente__venta__proyecto__id = project)))
        adelantos = cuotas_posterior - pagos_posterior

        array_pesos = np.array([cuotas_anteriores, pagos_anteriores, deuda, cuotas_posterior, pagos_posterior, adelantos])
        array_h = array_pesos/Constantes.objects.get(id = 7).valor

        data_project_b.append((proyecto, array_pesos, array_h, pagos_pesos))



    data = CuentaCorriente.objects.all()

    fecha_inicial_hoy = datetime.date.today()

    datos = []

    for c in data:

        pagos = sum(np.array(Pago.objects.values_list('pago_pesos').filter(fecha__lt = fecha_inicial_hoy, cuota__cuenta_corriente = c)))
        
        cuotas_anteriores_h = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__lt = fecha_inicial_hoy, constante__id = 7, cuenta_corriente = c)))*Constantes.objects.get(id = 7).valor
        cuotas_anteriores_usd = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__lt = fecha_inicial_hoy, constante__id = 1, cuenta_corriente = c)))*Constantes.objects.get(id = 1).valor
        pagos_anteriores_h = sum(np.array(Pago.objects.values_list('pago').filter(fecha__lt = fecha_inicial_hoy, cuota__constante__id = 7, cuota__cuenta_corriente = c)))*Constantes.objects.get(id = 7).valor
        pagos_anteriores_usd = sum(np.array(Pago.objects.values_list('pago').filter(fecha__lt = fecha_inicial_hoy, cuota__constante__id = 1, cuota__cuenta_corriente = c)))*Constantes.objects.get(id = 1).valor
        cuotas_posteriores_h = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__gt = fecha_inicial_hoy, constante__id = 7, cuenta_corriente = c)))*Constantes.objects.get(id = 7).valor
        cuotas_posteriores_usd = sum(np.array(Cuota.objects.values_list('precio').filter(fecha__gt = fecha_inicial_hoy, constante__id = 1, cuenta_corriente = c)))*Constantes.objects.get(id = 1).valor
        
        
        adeudado = cuotas_anteriores_h + cuotas_anteriores_usd - pagos_anteriores_h - pagos_anteriores_usd
        pendiente = cuotas_posteriores_h + cuotas_posteriores_usd
        
        datos.append((c, pagos, adeudado, pendiente))

    return render(request, 'ctacte_resumen.html', {"datos":datos, "data_project":data_project, "data_project_b":data_project_b, "array_total":array_total})

def calculadora (request):

    list_cuotas = 0

    if request.method == 'POST':

        cuota =  Cuota.objects.get(id = int(request.POST['cuotaorigen']))
        pago_cuota = sum(np.array(Pago.objects.filter(cuota = cuota).values_list("pago")))
        saldo = cuota.precio - pago_cuota

        cuota_destino = Cuota.objects.get(id = int(request.POST['cuotadestino']))
        cuota_destino.precio = cuota_destino.precio + saldo*(1+float(request.POST['interes']))
        cuota_destino.save()
        cuota.precio = cuota.precio - saldo
        cuota.save()

        return redirect('Calculadora')

    hoy = datetime.date.today()

    cuotas = Cuota.objects.filter(fecha__lt = hoy).exclude(pagada = "SI")

    datos = []

    for cuota in cuotas:

        pago_cuota = sum(np.array(Pago.objects.filter(cuota = cuota).values_list("pago", flat = True)))
        pago_pesos = sum(np.array(Pago.objects.filter(cuota = cuota).values_list("pago_pesos", flat = True)))
        saldo_cuota = cuota.precio - pago_cuota
        saldo_pesos = saldo_cuota*cuota.constante.valor

        if abs(saldo_pesos) < 100:
            cuota.precio = pago_cuota
            cuota.pagada = "SI"
            cuota.save()
            saldo_cuota = 0
            saldo_pesos = 0

        if saldo_cuota != 0:

            list_cuotas = Cuota.objects.filter(fecha__gt = cuota.fecha, cuenta_corriente = cuota.cuenta_corriente).exclude(pagada = "SI").order_by("fecha")[0:4]

            datos.append((cuota, saldo_cuota, list_cuotas))
  

    return render(request, 'calculadora.html', {'datos':datos})

class DescargarCuentacorriente(TemplateView):

    def get(self, request, id_cuenta, *args, **kwargs):
        wb = Workbook()

        cuenta = CuentaCorriente.objects.get(id = id_cuenta)
        cuota = Cuota.objects.filter(cuenta_corriente = cuenta).order_by("fecha")
        pagos = Pago.objects.filter(cuota__cuenta_corriente = cuenta)

        
        ws = wb.active
        ws.title = "Cuotas"
        ws["A1"] = "Resumen de cuenta corriente - Área Administración"
        ws["A1"].font = Font(bold = True)
        ws["A3"] = "Comprador:"
        ws["B3"] = cuenta.venta.comprador
        ws["A4"] = "Asignación:"
        ws["B4"] = cuenta.venta.unidad.asig
        ws["E3"] = "Precio venta:"
        ws["F3"] = cuenta.venta.precio_venta
        ws["F3"].number_format = '"$ "#,##0.00_-'
        ws["E4"] = "Anticipo:"
        ws["F4"] = cuenta.venta.anticipo
        ws["F4"].number_format = '"$ "#,##0.00_-'
        ws["A6"] = "Comentarios: {}".format(cuenta.venta.observaciones)


        cont = 8
        valor_t = 0
        pagado_t = 0
        saldo_t = 0
        saldo_p_t = 0

        for c in cuota:

            if cont == 8:
                ws = wb.active
                ws.title = "Resumen"
                ws["A"+str(cont)] = "FECHA"
                ws["B"+str(cont)] = "CONSTANTE"
                ws["C"+str(cont)] = "CONCEPTO"
                ws["D"+str(cont)] = "VALOR"
                ws["E"+str(cont)] = "PAGADO"
                ws["F"+str(cont)] = "COTIZACIÓN"
                ws["G"+str(cont)] = "SALDO"
                ws["H"+str(cont)] = "SALDO PESOS"


                ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont)].alignment = Alignment(horizontal = "center")

                ws["A"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["A"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["B"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["B"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["C"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["C"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["D"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["D"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["E"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["E"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["F"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["F"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["G"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["G"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["H"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["H"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")

                ws["A"+str(cont+1)] = c.fecha
                ws["B"+str(cont+1)] = c.constante.nombre
                ws["C"+str(cont+1)] = c.concepto
                ws["D"+str(cont+1)] = c.precio

                if len(Pago.objects.filter(cuota = c)) > 0:
                    pagos_pesos = sum(np.array(Pago.objects.filter(cuota = c).values_list("pago_pesos", flat = True)))
                    cotizacion = sum(np.array(Pago.objects.filter(cuota = c).values_list("pago_pesos", flat = True))/np.array(Pago.objects.filter(cuota = c).values_list("pago", flat = True)))
                    saldo = c.precio - sum(np.array(Pago.objects.filter(cuota = c).values_list("pago", flat = True)))
                    saldo_pesos = saldo * c.constante.valor
                    
                    ws["E"+str(cont+1)] = pagos_pesos
                    ws["F"+str(cont+1)] = cotizacion
                    ws["G"+str(cont+1)] = saldo
                    ws["H"+str(cont+1)] = saldo_pesos

                else:
                    ws["E"+str(cont+1)] = 0
                    ws["F"+str(cont+1)] = 0
                    ws["G"+str(cont+1)] = c.precio
                    ws["H"+str(cont+1)] = c.precio*c.constante.valor

                ws["A"+str(cont+1)].font = Font(bold = True)

                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")

                if c.pagada == "SI":
                    ws["C"+str(cont+1)].font = Font(color= "289E70")
                    ws["G"+str(cont+1)].font = Font(color= "289E70")

                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")

                ws["D"+str(cont+1)].number_format = '#,##0.00_-'
                ws["E"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["F"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["G"+str(cont+1)].number_format = '#,##0.00_-'
                ws["H"+str(cont+1)].number_format = '"$ "#,##0.00_-'

                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 20
                ws.column_dimensions['G'].width = 20
                ws.column_dimensions['H'].width = 20

                cont += 1

            else: 

                ws["A"+str(cont+1)] = c.fecha
                ws["B"+str(cont+1)] = c.constante.nombre
                ws["C"+str(cont+1)] = c.concepto
                ws["D"+str(cont+1)] = c.precio

                if len(Pago.objects.filter(cuota = c)) > 0:
                    pagos_pesos = sum(np.array(Pago.objects.filter(cuota = c).values_list("pago_pesos", flat = True)))
                    cotizacion = sum(np.array(Pago.objects.filter(cuota = c).values_list("pago_pesos", flat = True))/np.array(Pago.objects.filter(cuota = c).values_list("pago", flat = True)))
                    saldo = c.precio - sum(np.array(Pago.objects.filter(cuota = c).values_list("pago", flat = True)))
                    saldo_pesos = saldo * c.constante.valor
                    
                    ws["E"+str(cont+1)] = pagos_pesos
                    ws["F"+str(cont+1)] = cotizacion
                    ws["G"+str(cont+1)] = saldo
                    ws["H"+str(cont+1)] = saldo_pesos

                    valor_t += c.precio
                    pagado_t += pagos_pesos
                    saldo_t += saldo
                    saldo_p_t += saldo_pesos


                else:
                    ws["E"+str(cont+1)] = 0
                    ws["F"+str(cont+1)] = 0
                    ws["G"+str(cont+1)] = c.precio
                    ws["H"+str(cont+1)] = c.precio*c.constante.valor

                    valor_t += c.precio
                    saldo_t += c.precio
                    saldo_p_t += c.precio*c.constante.valor

                ws["A"+str(cont+1)].font = Font(bold = True)

                if c.pagada == "SI":
                    ws["C"+str(cont+1)].font = Font(color= "289E70")
                    ws["G"+str(cont+1)].font = Font(color= "289E70")

                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
 
                ws["D"+str(cont+1)].number_format = '#,##0.00_-'
                ws["E"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["F"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["G"+str(cont+1)].number_format = '#,##0.00_-'
                ws["H"+str(cont+1)].number_format = '"$ "#,##0.00_-'

                cont += 1

        ws["D"+str(cont+1)] = valor_t
        ws["E"+str(cont+1)] = pagado_t
        ws["G"+str(cont+1)] = saldo_t
        ws["H"+str(cont+1)] = saldo_p_t

        ws["D"+str(cont+1)].font = Font(bold = True)
        ws["D"+str(cont+1)].number_format = '#,##0.00_-'
        ws["E"+str(cont+1)].font = Font(bold = True)
        ws["E"+str(cont+1)].number_format = '"$ "#,##0.00_-'
        ws["G"+str(cont+1)].font = Font(bold = True)
        ws["G"+str(cont+1)].number_format = '#,##0.00_-'
        ws["H"+str(cont+1)].font = Font(bold = True)
        ws["H"+str(cont+1)].number_format = '"$ "#,##0.00_-'


        ws["A"+str(cont+3)] = "  *El valor representa la cuota en su moneda dura"
        ws["A"+str(cont+4)] = "  *Pagado es la suma historica en pesos"
        ws["A"+str(cont+5)] = "  *Cotización es pagado sobre cancelado en moneda"
        ws["A"+str(cont+6)] = "  *Saldo es en moneda dura"
        ws["A"+str(cont+7)] = "  *Saldo en pesos es según el valor vigente en Link-P"


        cont = 3
        for p in pagos:

            if cont == 3:
                ws = wb.create_sheet('Registro Pagos')

                ws["A1"] = "Resumen de pagos - Área Administración"
                ws["A1"].font = Font(bold = True)

                ws["A"+str(cont)] = "CUOTA FECHA"
                ws["B"+str(cont)] = "FECHA"
                ws["C"+str(cont)] = "PAGO (MD)"
                ws["D"+str(cont)] = "MONEDA"
                ws["E"+str(cont)] = "PAGO PESOS"
                ws["F"+str(cont)] = "COTIZACIÓN"
                ws["G"+str(cont)] = "DOCUMENTO 1"
                ws["H"+str(cont)] = "DOCUMENTO 2"

                ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont)].alignment = Alignment(horizontal = "center")

                ws["A"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["A"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["B"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["B"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["C"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["C"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["D"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["D"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["E"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["E"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["F"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["F"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["G"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["G"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")
                ws["H"+str(cont)].font = Font(bold = True, color= "E8F8F8")
                ws["H"+str(cont)].fill =  PatternFill("solid", fgColor= "2C9E9D")

                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 20
                ws.column_dimensions['G'].width = 20
                ws.column_dimensions['H'].width = 20

                ws["A"+str(cont+1)] = p.cuota.fecha
                ws["B"+str(cont+1)] = p.fecha
                ws["C"+str(cont+1)] = p.pago
                ws["D"+str(cont+1)] = p.cuota.constante.nombre
                ws["E"+str(cont+1)] = p.pago_pesos
                if p.pago:
                    ws["F"+str(cont+1)] = (p.pago_pesos/p.pago)
                else:
                    ws["F"+str(cont+1)] = "??????"
                if p.documento_1:
                    ws["G"+str(cont+1)] = p.documento_1
                else:
                    ws["G"+str(cont+1)] = "Sin documento"
                if p.documento_2:
                    ws["H"+str(cont+1)] = p.documento_2
                else:
                    ws["H"+str(cont+1)] = "Sin documento"

                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont+1)].number_format = '#,##0.00_-'
                ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["F"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["G"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].alignment = Alignment(horizontal = "center")

                cont += 1

            else:
                
                ws = wb["Registro Pagos"]

                ws["A"+str(cont+1)] = p.cuota.fecha
                ws["B"+str(cont+1)] = p.fecha
                ws["C"+str(cont+1)] = p.pago
                ws["D"+str(cont+1)] = p.cuota.constante.nombre
                ws["E"+str(cont+1)] = p.pago_pesos
                if p.pago:
                    ws["F"+str(cont+1)] = (p.pago_pesos/p.pago)
                else:
                    ws["F"+str(cont+1)] = "??????"
                if p.documento_1:
                    ws["G"+str(cont+1)] = p.documento_1
                else:
                    ws["G"+str(cont+1)] = "Sin documento"
                if p.documento_2:
                    ws["H"+str(cont+1)] = p.documento_2
                else:
                    ws["H"+str(cont+1)] = "Sin documento"

                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont+1)].number_format = '#,##0.00_-'
                ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["F"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["G"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].alignment = Alignment(horizontal = "center")

                cont += 1

        #Establecer el nombre del archivo
        nombre_archivo = "Cuenta.-{0}.xls".format(str(cuenta.venta.comprador))
        
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo).replace(',', '_')
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class DescargarResumen(TemplateView):

    def get(self, request, id_proyecto, *args, **kwargs):

        # --> Iniciamos el Workbook
        wb = Workbook()

        # --> Primeros calculos

        proyecto = Proyectos.objects.get(id = int(id_proyecto))
        today = datetime.date.today()
        total_cuentas = len(CuentaCorriente.objects.filter(venta__proyecto = proyecto))
        total_cuotas = len(Cuota.objects.filter(cuenta_corriente__venta__proyecto = proyecto))
        total_pagos =  len(Pago.objects.filter(cuota__cuenta_corriente__venta__proyecto = proyecto))

        cuotas_anteriores = sum(np.array(Cuota.objects.values_list('precio', flat=True).filter(fecha__lt = today, cuenta_corriente__venta__proyecto= proyecto))*np.array(Cuota.objects.values_list('constante__valor', flat=True).filter(fecha__lt = today, cuenta_corriente__venta__proyecto = proyecto)))
        pagos_anteriores = sum(np.array(Pago.objects.values_list('pago', flat=True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente__venta__proyecto = proyecto))*np.array(Pago.objects.values_list('cuota__constante__valor', flat=True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente__venta__proyecto = proyecto)))
        deuda = cuotas_anteriores - pagos_anteriores

        cuotas_posterior = sum(np.array(Cuota.objects.values_list('precio', flat=True).filter(fecha__gte = today, cuenta_corriente__venta__proyecto = proyecto))*np.array(Cuota.objects.values_list('constante__valor', flat=True).filter(fecha__gte = today, cuenta_corriente__venta__proyecto = proyecto)))
        pagos_posterior= sum(np.array(Pago.objects.values_list('pago', flat=True).filter(cuota__fecha__gte = today, cuota__cuenta_corriente__venta__proyecto = proyecto))*np.array(Pago.objects.values_list('cuota__constante__valor', flat=True).filter(cuota__fecha__gte = today, cuota__cuenta_corriente__venta__proyecto = proyecto)))
        pagos_pesos= sum(np.array(Pago.objects.values_list('pago_pesos', flat=True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente__venta__proyecto = proyecto)))
        adelantos = cuotas_posterior - pagos_posterior

        array_pesos = np.array([cuotas_anteriores, pagos_anteriores, deuda, cuotas_posterior, pagos_posterior, adelantos])
        array_h = array_pesos/Constantes.objects.get(id = 7).valor

        ws = wb.active
        ws.title = "Resumen"
        ws["A1"] = "Resumen del proyecto - Área Administración"
        ws["A1"].font = Font(bold = True)
        ws["A3"] = "Cuentas:"
        ws["A3"].font = Font(bold = True)
        ws["B3"] = total_cuentas
        ws["A4"] = "Cuotas:"
        ws["A4"].font = Font(bold = True)
        ws["B4"] = total_cuotas
        ws["A5"] = "Pagos:"
        ws["A5"].font = Font(bold = True)
        ws["B5"] = total_pagos

        


        ws.merge_cells("D9:G9")
        ws["D9"] = "En pesos"

        ws.merge_cells("H9:K9")
        ws["H9"] = "En hormigon"

        ws["D9"].font = Font(bold = True, color= "E8F8F8")
        ws["D9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["D9"].alignment = Alignment(horizontal = "center")

        ws["H9"].font = Font(bold = True, color= "E8F8F8")
        ws["H9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["H9"].alignment = Alignment(horizontal = "center")


        ws["A10"] = "Cliente"
        ws["B10"] = "Unidad"
        ws["C10"] = "Asignación"
        ws["D10"] = "Pagado"
        ws["E10"] = "Adeudado"
        ws["F10"] = "Adelantos"
        ws["G10"] = "Saldo"
        ws["H10"] = "Pagado Hº"
        ws["I10"] = "Adeudado Hº"
        ws["J10"] = "Adelantos Hº"
        ws["K10"] = "Saldo Hº"

        ws["A10"].font = Font(bold = True, color= "E8F8F8")
        ws["A10"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["B10"].font = Font(bold = True, color= "E8F8F8")
        ws["B10"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["C10"].font = Font(bold = True, color= "E8F8F8")
        ws["C10"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["D10"].font = Font(bold = True, color= "E8F8F8")
        ws["D10"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["E10"].font = Font(bold = True, color= "E8F8F8")
        ws["E10"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["F10"].font = Font(bold = True, color= "E8F8F8")
        ws["F10"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["G10"].font = Font(bold = True, color= "E8F8F8")
        ws["G10"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["H10"].font = Font(bold = True, color= "E8F8F8")
        ws["H10"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["I10"].font = Font(bold = True, color= "E8F8F8")
        ws["I10"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["J10"].font = Font(bold = True, color= "E8F8F8")
        ws["J10"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["K10"].font = Font(bold = True, color= "E8F8F8")
        ws["K10"].fill =  PatternFill("solid", fgColor= "2C9E9D")


        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 17
        ws.column_dimensions['D'].width = 17
        ws.column_dimensions['E'].width = 17
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15

        cuentas = CuentaCorriente.objects.filter(venta__proyecto = proyecto)

        cont = 11

        pagos_t = 0
        adeudado_t = 0
        adelantos_t = 0
        saldo_t = 0
        pagos_m3_t = 0
        adeudado_m3_t = 0
        adelantos_m3_t = 0
        saldo_m3_t = 0

        pagos_p = 0
        adeudado_p = 0
        adelantos_p = 0
        saldo_p = 0
        pagos_m3_p = 0
        adeudado_m3_p = 0
        adelantos_m3_p = 0
        saldo_m3_p = 0

        pagos_l = 0
        adeudado_l = 0
        adelantos_l = 0
        saldo_l = 0
        pagos_m3_l = 0
        adeudado_m3_l = 0
        adelantos_m3_l = 0
        saldo_m3_l = 0

        pagos_te = 0
        adeudado_te = 0
        adelantos_te = 0
        saldo_te = 0
        pagos_m3_te = 0
        adeudado_m3_te = 0
        adelantos_m3_te = 0
        saldo_m3_te = 0

        for c in cuentas:


            pagos = sum(np.array(Pago.objects.values_list('pago_pesos', flat = True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente = c)))           
            cuotas_anteriores = sum(np.array(Cuota.objects.values_list('precio', flat = True).filter(fecha__lt = today, cuenta_corriente = c))*np.array(Cuota.objects.values_list('constante__valor', flat = True).filter(fecha__lt = today, cuenta_corriente = c)))
            pagos_anteriores = sum(np.array(Pago.objects.values_list('pago', flat = True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente = c))*np.array(Pago.objects.values_list('cuota__constante__valor', flat = True).filter(cuota__fecha__lt = today, cuota__cuenta_corriente = c)))
            adeudado = cuotas_anteriores - pagos_anteriores
            adelantos = sum(np.array(Pago.objects.values_list('pago', flat = True).filter(cuota__fecha__gte = today, cuota__cuenta_corriente = c))*np.array(Pago.objects.values_list('cuota__constante__valor', flat = True).filter(cuota__fecha__gte = today, cuota__cuenta_corriente = c)))
            saldo = sum(np.array(Cuota.objects.values_list('precio', flat = True).filter(fecha__gte = today, cuenta_corriente = c))*np.array(Cuota.objects.values_list('constante__valor', flat = True).filter(fecha__gte = today, cuenta_corriente = c))) - adelantos
            
            ws = wb.active

            ws["A"+str(cont)] = c.venta.comprador
            ws["B"+str(cont)] = (str(c.venta.unidad.piso_unidad) + "-" + str(c.venta.unidad.nombre_unidad))
            ws["C"+str(cont)] = c.venta.unidad.asig
            ws["D"+str(cont)] = pagos
            ws["E"+str(cont)] = adeudado
            ws["F"+str(cont)] = adelantos
            ws["G"+str(cont)] = saldo
            ws["H"+str(cont)] = pagos_anteriores/Constantes.objects.get(id = 7).valor
            ws["I"+str(cont)] = adeudado/Constantes.objects.get(id = 7).valor
            ws["J"+str(cont)] = adelantos/Constantes.objects.get(id = 7).valor
            ws["K"+str(cont)] = saldo/Constantes.objects.get(id = 7).valor

            pagos_t += pagos
            adeudado_t += adeudado
            adelantos_t += adelantos
            saldo_t += saldo
            pagos_m3_t += pagos_anteriores/Constantes.objects.get(id = 7).valor
            adeudado_m3_t += adeudado/Constantes.objects.get(id = 7).valor
            adelantos_m3_t += adelantos/Constantes.objects.get(id = 7).valor
            saldo_m3_t += saldo/Constantes.objects.get(id = 7).valor

            if c.venta.unidad.asig == "PROYECTO":
                pagos_p += pagos
                adeudado_p += adeudado
                adelantos_p += adelantos
                saldo_p += saldo
                pagos_m3_p += pagos_anteriores/Constantes.objects.get(id = 7).valor
                adeudado_m3_p += adeudado/Constantes.objects.get(id = 7).valor
                adelantos_m3_p += adelantos/Constantes.objects.get(id = 7).valor
                saldo_m3_p += saldo/Constantes.objects.get(id = 7).valor

            if c.venta.unidad.asig == "HON. LINK":
                pagos_l += pagos
                adeudado_l += adeudado
                adelantos_l += adelantos
                saldo_l += saldo
                pagos_m3_l += pagos_anteriores/Constantes.objects.get(id = 7).valor
                adeudado_m3_l += adeudado/Constantes.objects.get(id = 7).valor
                adelantos_m3_l += adelantos/Constantes.objects.get(id = 7).valor
                saldo_m3_l += saldo/Constantes.objects.get(id = 7).valor

            if c.venta.unidad.asig == "TERRENO":
                pagos_te += pagos
                adeudado_te += adeudado
                adelantos_te += adelantos
                saldo_te += saldo
                pagos_m3_te += pagos_anteriores/Constantes.objects.get(id = 7).valor
                adeudado_m3_te += adeudado/Constantes.objects.get(id = 7).valor
                adelantos_m3_te += adelantos/Constantes.objects.get(id = 7).valor
                saldo_m3_te += saldo/Constantes.objects.get(id = 7).valor


            ws["D"+str(cont)].number_format = '"$"#,##0.00_-'
            ws["E"+str(cont)].number_format = '"$"#,##0.00_-'
            ws["F"+str(cont)].number_format = '"$"#,##0.00_-'
            ws["G"+str(cont)].number_format = '"$"#,##0.00_-'
            ws["H"+str(cont)].number_format = '#,##0.00_-"M3"'
            ws["I"+str(cont)].number_format = '#,##0.00_-"M3"'
            ws["J"+str(cont)].number_format = '#,##0.00_-"M3"'
            ws["K"+str(cont)].number_format = '#,##0.00_-"M3"'

            cont += 1

        ws["A"+str(cont)] = "Total"
        ws["D"+str(cont)] = pagos_t
        ws["E"+str(cont)] = adeudado_t
        ws["F"+str(cont)] = adelantos_t
        ws["G"+str(cont)] = saldo_t
        ws["H"+str(cont)] = pagos_m3_t
        ws["I"+str(cont)] = adeudado_m3_t
        ws["J"+str(cont)] = adelantos_m3_t
        ws["K"+str(cont)] = saldo_m3_t

        ws["A"+str(cont)].font = Font(bold = True)
        ws["D"+str(cont)].font = Font(bold = True)
        ws["E"+str(cont)].font = Font(bold = True)
        ws["F"+str(cont)].font = Font(bold = True)
        ws["G"+str(cont)].font = Font(bold = True)
        ws["H"+str(cont)].font = Font(bold = True)
        ws["I"+str(cont)].font = Font(bold = True)
        ws["J"+str(cont)].font = Font(bold = True)
        ws["K"+str(cont)].font = Font(bold = True)

        ws["D"+str(cont)].number_format = '"$"#,##0.00_-'
        ws["E"+str(cont)].number_format = '"$"#,##0.00_-'
        ws["F"+str(cont)].number_format = '"$"#,##0.00_-'
        ws["G"+str(cont)].number_format = '"$"#,##0.00_-'
        ws["H"+str(cont)].number_format = '#,##0.00_-"M3"'
        ws["I"+str(cont)].number_format = '#,##0.00_-"M3"'
        ws["J"+str(cont)].number_format = '#,##0.00_-"M3"'
        ws["K"+str(cont)].number_format = '#,##0.00_-"M3"'

        ws["A"+str(cont+1)] = "Proyecto"
        ws["D"+str(cont+1)] = pagos_p
        ws["E"+str(cont+1)] = adeudado_p
        ws["F"+str(cont+1)] = adelantos_p
        ws["G"+str(cont+1)] = saldo_p
        ws["H"+str(cont+1)] = pagos_m3_p
        ws["I"+str(cont+1)] = adeudado_m3_p
        ws["J"+str(cont+1)] = adelantos_m3_p
        ws["K"+str(cont+1)] = saldo_m3_p

        ws["A"+str(cont+1)].font = Font(bold = True)
        ws["D"+str(cont+1)].font = Font(bold = True)
        ws["E"+str(cont+1)].font = Font(bold = True)
        ws["F"+str(cont+1)].font = Font(bold = True)
        ws["G"+str(cont+1)].font = Font(bold = True)
        ws["H"+str(cont+1)].font = Font(bold = True)
        ws["I"+str(cont+1)].font = Font(bold = True)
        ws["J"+str(cont+1)].font = Font(bold = True)
        ws["K"+str(cont+1)].font = Font(bold = True)

        ws["D"+str(cont+1)].number_format = '"$"#,##0.00_-'
        ws["E"+str(cont+1)].number_format = '"$"#,##0.00_-'
        ws["F"+str(cont+1)].number_format = '"$"#,##0.00_-'
        ws["G"+str(cont+1)].number_format = '"$"#,##0.00_-'
        ws["H"+str(cont+1)].number_format = '#,##0.00_-"M3"'
        ws["I"+str(cont+1)].number_format = '#,##0.00_-"M3"'
        ws["J"+str(cont+1)].number_format = '#,##0.00_-"M3"'
        ws["K"+str(cont+1)].number_format = '#,##0.00_-"M3"'

        ws["A"+str(cont+2)] = "Link"
        ws["D"+str(cont+2)] = pagos_l
        ws["E"+str(cont+2)] = adeudado_l
        ws["F"+str(cont+2)] = adelantos_l
        ws["G"+str(cont+2)] = saldo_l
        ws["H"+str(cont+2)] = pagos_m3_l
        ws["I"+str(cont+2)] = adeudado_m3_l
        ws["J"+str(cont+2)] = adelantos_m3_l
        ws["K"+str(cont+2)] = saldo_m3_l

        ws["A"+str(cont+2)].font = Font(bold = True)
        ws["D"+str(cont+2)].font = Font(bold = True)
        ws["E"+str(cont+2)].font = Font(bold = True)
        ws["F"+str(cont+2)].font = Font(bold = True)
        ws["G"+str(cont+2)].font = Font(bold = True)
        ws["H"+str(cont+2)].font = Font(bold = True)
        ws["I"+str(cont+2)].font = Font(bold = True)
        ws["J"+str(cont+2)].font = Font(bold = True)
        ws["K"+str(cont+2)].font = Font(bold = True)

        ws["D"+str(cont+2)].number_format = '"$"#,##0.00_-'
        ws["E"+str(cont+2)].number_format = '"$"#,##0.00_-'
        ws["F"+str(cont+2)].number_format = '"$"#,##0.00_-'
        ws["G"+str(cont+2)].number_format = '"$"#,##0.00_-'
        ws["H"+str(cont+2)].number_format = '#,##0.00_-"M3"'
        ws["I"+str(cont+2)].number_format = '#,##0.00_-"M3"'
        ws["J"+str(cont+2)].number_format = '#,##0.00_-"M3"'
        ws["K"+str(cont+2)].number_format = '#,##0.00_-"M3"'

        ws["A"+str(cont+3)] = "Terreno"
        ws["D"+str(cont+3)] = pagos_te
        ws["E"+str(cont+3)] = adeudado_te
        ws["F"+str(cont+3)] = adelantos_te
        ws["G"+str(cont+3)] = saldo_te
        ws["H"+str(cont+3)] = pagos_m3_te
        ws["I"+str(cont+3)] = adeudado_m3_te
        ws["J"+str(cont+3)] = adelantos_m3_te
        ws["K"+str(cont+3)] = saldo_m3_te

        ws["A"+str(cont+3)].font = Font(bold = True)
        ws["D"+str(cont+3)].font = Font(bold = True)
        ws["E"+str(cont+3)].font = Font(bold = True)
        ws["F"+str(cont+3)].font = Font(bold = True)
        ws["G"+str(cont+3)].font = Font(bold = True)
        ws["H"+str(cont+3)].font = Font(bold = True)
        ws["I"+str(cont+3)].font = Font(bold = True)
        ws["J"+str(cont+3)].font = Font(bold = True)
        ws["K"+str(cont+3)].font = Font(bold = True)

        ws["D"+str(cont+3)].number_format = '"$"#,##0.00_-'
        ws["E"+str(cont+3)].number_format = '"$"#,##0.00_-'
        ws["F"+str(cont+3)].number_format = '"$"#,##0.00_-'
        ws["G"+str(cont+3)].number_format = '"$"#,##0.00_-'
        ws["H"+str(cont+3)].number_format = '#,##0.00_-"M3"'
        ws["I"+str(cont+3)].number_format = '#,##0.00_-"M3"'
        ws["J"+str(cont+3)].number_format = '#,##0.00_-"M3"'
        ws["K"+str(cont+3)].number_format = '#,##0.00_-"M3"'


        # Establecemos un rango para hacer el cash de ingreso

        fecha_inicial_hoy = datetime.date.today()

        fecha_inicial_2 = datetime.date(fecha_inicial_hoy.year, fecha_inicial_hoy.month, 1)

        fechas = []

        #Aqui buscamos la ultima fecha

        fecha_ultima = Cuota.objects.filter(cuenta_corriente__venta__proyecto = proyecto).order_by("-fecha")

        contador = 0
        
        contador_year = 1

        fecha_cargar = fecha_inicial_2

        while fecha_cargar < fecha_ultima[0].fecha:

            if (fecha_inicial_2.month + contador) == 13:
                
                year = fecha_inicial_2.year + contador_year
                
                fecha_cargar = date(year, 1, fecha_inicial_2.day)

                fechas.append(fecha_cargar)
                
                contador_year += 1

                contador = - (12 - contador)

            else:

                mes = fecha_inicial_2.month + contador

                year = fecha_inicial_2.year + contador_year - 1

                fecha_cargar = date(year, mes, fecha_inicial_2.day)

                fechas.append(fecha_cargar)

            contador += 1

        if (fecha_inicial_2.month + contador) == 13:
                
            year = fecha_inicial_2.year + contador_year
            
            fecha_cargar = date(year, 1, fecha_inicial_2.day)

            fechas.append(fecha_cargar)
            
            contador_year += 1

            contador = - (12 - contador)

        else:

            mes = fecha_inicial_2.month + contador

            year = fecha_inicial_2.year + contador_year - 1

            fecha_cargar = date(year, mes, fecha_inicial_2.day)

            fechas.append(fecha_cargar)

        ws = wb.create_sheet('Flujo')

        ws["A1"] = "Resumen de pagos - Área Administración"
        ws["A1"].font = Font(bold = True)


        ws.merge_cells("B8:C8")
        ws["B8"] = "Total"

        ws["B8"].font = Font(bold = True, color= "E8F8F8")
        ws["B8"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["B8"].alignment = Alignment(horizontal = "center")

        ws.merge_cells("D8:E8")
        ws["D8"] = "Proyecto"

        ws["D8"].font = Font(bold = True, color= "E8F8F8")
        ws["D8"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["D8"].alignment = Alignment(horizontal = "center")

        ws.merge_cells("F8:G8")
        ws["F8"] = "Link"

        ws["F8"].font = Font(bold = True, color= "E8F8F8")
        ws["F8"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["F8"].alignment = Alignment(horizontal = "center")

        ws.merge_cells("H8:I8")
        ws["H8"] = "Terreno"

        ws["H8"].font = Font(bold = True, color= "E8F8F8")
        ws["H8"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["H8"].alignment = Alignment(horizontal = "center")


        ws["A9"] = "Mes"
        ws["B9"] = "Ingreso total"
        ws["C9"] = "Ingreso total M3"
        ws["D9"] = "Ingreso proyecto"
        ws["E9"] = "Ingreso proyecto M3"
        ws["F9"] = "Ingreso Link"
        ws["G9"] = "Ingreso Link M3"
        ws["H9"] = "Ingreso Terreno"
        ws["I9"] = "Ingreso Terreno M3"

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20

        ws["A9"].font = Font(bold = True, color= "E8F8F8")
        ws["A9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["B9"].font = Font(bold = True, color= "E8F8F8")
        ws["B9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["C9"].font = Font(bold = True, color= "E8F8F8")
        ws["C9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["D9"].font = Font(bold = True, color= "E8F8F8")
        ws["D9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["E9"].font = Font(bold = True, color= "E8F8F8")
        ws["E9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["F9"].font = Font(bold = True, color= "E8F8F8")
        ws["F9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["G9"].font = Font(bold = True, color= "E8F8F8")
        ws["G9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["H9"].font = Font(bold = True, color= "E8F8F8")
        ws["H9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["I9"].font = Font(bold = True, color= "E8F8F8")
        ws["I9"].fill =  PatternFill("solid", fgColor= "2C9E9D")

        cont = 10

        cont_f = -1
        fecha_aux = 0
        for fecha in fechas:

            if cont_f <= 0:
                fecha_aux = fecha
                cont_f += 1
            else:

                fecha = fecha + timedelta(days=-1)
       
                # Total
                if len(Cuota.objects.values_list('pago', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto)) > 0:
                    cuotas_month = sum(np.array(Cuota.objects.values_list('precio', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto))*np.array(Cuota.objects.values_list('constante__valor', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto)))
                    pago_month = sum(np.array(Pago.objects.values_list('pago', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto))*np.array(Pago.objects.values_list('cuota__constante__valor', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto)))
                    saldo_month = cuotas_month - pago_month
                    saldo_month_h = saldo_month/Constantes.objects.get(id = 7).valor
                else:
                    saldo_month = 0
                    saldo_month_h = 0
                # Proyecto

                if len(Cuota.objects.values_list('pago', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "PROYECTO")) > 0:
                    cuotas_month = sum(np.array(Cuota.objects.values_list('precio', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "PROYECTO"))*np.array(Cuota.objects.values_list('constante__valor', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "PROYECTO")))
                    pago_month = sum(np.array(Pago.objects.values_list('pago', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "PROYECTO"))*np.array(Pago.objects.values_list('cuota__constante__valor', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "PROYECTO")))
                    saldo_month_p = cuotas_month - pago_month
                    saldo_month_p_h = saldo_month/Constantes.objects.get(id = 7).valor
                else:
                    saldo_month_p = 0
                    saldo_month_p_h = 0

                # Link

                if len(Cuota.objects.values_list('pago', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "HON. LINK")) > 0:
                    cuotas_month = sum(np.array(Cuota.objects.values_list('precio', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "HON. LINK"))*np.array(Cuota.objects.values_list('constante__valor', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "HON. LINK")))
                    pago_month = sum(np.array(Pago.objects.values_list('pago', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "HON. LINK"))*np.array(Pago.objects.values_list('cuota__constante__valor', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "HON. LINK")))
                    saldo_month_l = cuotas_month - pago_month
                    saldo_month_l_h = saldo_month/Constantes.objects.get(id = 7).valor
                else:
                    saldo_month_l = 0
                    saldo_month_l_h = 0

                # Terreno

                if len(Cuota.objects.values_list('pago', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "TERRENO")) > 0:
                    cuotas_month = sum(np.array(Cuota.objects.values_list('precio', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "TERRENO"))*np.array(Cuota.objects.values_list('constante__valor', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "TERRENO")))
                    pago_month = sum(np.array(Pago.objects.values_list('pago', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "TERRENO"))*np.array(Pago.objects.values_list('cuota__constante__valor', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "TERRENO")))
                    saldo_month_te = cuotas_month - pago_month
                    saldo_month_te_h = saldo_month/Constantes.objects.get(id = 7).valor
                else:
                    saldo_month_te = 0
                    saldo_month_te_h = 0

                ws["A"+str(cont)] = fecha_aux
                ws["B"+str(cont)] = saldo_month
                ws["C"+str(cont)] = saldo_month_h
                ws["D"+str(cont)] = saldo_month_p
                ws["E"+str(cont)] = saldo_month_p_h
                ws["F"+str(cont)] = saldo_month_l
                ws["G"+str(cont)] = saldo_month_l_h
                ws["H"+str(cont)] = saldo_month_te
                ws["I"+str(cont)] = saldo_month_te_h

                ws["B"+str(cont)].number_format = '"$"#,##0.00_-'
                ws["C"+str(cont)].number_format = '#,##0.00_-"M3"'
                ws["D"+str(cont)].number_format = '"$"#,##0.00_-'
                ws["E"+str(cont)].number_format = '#,##0.00_-"M3"'
                ws["F"+str(cont)].number_format = '"$"#,##0.00_-'
                ws["G"+str(cont)].number_format = '#,##0.00_-"M3"'
                ws["H"+str(cont)].number_format = '"$"#,##0.00_-'
                ws["I"+str(cont)].number_format = '#,##0.00_-"M3"'


                fecha = fecha + timedelta(days=+1)
                fecha_aux = fecha
                cont += 1

        # ---> BOLETO

        ws = wb.create_sheet('Flujo B')

        ws["A1"] = "Resumen de pagos - Área Administración"
        ws["A1"].font = Font(bold = True)


        ws.merge_cells("B8:C8")
        ws["B8"] = "Total"

        ws["B8"].font = Font(bold = True, color= "E8F8F8")
        ws["B8"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["B8"].alignment = Alignment(horizontal = "center")

        ws.merge_cells("D8:E8")
        ws["D8"] = "Proyecto"

        ws["D8"].font = Font(bold = True, color= "E8F8F8")
        ws["D8"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["D8"].alignment = Alignment(horizontal = "center")

        ws.merge_cells("F8:G8")
        ws["F8"] = "Link"

        ws["F8"].font = Font(bold = True, color= "E8F8F8")
        ws["F8"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["F8"].alignment = Alignment(horizontal = "center")

        ws.merge_cells("H8:I8")
        ws["H8"] = "Terreno"

        ws["H8"].font = Font(bold = True, color= "E8F8F8")
        ws["H8"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["H8"].alignment = Alignment(horizontal = "center")


        ws["A9"] = "Mes"
        ws["B9"] = "Ingreso total"
        ws["C9"] = "Ingreso total M3"
        ws["D9"] = "Ingreso proyecto"
        ws["E9"] = "Ingreso proyecto M3"
        ws["F9"] = "Ingreso Link"
        ws["G9"] = "Ingreso Link M3"
        ws["H9"] = "Ingreso Terreno"
        ws["I9"] = "Ingreso Terreno M3"

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20

        ws["A9"].font = Font(bold = True, color= "E8F8F8")
        ws["A9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["B9"].font = Font(bold = True, color= "E8F8F8")
        ws["B9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["C9"].font = Font(bold = True, color= "E8F8F8")
        ws["C9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["D9"].font = Font(bold = True, color= "E8F8F8")
        ws["D9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["E9"].font = Font(bold = True, color= "E8F8F8")
        ws["E9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["F9"].font = Font(bold = True, color= "E8F8F8")
        ws["F9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["G9"].font = Font(bold = True, color= "E8F8F8")
        ws["G9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["H9"].font = Font(bold = True, color= "E8F8F8")
        ws["H9"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["I9"].font = Font(bold = True, color= "E8F8F8")
        ws["I9"].fill =  PatternFill("solid", fgColor= "2C9E9D")

        cont = 10

        cont_f = -1
        fecha_aux = 0
        for fecha in fechas:

            if cont_f <= 0:
                fecha_aux = fecha
                cont_f += 1
            else:

                fecha = fecha + timedelta(days=-1)
       
                # Total
                if len(Cuota.objects.values_list('pago', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto).exclude(porc_boleto = None)) > 0:
                    cuotas_month = sum(np.array(Cuota.objects.values_list('precio', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto).exclude(porc_boleto = None))*np.array(Cuota.objects.values_list('constante__valor', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto).exclude(porc_boleto = None))*np.array(Cuota.objects.values_list('porc_boleto', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto).exclude(porc_boleto = None)))
                    pago_month = sum(np.array(Pago.objects.values_list('pago', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto).exclude(cuota__porc_boleto = None))*np.array(Pago.objects.values_list('cuota__constante__valor', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto).exclude(cuota__porc_boleto = None))*np.array(Pago.objects.values_list('cuota__porc_boleto', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto).exclude(cuota__porc_boleto = None)))
                    saldo_month = cuotas_month - pago_month
                    saldo_month_h = saldo_month/Constantes.objects.get(id = 7).valor
                else:
                    saldo_month = 0
                    saldo_month_h = 0
                # Proyecto

                if len(Cuota.objects.values_list('pago', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "PROYECTO").exclude(porc_boleto = None)) > 0:
                    cuotas_month = sum(np.array(Cuota.objects.values_list('precio', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "PROYECTO").exclude(porc_boleto = None))*np.array(Cuota.objects.values_list('constante__valor', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "PROYECTO").exclude(porc_boleto = None))*np.array(Cuota.objects.values_list('porc_boleto', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "PROYECTO").exclude(porc_boleto = None)))
                    pago_month = sum(np.array(Pago.objects.values_list('pago', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "PROYECTO").exclude(cuota__porc_boleto = None))*np.array(Pago.objects.values_list('cuota__constante__valor', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "PROYECTO").exclude(cuota__porc_boleto = None))*np.array(Pago.objects.values_list('cuota__porc_boleto', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "PROYECTO").exclude(cuota__porc_boleto = None)))
                    saldo_month_p = cuotas_month - pago_month
                    saldo_month_p_h = saldo_month/Constantes.objects.get(id = 7).valor
                else:
                    saldo_month_p = 0
                    saldo_month_p_h = 0

                # Link

                if len(Cuota.objects.values_list('pago', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "HON. LINK")) > 0:
                    cuotas_month = sum(np.array(Cuota.objects.values_list('precio', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "HON. LINK"))*np.array(Cuota.objects.values_list('constante__valor', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "HON. LINK")))
                    pago_month = sum(np.array(Pago.objects.values_list('pago', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "HON. LINK"))*np.array(Pago.objects.values_list('cuota__constante__valor', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "HON. LINK")))
                    saldo_month_l = cuotas_month - pago_month
                    saldo_month_l_h = saldo_month/Constantes.objects.get(id = 7).valor
                else:
                    saldo_month_l = 0
                    saldo_month_l_h = 0

                # Terreno

                if len(Cuota.objects.values_list('pago', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "TERRENO")) > 0:
                    cuotas_month = sum(np.array(Cuota.objects.values_list('precio', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "TERRENO"))*np.array(Cuota.objects.values_list('constante__valor', flat = True).filter(fecha__range = (fecha_aux, fecha), cuenta_corriente__venta__proyecto = proyecto, cuenta_corriente__venta__unidad__asig = "TERRENO")))
                    pago_month = sum(np.array(Pago.objects.values_list('pago', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "TERRENO"))*np.array(Pago.objects.values_list('cuota__constante__valor', flat = True).filter(cuota__fecha__range = (fecha_aux, fecha), cuota__cuenta_corriente__venta__proyecto = proyecto, cuota__cuenta_corriente__venta__unidad__asig = "TERRENO")))
                    saldo_month_te = cuotas_month - pago_month
                    saldo_month_te_h = saldo_month/Constantes.objects.get(id = 7).valor
                else:
                    saldo_month_te = 0
                    saldo_month_te_h = 0

                ws["A"+str(cont)] = fecha_aux
                ws["B"+str(cont)] = saldo_month
                ws["C"+str(cont)] = saldo_month_h
                ws["D"+str(cont)] = saldo_month_p
                ws["E"+str(cont)] = saldo_month_p_h
                ws["F"+str(cont)] = saldo_month_l
                ws["G"+str(cont)] = saldo_month_l_h
                ws["H"+str(cont)] = saldo_month_te
                ws["I"+str(cont)] = saldo_month_te_h

                ws["B"+str(cont)].number_format = '"$"#,##0.00_-'
                ws["C"+str(cont)].number_format = '#,##0.00_-"M3"'
                ws["D"+str(cont)].number_format = '"$"#,##0.00_-'
                ws["E"+str(cont)].number_format = '#,##0.00_-"M3"'
                ws["F"+str(cont)].number_format = '"$"#,##0.00_-'
                ws["G"+str(cont)].number_format = '#,##0.00_-"M3"'
                ws["H"+str(cont)].number_format = '"$"#,##0.00_-'
                ws["I"+str(cont)].number_format = '#,##0.00_-"M3"'


                fecha = fecha + timedelta(days=+1)
                fecha_aux = fecha
                cont += 1

        #Establecer el nombre del archivo
        nombre_archivo = "Resumen.-{}.xls".format(str(proyecto.nombre))
        
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo).replace(',', '_')
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class DescargarRegistroContable(TemplateView):

    def get(self, request, *args, **kwargs):

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"
        ws["A1"] = "Resgistro de cargas realizadas"
        ws["A1"].font = Font(bold = True)

        ws["A5"] = "Destino"
        ws["B5"] = "Cargo"
        ws["C5"] = "Caja"
        ws["D5"] = "Tipo"
        ws["E5"] = "Cuenta"
        ws["F5"] = "Categoria"
        ws["G5"] = "Monto"

        ws["A5"].font = Font(bold = True, color= "E8F8F8")
        ws["A5"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["B5"].font = Font(bold = True, color= "E8F8F8")
        ws["B5"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["C5"].font = Font(bold = True, color= "E8F8F8")
        ws["C5"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["D5"].font = Font(bold = True, color= "E8F8F8")
        ws["D5"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["E5"].font = Font(bold = True, color= "E8F8F8")
        ws["E5"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["F5"].font = Font(bold = True, color= "E8F8F8")
        ws["F5"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["G5"].font = Font(bold = True, color= "E8F8F8")
        ws["G5"].fill =  PatternFill("solid", fgColor= "2C9E9D")

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 17
        ws.column_dimensions['D'].width = 17
        ws.column_dimensions['E'].width = 17
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15

        cont = 6

        datos_primeros = Regis


class DescargarTotalCuentas(TemplateView):

    def get(self, request, *args, **kwargs):


        #color = '#%02x%02x%02x' % (0, 128, 64)
        
        wb = Workbook()
        #Establecemos una lista

        proyectos = Proyectos.objects.all()

        datos_primeros = []

        listado = []

        for proyecto in proyectos:

            cuotas = Cuota.objects.filter(cuenta_corriente__venta__proyecto = proyecto)

            if len(cuotas)>0:
                datos_primeros.append(proyecto)
                
                for c in cuotas:

                    listado.append(c.cuenta_corriente.venta.proyecto)

        listado = set(list(listado))


        #En la primera vuelta sacamos el totalizado

        proy = 0

        cantidad_cuentas = len(CuentaCorriente.objects.all())
        
        #Establecemos un rango para hacer el cash de ingreso
        
        fecha_inicial_hoy = datetime.date.today()

        fecha_inicial_2 = datetime.date(fecha_inicial_hoy.year, fecha_inicial_hoy.month, 1)

        fechas = []

        #Aqui buscamos la ultima fecha

        fecha_ultima = Cuota.objects.order_by("-fecha")

        contador = 0
        
        contador_year = 1

        fecha_cargar = fecha_inicial_2

        while fecha_cargar < fecha_ultima[0].fecha:

            if (fecha_inicial_2.month + contador) == 13:
                
                year = fecha_inicial_2.year + contador_year
                
                fecha_cargar = date(year, 1, fecha_inicial_2.day)

                fechas.append(fecha_cargar)
                
                contador_year += 1

                contador = - (12 - contador)

            else:

                mes = fecha_inicial_2.month + contador

                year = fecha_inicial_2.year + contador_year - 1

                fecha_cargar = date(year, mes, fecha_inicial_2.day)

                fechas.append(fecha_cargar)

            contador += 1

        #Aqui hacemos los totalizadores generales

        cuotas_anteriores = Cuota.objects.filter(fecha__lt = fecha_inicial_hoy)
        pagos_anteriores = Pago.objects.filter(fecha__lt = fecha_inicial_hoy)
        cuotas_posteriores = Cuota.objects.filter(fecha__gt = fecha_inicial_hoy)

        total_original = 0
        total_cobrado = 0
        total_pendiente = 0
        total_acobrar= 0

        for cuo in cuotas_anteriores:

            total_original = total_original + cuo.precio

        for pag in pagos_anteriores:

            total_cobrado = total_cobrado +  pag.pago

        for cuot in cuotas_posteriores:

            total_acobrar= total_acobrar + cuot.precio

        total_pendiente = total_original - total_cobrado

        otros_datos = [total_cobrado, total_pendiente, total_acobrar]

        #Aqui buscamos agrupar proyecto - sumatorias de cuotas y pagos - mes
        
        datos_segundos = []

        total_fecha = []

        fecha_inicial = 0

        for f in fechas:

            total = 0
            total_link = 0

            datos_terceros = []

            if fecha_inicial == 0:

                    fecha_inicial = fecha_inicial_hoy

            else:

                cuotas = Cuota.objects.filter(fecha__range = (fecha_inicial, f))
                    
                pagos = Pago.objects.filter(fecha__range = (fecha_inicial, f))

                total_cuotas = 0
                total_cuotas_link = 0
                total_pagado = 0
                total_pagado_link = 0
                saldo = 0
                saldo_link = 0

                if len(cuotas)>0:

                    for c in cuotas:

                        total_cuotas = total_cuotas + c.precio*c.constante.valor

                        if c.cuenta_corriente.venta.unidad.asig == "HON. LINK" or c.cuenta_corriente.venta.unidad.asig == "TERRENO":

                            total_cuotas_link = total_cuotas_link + c.precio*c.constante.valor 

                if len(pagos)>0:

                    for p in pagos:

                        total_pagado = total_pagado + p.pago*p.cuota.constante.valor

                        if p.cuota.cuenta_corriente.venta.unidad.asig == "HON. LINK" or p.cuota.cuenta_corriente.venta.unidad.asig == "TERRENO":

                            total_pagado_link = total_pagado_link + p.pago*p.cuota.constante.valor 

                saldo = total_cuotas-total_pagado

                total = total + saldo

                #Aqui calculamos el saldo de cuotas de LINK

                saldo_link = total_cuotas_link-total_pagado_link

                total_link = total_link + saldo_link
                
                datos_terceros.append((fecha_inicial, saldo, saldo_link))

                fecha_inicial = f

                horm = Constantes.objects.get(nombre = "Hº VIVIENDA")
                
                total_horm = total/horm.valor

                total_horm_link = total_link/horm.valor

                datos_segundos.append((datos_terceros, total, total_horm, total_link, total_horm_link))


        #Aqui termina la primera vuelta -> Total por proyecto

        cont = 10

        ws = wb.active
        ws.title = "Cuotas"
        ws["A"+str(2)] = "RESUMEN MES A MES DE INGRESOS POR CUOTAS A COBRAR"
        ws["A"+str(5)] = "CUENTAS"
        ws["A"+str(6)] = "COBRADO"
        ws["A"+str(7)] = "ADEUDADO"
        ws["A"+str(8)] = "PENDIENTE"

        ws["A"+str(2)].font = Font(bold = True)
        ws["A"+str(5)].font = Font(bold = True, color= "FDFFFF")
        ws["A"+str(5)].fill =  PatternFill("solid", fgColor= "33353B")
        ws["A"+str(6)].font = Font(bold = True, color= "FDFFFF")
        ws["A"+str(6)].fill =  PatternFill("solid", fgColor= "33353B")
        ws["A"+str(7)].font = Font(bold = True, color= "FDFFFF")
        ws["A"+str(7)].fill =  PatternFill("solid", fgColor= "33353B")
        ws["A"+str(8)].font = Font(bold = True, color= "FDFFFF")
        ws["A"+str(8)].fill =  PatternFill("solid", fgColor= "33353B")


        ws["B"+str(5)] = cantidad_cuentas
        ws["B"+str(6)] = otros_datos[0]
        ws["B"+str(7)] = otros_datos[1]
        ws["B"+str(8)] = otros_datos[2]

        ws["B"+str(6)].number_format = '#,##0.00_-"M3"'
        ws["B"+str(7)].number_format = '#,##0.00_-"M3"'
        ws["B"+str(8)].number_format = '#,##0.00_-"M3"'

        for dato in datos_segundos:

            if cont == 10:
                ws = wb.active
                ws["A"+str(cont)] = "FECHA"
                ws["B"+str(cont)] = "TOTAL $"
                ws["C"+str(cont)] = "TOTAL Hº"
                ws["D"+str(cont)] = "TOTAL LINK $"
                ws["E"+str(cont)] = "TOTAL LINK Hª"


                ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont)].alignment = Alignment(horizontal = "center")

                ws["A"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["B"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["C"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["D"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["E"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["A"+str(cont)].fill =  PatternFill("solid", fgColor= "33353B")
                ws["B"+str(cont)].fill =  PatternFill("solid", fgColor= "33353B")
                ws["C"+str(cont)].fill =  PatternFill("solid", fgColor= "33353B")
                ws["D"+str(cont)].fill =  PatternFill("solid", fgColor= "33353B")
                ws["E"+str(cont)].fill =  PatternFill("solid", fgColor= "33353B")


                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15


                ws["A"+str(cont+1)] = dato[0][0][0]
                ws["B"+str(cont+1)] = dato[1]
                ws["C"+str(cont+1)] = dato[2]
                ws["D"+str(cont+1)] = dato[3]
                ws["E"+str(cont+1)] = dato[4]

                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["C"+str(cont+1)].number_format = '#,##0.00_-"M3"'
                ws["E"+str(cont+1)].number_format = '#,##0.00_-"M3"'
                ws["D"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].alignment = Alignment(horizontal = "center")


                cont += 1

            else: 

                ws = wb.active
                ws["A"+str(cont+1)] = dato[0][0][0]
                ws["B"+str(cont+1)] = dato[1]
                ws["C"+str(cont+1)] = dato[2]
                ws["D"+str(cont+1)] = dato[3]
                ws["E"+str(cont+1)] = dato[4]

                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["C"+str(cont+1)].number_format = '#,##0.00_-"M3"'
                ws["E"+str(cont+1)].number_format = '#,##0.00_-"M3"'
                ws["D"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].alignment = Alignment(horizontal = "center")

                cont += 1
        cont = 1


        for proyect in listado:

            #En la primera vuelta sacamos el totalizado

            cantidad_cuentas = len(CuentaCorriente.objects.filter(venta__proyecto = proyect))
            
            #Establecemos un rango para hacer el cash de ingreso
            
            fecha_inicial_hoy = datetime.date.today()

            fecha_inicial_2 = datetime.date(fecha_inicial_hoy.year, fecha_inicial_hoy.month, 1)

            fechas = []

            #Aqui buscamos la ultima fecha

            fecha_ultima = Cuota.objects.filter(cuenta_corriente__venta__proyecto = proyect).order_by("-fecha")

            contador = 0
            
            contador_year = 1

            fecha_cargar = fecha_inicial_2

            while fecha_cargar < fecha_ultima[0].fecha:

                if (fecha_inicial_2.month + contador) == 13:
                    
                    year = fecha_inicial_2.year + contador_year
                    
                    fecha_cargar = date(year, 1, fecha_inicial_2.day)

                    fechas.append(fecha_cargar)
                    
                    contador_year += 1

                    contador = - (12 - contador)

                else:

                    mes = fecha_inicial_2.month + contador

                    year = fecha_inicial_2.year + contador_year - 1

                    fecha_cargar = date(year, mes, fecha_inicial_2.day)

                    fechas.append(fecha_cargar)

                contador += 1

            #Aqui hacemos los totalizadores generales

            cuotas_anteriores = Cuota.objects.filter(fecha__lt = fecha_inicial_hoy, cuenta_corriente__venta__proyecto = proyect)
            pagos_anteriores = Pago.objects.filter(fecha__lt = fecha_inicial_hoy, cuota__cuenta_corriente__venta__proyecto = proyect)
            cuotas_posteriores = Cuota.objects.filter(fecha__gt = fecha_inicial_hoy, cuenta_corriente__venta__proyecto = proyect)

            total_original = 0
            total_cobrado = 0
            total_pendiente = 0
            total_acobrar= 0

            for cuo in cuotas_anteriores:

                total_original = total_original + cuo.precio

            for pag in pagos_anteriores:

                total_cobrado = total_cobrado +  pag.pago

            for cuot in cuotas_posteriores:

                total_acobrar= total_acobrar + cuot.precio

            total_pendiente = total_original - total_cobrado

            otros_datos = [total_cobrado, total_pendiente, total_acobrar]
        
            #Aqui buscamos agrupar proyecto - sumatorias de cuotas y pagos - mes
            
            datos_segundos = []

            total_fecha = []

            fecha_inicial = 0

            for f in fechas:

                total = 0
                total_link = 0

                datos_terceros = []

                if fecha_inicial == 0:

                        fecha_inicial = fecha_inicial_hoy

                else:

                    cuotas = Cuota.objects.filter(fecha__range = (fecha_inicial, f), cuenta_corriente__venta__proyecto = proyect)
                        
                    pagos = Pago.objects.filter(fecha__range = (fecha_inicial, f), cuota__cuenta_corriente__venta__proyecto = proyect)

                    total_cuotas = 0
                    total_cuotas_link = 0
                    total_pagado = 0
                    total_pagado_link = 0
                    saldo = 0
                    saldo_link = 0

                    if len(cuotas)>0:

                        for c in cuotas:

                            total_cuotas = total_cuotas + c.precio*c.constante.valor

                            if c.cuenta_corriente.venta.unidad.asig == "HON. LINK" or c.cuenta_corriente.venta.unidad.asig == "TERRENO":

                                total_cuotas_link = total_cuotas_link + c.precio*c.constante.valor 

                    if len(pagos)>0:

                        for p in pagos:

                            total_pagado = total_pagado + p.pago*p.cuota.constante.valor

                            if p.cuota.cuenta_corriente.venta.unidad.asig == "HON. LINK" or p.cuota.cuenta_corriente.venta.unidad.asig == "TERRENO":

                                total_pagado_link = total_pagado_link + p.pago*p.cuota.constante.valor 

                    saldo = total_cuotas-total_pagado

                    total = total + saldo

                    #Aqui calculamos el saldo de cuotas de LINK

                    saldo_link = total_cuotas_link-total_pagado_link

                    total_link = total_link + saldo_link
                    
                    datos_terceros.append((fecha_inicial, saldo, saldo_link))

                    fecha_inicial = f

                    horm = Constantes.objects.get(nombre = "Hº VIVIENDA")
                    
                    total_horm = total/horm.valor

                    total_horm_link = total_link/horm.valor

                    datos_segundos.append((datos_terceros, total, total_horm, total_link, total_horm_link))


            #Aqui termina la primera vuelta -> Total por proyecto

            cont = 10

            ws = wb.create_sheet("My sheet")
            ws.title = "{0}".format(proyect.nombre)
            ws["A"+str(2)] = "RESUMEN MES A MES DE INGRESOS POR CUOTAS A COBRAR"
            ws["A"+str(5)] = "CUENTAS"
            ws["A"+str(6)] = "COBRADO"
            ws["A"+str(7)] = "ADEUDADO"
            ws["A"+str(8)] = "PENDIENTE"

            ws["A"+str(2)].font = Font(bold = True)
            ws["A"+str(5)].font = Font(bold = True, color= "FDFFFF")
            ws["A"+str(5)].fill =  PatternFill("solid", fgColor= "33353B")
            ws["A"+str(6)].font = Font(bold = True, color= "FDFFFF")
            ws["A"+str(6)].fill =  PatternFill("solid", fgColor= "33353B")
            ws["A"+str(7)].font = Font(bold = True, color= "FDFFFF")
            ws["A"+str(7)].fill =  PatternFill("solid", fgColor= "33353B")
            ws["A"+str(8)].font = Font(bold = True, color= "FDFFFF")
            ws["A"+str(8)].fill =  PatternFill("solid", fgColor= "33353B")


            ws["B"+str(5)] = cantidad_cuentas
            ws["B"+str(6)] = otros_datos[0]
            ws["B"+str(7)] = otros_datos[1]
            ws["B"+str(8)] = otros_datos[2]

            ws["B"+str(6)].number_format = '#,##0.00_-"M3"'
            ws["B"+str(7)].number_format = '#,##0.00_-"M3"'
            ws["B"+str(8)].number_format = '#,##0.00_-"M3"'

            for dato in datos_segundos:

                if cont == 10:
                    ws = wb["{0}".format(proyect.nombre)]
                    ws["A"+str(cont)] = "FECHA"
                    ws["B"+str(cont)] = "TOTAL $"
                    ws["C"+str(cont)] = "TOTAL Hº"
                    ws["D"+str(cont)] = "TOTAL LINK $"
                    ws["E"+str(cont)] = "TOTAL LINK Hª"


                    ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                    ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                    ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                    ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                    ws["E"+str(cont)].alignment = Alignment(horizontal = "center")

                    ws["A"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                    ws["B"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                    ws["C"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                    ws["D"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                    ws["E"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                    ws["A"+str(cont)].fill =  PatternFill("solid", fgColor= "33353B")
                    ws["B"+str(cont)].fill =  PatternFill("solid", fgColor= "33353B")
                    ws["C"+str(cont)].fill =  PatternFill("solid", fgColor= "33353B")
                    ws["D"+str(cont)].fill =  PatternFill("solid", fgColor= "33353B")
                    ws["E"+str(cont)].fill =  PatternFill("solid", fgColor= "33353B")


                    ws.column_dimensions['A'].width = 15
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['E'].width = 15


                    ws["A"+str(cont+1)] = dato[0][0][0]
                    ws["B"+str(cont+1)] = dato[1]
                    ws["C"+str(cont+1)] = dato[2]
                    ws["D"+str(cont+1)] = dato[3]
                    ws["E"+str(cont+1)] = dato[4]

                    ws["A"+str(cont+1)].font = Font(bold = True)
                    ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                    ws["B"+str(cont+1)].number_format = '"$"#,##0.00_-'
                    ws["C"+str(cont+1)].number_format = '#,##0.00_-"M3"'
                    ws["E"+str(cont+1)].number_format = '#,##0.00_-"M3"'
                    ws["D"+str(cont+1)].number_format = '"$"#,##0.00_-'
                    ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                    ws["E"+str(cont+1)].alignment = Alignment(horizontal = "center")


                    cont += 1

                else: 

                    ws = wb["{0}".format(proyect.nombre)]
                    ws["A"+str(cont+1)] = dato[0][0][0]
                    ws["B"+str(cont+1)] = dato[1]
                    ws["C"+str(cont+1)] = dato[2]
                    ws["D"+str(cont+1)] = dato[3]
                    ws["E"+str(cont+1)] = dato[4]

                    ws["A"+str(cont+1)].font = Font(bold = True)
                    ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                    ws["B"+str(cont+1)].number_format = '"$"#,##0.00_-'
                    ws["C"+str(cont+1)].number_format = '#,##0.00_-"M3"'
                    ws["E"+str(cont+1)].number_format = '#,##0.00_-"M3"'
                    ws["D"+str(cont+1)].number_format = '"$"#,##0.00_-'
                    ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                    ws["E"+str(cont+1)].alignment = Alignment(horizontal = "center")

                    cont += 1
            cont = 1

        #Establecer el nombre del archivo
        nombre_archivo = "ResumenCuentaCorriente.xls"
        
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo).replace(',', '_')
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response