import io
from random import sample
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Proveedores, Certificados
from .models import StockComprasAnticipadas, Compras, Proyectos, Proveedores, Retiros, Comparativas, ComparativasMensaje, Contratos, AdjuntosContratos
from rrhh.models import datosusuario
from users.models import VariablesGenerales
from .form import StockAntForm
from .filters import CertificadoFilter
from presupuestos.models import Articulos, Constantes, Capitulos
import sqlite3
import operator
import datetime
import smtplib
import requests
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import dateutil.parser
from agenda import settings
from datetime import date
from dateutil.relativedelta import relativedelta
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side 
from django.views.generic.base import TemplateView 
from rest_framework.generics import ListAPIView
from .functions_comparativas import mensajeCierreOc, mandarEmail
from .funciones.f_g_mandar_email import *
from django.db.models import Q


def contratos(request):

    data = Contratos.objects.all()

    return render(request, 'contratos.html', {"data":data})

def contratosdescripcion(request, id_contrato):

    if request.method == "POST":

        try:
            b = AdjuntosContratos(
                contrato = Contratos.objects.get(id = id_contrato),
                nombre = request.POST['nombre'],
                adjunto = request.FILES['adjunto'],
                fecha_c = request.POST['fecha'],
            )
            b.save()

        except:
            pass

    data = Contratos.objects.get(id = id_contrato)

    adjuntos = AdjuntosContratos.objects.filter(contrato = data)

    pagos = Comparativas.objects.filter(contrato = data)

    return render(request, 'contratosdescripcion.html', {"data":data, "adjuntos":adjuntos, "pagos":pagos})

def editarcomparativas(request, id_comp):

    proveedores = Proveedores.objects.all()
    contratos = Contratos.objects.all()

    comparativa = Comparativas.objects.get(id = id_comp)

    if request.method == "POST":

        try:
            comparativa = Comparativas.objects.get(id = request.POST['borrar'])
            comparativa.delete()

        except:

            comparativa.proveedor = Proveedores.objects.get(name=request.POST['proveedor'])
            comparativa.proyecto = request.POST['proyecto']
            comparativa.numero  = request.POST['referencia']
            comparativa.monto = float(request.POST['valor'])
            comparativa.o_c = request.POST['numerooc']
            comparativa.autoriza = request.POST['autoriza']
            comparativa.publica = request.POST['publica']
            comparativa.tipo_oc = request.POST['tipo_oc']
            if request.POST['gerente'] != "":
                comparativa.gerente_autoriza = datosusuario.objects.get(identificacion = request.POST['gerente']) 
            else:
                comparativa.gerente_autoriza = None
            try:
                comparativa.contrato = Contratos.objects.get(id=request.POST['contrato'])
            except:
                comparativa.contrato = None
            try:
                comparativa.adjunto = request.FILES['imagen']
                comparativa.save()
            except:
                comparativa.save()
            
            try:
                comparativa.adj_oc = request.FILES['oc']
                comparativa.save()
            except:
                pass

            if comparativa.estado == "NO AUTORIZADA":

                comparativa.estado = "ESPERA"
                comparativa.save()
        
        return redirect(f'/compras/comparativas/{20}/{0}/{0}#{comparativa.id}')

    context = {}
    context['contratos'] = contratos
    context['proveedores'] = proveedores
    context['comparativa'] = comparativa
    context['gerentes'] = datosusuario.objects.filter(cargo = "GERENTE").exclude(estado = "NO ACTIVO")
    context['monto_minimo'] = VariablesGenerales.objects.get(id = 1).monto_minimo

    return render(request, 'comparativas_editar.html', context)

def funcionstock():

    compras = Compras.objects.filter(tipo = "ANT")
    retiros = Retiros.objects.all()   

    for i in compras:

        for c in retiros:

            if i.nombre == c.compra.nombre:

                if i.articulo == c.articulo:

                    i.cantidad = i.cantidad - c.cantidad
    stock = []
    
    for i in compras:

        if not i.cantidad == 0:
            
            stock.append(i)
    
    return stock

def detalleinforme(request, fecha_i, fecha_f, proyecto):

    proyecto = Proyectos.objects.get(id = proyecto)
    fecha_inicial = datetime.date(year = int(fecha_i[0:4]), month=int(fecha_i[4:6]), day=int(fecha_i[6:8]))
    fecha_final = datetime.date(year = int(fecha_f[0:4]), month=int(fecha_f[4:6]), day=int(fecha_f[6:8]))

    datos = Compras.objects.filter(fecha_c__range=(fecha_inicial, fecha_final), proyecto = proyecto)

    datos_compra = []

    for d in datos:

        if d.precio_presup != 0 or d.precio_presup != None:
            
            datos_compra.append((d, d.cantidad*d.precio, (d.cantidad*d.precio_presup - d.cantidad*d.precio)))

    datos_compra = sorted(datos_compra, key=lambda x: x[2], reverse=True)

    datos = {"datos_compra":datos_compra,
    "fecha_inicial":fecha_inicial,
    "fecha_final":fecha_final,
    "proyecto":proyecto}

    return render(request, 'detalle_informe.html', {"datos":datos})

def informecompras(request):

    datos = 0

    if request.method == "POST":

        datos = request.POST.items()

        for dato in datos:

            if dato[0] == "fechainicial":
                fechainicial = dato[1]

            if dato[0] == "fechafinal":
                fechafinal = dato[1]

        datos_compra = Compras.objects.filter(fecha_c__range=(fechainicial, fechafinal))

        cantidad_doc = []
        proyectos = []

        #Listado d rubros

        materiales_electricos = 0
        materiales_electricos_estimado = 0
        materiales_sanitarios = 0
        materiales_sanitarios_esimado = 0
        materiales_pintura = 0
        materiales_pintura_esimado = 0


        monto_total = 0

        monto_estimado = 0

        for d in datos_compra:

            if "FONDO DE REPARO" in str(d.articulo.nombre) or "ANTICIPO" in str(d.articulo.nombre) or "DIANCO" in str(d.proyecto.nombre):
                basura = 1

            else:

                cantidad_doc.append((d.proyecto, d.proveedor, d.documento))
                proyectos.append(d.proyecto)
                monto_total = monto_total + d.precio*d.cantidad
                monto_estimado = monto_estimado + d.precio_presup*d.cantidad

                #Listado de los rubros mas importantes

                if "30900" in str(d.articulo.codigo):
                    materiales_electricos = materiales_electricos + d.precio*d.cantidad
                    materiales_electricos_estimado = materiales_electricos_estimado + d.precio_presup*d.cantidad

                
                if "31400" in str(d.articulo.codigo):
                    materiales_sanitarios = materiales_sanitarios + d.precio*d.cantidad
                    materiales_sanitarios_esimado = materiales_sanitarios_esimado + d.precio_presup*d.cantidad

                if "30700" in str(d.articulo.codigo):
                    materiales_pintura = materiales_pintura + d.precio*d.cantidad
                    materiales_pintura_esimado = materiales_pintura_esimado + d.precio_presup*d.cantidad

        
        #Aqui terminamos de armar la lista

        materiales_rubros = []

        materiales_rubros.append(("Materiales electricos", materiales_electricos, materiales_electricos_estimado, 0))
        materiales_rubros.append(("Materiales sanitarios", materiales_sanitarios, materiales_sanitarios_esimado, 0))
        materiales_rubros.append(("Pintura y afines", materiales_pintura, materiales_pintura_esimado, 0))
        
        
        cantidad_doc = len(set(cantidad_doc))

        datos_proyecto = []

        lista_proyectos = set(proyectos)

        for proyecto in lista_proyectos:

            monto_mat_p = 0
            monto_mo_p = 0
            monto_total_p = 0

            monto_mat_p_est = 0
            monto_mo_p_est = 0
            monto_total_p_est = 0

            for d in datos_compra:

                if proyecto == d.proyecto:

                    if "FONDO DE REPARO" in str(d.articulo.nombre) or "ANTICIPO" in str(d.articulo.nombre):
                        basura = 1

                    else:

                        monto_total_p = monto_total_p + d.cantidad*d.precio
                        monto_total_p_est= monto_total_p_est + d.cantidad*d.precio_presup

                    if str(d.articulo.codigo)[0] == "3":
                        monto_mat_p = monto_mat_p + d.cantidad*d.precio
                        monto_mat_p_est = monto_mat_p_est + d.cantidad*d.precio_presup

                    else:

                        if "FONDO DE REPARO" in str(d.articulo.nombre) or "ANTICIPO" in str(d.articulo.nombre):

                            basura = 1

                        else:

                            monto_mo_p = monto_mo_p + d.cantidad*d.precio
                            monto_mo_p_est = monto_mo_p_est + d.cantidad*d.precio_presup


            


            ahorro_total_p = monto_total_p_est - monto_total_p
            ahorro_mat_p = monto_mat_p_est - monto_mat_p
            ahorro_mo_p = monto_mo_p_est - monto_mo_p

            datos_proyecto.append((proyecto, monto_total_p, monto_total_p_est, monto_mat_p, monto_mat_p_est, monto_mo_p, monto_mo_p_est, ahorro_total_p, ahorro_mat_p, ahorro_mo_p ))


        try:
            diferencia = (monto_total/monto_estimado-1)*100
            diferencia_plata = monto_estimado - monto_total

        except:
            diferencia = 0
            diferencia_plata = 0

        cantidad_compras = len(datos_compra)

        fecha_i = str(fechainicial[0:4])+str(fechainicial[5:7])+str(fechainicial[8:10])
        fecha_f = str(fechafinal[0:4])+str(fechafinal[5:7])+str(fechafinal[8:10])

        datos = {"cantidad_compras":cantidad_compras, "cantidad_doc":cantidad_doc, "monto_total":monto_total,
        "fechafinal":fechafinal, "fechainicial":fechainicial, "monto_estimado":monto_estimado,
        "datos_proyecto":datos_proyecto, "diferencia":diferencia, "diferencia_plata":diferencia_plata,
        "materiales_rubros":materiales_rubros,
        "fecha_i":fecha_i, "fecha_f":fecha_f}

    return render(request, 'informe_compra_semana.html', {"datos":datos})

def listaretiros(request):

    datos = Retiros.objects.all()

    return render(request, 'retirolist.html', {"datos":datos})

def comprasdisponibles(request):

    if request.method == "POST":

        datos = request.POST.items()

        for i in datos:
        
            if i[0] == "compra":

                compra_c = i

                if i[1] != "":

                    stock = funcionstock()

                    compra = []

                    for c in stock:

                        if c.nombre == i[1]:
                            
                            compra.append(c)
                    
                    return render(request, 'retiros.html', {"compra":compra})
            
            elif i[0] == "documento":

                pass


    lista = []

    compras = Compras.objects.all()

    for i in compras:

        if i.tipo == "ANT":

            lista.append((i.nombre, i.proyecto, i.proveedor, i.documento))
   
    datos = list(set(lista))

    return render(request, 'retiros.html', {"datos":datos})

def principalautorizacion(request):
    # Saludo de bienvenida
    hora_actual = datetime.datetime.now()
    if hora_actual.hour >= 20:
        mensaje_bievenida = "¡Buenas noches {}!".format(request.user.first_name)
    elif hora_actual.hour >= 13:
        mensaje_bievenida = "¡Buenas tardes {}!".format(request.user.first_name)
    else:
        mensaje_bievenida = "¡Buen dia {}!".format(request.user.first_name)

    # Titulo del div principal
    oc_pend_sp = Comparativas.objects.filter(autoriza = "SP").exclude(estado = "AUTORIZADA")
    op_pend_sp = Comparativas.objects.filter(fecha_c__gte = "2021-02-01").exclude(visto = "VISTO", estado = "AUTORIZADA")
    cant_oc_pend_sp = len(oc_pend_sp)
    cant_op_pend_sp = 0
    for i in op_pend_sp:
        if not i.autoriza == "PL" and not i.publica == "NO":
            cant_op_pend_sp += 1

    if cant_oc_pend_sp > 0 and cant_op_pend_sp > 0:
        mensaje = "Hay mucho trabajo por hacer!"
    elif cant_oc_pend_sp > 0:
        mensaje = "Algunas OC no estan listas, podrias revisarlas"
    elif cant_op_pend_sp > 0:
        mensaje = "Las OC están listas pero algunas OP quedaron en el tintero"
    else:
        mensaje = "No tienes pendientes por aqui, pero siempre es bueno revisar"
    return render(request, "oc_principal_autorizacion.html", {"mensaje_bievenida":mensaje_bievenida, "mensaje":mensaje})

def ocautorizargerente1(request, estado, creador):

    if request.method == 'POST':

        datos_post = request.POST.items()
        id_selec = 0

        for d in datos_post:

            if d[0] == 'APROBADA':

                id_selec = d[1]
                comparativa = Comparativas.objects.get(id = id_selec)
                comparativa.estado = "AUTORIZADA"
                comparativa.visto = "VISTO"
                comparativa.quien_autorizo = request.user.username
                
                date = datetime.datetime.now() - datetime.timedelta(hours=3)
                comparativa.fecha_autorizacion = date
                comparativa.save()

                try:

                    recibe = datosusuario.objects.get(identificacion = comparativa.creador).email
                    subject = f"Tu O.C {comparativa.o_c} para {comparativa.proveedor.name} esta autorizada!"
                    mandar_email(1, recibe, subject)

                except:

                    pass

            if d[0] == 'NO APROBADA':
                id_selec = d[1]
                comparativa = Comparativas.objects.get(id = id_selec)
                comparativa.estado = "NO AUTORIZADA"
                comparativa.save()
                
                try:
                    recibe = datosusuario.objects.get(identificacion = comparativa.creador).email
                    subject = f"Atención! La OC {comparativa.o_c} para {comparativa.proveedor.name} fue rechazada!"
                    mandar_email(2, recibe, subject)

                except:

                    pass

    context = {}
    con_principal = Comparativas.objects.filter(autoriza = "SP")
    dic_estados = {
        '0':'Todas',
        '1':'Autorizada',
        '2': 'No autorizada',
        '3': 'Espera',
        '4': 'Adjunto ✓',
    }

    estado_selec = dic_estados[estado]

    # Datos de las cantidades de cada uno

    cant_todas = len(con_principal)
    cant_espera = len(con_principal.filter(estado = dic_estados['3'].upper()))
    cant_autorizada = len(con_principal.filter(estado = dic_estados['1'].upper()))
    cant_no_autorizada = len(con_principal.filter(estado = dic_estados['2'].upper()))
    cant_adjunto = len(con_principal.filter(estado = dic_estados['4'].upper()))

    cant_oc_sp = {
        "cant_todas":cant_todas,
        "cant_espera":cant_espera,
        "cant_autorizada":cant_autorizada,
        "cant_no_autorizada":cant_no_autorizada,
        "cant_adjunto":cant_adjunto,
    }

    context["cant_oc_sp"] = cant_oc_sp
    context["estado_selec"] = estado_selec
    context["estado"] = estado

    if estado == "0":
        con_filtro_estado = con_principal.order_by("-fecha_c")
    else:
        con_filtro_estado = con_principal.filter(estado = estado_selec.upper()).order_by("-fecha_c")

    context["creadores"] = con_principal.exclude(creador = "").values_list("creador", flat = True).distinct()
    context["creadores"].order_by("creador")

    if creador == "0":
        context["estado_creador"] = "Creador"
    else:
        context["estado_creador"] = creador

    context["creador"] = creador

    if creador != "0":
        con_filtro_estado = con_filtro_estado.filter(creador = creador)
    
    datos_render = []

    for d in con_filtro_estado:
        mensajes = ComparativasMensaje.objects.filter(comparativa = d)

        if d.creador:
            usuario = datosusuario.objects.get(identificacion = d.creador)

        else:
            usuario = 0

        datos_render.append((usuario, mensajes, d))

    context["datos"] = datos_render

    return render(request, "comparativas/comparativa_SP.html", context)

def panelvisto(request, estado, creador):

    if request.method == 'POST':

        datos_post = request.POST.items()

        id_selec = 0

        for d in datos_post:

            if d[0] == 'APROBADA':
                id_selec = d[1]

                comparativa = Comparativas.objects.get(id = id_selec)

                comparativa.visto = "VISTO"

                comparativa.save()


            if d[0] == 'ADJAPROB':

                id_selec = d[1]
                comparativa = Comparativas.objects.get(id = id_selec)

                comparativa.visto = "VISTO NO CONFORME"

                comparativa.save()

                recibe = datosusuario.objects.get(identificacion = comparativa.creador).email
                subject = f"Tu OC {comparativa.o_c} de {comparativa.proveedor.name} esta observada por SP"
                mandar_email(3, recibe, subject)

    context = {}

    con_principal = Comparativas.objects.all()

    dic_estados = {
        '0':'Todas',
        '1':'Visto',
        '2': 'No_visto',
        '3': 'Visto no conforme',
    }

    cant_todas = con_principal.count()
    cant_vistas= con_principal.filter(fecha_c__gte = "2021-02-01", estado = "AUTORIZADA", visto = dic_estados['1'].upper()).exclude(Q(autoriza = "PL") & Q(publica = "NO")).count()
    cant_no_vistas = con_principal.filter(fecha_c__gte = "2021-02-01", estado = "AUTORIZADA", visto = dic_estados['2'].upper()).exclude(Q(autoriza = "PL") & Q(publica = "NO")).count()
    cant_no_conforme = con_principal.filter(fecha_c__gte = "2021-02-01", estado = "AUTORIZADA", visto = dic_estados['3'].upper()).exclude(Q(autoriza = "PL") & Q(publica = "NO")).count()

    cant_oc_sp = {
        "cant_todas":cant_todas,
        "cant_vistas":cant_vistas,
        "cant_no_vistas":cant_no_vistas,
        "cant_no_conforme":cant_no_conforme,

    }

    context["cant_oc_sp"] = cant_oc_sp

    estado_selec = dic_estados[estado]
    
    datos = 0

    if estado == "0":

        con_filtro_estado = con_principal.filter(fecha_c__gte = "2021-02-01", estado = "AUTORIZADA").order_by("-fecha_c")

    else:
    
        con_filtro_estado = con_principal.filter(visto = estado_selec.upper(), estado = "AUTORIZADA", fecha_c__gte = "2021-02-01").order_by("-fecha_c")

    context["creadores"] = con_principal.exclude(creador = "").values_list("creador", flat = True).distinct()
    context["creadores"].order_by("creador")

    if creador == "0":
        context["estado_creador"] = "Creador"
    else:
        context["estado_creador"] = creador

    context["creador"] = creador

    if creador != "0":

        con_filtro_estado = con_filtro_estado.filter(creador = creador)

    context['estado_selec'] = estado_selec
    context['estado'] = estado

    datos = []

    for d in con_filtro_estado:

        if not (d.autoriza == "PL" and d.publica == "NO"):

            mensajes = ComparativasMensaje.objects.filter(comparativa = d)

            if d.creador:
                usuario = datosusuario.objects.get(identificacion = d.creador)

            else:
                usuario = 0

            datos.append((usuario, mensajes, d))

    context['datos'] = datos

    return render(request, 'comparativas/comparativa_check.html', context)

def mensajescomparativas(request, id_comparativa):

    if request.method == 'POST':

        datos_post = request.POST.items()

        for i in datos_post:

            if i[0] == "mensaje" and i[1] != "" :

                b = ComparativasMensaje(
                        usuario = datosusuario.objects.get(identificacion = request.user),
                        comparativa = Comparativas.objects.get(id = id_comparativa),
                        mensaje = i[1],

                        )

                b.save()

                if "@Pablo" in b.mensaje or "@pablo" in b.mensaje:

                    # Prueba Telegram

                    send = "{} te respondio: '{}' en la OC {}".format(b.usuario.nombre, b.mensaje, b.comparativa.o_c)

                    id = "-585663986"

                    token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"

                    url = "https://api.telegram.org/bot" + token + "/sendMessage"

                    params = {
                        'chat_id' : id,
                        'text' : send
                    }

                    requests.post(url, params=params)


    datos = Comparativas.objects.get(id = id_comparativa)

    mensajes = ComparativasMensaje.objects.filter(comparativa__id = id_comparativa).order_by("fecha")

    return render(request, 'mensajescomparativas.html', {'datos':datos, 'mensajes':mensajes})

def descargacomparativas(request):

    if request.method == 'POST':

        fechai = str(request.POST['fechai'])[0:4]+str(request.POST['fechai'])[5:7]+str(request.POST['fechai'])[8:]
        fechaf = str(request.POST['fechaf'])[0:4]+str(request.POST['fechaf'])[5:7]+str(request.POST['fechaf'])[8:]

        return redirect('Descargar estado', fechai = fechai, fechaf = fechaf)

    return render(request, 'descargacom.html')

def comparativas(request, estado, creador, autoriza):

    context = {}

    if 10 <= int(estado) < 20:
        mensaje = 1
        estado = str(int(estado) - 10)

    elif 20 <= int(estado) < 30:

        mensaje = 2
        estado = str(int(estado) - 20)

    else:

        estado = estado

    # Consultas necesarias

    con_comparativas = Comparativas.objects.all()

    usuarios=datosusuario.objects.all()
    list_autoriza = usuarios.filter(Q(identificacion='PL') | Q(identificacion='SP') | Q(cargo = "GERENTE"))

    # Codigo para fecha de pagos

    fecha_inicial = datetime.date.today()
    fecha_pago = datetime.date(2021, 4, 16)
    mensaje_PL_SP='Autoriza'

    while fecha_pago <= fecha_inicial:
        fecha_pago = fecha_pago + datetime.timedelta(days=14)
   
    if creador == "0":
        mensaje_creador = "Creador"

    else:

        mensaje_creador = datosusuario.objects.get(id = creador).identificacion

    if request.method == 'POST':

        datos_post = request.POST.dict()

        if 'visto_bueno_gerente' in datos_post:

            try:

                comparativa_modificar = Comparativas.objects.get(id = int(request.POST['visto_bueno_gerente']))
                comparativa_modificar.visto_gerente = True
                comparativa_modificar.save()

                context['mensaje_accion'] = [1, "Todo listo!"]

            except:

                context['mensaje_accion'] = [0, "Error inesperado"]

        datos_post = request.POST.items()

        id_selec = 0

        for d in datos_post:

            if d[0] == 'APROBADA':

                id_selec = d[1]

                comparativa = Comparativas.objects.get(id = id_selec)

                comparativa.estado = "AUTORIZADA"

                if request.user.username == "SP":

                    comparativa.visto = "VISTO"

                if comparativa.publica == "NO":

                    comparativa.visto = "VISTO"

                # El servidor no esta ubicado en el mismo lugar que los trabajadores, por lo cual debo ajustarlo

                date = datetime.datetime.now() - datetime.timedelta(hours=3)

                comparativa.fecha_autorizacion = date

                comparativa.save()

                try:

                    mandarEmail(comparativa, 1)

                    if comparativa.creador == "AT" or comparativa.creador == "LG":

                        send = "Han aprobado la OC {} de {}".format(comparativa.creador, comparativa.o_c)

                        id = "-455382561"

                        token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"

                        url = "https://api.telegram.org/bot" + token + "/sendMessage"

                        params = {
                            'chat_id' : id,
                            'text' : send
                        }

                        requests.post(url, params=params)

                except:

                    pass

            if d[0] == 'NO APROBADA':
                id_selec = d[1]

                comparativa = Comparativas.objects.get(id = id_selec)

                comparativa.estado = "NO AUTORIZADA"

                comparativa.save()
                
                try:

                    mandarEmail(comparativa, 2)

                    if comparativa.creador == "AT" or comparativa.creador == "LG":

                        send = "Han rechazado la OC {} de {}".format(comparativa.creador, comparativa.o_c)

                        id = "-455382561"

                        token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"

                        url = "https://api.telegram.org/bot" + token + "/sendMessage"

                        params = {
                            'chat_id' : id,
                            'text' : send
                        }

                        requests.post(url, params=params)

                except:

                    pass

            if d[0] == 'ADJAPROB':

                id_selec = d[1]

                comparativa = Comparativas.objects.get(id = id_selec)

                comparativa.estado = "ADJUNTO ✓"

                comparativa.save()

            if d[0] == 'MENSAJE':
                

                if d[1] != "":

                    comparativa = Comparativas.objects.get(id = id_selec)

                    comparativa.comentario = str(request.user.username) + ": " + str(d[1])

                    comparativa.save()

                    mensaje = str(d[0]) + ": " + str(d[1])


                    b = ComparativasMensaje(
                            usuario = datosusuario.objects.get(identificacion = request.user),
                            comparativa = Comparativas.objects.get(id = comparativa.id),
                            mensaje = mensaje,

                            )

                    b.save()

                else:

                    comparativa = Comparativas.objects.get(id = id_selec)

                    comparativa.comentario = str((request.user.username) + ": Sin motivo")

                    comparativa.save()


    if estado == "0":
        consulta = con_comparativas
        mensaje_aux = "Totales"

    if estado == "1":
        consulta = con_comparativas.filter(estado = "ESPERA")
        mensaje_aux = "Espera" 

    if estado == "2":
        consulta = con_comparativas.filter(estado = "ADJUNTO ✓")
        mensaje_aux = "Adjunto ✓" 

    if estado == "3":
        consulta = con_comparativas.filter(estado = "NO AUTORIZADA")
        mensaje_aux = "Rechazadas"

    if estado == "4":
        consulta = con_comparativas.filter(estado = "AUTORIZADA")
        mensaje_aux = "Autorizadas"

    if estado == "5":
        consulta = con_comparativas.filter(autoriza = "SP").exclude(estado = "AUTORIZADA")
        mensaje_aux = "Estado SP"

    if estado == "6":
        consulta = con_comparativas.exclude(estado = "AUTORIZADA").exclude(adj_oc = '').order_by("-fecha_c")
        consulta = consulta.exclude(estado = "NO AUTORIZADA")
        mensaje_aux = "Comp con OC"

    if estado == "7":
        consulta = con_comparativas.order_by("-fecha_c")
        mensaje_aux = "Sin filtro"

    if autoriza=='0':
            mensaje_PL_SP = 'Autoriza'
            consult_totales = con_comparativas
            
    else:
        usuario = usuarios.get(pk=autoriza)
        consulta=consulta.filter(Q(autoriza = usuario.identificacion) | Q(gerente_autoriza = usuario.id))
        mensaje_aux = f"{mensaje_aux}: {str(len(consulta))}, Privadas: {len(consulta.filter(publica = 'NO'))}"
        
        consult_totales = con_comparativas.filter(Q(autoriza = usuario.identificacion) | Q(gerente_autoriza = usuario.id))
        mensaje_PL_SP = usuario.identificacion
        
        if estado=='0':
            mensaje_aux=mensaje_aux

    datos_base = consulta.order_by("-fecha_c")
    
    if creador != "0":
        datos_base = consulta.filter(creador = mensaje_creador).order_by("-fecha_c")
        
    creadores = list(set(consulta.values_list('creador').order_by('creador')))   
    list_creadores = []

    for c in creadores:
        try:
            usuario = datosusuario.objects.get(identificacion = c[0])
            list_creadores.append(usuario)

        except:
            None

    datos = []

    for d in datos_base:
        mensajes = ComparativasMensaje.objects.filter(comparativa = d)

        if d.creador:
            usuario = datosusuario.objects.get(identificacion = d.creador)

        else:
            usuario = 0

        datos.append((usuario, mensajes, d))
            
    # Reordenar la lista
  
    list_creadores = sorted(list_creadores, key=lambda creador : creador.identificacion)

    # Recortamos para evitar problemas para renderizar

    if estado != "7":
        datos = datos[0:150]
    
    
    context['mensaje_creador'] = mensaje_creador
    context['list_creadores'] = list_creadores
    context['datos'] = datos
    context['estado'] = estado
    context['creador'] = creador
    context['autoriza'] = autoriza
    context['mensaje'] = mensaje_aux
    context['espera'] = len(consult_totales.filter(estado = "ESPERA"))
    context['autorizada'] = len(consult_totales.filter(estado = "AUTORIZADA"))
    context['rechazada'] = len(consult_totales.filter(estado = "NO AUTORIZADA"))
    context['sin_filtro'] = len(consult_totales)
    context['comparativa_oc'] = len(consult_totales.exclude(estado = "AUTORIZADA").exclude(estado = "NO AUTORIZADA").exclude(adj_oc = ''))
    context['adjunto'] = len(consult_totales.filter(estado = "ADJUNTO ✓"))
    context['fecha_pago'] = fecha_pago
    context['aviso'] = mensajeCierreOc()[0]
    context['fecha_cierre'] = mensajeCierreOc()[1]
    context['mensaje_PL_SP']=mensaje_PL_SP
    context['list_autoriza']=list_autoriza
    context['monto_minimo'] = VariablesGenerales.objects.get(id = 1).monto_minimo

    try:

        if mensaje == 1:
            context['mensaje_s']="Su OC fue cargada correctamente!"

        if mensaje == 2:
            context['mensaje_s']="Su OC se edito correctamente"

    except:
        
        pass

    return render(request, 'comparativas.html', context)

def compras(request, id_proyecto):

    # if id_proyecto == "0":
    #     proyecto = 0
    # else:
    #     proyecto = Proyectos.objects.get(id = id_proyecto)

    # #Aqui armamos un listado

    # if request.user.username == "HC":
    #     proyectos = Proyectos.objects.filter(nombre = "DIANCO - LAMADRID 1137")
    # else:
    #     proyectos = Proyectos.objects.order_by("nombre")

    # if id_proyecto == "0":

    #     if request.user.username == "HC":
    #         datos = Compras.objects.filter(proyecto__nombre = "DIANCO - LAMADRID 1137").order_by("-fecha_c")
    #     else:
    #         datos = Compras.objects.all().order_by("-fecha_c")[0:500]
    # else:
        
    #     datos = Compras.objects.filter(proyecto = id_proyecto).order_by("-fecha_c")

    # compras = []
    
    # #proyectos filtrados por id
    # for dato in datos:
    #     if dato.precio_presup > dato.precio:
    #         total = dato.cantidad*dato.precio
    #         v = (1 - (dato.precio/dato.precio_presup))*100 
    #         compras.append((0,dato, total, -v ))

    #     elif dato.precio_presup == dato.precio:
    #         total = dato.cantidad*dato.precio
    #         compras.append((1,dato, total, 0))

    #     else:
    #         total = dato.cantidad*dato.precio
    #         if dato.precio_presup != 0:
    #             v = -((dato.precio/dato.precio_presup) - 1)*100
    #         else:
    #             v = 0
            # compras.append((2,dato, total, v))


    context = {}
    context['proveedores'] = Proveedores.objects.all()
    context['articulos'] = Articulos.objects.all()
    context['proyectos'] = Proyectos.objects.all()

    return render(request, 'registro_compras/compras_registro.html', context)

def certificados(request):

    datos = Certificados.objects.all()

    myfilter = CertificadoFilter(request.GET, queryset=datos)

    datos = myfilter.qs

    datos_enviados = {'datos':datos, 'myfilter':myfilter}

    return render(request, 'certificados.html', datos_enviados )

# ----------------------------------------------------- VISTAS PARA PROVEEDORES ---------------------------------------------- 


# ----------------------------------------------------- VISTAS STOCK ----------------------------------------------
 
def stockproveedores(request):

    compras = Compras.objects.filter(tipo = "ANT")
    retiros = Retiros.objects.all()   

    for i in compras:

        for c in retiros:

            if i.nombre == c.compra.nombre:

                if i.articulo == c.articulo:

                    i.cantidad = i.cantidad - c.cantidad
    datos = []
    
    for i in compras:

        if not i.cantidad == 0:
            
            datos.append(i)

    #Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        datos_viejos = datos

        datos = []   

        for i in palabra_buscar:

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                lista_palabra = palabra.split()

                buscar = (str(i.proyecto)+str(i.proveedor)+str(i.articulo)+str(i.cantidad))

                contador = 0

                for palabra in lista_palabra:

                    contador2 = 0

                    if palabra.lower() in buscar.lower():
  
                        contador += 1

                if contador == len(lista_palabra):

                    datos.append(i)


    #Aqui termina el filtro

    return render(request, 'stockprov.html', {'datos':datos})

# ----------------------------------------------------- VISTAS PARA ANALISIS DE COMPRA ----------------------------------------------
 
def analisiscompras(request):

    #Traemos los datos de las compras

    datos = Compras.objects.filter(documento__startswith="O")
    proyectos = Proyectos.objects.all()
    proyecto = 0

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        for i in palabra_buscar:
            if i[0] != "csrfmiddlewaretoken":
                pro = i[1]
        datos = Compras.objects.filter(documento__startswith="O", proyecto = pro)

        proyecto = Proyectos.objects.get(id = pro)

    #Establecemos el periodo de tiempo

    inicio_fecha = date.today() - datetime.timedelta(days = 365)

    fechas = []

    contador = 0

    for fecha in range(14):
        fecha_agregar = inicio_fecha + datetime.timedelta(days = ((365*contador/12)))
        fechas.append(fecha_agregar)
        contador +=1 


    fechas_compras = []

    monto_compras = 0
    monto_presupuesto = 0

    contador = 0

    for dato in range(13):

        volumen_comprado = 0
        volumen_presupuesto = 0
        rendimiento = 100

        for compra in datos:

            date_object = datetime.datetime.strptime(str(compra.fecha_c), '%Y-%m-%d')
            date_object = dateutil.parser.parse(str(date_object)).date()

            if date_object >= fechas[contador] and date_object < fechas[contador +1]:
                volumen_comprado = volumen_comprado + (compra.cantidad*compra.precio)/1000
                monto_compras = monto_compras + (compra.cantidad*compra.precio)/1000

                if compra.precio_presup != None:
                    volumen_presupuesto = volumen_presupuesto + (compra.cantidad*compra.precio_presup)/1000
                    monto_presupuesto = monto_presupuesto + (compra.cantidad*compra.precio_presup)/1000
        
        
        if volumen_presupuesto != 0:
            rendimiento = (volumen_comprado/volumen_presupuesto)*100

        fechas_compras.append((fechas[contador], volumen_comprado, volumen_presupuesto, rendimiento))
        contador += 1

    inc_total = (monto_compras/monto_presupuesto)*100

    datos = {"datos":fechas_compras,
    "montocompras":monto_compras,
    "inc":inc_total,
    "proyectos":proyectos,
    "proyecto":proyecto}

    return render(request, 'analisiscompras.html', {"datos":datos} )

# ----------------------------------------------------- VISTAS PARA INFORME ----------------------------------------------
 
def informe(request):

    # --> Modelos necesarios

    compras = Compras.objects.all()
    compras_ant = Compras.objects.filter(tipo = "ANT")
    retiros = Retiros.objects.all() 
    constantes = Constantes.objects.get(nombre="USD")  
    proveedores = Proveedores.objects.all()


    # --> Metodo para calcular la cantidad de compras

    lista_compras = []

    for i in compras:

        lista_compras.append(i.nombre)

    lista_compras = len(list(set(lista_compras)))
    
    compras_nominal = 0
    
    
    # --> Metodo para calcular el valor nominal de las compras
    
    for i in compras:
        
        compras_nominal = compras_nominal + (i.precio*i.cantidad)/1000000

    # --> Metodo para calcular el valor actual de las compras

    compras_actualizado = 0
    
    for i in compras:

        articulo = Articulos.objects.get(codigo = i.articulo.codigo) 
        
        compras_actualizado = compras_actualizado + (articulo.valor*i.cantidad)/1000000
        

    # --> Metodo para calcular el stock

    stock = compras_ant

    for i in stock:

        for c in retiros:

            if i.nombre == c.compra.nombre:

                if i.articulo == c.articulo:

                    i.cantidad = i.cantidad - c.cantidad

    stock_valorizado = 0
    
    # --> Metodo para valorizar el stock

    stock_pesos = 0
    stock_horm = 0
    stock_usd = 0

    for i in stock:

        valor = i.articulo.valor
        cantidad = i.cantidad

        valor_act = valor*cantidad

        if "USD" in str(i.articulo.constante):
            stock_usd = stock_usd + valor_act

        elif "HORMIG" in str(i.articulo.nombre):
            stock_horm = stock_horm + valor_act
        else:
            stock_pesos = stock_pesos + valor_act
        
        stock_valorizado = stock_valorizado + valor_act
        stock_valorizado_m = stock_valorizado/1000000

    stock_pesos = (stock_pesos/1000000)/stock_valorizado_m*100
    stock_horm = (stock_horm/1000000)/stock_valorizado_m*100
    stock_usd = (stock_usd/1000000)/stock_valorizado_m*100

    # --> Metodo para valorizar en USD el stock

    usd = Constantes.objects.get(nombre="USD")

    stock_valorizado_usd = stock_valorizado/usd.valor
    stock_valorizado_usd_m = stock_valorizado_usd/1000000


    # --> Modelo para armar la lista de proveedores/activo
    

    lista = proveedores

    for i in lista:

        i.phone = 0

        for c in stock:

            if i == c.proveedor:

                i.phone = i.phone + (c.articulo.valor*c.cantidad)/1000000

    listas = []
    for i in lista:

        listas.append((str(i.name), float(i.phone)))   

    listas = sorted(listas, key=lambda tup: tup[1], reverse=True)

    # --> Modelo para armar el listado de los articulos

    lista_articulos = Articulos.objects.all()

    for i in lista_articulos:

        i.valor = 0

        for c in stock:

            if i == c.articulo:

                i.valor = i.valor + (c.articulo.valor*c.cantidad)/1000000

    listas_art = []

    for i in lista_articulos:

        listas_art.append((str(i.nombre), float(i.valor)))

    
    listas_art = sorted(listas_art, key=lambda tup: tup[1], reverse=True)

    # --> Modelo para armar el stock por fideicomiso

    lista_proyectos = Proyectos.objects.all()

    for i in lista_proyectos:

        i.m2 = 0

        for c in stock:

            if i == c.proyecto:

                i.m2 = i.m2 + (c.articulo.valor*c.cantidad)/1000000

    listas_pro = []

    for i in lista_proyectos:

        if i.m2 != 0:

            listas_pro.append((str(i.nombre), float(i.m2)))

    
    listas_pro = sorted(listas_pro, key=lambda tup: tup[1], reverse=True)

    # --> Comprasas Nominales por fideicomiso

    compras_fidei = []

    for i in lista_proyectos:
        datos_compras = Compras.objects.filter(proyecto = i)
        
        valor_nominal_compras = 0

        for dato in datos_compras:
            if str(dato.articulo.nombre) != "FONDO DE REPARO ACT. UOCRA" and str(dato.articulo.nombre) != "ANTICIPO FINANCIERO ACT. UOCRA":
                valor_nominal_compras = valor_nominal_compras + dato.precio*dato.cantidad

        if valor_nominal_compras != 0:
        
            compras_fidei.append((i, valor_nominal_compras))

    
    datos = {"stock_valorizado":stock_valorizado,
    "stock_valorizado_m":stock_valorizado_m,
    "stock_valorizado_usd":stock_valorizado_usd,
    "stock_valorizado_usd_m":stock_valorizado_usd_m,
    "lista_compras":lista_compras,
    "constantes":constantes,
    "listas":listas,
    "listas_art":listas_art,
    "listas_pro":listas_pro,
    "compras_nominal":compras_nominal,
    "compras_actualizado":compras_actualizado,
    "stock_pesos":stock_pesos,
    "stock_usd":stock_usd,
    "stock_horm":stock_horm,
    "compras_fidei":compras_fidei

    }

    return render(request, 'stockant.html', {'datos': datos})

# ----------------------------------------------------- VISTAS PARA CARGA DE RETIROS----------------------------------------------

def cargaretiro(request, nombre_compra):

    compras = Compras.objects.all()

    for i in compras:
        if str(i.nombre) == "29032019.PUERTAS":
            pass

    return render(request, 'cargaretiro.html',)

# ----------------------------------------------------- VISTAS PARA CARGA DE RETIROS----------------------------------------------

class Reegistrodecompras(TemplateView):

    def get(self, request, *args, **kwargs):
        
        wb = Workbook()

        #Aqui coloco la formula para calcular

        datos = Compras.objects.order_by("fecha_c")

        ws = wb.active
        ws.title = "ADVERTENCIA"

        ws["A2"] = "UNA ADVERTENCIA ANTES DE AVANZAR"

        ws["A2"].alignment = Alignment(horizontal = "left")
        ws["A2"].font = Font(bold = True, color= "23346D", size = 20)

        ws.merge_cells("A5:K25")
        ws["A5"] = """
        La información que contiene este documento se considera de caracter CONFIDENCIAL.
        
         Esto quiere decir debes garantizar su protección y no debe ser divulgada sin el consentimiento de Link Inversiones S.R.L.
         
         Algunas recomendaciones:

         1 - Habla con tu responsable de área antes de pasar este documento
         2 - Si usas este documento fuera de las computadoras de la empresa, borra el archivo y vacia la papelera
         
         Gracias por tu atención

         Saludos!
         """

        ws["A5"].alignment = Alignment(horizontal = "left", vertical = "center", wrap_text=True)
        ws["A5"].font = Font(bold = True)
        
        cont = 1
        
        for d in datos:

            if cont == 1:
                ws = wb.create_sheet("My sheet")
                ws.title = "Registrodecompras"
                ws["A"+str(cont)] = "PROYECTO"
                ws["B"+str(cont)] = "ARTICULO"
                ws["C"+str(cont)] = "UNIDAD"
                ws["D"+str(cont)] = "VALOR"
                ws["E"+str(cont)] = "CANTIDAD"
                ws["F"+str(cont)] = "PROVEEDOR"
                ws["G"+str(cont)] = "FECHA"
                ws["H"+str(cont)] = "DOCUMENTO"


                ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont)].alignment = Alignment(horizontal = "center")


                ws["A"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["A"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["B"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["B"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["C"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["C"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["D"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["D"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["E"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["E"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["F"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["F"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["G"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["G"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["H"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["H"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")


                ws.column_dimensions['A'].width = 23.29
                ws.column_dimensions['B'].width = 63.86
                ws.column_dimensions['C'].width = 7.57
                ws.column_dimensions['D'].width = 12.14
                ws.column_dimensions['E'].width = 18.57
                ws.column_dimensions['F'].width = 46.71
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 31

                ws["A"+str(cont+1)] = d.proyecto.nombre
                ws["B"+str(cont+1)] = d.articulo.nombre
                ws["C"+str(cont+1)] = d.articulo.unidad
                ws["D"+str(cont+1)] = d.precio
                ws["E"+str(cont+1)] = d.cantidad
                ws["F"+str(cont+1)] = d.proveedor.name
                ws["G"+str(cont+1)] = d.fecha_c
                ws["H"+str(cont+1)] = d.documento


                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["E"+str(cont+1)].number_format = '#,##0.00_-'
                ws["F"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].alignment = Alignment(horizontal = "center")
  

                cont += 1

            else:
                ws = wb["Registrodecompras"]

                ws["A"+str(cont+1)] = d.proyecto.nombre
                ws["B"+str(cont+1)] = d.articulo.nombre
                ws["C"+str(cont+1)] = d.articulo.unidad
                ws["D"+str(cont+1)] = d.precio
                ws["E"+str(cont+1)] = d.cantidad
                ws["F"+str(cont+1)] = d.proveedor.name
                ws["G"+str(cont+1)] = d.fecha_c
                ws["H"+str(cont+1)] = d.documento


                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["E"+str(cont+1)].number_format = '#,##0.00_-'
                ws["F"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].alignment = Alignment(horizontal = "center")


                cont += 1

        #Establecer el nombre del archivo
        nombre_archivo = "Registrodecompras.xls"
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class CompOCestado(TemplateView):

    def get(self, request, fechai, fechaf, *args, **kwargs):
        
        wb = Workbook()

        #Aqui coloco la formula para calcular

        fecha_inicial = datetime.date(int(fechai[0:4]), int(fechai[4:6]), int(fechai[6:8]))
        fecha_final = datetime.date(int(fechaf[0:4]), int(fechaf[4:6]), int(fechaf[6:8]))

        datos = Comparativas.objects.filter(fecha_c__gte=fecha_inicial, fecha_c__lte=fecha_final,).order_by("-fecha_c")

        ws = wb.active
        ws.title = "ADVERTENCIA"

        ws["A2"] = "UNA ADVERTENCIA ANTES DE AVANZAR"

        ws["A2"].alignment = Alignment(horizontal = "left")
        ws["A2"].font = Font(bold = True, color= "23346D", size = 20)

        ws.merge_cells("A5:K25")
        ws["A5"] = """
        La información que contiene este documento se considera de caracter CONFIDENCIAL.
        
         Esto quiere decir debes garantizar su protección y no debe ser divulgada sin el consentimiento de Link Inversiones S.R.L.
         
         Algunas recomendaciones:

         1 - Habla con tu responsable de área antes de pasar este documento
         2 - Si usas este documento fuera de las computadoras de la empresa, borra el archivo y vacia la papelera
         
         Gracias por tu atención

         Saludos!
         """

        ws["A5"].alignment = Alignment(horizontal = "left", vertical = "center", wrap_text=True)
        ws["A5"].font = Font(bold = True)
        cont = 5
        
        for d in datos:

            if cont == 5:
                ws = wb.create_sheet("My sheet")
                ws.title = "Listado"

                ws["A2"] = "FECHA INICIAL"
                ws["A3"] = "FECHA FINAL"
                ws["A2"].font = Font(bold = True)
                ws["A3"].font = Font(bold = True)

                ws["B2"] = fecha_inicial
                ws["B3"] = fecha_final
                ws["B2"].font = Font(bold = True)
                ws["B2"].alignment = Alignment(horizontal = "center")
                ws["B3"].font = Font(bold = True)
                ws["B3"].alignment = Alignment(horizontal = "center")

                ws["A"+str(cont)] = "FECHA"
                ws["B"+str(cont)] = "CREADOR"
                ws["C"+str(cont)] = "PROYECTO"
                ws["D"+str(cont)] = "PROVEEDOR"
                ws["E"+str(cont)] = "REFERENCIA"
                ws["F"+str(cont)] = "OC"
                ws["G"+str(cont)] = "MONTO"
                ws["H"+str(cont)] = "GERENTE"
                ws["I"+str(cont)] = "DIRECTOR"
                ws["J"+str(cont)] = "ESTADO"
                ws["K"+str(cont)] = "AUTORIZO"
                ws["L"+str(cont)] = "VISTO SP"
                ws["M"+str(cont)] = "AUTORIZADO EL"
                ws["N"+str(cont)] = "OBSERVACIONES"

                ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["I"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["J"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["K"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["L"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["M"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["N"+str(cont)].alignment = Alignment(horizontal = "center")


                ws["A"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["A"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["B"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["B"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["C"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["C"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["D"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["D"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["E"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["E"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["F"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["F"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["G"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["G"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["H"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["H"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["I"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["I"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["J"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["J"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["K"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["K"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["L"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["L"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["M"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["M"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["N"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["N"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")


                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 40
                ws.column_dimensions['E'].width = 40
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 20
                ws.column_dimensions['I'].width = 20
                ws.column_dimensions['J'].width = 20
                ws.column_dimensions['K'].width = 20
                ws.column_dimensions['L'].width = 30
                ws.column_dimensions['M'].width = 40
                ws.column_dimensions['N'].width = 60

                ws["A"+str(cont+1)] = d.fecha_c
                ws["B"+str(cont+1)] = d.creador
                ws["C"+str(cont+1)] = d.proyecto
                ws["D"+str(cont+1)] = d.proveedor.name
                ws["E"+str(cont+1)] = d.numero
                ws["F"+str(cont+1)] = d.o_c
                ws["G"+str(cont+1)] = d.monto

                if d.gerente_autoriza:
                    ws["H"+str(cont+1)] = d.gerente_autoriza.identificacion
                else:
                    ws["H"+str(cont+1)] = "No asignado"

                ws["I"+str(cont+1)] = d.autoriza
                ws["J"+str(cont+1)] = d.estado
                
                if d.estado == 'AUTORIZADA':
                    ws["K"+str(cont+1)] = d.quien_autorizo
                else:
                    ws["K"+str(cont+1)] = "Sin autorizar"
                
                ws["L"+str(cont+1)] = d.visto
                
                if d.fecha_autorizacion:
                    ws["M"+str(cont+1)] = str(d.fecha_autorizacion)
                
                ws["N"+str(cont+1)] = d.comentario
                
                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].font = Font(bold = True)
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont+1)].font = Font(bold = True)
                ws["G"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["H"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["I"+str(cont+1)].alignment = Alignment(horizontal = "center")

                ws["J"+str(cont+1)].alignment = Alignment(horizontal = "center")
                if d.estado == "AUTORIZADA":

                    ws["J"+str(cont+1)].font = Font(bold = True, color= "236D25")

                elif d.estado == "ADJUNTO ✓":

                    ws["J"+str(cont+1)].font = Font(bold = True, color= "B86914")

                elif d.estado == "NO AUTORIZADA":

                    ws["J"+str(cont+1)].font = Font(bold = True, color= "99221D")

                ws["K"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["L"+str(cont+1)].alignment = Alignment(horizontal = "center")

                if d.visto == "VISTO":

                    ws["L"+str(cont+1)].font = Font(bold = True, color= "236D25")

                elif d.visto == "VISTO NO CONFORME":

                    ws["L"+str(cont+1)].font = Font(bold = True, color= "B86914")

                elif d.visto == "NO_VISTO":

                    ws["L"+str(cont+1)].font = Font(bold = True, color= "99221D")
                    
                ws["M"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["N"+str(cont+1)].alignment = Alignment(horizontal = "left")
  

                cont += 1

            else:
                ws = wb["Listado"]

                ws["A"+str(cont+1)] = d.fecha_c
                ws["B"+str(cont+1)] = d.creador
                ws["C"+str(cont+1)] = d.proyecto
                ws["D"+str(cont+1)] = d.proveedor.name
                ws["E"+str(cont+1)] = d.numero
                ws["F"+str(cont+1)] = d.o_c
                ws["G"+str(cont+1)] = d.monto

                if d.gerente_autoriza:
                    ws["H"+str(cont+1)] = d.gerente_autoriza.identificacion
                else:
                    ws["H"+str(cont+1)] = "No asignado"

                ws["I"+str(cont+1)] = d.autoriza
                ws["J"+str(cont+1)] = d.estado
                
                if d.estado == 'AUTORIZADA':
                    ws["K"+str(cont+1)] = d.quien_autorizo
                else:
                    ws["K"+str(cont+1)] = "Sin autorizar"
                
                ws["L"+str(cont+1)] = d.visto

                if d.fecha_autorizacion:
                    ws["M"+str(cont+1)] = str(d.fecha_autorizacion)

                ws["N"+str(cont+1)] = d.comentario


                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].font = Font(bold = True)
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont+1)].font = Font(bold = True)
                ws["G"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["H"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["I"+str(cont+1)].alignment = Alignment(horizontal = "center")

                ws["J"+str(cont+1)].alignment = Alignment(horizontal = "center")
                if d.estado == "AUTORIZADA":

                    ws["J"+str(cont+1)].font = Font(bold = True, color= "236D25")

                elif d.estado == "ADJUNTO ✓":

                    ws["J"+str(cont+1)].font = Font(bold = True, color= "B86914")

                elif d.estado == "NO AUTORIZADA":

                    ws["J"+str(cont+1)].font = Font(bold = True, color= "99221D")

                ws["K"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["L"+str(cont+1)].alignment = Alignment(horizontal = "center")

                if d.visto == "VISTO":

                    ws["L"+str(cont+1)].font = Font(bold = True, color= "236D25")

                elif d.visto == "VISTO NO CONFORME":

                    ws["L"+str(cont+1)].font = Font(bold = True, color= "B86914")

                elif d.visto == "NO_VISTO":

                    ws["L"+str(cont+1)].font = Font(bold = True, color= "99221D")
                    
                ws["M"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["N"+str(cont+1)].alignment = Alignment(horizontal = "left")

                cont += 1

        #Establecer el nombre del archivo
        nombre_archivo = "EstadoOcComparativas.xls"
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
