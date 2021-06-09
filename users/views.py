import os
from django.shortcuts import render, redirect
from django.contrib.auth import logout as do_logout
from django.contrib.auth import authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login as do_login
from django.contrib.auth.forms import UserCreationForm
from django.http import HttpResponse
from django.views.generic.base import TemplateView 
from django.views.generic import View
from django.template.loader import get_template 
from finanzas.models import Almacenero, RegistroAlmacenero, Arqueo, RetirodeSocios, Honorarios
from presupuestos.models import Presupuestos, InformeMensual, TareasProgramadas, Bitacoras
from proyectos.models import Proyectos, Unidades
from ventas.models import VentasRealizadas
from compras.models import Compras, Comparativas
from registro.models import RegistroValorProyecto
from rrhh.models import datosusuario, mensajesgenerales, NotaDePedido, Vacaciones, MonedaLink, EntregaMoneda, Anuncios, Seguimiento, Minutas, Acuerdos, PremiosMonedas, Logros, RegistroContable, CanjeMonedas, Sugerencia, DicRegistroContable
import datetime
import requests
from datetime import date
import pandas as pd
import numpy as np
from django.contrib.auth.models import User
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from agenda import settings
from django.contrib.auth import models
from statistics import mode
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class PdfMinutas(View):

    def link_callback(self, uri, rel):
            """
            Convert HTML URIs to absolute system paths so xhtml2pdf can access those
            resources
            """
            sUrl = settings.STATIC_URL        # Typically /static/
            sRoot = settings.STATIC_ROOT      # Typically /home/userX/project_static/
            mUrl = settings.MEDIA_URL         # Typically /media/
            mRoot = settings.MEDIA_ROOT       # Typically /home/userX/project_static/media/

            if uri.startswith(mUrl):
                path = os.path.join(mRoot, uri.replace(mUrl, ""))
            elif uri.startswith(sUrl):
                path = os.path.join(sRoot, uri.replace(sUrl, ""))
            else:
                return uri

            # make sure that file exists
            if not os.path.isfile(path):
                raise Exception(
                        'media URI must start with %s or %s' % (sUrl, mUrl)
                )
            return path

    def get(self, request, id_minuta, *args, **kwargs):

        #Creamos la información

        minuta = Minutas.objects.get(id = id_minuta)

        acuerdos = Acuerdos.objects.filter(minuta = minuta)

        # Aqui llamamos y armamos el PDF
      
        template = get_template('minutas/pdfminuta.html')
        contexto = {'minuta':minuta, 
        'acuerdos':acuerdos, 
        'fecha':datetime.date.today(),
        'logo':'{}{}'.format(settings.STATIC_URL, 'img/Linkp.png'),
        'cabecera':'{}{}'.format(settings.STATIC_URL, 'img/fondo.png'),
        'fondo':'{}{}'.format(settings.STATIC_URL, 'img/fondo.jpg')}
        html = template.render(contexto)
        response = HttpResponse(content_type = "application/pdf")
        
        #response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'
        
        pisaStatus = pisa.CreatePDF(html, dest=response, link_callback=self.link_callback)
        
        if pisaStatus.err:
            
            return HttpResponse("Hay un error")

        return response

def linkp(request):

    try:
        user = datosusuario.objects.get(identificacion = request.user.username)

        if len(Logros.objects.filter(usuario = user, nombre = "Curioso")) == 0:

            b = Logros(
                usuario = user,
                nombre = "Curioso",
                descrip = "Encontraste una pagina secreta ;)"
            )

            b.save()
    except:
        pass

    return render(request, 'users/linkp.html')

def sugerencias(request):

    if request.method == 'POST':

        try:

            sugerencia_selec = Sugerencia.objects.get(id = int(request.POST['id']))

            sugerencia_selec.nombre = request.POST['nombre']
            sugerencia_selec.descripcion = request.POST['descripcion']

            try:
                sugerencia_selec.adjunto = request.FILES['adjunto']
                sugerencia_selec.save()
            except:
                sugerencia_selec.save()
           
        except:
            pass

        try:

            usuario = datosusuario.objects.get(identificacion = request.user.username)

            b = Sugerencia(
                usuario = usuario,
                nombre = request.POST['nombre'],
                descripcion = request.POST['descripcion'],
            )

            b.save()

            try:
                b.adjunto = request.FILES['adjunto']
            except:
                pass
        except:
            pass
        try:
            today  = date.today()
            sugerencia = Sugerencia.objects.get(id = int(request.POST['ENTREGADO']))
            sugerencia.estado = "LISTO"
            sugerencia.fecha_listo = today
            sugerencia.save()
        except:
            pass

    data = Sugerencia.objects.all().order_by("-id")


    return render (request, 'users/sugerencias.html', {'data':data})

def monedalink(request):

    usuario = datosusuario.objects.get(identificacion = request.user)

    if request.method == 'POST':

        monedas = MonedaLink.objects.filter(usuario_portador = usuario)

        monedas_disponibles = []

        for m in monedas:

            if len(EntregaMoneda.objects.filter(moneda = m)) == 0:

                monedas_disponibles.append(m)

        index_num = 0

        for i in range(int(request.POST["cantidad"])):

            b = EntregaMoneda(
                moneda = monedas_disponibles[index_num],
                usuario_recibe = datosusuario.objects.get(id = int(request.POST["usuario"])),
                mensaje = request.POST["mensaje"])
                

            b.save()

            index_num += 1

        try:

            # Establecemos conexion con el servidor smtp de gmail
            mailServer = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
            mailServer.ehlo()
            mailServer.starttls()
            mailServer.ehlo()
            mailServer.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

            # Construimos el mensaje simple
            
            mensaje = MIMEText("""
            
            Recibiste una moneda!,

            "{}"

            - {}

            """.format(b.mensaje, request.user.username))
            mensaje['From']=settings.EMAIL_HOST_USER
            mensaje['To']=b.usuario_recibe.email
            mensaje['Subject']="Recibiste una moneda!!"


            # Envio del mensaje

            mailServer.sendmail(settings.EMAIL_HOST_USER,
                            b.usuario_recibe.email,
                            mensaje.as_string())

        except:

            pass


    list_usuarios = datosusuario.objects.all().exclude(identificacion = request.user)

    monedas = MonedaLink.objects.filter(usuario_portador = usuario)

    monedas_disponibles = 0

    for m in monedas:

        if len(EntregaMoneda.objects.filter(moneda = m)) == 0:

            monedas_disponibles += 1

    monedas_recibidas = len(EntregaMoneda.objects.filter(usuario_recibe = usuario))
    recibidas_list = EntregaMoneda.objects.filter(usuario_recibe = usuario).values_list("mensaje", flat = True)

    recibidas_list = list(set(recibidas_list))

    recibidas = []

    for r in recibidas_list:

        data = EntregaMoneda.objects.filter(usuario_recibe = usuario, mensaje = r)

        usuarios_entrega = ""

        for d in data:

            if str(d.moneda.usuario_portador.identificacion) not in usuarios_entrega:
                usuarios_entrega = usuarios_entrega + str(d.moneda.usuario_portador.identificacion) + ""

        recibidas.append((len(data), r, usuarios_entrega))
  
    return render(request, 'users/monedaslink.html', {"recibidas":recibidas, "monedas_recibidas":monedas_recibidas, "monedas_disponibles":monedas_disponibles, "list_usuarios":list_usuarios})

def vacaciones(request):

    # Determinamos las fechas

    hoy = datetime.date.today()

    if hoy.month < 4:

        abril = datetime.date(hoy.year, 4, 1)
        marzo = datetime.date(hoy.year, 3, 1)
        febrero = datetime.date(hoy.year, 2, 1)
        enero = datetime.date(hoy.year, 1, 1)
        diciembre = datetime.date((hoy.year - 1), 12, 1)
        noviembre = datetime.date((hoy.year - 1), 11, 1)
        octubre = datetime.date((hoy.year - 1), 10, 1)

        fechas = [octubre, noviembre, diciembre, enero, febrero, marzo, abril]

        datos = Vacaciones.objects.filter(fecha_inicio__gte = octubre, fecha_final__lte = abril)
           
    return render(request, "users/holidays.html", {'fechas':fechas, 'datos':datos})

def password(request):

    mensaje = 0

    if request.method == 'POST':

        usuario = User.objects.get(id = request.user.id)
        
        usuario.set_password(request.POST["password"])

        usuario.save()

        mensaje = 1

    return render(request, "users/password.html", {'mensaje':mensaje})

def guia(request):

    amor = 0
    rey = 0
    otros_datos = 0
    monedas_disponibles_canje = 0

    try:
        usuario = datosusuario.objects.get(identificacion = request.user.username)

    except:

        usuario = 0

    print("Complete bien la busqueda del usuario")

    if request.method == 'POST':

        monedas = MonedaLink.objects.filter(usuario_portador = usuario)

        monedas_disponibles = []

        for m in monedas:

            if len(EntregaMoneda.objects.filter(moneda = m)) == 0:

                monedas_disponibles.append(m)
        
        index_num = 0

        for i in range(int(request.POST["cantidad"])):

            b = EntregaMoneda(
                moneda = monedas_disponibles[index_num],
                usuario_recibe = datosusuario.objects.get(id = int(request.POST["usuario"])),
                mensaje = request.POST["mensaje"])
                

            b.save()

            index_num += 1

        try:

            # Establecemos conexion con el servidor smtp de gmail
            mailServer = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
            mailServer.ehlo()
            mailServer.starttls()
            mailServer.ehlo()
            mailServer.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

            # Construimos el mensaje simple
            
            mensaje = MIMEText("""
            
            Recibiste una moneda!,

            "{}"

            - {}

            """.format(b.mensaje, request.user.username))
            mensaje['From']=settings.EMAIL_HOST_USER
            mensaje['To']=b.usuario_recibe.email
            mensaje['Subject']="Recibiste una moneda!!"


            # Envio del mensaje

            mailServer.sendmail(settings.EMAIL_HOST_USER,
                            b.usuario_recibe.email,
                            mensaje.as_string())

        except:

            pass

    try:

        list_usuarios = datosusuario.objects.all().exclude(identificacion = request.user.username).order_by("identificacion").exclude(estado = "NO ACTIVO")

        ########################################
        # Calculo de monedasentregadas (Listado)
        ########################################
        
        list_user_unique = EntregaMoneda.objects.filter(moneda__usuario_portador__identificacion = request.user.username).values_list("usuario_recibe__id", flat = True).distinct()

        info_coins_entregadas = []

        for user in list_user_unique:

            coins_entregadas = datosusuario.objects.get(id = user)
            coins_cantidad = len(EntregaMoneda.objects.filter(moneda__usuario_portador__identificacion = request.user.username, usuario_recibe__id = user))
            info_coins_entregadas.append((coins_entregadas, coins_cantidad))
        
        ########################################
        # Calculo de monedas disponibles para dar
        ########################################

        monedas = MonedaLink.objects.filter(usuario_portador = usuario)

        monedas_disponibles = 0
        
        for m in monedas:

            if len(EntregaMoneda.objects.filter(moneda = m)) == 0:

                monedas_disponibles += 1

        print("Complete bien la parte de calcular monedas")


        ########################################
        # Precio por DAR
        ########################################

        if len(monedas) == monedas_disponibles:
            amor = 0
        else:
            amor = 1

        ########################################
        # Premio al puesto numero 1 y 2
        ########################################

        rey_l = EntregaMoneda.objects.all().values_list("usuario_recibe", flat = True)

        try:
            if int(usuario.id) == int(mode(rey_l)):
                rey = 1

                
            rey_2 = EntregaMoneda.objects.all().values_list("usuario_recibe", flat = True).exclude(usuario_recibe__id = int(mode(rey_l)))

            if int(usuario.id) == int(mode(rey_2)):
                rey = 2
        except:
            rey = 0

        print("Complete bien la parte de premios")

        ########################################
        # Calculo de monedas recibidas 
        ########################################
  
        monedas_recibidas = len(EntregaMoneda.objects.filter(usuario_recibe = usuario))
        monedas_disponibles_canje = monedas_recibidas - sum(CanjeMonedas.objects.filter(usuario = usuario).values_list("monedas", flat=True))
        recibidas_list = EntregaMoneda.objects.filter(usuario_recibe = usuario).values_list("mensaje", flat = True)

        recibidas_list = list(set(recibidas_list))

        recibidas = []
   
        for r in recibidas_list:

            data = EntregaMoneda.objects.filter(usuario_recibe = usuario, mensaje = r)

            usuarios_entrega = ""

            for d in data:

                if str(d.moneda.usuario_portador.identificacion) not in usuarios_entrega:
                    usuarios_entrega = usuarios_entrega + str(d.moneda.usuario_portador.identificacion) + ""

            recibidas.append((len(data), r, usuarios_entrega))

        datos = 0

        otros_datos = 0
 
    except:
        recibidas = 0
        monedas_recibidas = 0
        monedas_disponibles = 0
        list_usuarios = 0

    try:
        datos = datosusuario.objects.get(identificacion = request.user)

        if datos:

            areas = datosusuario.objects.values_list("area").exclude(estado = "NO ACTIVO")

            areas = list(set(areas))

            otros_datos = []

            for a in areas:

                miembros = datosusuario.objects.filter(area = a[0]).order_by("identificacion").exclude(estado = "NO ACTIVO")

                otros_datos.append((a, miembros))

    except:

        datos = 0

    try:

        usuario = datosusuario.objects.get(identificacion = request.user)

    except:
        usuario = 0

    monedas_recibidas = len(EntregaMoneda.objects.filter(usuario_recibe = usuario))

    ########################################
    # Logros Argentino
    ########################################

    argentino = len(EntregaMoneda.objects.filter(moneda__usuario_portador__identificacion = request.user.username, mensaje__icontains = "bolud"))


    ########################################
    # Logros
    ########################################

    try:
        logros = Logros.objects.filter(usuario = datosusuario.objects.get(identificacion = request.user))
    
    except:
        logros = 0  

    return render(request, "users/guia.html", {"argentino":argentino, "logros":logros, "rey":rey, "amor":amor, "datos":datos, "otros_datos":otros_datos, "recibidas":recibidas, "monedas_recibidas":monedas_recibidas, "monedas_disponibles":monedas_disponibles, "monedas_disponibles_canje":monedas_disponibles_canje, "list_usuarios":list_usuarios, "info_coins_entregadas":info_coins_entregadas})

def canjemonedas(request):

    mensaje = ""

    usuario = datosusuario.objects.get(identificacion = request.user)

    monedas_recibidas = len(EntregaMoneda.objects.filter(usuario_recibe = usuario))

    monedas_canjear = monedas_recibidas - sum(CanjeMonedas.objects.filter(usuario = usuario).values_list("monedas", flat=True))

    if request.method == 'POST':

        try:

            today = date.today()

            if today.day <= 10:

                rrhh = "am@linkinversiones.com.ar"

                premio_solicitado = PremiosMonedas.objects.get(id = int(request.POST['premio']))

                if monedas_canjear >= premio_solicitado.cantidad:

                    canje = CanjeMonedas(
                        usuario = usuario,
                        fecha = datetime.date.today(),
                        premio = str(premio_solicitado.nombre),
                        monedas = int(premio_solicitado.cantidad),
                    )

                    canje.save()

                    # Establecemos conexion con el servidor smtp de gmail
                    mailServer = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
                    mailServer.ehlo()
                    mailServer.starttls()
                    mailServer.ehlo()
                    mailServer.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

                    # Construimos el mensaje simple
                    
                    mensaje = MIMEText("""
                
Hola!,

{} acaba de canjear Linkcoins, el premio es {} que costó {} monedas.

Podrás visualizarlo en el panel de seguimiento. Cualquier duda, comunicate con el equipo de IT.

Saludos!

-- Link-Help 


                    
                    """.format(usuario, canje.premio, canje.monedas))
                    mensaje['From']=settings.EMAIL_HOST_USER
                    mensaje['To']= rrhh
                    mensaje['Subject']="{} realizo un canje".format(usuario)


                    # Envio del mensaje

                    mailServer.sendmail(settings.EMAIL_HOST_USER,
                                    rrhh,
                                    mensaje.as_string())

                    # Establecemos conexion con el servidor smtp de gmail
                    mailServer = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
                    mailServer.ehlo()
                    mailServer.starttls()
                    mailServer.ehlo()
                    mailServer.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

                    # Construimos el mensaje simple
                    
                    mensaje = MIMEText("""
                    
¡Hola!,

Acabas canjear {}  Linkcoins por el siguiente premio: {}.

El equipo de RRHH te notificará cuando el mismo esté disponible para retirarlo (esta gestión puede tomar hasta 10 días hábiles posteriores a la fecha límite de canje).

Si hubiera algún problema del sistema, comunicate con el equipo de IT para solucionarlo.

Saludos,

-- Link-Help 

                
                    """.format(canje.monedas, canje.premio))
                    mensaje['From']=settings.EMAIL_HOST_USER
                    mensaje['To']= usuario.email
                    mensaje['Subject']="Realizaste un canje de Linkcoins".format(usuario)


                    # Envio del mensaje

                    mailServer.sendmail(settings.EMAIL_HOST_USER,
                                    usuario.email,
                                    mensaje.as_string())


                    return redirect('Canje de monedas')

                else:
                    mensaje = "No tienes suficientes monedas para canjear el premio seleccionado"

            else:

                mensaje = "Solo puedes canjear hasta el dia 10 de cada mes"

        except:

            pass

        try:

            if request.POST['id'] != "0":
                premio = PremiosMonedas.objects.get(id = int(request.POST['id']))
                premio.nombre = request.POST['nombre']
                premio.cantidad = request.POST['cantidad']
                premio.save()

            

            else:
                b = PremiosMonedas(
                    nombre = request.POST['nombre'],
                    cantidad = int(request.POST['cantidad']),
                )

                b.save()

        except:

            pass

        try:

            premio = PremiosMonedas.objects.get(id = int(request.POST['borrar']))
            premio.delete()


        except:

            pass


    premios = PremiosMonedas.objects.all().order_by("nombre")

    monedas = MonedaLink.objects.filter(usuario_portador = usuario)

    monedas_recibidas = len(EntregaMoneda.objects.filter(usuario_recibe = usuario))

    monedas_canjear = monedas_recibidas - sum(CanjeMonedas.objects.filter(usuario = usuario).values_list("monedas", flat=True))

    monedas_disponibles = 0
    
    for m in monedas:

        if len(EntregaMoneda.objects.filter(moneda = m)) == 0:

            monedas_disponibles += 1

    dato_monedas = {'monedas_recibidas':monedas_recibidas, 'monedas_disponibles':monedas_disponibles, 'monedas':monedas, 'monedas_canjear':monedas_canjear}
  
    return render(request, "users/canjemonedas.html", {'mensaje':mensaje, "premios":premios, 'dato_monedas':dato_monedas})

def canjerealizados(request):

    if request.method == 'POST':

        canje = CanjeMonedas.objects.get(id = int(request.POST['ENTREGADO']))
        canje.entregado = "SI"
        canje.save()

    data = CanjeMonedas.objects.all()

    
    return render(request, "users/canjesrealizados.html", {'data':data})

def dashboard(request):

    # En esta parte resolvemos el grafico del arqueo

    data_cruda = Arqueo.objects.order_by("-fecha")

    fecha = data_cruda[0].fecha

    data = data_cruda[0]

    data_frame = pd.read_excel(data.arqueo)

    array_usd = np.array(data_frame['USD'])

    usd = sum(array_usd)

    array_euro = np.array(data_frame['EUROS'])

    euro = sum(array_euro)

    nombre_columnas = data_frame.columns

    banco = 0

    for n in nombre_columnas:

        if "BANCO" in n:

            banco = banco + sum(np.array(data_frame[n]))


    array_pesos = np.array(data_frame['EFECTIVO'])

    pesos = sum(array_pesos)

    array_cheques = np.array(data_frame['CHEQUES'])

    cheques = sum(array_cheques)

    array_me = np.array(data_frame['MONEDA EXTRANJERA'])

    me = sum(array_me)

    consolidado = pesos + cheques + me + banco

    porcentaje_pesos = pesos/consolidado*100
    porcentaje_banco = banco/consolidado*100
    porcentaje_cheques = cheques/consolidado*100
    porcentaje_me = me/consolidado*100

    arqueo = [fecha, consolidado, porcentaje_pesos, porcentaje_me, porcentaje_cheques, porcentaje_banco, usd, euro]


    #Calculos para tablero de avance de presupuesto

    barras = []

    datos_barras = Presupuestos.objects.order_by("-saldo").exclude(proyecto__nombre = "DIANCO - LAMADRID 1137")

    for db in datos_barras:

        if db.valor != 0:

            avance = (100 - db.saldo/db.valor*100)

            barras.append((db, int(avance)))

    barras = sorted(barras,reverse=True, key=lambda tup: tup[1])


    #Calculos para tablero de ventas

    proyectos = Proyectos.objects.all()

    ventas_barras = []

    for p in proyectos:

        total_unidades = Unidades.objects.filter(proyecto = p, asig = "PROYECTO")

        unidades_vendidas = Unidades.objects.filter(proyecto = p, asig = "PROYECTO").exclude(estado = "DISPONIBLE")

        if len(total_unidades) != 0:

            avance = (len(unidades_vendidas)/len(total_unidades))*100

            ventas_barras.append((p, int(avance)))

    ventas_barras = sorted(ventas_barras, reverse=True, key=lambda tup: tup[1])

    date = datetime.date.today()

    fecha_inicio = date - datetime.timedelta(days=90)

    ventas_realizadas = len(VentasRealizadas.objects.filter(fecha__gt = fecha_inicio))


    #Calculo para compras

    compras = Compras.objects.filter(fecha_c__gt = fecha_inicio)

    comprado = 0
    estimado = 0

    for c in compras:
        comprado = comprado + c.cantidad*c.precio
        estimado = estimado + c.cantidad*c.precio_presup

    if estimado == 0:

        diferencia = 0

    else:

        diferencia = (comprado/estimado-1)*100

    datos_compras = [comprado, estimado, diferencia]


    #Calculo para unidades

    deptos_disp = len(Unidades.objects.filter(estado = "DISPONIBLE", tipo = "DEPARTAMENTO", asig = "PROYECTO"))
    cocheras_disp = len(Unidades.objects.filter(estado = "DISPONIBLE", tipo = "COCHERA", asig = "PROYECTO"))

    datos_unidades = [deptos_disp, cocheras_disp]


    # -----------------> Aqui calculo indice LINK

    datos = Almacenero.objects.all()

    datos_completos = []
    datos_finales = []

    saldo_caja_total = 0
    pendiente_gastar_total = 0
    ingresos_total = 0
    descuento_total = 0

    for dato in datos:

        presupuesto = "NO"

        pricing = "NO"

        almacenero = dato

        presupuesto = Presupuestos.objects.get(proyecto = dato.proyecto)

        # Aqui calculo el IVA sobre compras

        iva_compras = (presupuesto.imprevisto + presupuesto.saldo_mat + presupuesto.saldo_mo + presupuesto.credito + presupuesto.fdr)*0.07875

        almacenero.pendiente_iva_ventas = iva_compras

        almacenero.save()

        # Calculo el resto de las cosas

        retiro_socios = sum(np.array(RetirodeSocios.objects.values_list('monto_pesos').filter(proyecto = dato.proyecto)))
        saldo_caja = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados
        saldo_caja_total = saldo_caja_total + saldo_caja
        pend_gast = almacenero.pendiente_admin + almacenero.pendiente_comision + presupuesto.saldo_mat + presupuesto.saldo_mo + presupuesto.imprevisto + presupuesto.credito + presupuesto.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos
        pendiente_gastar_total = pendiente_gastar_total + pend_gast
        prest_cobrar = almacenero.prestamos_proyecto + almacenero.prestamos_otros
        total_ingresos = prest_cobrar + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + saldo_caja
        ingresos_total = ingresos_total + total_ingresos
        margen = total_ingresos - pend_gast
        descuento = almacenero.ingreso_ventas*0.06
        descuento_total = descuento_total + descuento
        margen_2 = margen - descuento 
               
        datos_completos.append((dato, saldo_caja, pend_gast, total_ingresos, margen, descuento, margen_2))

    # -----------------> Aqui calculo los totalizadores

    margen1_total = ingresos_total - pendiente_gastar_total
    margen2_total = margen1_total - descuento_total


    # -----------------> Aqui calculo la parte de honorarios

    honorarios = Honorarios.objects.order_by("-fecha")
    caja_actual = honorarios[0].caja_actual
    subtotal_1 = honorarios[0].cuotas + honorarios[0].ventas
    ingresos = subtotal_1 + honorarios[0].creditos
    comision = honorarios[0].comision_venta*honorarios[0].ventas
    subtotal_2 = honorarios[0].estructura_gio + honorarios[0].aportes + honorarios[0].socios + comision
    costos = subtotal_2  + honorarios[0].deudas
    honorario = ingresos - costos + honorarios[0].caja_actual
    honorarios2 = honorario

    # -----------------> Aqui calculo los totalizadores con los honorarios

    caja_total = caja_actual + saldo_caja_total
    costos_totales = pendiente_gastar_total + costos
    ingresos_totales = ingresos_total + ingresos
    margen1_completo = margen1_total

    # -----------------> Información para graficos

    retiro_honorarios = 0
    honorarios_beneficio2 = 0
    honorarios_beneficio1 = 0

    datos_finales.append((saldo_caja_total , pendiente_gastar_total, ingresos_total, descuento_total, margen1_total, margen2_total))

    datos_finales_2 = [margen1_completo]


    # -----------------> Esta es la parte del historico

    datos_historicos = RegistroAlmacenero.objects.order_by("fecha")

    fechas = []

    for d in datos_historicos:

        if not d.fecha in fechas:

            fechas.append(d.fecha)

    datos_registro = []

    for fecha in fechas:

        datos = RegistroAlmacenero.objects.filter(fecha = fecha)

        saldo_caja_total = 0
        pendiente_gastar_total = 0
        ingresos_total = 0
        descuento_total = 0
        honorario = 0


        for dato in datos:

            almacenero = dato

            #Calculo el resto de las cosas

            retiro_socios = almacenero.retiro_socios
            saldo_caja = almacenero.cuotas_cobradas - almacenero.gastos_fecha - almacenero.Prestamos_dados
            
            pend_gast = almacenero.pendiente_admin + almacenero.pendiente_comision + almacenero.saldo_mat + almacenero.saldo_mo + almacenero.imprevisto + almacenero.credito + almacenero.fdr - almacenero.pendiente_adelantos + almacenero.pendiente_iva_ventas + almacenero.pendiente_iibb_tem +almacenero.cheques_emitidos
            
            prest_cobrar = almacenero.prestamos_proyecto + almacenero.prestamos_otros
            total_ingresos = prest_cobrar + almacenero.cuotas_a_cobrar + almacenero.ingreso_ventas + saldo_caja
            
            margen = total_ingresos - pend_gast
            descuento = almacenero.ingreso_ventas*0.06
            
            margen_2 = margen - descuento
            honorario = almacenero.honorarios

            pendiente_gastar_total = pendiente_gastar_total + pend_gast
            saldo_caja_total = saldo_caja_total + saldo_caja
            descuento_total = descuento_total + descuento
            ingresos_total = ingresos_total + total_ingresos

            # Me falta calcular la parte de honorarios


        margen1 = ingresos_total - pendiente_gastar_total

        datos_registro.append(margen1)


    return render(request, "users/dashboard.html", {"fechas":fechas, "datos_finales_2":datos_finales_2, "indice":margen1_completo, "datos_registro":datos_registro, "datos_barras":barras, "ventas_barras":ventas_barras, "ventas":ventas_realizadas, "datos_compras":datos_compras, "datos_unidades":datos_unidades, "arqueo":arqueo})

def inicio(request):

    # La creación de monedas

    usuarios = datosusuario.objects.all().exclude(estado = "NO ACTIVO")

    hoy = datetime.date.today()
    fecha_control = datetime.date(hoy.year, hoy.month, 1)

    for u in usuarios:

        moneda = MonedaLink.objects.filter(usuario_portador = u, fecha__gte = fecha_control)

        if len(moneda) == 0:

            numero = 0

            for i in range(10):

                b = MonedaLink(
                    nombre = str(fecha_control)+str(u.identificacion)+str(numero),
                    usuario_portador = u,
                    fecha = fecha_control,
                    tipo = "NORMAL"
                )

                b.save()
                
                numero += 1

    # Esta es la parte de los permisos

    lista_grupos = 0

    grupos = request.user.groups.all()

    for g in grupos :

        lista_grupos = str(lista_grupos)+"-"+str(g.name)

    # Esta parte es para Pablo
    
    mensajesdeldia = mensajesgenerales.objects.all()

    # -----> Aqui para decirte si tenes pendiente firmar correspondencia

    usuario = request.user.username

    datos_mensajeria = len(NotaDePedido.objects.filter(copia__icontains = str(usuario)).exclude(visto__icontains = str(usuario))) + len(NotaDePedido.objects.filter(destinatario__icontains = str(usuario)).exclude(visto__icontains = str(usuario)))
    
    # -----> Aqui para decirte si tenes pendiente firmar OC

    mensaje_oc = 0

    if usuario == "PL":
        
        compras_espera = Comparativas.objects.filter(estado = "ESPERA").exclude(creador = "MES", numero__contains = "POSTV")
        compras_adjunto_ok = Comparativas.objects.filter(estado = "ADJUNTO ✓").exclude(creador = "MES", numero__contains = "POSTV")

        if len(compras_espera) > 0 or len(compras_adjunto_ok) > 0:
        
            mensaje_oc = [compras_espera, compras_adjunto_ok, (len(compras_espera) + len(compras_adjunto_ok))]
        
    elif usuario == "SP":

        compras = Comparativas.objects.filter(creador = "MES").exclude(estado = "AUTORIZADA")
        compras_2 = Comparativas.objects.filter(numero__contains = "POSTV").exclude(estado = "AUTORIZADA")

        if len(compras) > 0 or len(compras_2) > 0:
        
            mensaje_oc = [compras, compras_2, (len(compras) + len(compras_2))]

    else:

        data_1 = Comparativas.objects.filter(estado = "NO AUTORIZADA", creador = usuario)
        data_2 = Comparativas.objects.filter(visto = "VISTO NO CONFORME", creador = usuario, fecha_c__gte = "2021-02-01")

        if len(data_1) > 0 or len(data_2) > 0:

            mensaje_oc = [data_1, data_2, (len(data_1) + len(data_2))]
        

    datos_logo = 0

    try:
        datos_logo = datosusuario.objects.get(identificacion = request.user)

    except:

        datos_logo = 0

    date = datetime.date.today()

    Registros = RegistroAlmacenero.objects.filter(fecha =date)

    # -----------------> Aqui traigo la parte de honorarios para guardar en los historicos

    honorarios = Honorarios.objects.order_by("-fecha")

    subtotal_1 = honorarios[0].cuotas + honorarios[0].ventas
    ingresos = subtotal_1 + honorarios[0].creditos
    comision = honorarios[0].comision_venta*honorarios[0].ventas
    subtotal_2 = honorarios[0].estructura_gio + honorarios[0].aportes + honorarios[0].socios + comision
    costos = subtotal_2  + honorarios[0].deudas
    honorario = ingresos - costos + honorarios[0].caja_actual
    retiros = honorarios[0].retiro_socios
    honorarios2 = honorario - retiros

    if len(Registros) == 0:
    
        almacenero = Almacenero.objects.all()

        for alma in almacenero:

            presupuesto = Presupuestos.objects.get(proyecto = alma.proyecto)

            b = RegistroAlmacenero(
                fecha = date,
                proyecto = alma.proyecto,
                cheques_emitidos = alma.cheques_emitidos,
                gastos_fecha = alma.gastos_fecha,
                pendiente_admin = alma.pendiente_admin,
                pendiente_comision = alma.pendiente_comision,
                pendiente_adelantos = alma.pendiente_adelantos,
                pendiente_iva_ventas = alma.pendiente_iva_ventas,
                pendiente_iibb_tem = alma.pendiente_iibb_tem,
                pendiente_iibb_tem_link = alma.pendiente_iibb_tem_link,
                prestamos_proyecto = alma.prestamos_proyecto,
                prestamos_otros = alma.prestamos_otros,
                cuotas_cobradas = alma.cuotas_cobradas,
                cuotas_a_cobrar = alma.cuotas_a_cobrar,
                ingreso_ventas = alma.ingreso_ventas,
                ingreso_ventas_link = alma.ingreso_ventas_link,
                Prestamos_dados = alma.Prestamos_dados,
                unidades_socios = alma.unidades_socios,
                saldo_mat = presupuesto.saldo_mat,
                saldo_mo = presupuesto.saldo_mo,
                imprevisto = presupuesto.imprevisto,
                credito = presupuesto.credito, 
                fdr = presupuesto.fdr,
                retiro_socios= sum(np.array(RetirodeSocios.objects.values_list('monto_pesos').filter(proyecto = alma.proyecto))),
                retiro_socios_honorarios = retiros,
                honorarios = honorario,
                tenencia = alma.tenencia,
                financiacion = alma.financiacion,
                inmuebles = alma.inmuebles

            )

            b.save()

    Registro_presupuestos = RegistroValorProyecto.objects.filter(fecha =date)

    # -----------------> Aqui guardo por dia la información de presupuesto

    if len(Registro_presupuestos) == 0:

        presupuestos = Presupuestos.objects.all()

        for p in presupuestos:

            b = RegistroValorProyecto(

                proyecto = p.proyecto,
                fecha = date,
                precio_proyecto = p.valor,

            )

            b.save()

        if settings.ALLOWED_HOSTS:

            # -----------------> Aprovecho para avisar a PL sobre las OC

            today_h = datetime.date.today()

            fecha_pago = datetime.date(2021, 4, 16)

            while fecha_pago <= today_h:
                fecha_pago = fecha_pago + datetime.timedelta(days=14)
        
            fecha_alerta = fecha_pago - datetime.timedelta(days=3)

            if today_h == fecha_alerta:

                compras_espera = len(Comparativas.objects.filter(estado = "ESPERA").exclude(creador = "MES", numero__contains = "POSTV"))
                compras_adjunto_ok = len(Comparativas.objects.filter(estado = "ADJUNTO ✓").exclude(creador = "MES", numero__contains = "POSTV"))
                cantidad_oc = compras_espera + compras_adjunto_ok

                if cantidad_oc == 0:

                    send = "@Pablo, hasta el momento todas las OC estan firmadas para mañana"

                    id = "-585663986"

                    token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"

                    url = "https://api.telegram.org/bot" + token + "/sendMessage"

                    params = {
                        'chat_id' : id,
                        'text' : send
                    }

                    requests.post(url, params=params)

                else:

                    send = "@Pablo mañana a las 13 horas deben estar firmadas las OC. Cantidad pendientes {} ".format(cantidad_oc)

                    id = "-585663986"

                    token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"

                    url = "https://api.telegram.org/bot" + token + "/sendMessage"

                    params = {
                        'chat_id' : id,
                        'text' : send
                    }

                    requests.post(url, params=params)
            
            # -----------------> Saludamos al grupo
            
            if today_h.weekday() == 0:

                send = "Buen dia y buena semana grupo!, fui activado por {}. Listo para empezar el dia".format(request.user.first_name)

                id = "-455382561"

                token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"

                url = "https://api.telegram.org/bot" + token + "/sendMessage"

                params = {
                    'chat_id' : id,
                    'text' : send
                }

                requests.post(url, params=params)

            elif today_h.weekday() == 2:

                send = "Buen miercoles!, fui activado por {}. Listo para ayudar".format(request.user.first_name)

                id = "-455382561"

                token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"

                url = "https://api.telegram.org/bot" + token + "/sendMessage"

                params = {
                    'chat_id' : id,
                    'text' : send
                }

                requests.post(url, params=params)

            elif today_h.weekday() == 4:

                send = "Buen viernes!, fui activado por {}. Ultimo tiron".format(request.user.first_name)

                id = "-455382561"

                token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"

                url = "https://api.telegram.org/bot" + token + "/sendMessage"

                params = {
                    'chat_id' : id,
                    'text' : send
                }

                requests.post(url, params=params)

            elif today_h.weekday() > 4:

                send = "Buen fin de semana!, fui activado por {} por alguna razón. Lamento molestar".format(request.user.first_name)

                id = "-455382561"

                token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"

                url = "https://api.telegram.org/bot" + token + "/sendMessage"

                params = {
                    'chat_id' : id,
                    'text' : send
                }

                requests.post(url, params=params)

        #if hoy.weekday() == 0 or hoy.weekday() == 5:
            

    barras = []

    datos_barras = Presupuestos.objects.order_by("-saldo")

    for db in datos_barras:

        if db.valor != 0:

            avance = (100 - db.saldo/db.valor*100)

            barras.append((db, int(avance)))

    barras = sorted(barras,reverse=True, key=lambda tup: tup[1])


    miembros = datosusuario.objects.all().order_by("identificacion").exclude(estado = "NO ACTIVO")

    cantidad_m = len(datosusuario.objects.all())
    cantidad_p = len(Proyectos.objects.all())

    hoy = datetime.date.today()
    inicio = datetime.date(2020, 5, 1)
    dias_funcionando = (hoy - inicio).days
    monedas = len(EntregaMoneda.objects.filter(fecha__gte = datetime.date.today(), usuario_recibe__identificacion = request.user))
    anuncios = Anuncios.objects.all().exclude(activo = "NO").order_by("-id")

    #######################################
    # Parte de minutas
    #######################################

    minutas_cantidad = len(Acuerdos.objects.filter(responsable__identificacion = request.user.username, estado="NO CHECK"))

    minutas_cantidad_data = Acuerdos.objects.filter(responsable__identificacion = request.user.username, estado="NO CHECK")
    ########################################
    # OC observadas por SP
    ########################################

    sp_oc = []

    aux_sp = len(Comparativas.objects.filter(creador = request.user.username, visto = "VISTO NO CONFORME"))
    aux_sp_2 = Comparativas.objects.filter(creador = request.user.username, visto = "VISTO NO CONFORME").values_list("o_c")

    sp_oc.append(aux_sp)
    sp_oc.append(aux_sp_2)

    return render(request, "users/inicio2.html", {"minutas_cantidad_data":minutas_cantidad_data, "sp_oc":sp_oc, "minutas_cantidad":minutas_cantidad, "anuncios":anuncios, "monedas":monedas, "dias_funcionando":dias_funcionando, "cantidad_p":cantidad_p, "cantidad_m":cantidad_m, "datos_barras":barras, "datos_logo":datos_logo, "mensaje_oc":mensaje_oc, "mensajesdeldia":mensajesdeldia, "datos_mensajeria":datos_mensajeria, "lista_grupos":lista_grupos, "miembros":miembros})

def welcome(request):
    # Si estamos identificados devolvemos la portada
    if request.user.is_authenticated:

        return render(request, "users/welcome.html")
    # En otro caso redireccionamos al login
    return redirect('/login')

def register(request):
    # Creamos el formulario de autenticación vacío
    form = UserCreationForm()
    if request.method == "POST":
        # Añadimos los datos recibidos al formulario
        form = UserCreationForm(data=request.POST)
        # Si el formulario es válido...
        if form.is_valid():

            # Creamos la nueva cuenta de usuario
            user = form.save()

            # Si el usuario se crea correctamente 
            if user is not None:
                # Hacemos el login manualmente
                do_login(request, user)
                # Y le redireccionamos a la portada
                return redirect('/')

    # Si llegamos al final renderizamos el formulario
    return render(request, "users/register.html", {'form': form})

def login(request):
    # Creamos el formulario de autenticación vacío
    form = AuthenticationForm()
    if request.method == "POST":
        # Añadimos los datos recibidos al formulario
        form = AuthenticationForm(data=request.POST)
        # Si el formulario es válido...
        if form.is_valid():
            # Recuperamos las credenciales validadas
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']

            # Verificamos las credenciales del usuario
            user = authenticate(username=username, password=password)

            # Si existe un usuario con ese nombre y contraseña
            if user is not None:
                # Hacemos el login manualmente
                do_login(request, user)
                # Y le redireccionamos a la portada
                return redirect('/inicio')

        else:

            return render(request, "users/logine.html", {'form': form}) 

    # Si llegamos al final renderizamos el formulario

    return render(request, "users/login.html", {'form': form})

def logout(request):
    # Finalizamos la sesión
    do_logout(request)
    # Redireccionamos a la portada
    return redirect('/')

def informes(request):

    if request.method == 'POST':

        data_post = request.POST.items()

        for d in data_post:

            if d[0] == "LISTO":
                    tarea = TareasProgramadas.objects.get(id = int(d[1]))
                    tarea.estado = "LISTO"
                    tarea.save()
            if d[0] == "TRABAJANDO":
                tarea = TareasProgramadas.objects.get(id = int(d[1]))
                tarea.estado = "TRABAJANDO"
                tarea.save()
            if d[0] == "PROBLEMAS":
                tarea = TareasProgramadas.objects.get(id = int(d[1]))
                tarea.estado = "PROBLEMAS"
                tarea.save()

        
    groups = request.user.groups.all().values_list('name', flat=True)

    if "MANDO MEDIO" in groups:
         informes_data = InformeMensual.objects.all().order_by('-fecha')
    else:
        informes_data = InformeMensual.objects.filter(user__identificacion = request.user).order_by('-fecha')
    tareas_data = TareasProgramadas.objects.filter(informe__user__identificacion = request.user).exclude(estado = "LISTO")

    return render(request, 'informes.html', {'tareas_data':tareas_data, 'informes_data':informes_data})

def verinforme(request, id_informe):

    project_list = Proyectos.objects.all()

    informes_data = InformeMensual.objects.get(id = id_informe)

    if request.method == 'POST':

        data_post = request.POST.items()

        for d in data_post:

            if d[0] == "mensaje":
                informes_data.informe = d[1]
                informes_data.save()

            if d[0] == "tareas":
                b = TareasProgramadas(
                    tarea = request.POST['tareas'],
                    informe = informes_data,
                    estado = "ESPERA",
                )
                b.save()

            if d[0] == "LISTO":
                tarea = TareasProgramadas.objects.get(id = int(d[1]))
                tarea.estado = "LISTO"
                tarea.save()
            if d[0] == "TRABAJANDO":
                tarea = TareasProgramadas.objects.get(id = int(d[1]))
                tarea.estado = "TRABAJANDO"
                tarea.save()
            if d[0] == "PROBLEMAS":
                tarea = TareasProgramadas.objects.get(id = int(d[1]))
                tarea.estado = "PROBLEMAS"
                tarea.save()
            if d[0] == "titulo":
                d = Bitacoras(
                    titulo = request.POST['titulo'],
                    proyecto = Proyectos.objects.get(id = int(request.POST['proyecto'])),
                    descrip = request.POST['descrip'],
                    informe = informes_data,
                )

                d.save()
            try:
                if d[0] == "descrip2":
                    bitacora = Bitacoras.objects.get(id = int(request.POST['bitacora']))
                    bitacora.descrip = request.POST['descrip2']
                    bitacora.save()
            except:
                pass


    tareas_data = TareasProgramadas.objects.filter(informe = informes_data)
    bitacoras_data = Bitacoras.objects.filter(informe = informes_data).order_by("-fecha")

    return render(request, 'informes_informe.html', {'bitacoras_data':bitacoras_data, 'project_list':project_list, 'informes_data':informes_data, 'tareas_data':tareas_data})

def informescrear(request):

    usuarios = datosusuario.objects.all().order_by("identificacion")


    if request.method == "POST":

        user = datosusuario.objects.get(id = request.POST['user'])

        b = InformeMensual(
            fecha = request.POST['fecha'],
            user = user,
        )

        b.save()

        return redirect('Informes')
    return render(request, 'informes_crear.html', {'usuarios':usuarios})

def tablerorega(request, id_proyecto, id_area, id_estado):

    if request.method == "POST":

        tarea = Seguimiento.objects.get(id = int(request.POST['id']))
        tarea.area = request.POST['area']
        tarea.nombre = request.POST['nombre']
        tarea.estado = request.POST['estado']
        tarea.proyecto = Proyectos.objects.get(nombre = request.POST['proyecto'])
        tarea.responsable = datosusuario.objects.get(id = int(request.POST['responsable']))
        tarea.save()

        try:
            tarea.fecha_inicio = request.POST['fechai']
            tarea.save()
        except:
            pass
        try:
            tarea.fecha_final = request.POST['fechaf']
            tarea.save()
        except:
            pass

    list_project_all = Proyectos.objects.all()

    group=models.Group.objects.get(name='REGA NIVEL 1')
    users=group.user_set.all()
    list_users = []
    for user in users:
        try:
            us = datosusuario.objects.get(identificacion = user.username)
            list_users.append(us)
        except:
            pass

    if id_proyecto != "0":
        proyecto_el = Proyectos.objects.get(id = int(id_proyecto))
    else:
        proyecto_el = 0

    estado = "Estado"

    areas = ["ADMINISTRACIÓN",
    "COMERCIALIZACIÓN",
    "PRODUCCIÓN"]

    diccionario = {'1': ("ADMINISTRACIÓN", "55, 172, 99 "),
    '2': ("PRODUCCIÓN", "161, 200, 58")}

    if id_area == "0":
        area = ["Área",""]
    else:
        area = diccionario[id_area]

    if id_area == "0":
        list_areas = Seguimiento.objects.all().values_list('area')
        list_areas = list(set(list_areas))
    else:
        list_areas = Seguimiento.objects.filter(area = diccionario[id_area][0]).values_list('area')
        list_areas = list(set(list_areas))
    list_project_dummy = Seguimiento.objects.all().values_list('proyecto')
    list_project_dummy = list(set(list_project_dummy))
    list_project = []

    for l in list_project_dummy:
        proyecto = Proyectos.objects.get(id = int(l[0]))
        list_project.append(proyecto)

    data = []

    for l in list_areas:

        if id_proyecto == "0":
            if id_estado == "0":
                data_list = Seguimiento.objects.filter(area = l[0])
                data.append((l, data_list))
            elif id_estado == "1":
                data_list = Seguimiento.objects.filter(area = l[0]).exclude(estado = "LISTO")
                data.append((l, data_list))
                estado = "Activos"
            elif id_estado == "2":
                data_list = Seguimiento.objects.filter(area = l[0], estado = "TRABAJANDO")
                data.append((l, data_list))
                estado = "Trabajando"
            elif id_estado == "3":
                data_list = Seguimiento.objects.filter(area = l[0], estado = "PROBLEMAS")
                data.append((l, data_list))
                estado = "Problemas"
            elif id_estado == "4":
                data_list = Seguimiento.objects.filter(area = l[0], estado = "LISTO")
                data.append((l, data_list))
                estado = "Listo"
            else:
                data_list = Seguimiento.objects.filter(area = l[0], estado = "ESPERA")
                data.append((l, data_list))
                estado = "Espera"

        else:
            if id_estado == "0":
                data_list = Seguimiento.objects.filter(area = l[0], proyecto__id = int(id_proyecto))
                data.append((l, data_list))
            elif id_estado == "1":
                data_list = Seguimiento.objects.filter(area = l[0], proyecto__id = int(id_proyecto)).exclude(estado = "LISTO")
                data.append((l, data_list))
                estado = "Activos"
            elif id_estado == "2":
                data_list = Seguimiento.objects.filter(area = l[0], proyecto__id = int(id_proyecto), estado = "TRABAJANDO")
                data.append((l, data_list))
                estado = "Trabajando"
            elif id_estado == "3":
                data_list = Seguimiento.objects.filter(area = l[0], proyecto__id = int(id_proyecto), estado = "PROBLEMAS")
                data.append((l, data_list))
                estado = "Problemas"
            elif id_estado == "4":
                data_list = Seguimiento.objects.filter(area = l[0], proyecto__id = int(id_proyecto), estado = "LISTO")
                data.append((l, data_list))
                estado = "Listo"
            else:
                data_list = Seguimiento.objects.filter(area = l[0], proyecto__id = int(id_proyecto), estado = "ESPERA")
                data.append((l, data_list))
                estado = "Espera"

    return render(request, 'seguimiento.html', {'list_project_all':list_project_all, 'list_users':list_users, 'area':area, 'estado':estado, 'proyecto':proyecto_el, 'data':data, 'list_project':list_project, 'id_estado':id_estado, 'id_area':id_area, 'id_proyecto':id_proyecto, 'areas':areas})

def tableroregaadd(request):

    group=models.Group.objects.get(name='REGA NIVEL 1')
    users=group.user_set.all()
    list_users = []
    for user in users:
        try:
            us = datosusuario.objects.get(identificacion = user.username)
            list_users.append(us)
        except:
            pass

    areas = ["ADMINISTRACIÓN",
    "COMERCIALIZACIÓN",
    "PRODUCCIÓN"]

    proyecto = Proyectos.objects.all().order_by("nombre")

    if request.method == "POST":

        datos_p = request.POST.items()
        
        tarea = Seguimiento(
            orden = len(Seguimiento.objects.filter(area = request.POST['area'])),
            area = request.POST['area'],
            proyecto = Proyectos.objects.get(id = int(request.POST['proyecto'])),
            nombre = request.POST['nombre'],
            responsable = datosusuario.objects.get(identificacion = request.POST['responsable'])

        )

        tarea.save()

        try:
            tarea.fecha_inicio = request.POST['fechai']
            tarea.save()
        except:
            pass

        try:
            tarea.fecha_final = request.POST['fechaf']
            tarea.save()
        except:
            pass

        return redirect('Tablero Rega', id_proyecto = 0, id_area = 0, id_estado = 1)


    return render(request, 'seguimiento_add.html', {'areas':areas, 'proyecto':proyecto, 'list_users':list_users})

def anuncios(request):

    if request.method == 'POST':

        datos_p = request.POST.items()

        for d in datos_p:
            if d[0] == "ID":
                anuncio = Anuncios.objects.get(id = int(request.POST["ID"]))
                anuncio.titulo = request.POST["titulo"]
                anuncio.categoria = request.POST["categoria"]
                anuncio.descrip = request.POST["descrip"]
                anuncio.activo = request.POST["activo"]
                try:
                    anuncio.imagen = request.FILES['imagen']
                    anuncio.save()
                except:
                    anuncio.save()
            
            if d[0] == "NUEVO":

                b = Anuncios(
                    titulo = request.POST["titulo"],
                    categoria = request.POST["categoria"],
                    descrip = request.POST["descrip"],
                    activo = "SI",
                    imagen = request.FILES['imagen']
                )

                b.save()

            if d[0] == "delete":
                anuncio = Anuncios.objects.get(id = int(request.POST["delete"]))
                anuncio.delete()


    data = Anuncios.objects.all().order_by("-id")

    return render(request, 'users/anuncios.html', {"data":data})

def minutas(request):

    if request.method == 'POST':
        datos_p = request.POST.items()
        for d in datos_p:

            if d[0] == "delete":
                minuta = Minutas.objects.get(id = int(request.POST['delete']))
                minuta.delete()

    minutas_activas = []

    minutas_archivadas = []

    nombre_reuniones = Minutas.objects.values_list("reunion", flat = True).exclude(reunion = None).order_by("-fecha")

    nombre_reuniones = list(set(nombre_reuniones))    

    for name in nombre_reuniones:
        aux = Minutas.objects.filter(reunion = name).order_by("-fecha")[0:1]
        for a in aux:
            minutas_activas.append(a)
        aux_2 = Minutas.objects.filter(reunion = name).order_by("-fecha")[1:]
        for b in aux_2:
            minutas_archivadas.append(b)

    data_2 = Minutas.objects.filter(reunion = None)

    for d in data_2:

        minutas_activas.append(d)

    minutas_activas = sorted(minutas_activas, key = lambda x : x.fecha, reverse = True)

    return render(request, 'minutas/minutasLista.html', {'minutas_activas':minutas_activas, 'minutas_archivadas':minutas_archivadas})

def minutascrear(request):

    mensaje = 0

    if request.method == 'POST':

        datos_p = request.POST.items()

        try:

            minuta = Minutas(
                nombre = request.POST['nombre'],
                integrantes = request.POST['integrantes'],
                fecha = request.POST['fecha'],
                reunion = request.POST['reunion'],
                creador = datosusuario.objects.get(identificacion = request.user.username),
            )

            minuta.save()

            return redirect('Minutas Id', id_minuta = minuta.id)

        except:

            mensaje = "Algún dato de la minuta no esta completo o el creador no esta registrado"

    return render(request, 'minutas/minutasCrear.html', {'mensaje':mensaje})

def minutasmodificar(request, id_minuta):


    #group=models.Group.objects.get(name='REGA NIVEL 1')
    #users=group.user_set.all()

    list_users = datosusuario.objects.filter(estado = "ACTIVO")

    data = Minutas.objects.get(id = int(id_minuta))

    if request.method == 'POST':
        data.creador = datosusuario.objects.get(identificacion = request.POST['creador'])
        data.nombre = request.POST['nombre']
        data.fecha = request.POST['fecha']
        data.integrantes = request.POST['integrantes']
        data.reunion = request.POST['reunion']
        data.save()
        return redirect('Minutas Id', id_minuta = data.id)


    return render(request, 'minutas/minutasModificar.html', {'data':data, 'list_users':list_users})

def minutasid(request, id_minuta):

    group=models.Group.objects.get(name='REGA NIVEL 1')
    users=group.user_set.all()
    list_users = []
    for user in users:
        try:
            us = datosusuario.objects.get(identificacion = user.username)
            list_users.append(us)
        except:
            pass

    if request.method == 'POST':
        datos_p = request.POST.items()
        for d in datos_p:
            if d[0] == "tema":
                acuerdo = Acuerdos.objects.get(id = int(request.POST['id']))
                acuerdo.tema = request.POST['tema']
                try:
                    acuerdo.responsable = datosusuario.objects.get(identificacion = request.POST['responsable'])
                    acuerdo.save()
                except:
                    acuerdo.responsable = None
                    acuerdo.save()

                try:
                    acuerdo.fecha_limite = request.POST['fecha_limite']
                    acuerdo.save()
                except:
                    pass

            if d[0] == "check":
                acuerdo = Acuerdos.objects.get(id = int(request.POST['check']))
                acuerdo.estado = "CHECK"
                acuerdo.save()

            if d[0] == "no_check":
                acuerdo = Acuerdos.objects.get(id = int(request.POST['no_check']))
                acuerdo.estado = "NO CHECK"
                acuerdo.save()

            if d[0] == "acuerdo":
                b = Acuerdos(
                    tema = request.POST['acuerdo'],
                    minuta = Minutas.objects.get(id = int(id_minuta))
                )
                try:
                    b.responsable = datosusuario.objects.get(identificacion = request.POST['responsable'])
                    b.save()
                except:
                    b.save()

            if d[0] == "delete":
                acuerdo = Acuerdos.objects.get(id = int(request.POST['delete']))
                acuerdo.delete()

    data = Minutas.objects.get(id = int(id_minuta))

    acuerdos = Acuerdos.objects.filter(minuta = data)

    acuerdos_viejos = Acuerdos.objects.filter(minuta__id__lt = data.id, estado = "NO CHECK", minuta__reunion = data.reunion)

    return render(request, 'minutas/minutasId.html', {'data':data, 'acuerdos':acuerdos, 'acuerdos_viejos':acuerdos_viejos, 'list_users':list_users})

def registro_contable(request, date_i):

    hoy = datetime.date(int(date_i[0:4]), int(date_i[4:]), 1)

    user = datosusuario.objects.get(identificacion = request.user.username)

    registros_totales = RegistroContable.objects.filter(usuario = user)

    if request.method == 'POST':

        try:
            if request.POST['carga_archivo'] == "1":
                archivo_pandas = pd.read_excel(request.FILES['archivo'])
                '''
                # ---> Editar registros
                registros_editar = archivo_pandas[archivo_pandas["Acción"] == "EDITAR"].reset_index(drop=True)
                numero = 0
                for row in range(registros_editar.shape[0]):
                    importe_data_aux = str(registros_editar.loc[numero, "Importe"]).split(sep=" ")
                    data_aux = RegistroContable.objects.get(id = int(registros_editar.loc[numero, "Id"]))
                    data_aux.creador = registros_editar.loc[numero, "Creador"]
                    data_aux.fecha = registros_editar.loc[numero, "Fecha"]
                    data_aux.estado = registros_editar.loc[numero, "Tipo"]
                    data_aux.caja = registros_editar.loc[numero, "Caja"]
                    data_aux.cuenta = registros_editar.loc[numero, "Cuenta"]
                    data_aux.categoria = registros_editar.loc[numero, "Categoria"]
                    data_aux.importe = abs(float(importe_data_aux[0]))
                    data_aux.nota = registros_editar.loc[numero, "Nota"]
                    data_aux.save()
                    numero += 1

                registros_borrar = archivo_pandas[archivo_pandas["Acción"] == "BORRAR"].reset_index(drop=True)
                numero = 0
                for row in range(registros_borrar.shape[0]):
                    data_aux = RegistroContable.objects.get(id = int(registros_borrar.loc[numero, "Id"]))
                    data_aux.delete()
                    numero += 1

                registros_nuevo = archivo_pandas[archivo_pandas["Acción"] == "NUEVO"].reset_index(drop=True)
                '''
                registros_nuevo = archivo_pandas
                numero = 0

                # Bucle para cargar registros

                for row in range(registros_nuevo.shape[0]):
                    usuario = datosusuario.objects.get(identificacion = str(registros_nuevo.loc[numero, "Usuario"]))
                    if registros_nuevo.loc[numero, "Saldo (CTE)"] >= 0:
                        estado = "INGRESOS"
                        importe = abs(registros_nuevo.loc[numero, "Saldo (CTE)"])
                    else:
                        estado = "GASTOS"
                        importe = abs(registros_nuevo.loc[numero, "Saldo (CTE)"])

                    if len(DicRegistroContable.objects.filter(entrada = registros_nuevo.loc[numero, "Auxiliar"])) > 0:
                        caja = DicRegistroContable.objects.filter(entrada = registros_nuevo.loc[numero, "Auxiliar"])[0].salida
                    else:
                        caja = registros_nuevo.loc[numero, "Auxiliar"]

                    if len(DicRegistroContable.objects.filter(entrada = registros_nuevo.loc[numero, "Desc. cuenta"])) > 0:
                        cuenta = DicRegistroContable.objects.filter(entrada = registros_nuevo.loc[numero, "Desc. cuenta"])[0].salida
                    else:
                        cuenta = registros_nuevo.loc[numero, "Desc. cuenta"]

                    if len(DicRegistroContable.objects.filter(entrada = registros_nuevo.loc[numero, "Desc. auxiliar"])) > 0:
                        categoria = DicRegistroContable.objects.filter(entrada = registros_nuevo.loc[numero, "Desc. auxiliar"])[0].salida
                    else:
                        categoria = registros_nuevo.loc[numero, "Desc. auxiliar"]
                    
                    if len(DicRegistroContable.objects.filter(entrada = registros_nuevo.loc[numero, "Subauxiliar"])) > 0:
                        nota = DicRegistroContable.objects.filter(entrada = registros_nuevo.loc[numero, "Subauxiliar"])[0].salida
                    else:
                        nota = registros_nuevo.loc[numero, "Subauxiliar"]

                    b = RegistroContable(

                        usuario = usuario,
                        creador = registros_nuevo.loc[numero, "Creador"],
                        fecha = registros_nuevo.loc[numero, "Fecha de emisión"],
                        estado = estado,
                        caja = caja,
                        cuenta = cuenta,
                        categoria = categoria,
                        importe = importe,
                        nota = nota,


                    )

                    b.save()
                    numero += 1

        except:
            pass

        try:
 
            if request.POST['fecha_m'] == "1":
                if hoy.month != 12:
                    new_date_i = str(hoy.year)+str(hoy.month + 1)
                if hoy.month == 12:
                    new_date_i = str(hoy.year + 1)+str(1)

                return redirect('Registro Contable', date_i = new_date_i)

            if request.POST['fecha_m'] == "0":
                if hoy.month != 1:
                    new_date_i = str(hoy.year)+str(hoy.month - 1)
                if hoy.month == 1:
                    new_date_i = str(hoy.year - 1)+str(12)

                return redirect('Registro Contable', date_i = new_date_i)

        except:
            pass
        try:
            
            b = RegistroContable(
                usuario = user,
                creador = request.user.username,
                fecha = request.POST['fecha'],
                estado = request.POST['tipo'],
                caja = request.POST['caja'],
                cuenta = request.POST['cuenta'],
                categoria = request.POST['categoria'],
                importe = float(request.POST['importe']),
                nota = request.POST['nota'],
            )

            try:
                b.adjunto = request.FILES['adjunto']
                b.save()

            except:
                b.save()

        except:
            pass

        try:

            if request.POST['editar']:
                registro = RegistroContable.objects.get(id = request.POST['editar'])
                registro.fecha = request.POST['fecha']
                registro.cuenta = request.POST['cuenta']
                registro.categortia = request.POST['categoria']
                registro.importe = request.POST['importe']
                registro.nota = request.POST['nota']
                try:
                    registro.adjunto = request.FILES['adjunto']
                    registro.save()

                except:
                    registro.save()
        except:
            pass
        try:
            if request.POST['eliminar']:
                registro = RegistroContable.objects.get(id = request.POST['eliminar'])
                registro.delete()
        except:
            pass

    ##### Esquema diario

    fecha_inicial = date(hoy.year, hoy.month, 1)

    if hoy.month == 12:

        fecha_final = date(hoy.year + 1, 1 , 1)

    else:

        fecha_final = date(hoy.year, hoy.month + 1, 1)

    fechas = RegistroContable.objects.filter(usuario = user, fecha__range=[fecha_inicial, fecha_final]).values_list("fecha", flat=True).order_by("-fecha").distinct()

    datos = []

    for f in fechas:

        ingresos_f = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha = f).values_list("importe", flat=True)))
        gastos_f = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha = f).values_list("importe", flat=True)))

        data = RegistroContable.objects.filter(usuario = user, fecha = f)

        datos.append([f, data, ingresos_f, gastos_f])

    

    ### Cuadros generales

    ##### Idea de la caja

    total_cajas = []

    cajas_cargadas =  RegistroContable.objects.filter(usuario = user).values_list("caja", flat=True).distinct()
    for c in cajas_cargadas:
        ingresos = sum(RegistroContable.objects.filter(usuario = user, caja = c, estado = "INGRESOS").values_list("importe", flat=True))
        gastos = sum(RegistroContable.objects.filter(usuario = user, caja = c, estado = "GASTOS").values_list("importe", flat=True))
        balance = ingresos - gastos
        total_cajas.append((c, ingresos, gastos, balance))

    cat_ingresos = RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha__range=[fecha_inicial, fecha_final]).values_list("categoria", flat=True).distinct()

    pie_ingresos = []

    for ci in cat_ingresos:
        aux = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha__range=[fecha_inicial, fecha_final], categoria = ci).values_list("importe", flat=True)))
        pie_ingresos.append([ci, aux])

    pie_ingresos = sorted(pie_ingresos, key=lambda X : -X[1])

    aux_color = 0
    for pie in pie_ingresos: 
        color = ((200- aux_color), (200- aux_color), 253)
        pie.append(color)
        aux_color += (200 - 20)/len(pie_ingresos)


    cat_gastos = RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha__range=[fecha_inicial, fecha_final]).values_list("categoria", flat=True).distinct()

    pie_gastos = []

    aux_color = 0

    for cg in cat_gastos:
        aux = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha__range=[fecha_inicial, fecha_final], categoria = cg).values_list("importe", flat=True)))        
        pie_gastos.append([cg, aux])

    pie_gastos = sorted(pie_gastos, key=lambda X : -X[1])

    aux_color = 0
    for pie in pie_gastos: 
        color = (253, (200- aux_color), (200- aux_color))
        pie.append(color)
        aux_color += (200 - 20)/len(pie_gastos)


    list_cat_gasto = RegistroContable.objects.filter(usuario = user, estado = "GASTOS").values_list("categoria", flat=True).distinct()
    list_cat_ing = RegistroContable.objects.filter(usuario = user, estado = "INGRESOS").values_list("categoria", flat=True).distinct()


    ## Esquema mensual

    data_month = []

    try:

        fecha_1 = RegistroContable.objects.all().order_by("fecha")[0].fecha
        fecha_f = RegistroContable.objects.all().order_by("-fecha")[0].fecha
        fecha_f_auxiliar = datetime.date(fecha_f.year, fecha_f.month, 1)
        fechas = []

        fecha_auxiliar = datetime.date(fecha_1.year, fecha_1.month, 1)

        if fecha_1.month == 12:

            fecha_auxiliar_2 = datetime.date(fecha_1.year + 1, 1, 1)

        else:

            fecha_auxiliar_2 = datetime.date(fecha_1.year, fecha_1.month + 1, 1)

        while fecha_auxiliar <= fecha_f_auxiliar:

            mes = fecha_auxiliar
            ingresos = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha__range=[fecha_auxiliar, fecha_auxiliar_2]).values_list("importe", flat=True)))
            gastos = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha__range=[fecha_auxiliar, fecha_auxiliar_2]).values_list("importe", flat=True)))
            balance = ingresos - gastos
            data_month.append((mes, ingresos, gastos, balance))

            if fecha_auxiliar.month == 12:

                fecha_auxiliar = datetime.date(fecha_auxiliar.year + 1, 1, 1)

            else:

                fecha_auxiliar = datetime.date(fecha_auxiliar.year, fecha_auxiliar.month + 1, 1)

            if fecha_auxiliar_2.month == 12:

                fecha_auxiliar_2 = datetime.date(fecha_auxiliar_2.year + 1, 1, 1)

            else:

                fecha_auxiliar_2 = datetime.date(fecha_auxiliar_2.year, fecha_auxiliar_2.month + 1, 1)

    except:
        pass


    ##### Generales

    ingresos = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha__range=[fecha_inicial, fecha_final]).values_list("importe", flat=True)))

    gastos = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha__range=[fecha_inicial, fecha_final]).values_list("importe", flat=True)))
  
    balance = ingresos - gastos

    ##### Semana

    semana = hoy.weekday()

    semana_1 = hoy - datetime.timedelta(days=hoy.weekday())

    semana_2 = semana_1

    datos_week = []

    for i in range(5):
        semana_2 = semana_2 + datetime.timedelta(days=7)
        ingresos_week = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha__range=[semana_1, semana_2]).values_list("importe", flat=True)))
        gastos_week = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha__range=[semana_1, semana_2]).values_list("importe", flat=True)))
        balance_week = ingresos - gastos
        datos_week.append((semana_1, semana_2, ingresos_week, gastos_week, balance_week))
        semana_1 = semana_1 + datetime.timedelta(days=7)    


    return render(request, "users/registro_contable.html", {'total_cajas':total_cajas, 'registros_totales':registros_totales,'datos_week':datos_week, 'data_month':data_month, 'list_cat_ing':list_cat_ing, 'list_cat_gasto':list_cat_gasto, 'pie_gastos':pie_gastos, 'pie_ingresos':pie_ingresos, 'hoy':hoy, 'datos':datos, "ingresos":ingresos, "gastos":gastos, "balance":balance})

class DescargarRegistroContable(TemplateView):

    def get(self, request, *args, **kwargs):

        # --> Iniciamos el Workbook
        wb = Workbook()

        # --> Primeros calculos

        data = RegistroContable.objects.all()

        ws = wb.active
        ws.title = "Resumen"
        ws["A1"] = "Registros contables"
        ws["A1"].font = Font(bold = True)
        ws["A2"] = "Para usar este archivo para modificar, elimina las primeras 3 filas y en acción agrega 'EDITAR' , 'NUEVO', 'BORRAR' "

        ws["A4"] = "Usuario"
        ws["B4"] = "Creador"
        ws["C4"] = "Fecha"
        ws["D4"] = "Tipo"
        ws["E4"] = "Caja"
        ws["F4"] = "Cuenta"
        ws["G4"] = "Categoria"
        ws["H4"] = "Importe"
        ws["I4"] = "Nota"
        ws["J4"] = "Acción"
        ws["K4"] = "Id"


        ws["A4"].font = Font(bold = True, color= "E8F8F8")
        ws["A4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["B4"].font = Font(bold = True, color= "E8F8F8")
        ws["B4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["C4"].font = Font(bold = True, color= "E8F8F8")
        ws["C4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["D4"].font = Font(bold = True, color= "E8F8F8")
        ws["D4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["E4"].font = Font(bold = True, color= "E8F8F8")
        ws["E4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["F4"].font = Font(bold = True, color= "E8F8F8")
        ws["F4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["G4"].font = Font(bold = True, color= "E8F8F8")
        ws["G4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["H4"].font = Font(bold = True, color= "E8F8F8")
        ws["H4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["I4"].font = Font(bold = True, color= "E8F8F8")
        ws["I4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["J4"].font = Font(bold = True, color= "E8F8F8")
        ws["J4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["K4"].font = Font(bold = True, color= "E8F8F8")
        ws["K4"].fill =  PatternFill("solid", fgColor= "2C9E9D")



        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 17
        ws.column_dimensions['D'].width = 17
        ws.column_dimensions['E'].width = 17
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 7

        cont = 5

        for d in data:

            if request.user.username in d.usuario.identificacion or request.user.username in d.creador:

                ws = wb.active

                ws["A"+str(cont)] = d.usuario.identificacion
                ws["B"+str(cont)] = d.creador
                ws["C"+str(cont)] = d.fecha
                ws["D"+str(cont)] = d.estado
                ws["E"+str(cont)] = d.caja
                ws["F"+str(cont)] = d.cuenta
                ws["G"+str(cont)] = d.categoria
                if d.estado == "INGRESOS":
                    ws["H"+str(cont)] = d.importe
                else:
                    ws["H"+str(cont)] = - d.importe
                ws["I"+str(cont)] = d.nota
                ws["K"+str(cont)] = d.id

                ws["H"+str(cont)].number_format = '"$"" "#,##0.00_-'
                

                cont += 1


        #Establecer el nombre del archivo
        nombre_archivo = "RegistroContable.-{}.xls".format(str(d.usuario.identificacion))
        
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo).replace(',', '_')
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

