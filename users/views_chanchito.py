from django.shortcuts import render, redirect
from django.views.generic.base import TemplateView 
from django.http import HttpResponse
import datetime
import pandas as pd
import numpy as np
from datetime import date
from rrhh.models import datosusuario, DicRegistroContable, RegistroContable, ArqueoChanchito
from finanzas.models import Arqueo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from .functions_chanchito import cajasDerivadas, calcularResumenIngresos, cajasActivas, cajasAdministras
from .functions import saludo

def registro_contable_registro(request):

    usuario = datosusuario.objects.get(identificacion = request.user.username)

    context = {}
    context['mensaje_bievenida'] = saludo().format(request.user.first_name)
    context['datos'] = calcularResumenIngresos(usuario)

    return render(request, "chanchito/registro_contable_reporte.html", context)

def registro_contable_home(request):

    context = {}
    context['mensaje_bievenida'] = saludo().format(request.user.first_name)
    context['fecha'] = datetime.date.today()

    return render(request, 'chanchito/registro_contable_home.html', context)

def registro_contable_cajas(request):

    context = {}
    context["mensaje"] = "no"

    if request.method == 'POST':
        try:
            if request.POST['actualizar_cajas'] == "1":

                usuario = datosusuario.objects.get(identificacion = request.user.username)
                consulta_principal = RegistroContable.objects.filter(creador = usuario).values_list("usuario", flat = True)
                for consulta in consulta_principal:
                    usuario = datosusuario.objects.get(id = consulta)
                    cajasDerivadas(usuario, "RETIROS PERSONALES", "PERSONAL")
                    cajasDerivadas(usuario, "DEPOSITO BANCO", "CAJA BANCO")

                context["mensaje"] = "ok"

        except:
            
            try:
                if request.POST['carga_archivo'] == "1":

                    archivo_pandas = pd.read_excel(request.FILES['archivo'])
                    registros_nuevo = archivo_pandas
                    numero = 0

                    columnas_necesarias = ["Usuario", "Creador", "Auxiliar", "Desc. cuenta", "Desc. auxiliar", "Saldo (CTE)", "SALDO USD", "Subauxiliar", "Fecha de emisión"]

                    for c in columnas_necesarias:

                        if c not in registros_nuevo.columns:
                            context["mensaje"] = "Columna {} no detectada, revise".format(c)


                    # Bucle para cargar registros

                    try:

                        con_dicc = DicRegistroContable.objects.all()

                        for row in range(registros_nuevo.shape[0]):
                            usuario = datosusuario.objects.get(identificacion = str(registros_nuevo.loc[numero, "Usuario"]))
                            if registros_nuevo.loc[numero, "Saldo (CTE)"] >= 0:
                                estado = "INGRESOS"
                                importe = abs(registros_nuevo.loc[numero, "Saldo (CTE)"])
                            else:
                                estado = "GASTOS"
                                importe = abs(registros_nuevo.loc[numero, "Saldo (CTE)"])

                            importe_usd = abs(importe/float(registros_nuevo.loc[numero, "SALDO USD"]))
                            
                            if len(con_dicc.filter(entrada = registros_nuevo.loc[numero, "Auxiliar"])) > 0:
                                caja = con_dicc.filter(entrada = registros_nuevo.loc[numero, "Auxiliar"])[0].salida
                            else:
                                caja = registros_nuevo.loc[numero, "Auxiliar"]

                            if len(con_dicc.filter(entrada = registros_nuevo.loc[numero, "Desc. cuenta"])) > 0:
                                cuenta = con_dicc.filter(entrada = registros_nuevo.loc[numero, "Desc. cuenta"])[0].salida
                            else:
                                cuenta = registros_nuevo.loc[numero, "Desc. cuenta"]

                            if len(con_dicc.filter(entrada = registros_nuevo.loc[numero, "Desc. auxiliar"])) > 0:
                                categoria = con_dicc.filter(entrada = registros_nuevo.loc[numero, "Desc. auxiliar"])[0].salida
                            else:
                                categoria = registros_nuevo.loc[numero, "Desc. auxiliar"]
                            
                            if len(con_dicc.filter(entrada = registros_nuevo.loc[numero, "Subauxiliar"])) > 0:
                                nota = con_dicc.filter(entrada = registros_nuevo.loc[numero, "Subauxiliar"])[0].salida
                            else:
                                nota = registros_nuevo.loc[numero, "Subauxiliar"]

                            try:
                                nuevo_registro = RegistroContable(

                                    usuario = usuario,
                                    creador = registros_nuevo.loc[numero, "Creador"],
                                    fecha = registros_nuevo.loc[numero, "Fecha de emisión"],
                                    estado = estado,
                                    caja = caja,
                                    cuenta = cuenta,
                                    categoria = categoria,
                                    importe = importe,
                                    importe_usd = importe_usd,
                                    nota = nota,

                                )

                                nuevo_registro.save()
                                numero += 1
                            except:
                                mensaje = "Error en la fila {}".format(numero)
                                numero += 1

                        context["mensaje"] = "ok"

                    except:
                        pass

            except:

                try:
                    data_caja = request.POST['borrar_selec'].split("&")
                    fecha_i = request.POST['fecha_i']
                    fecha_f = request.POST['fecha_f']
                    cajas_eliminar = RegistroContable.objects.filter(creador = request.user.username, caja = data_caja[0], usuario__identificacion = data_caja[1], fecha__renge = (fecha_i, fecha_f))
                    for caja in cajas_eliminar:
                        caja.delete()

                except:
                    data_caja = request.POST['borrar'].split("&")
                    cajas_eliminar = RegistroContable.objects.filter(creador = request.user.username, caja = data_caja[0], usuario__identificacion = data_caja[1])
                    for caja in cajas_eliminar:
                        caja.delete()

    user = datosusuario.objects.get(identificacion = request.user.username)

    context["total_cajas"] = cajasActivas(user)
    context["cajas_administras"] = cajasAdministras(user)
    
    context["user"] = user


    return render(request, 'chanchito/registro_contable_cajas.html', context)

def registro_contable_caja(request, caja, user_caja, estado, mes, year):
    mes = mes
    year = year

    user = datosusuario.objects.get(identificacion = request.user.username)
    list_year = list(set(RegistroContable.objects.filter(usuario = user, caja = caja).values_list("fecha__year", flat=True)))
    if request.method == 'POST':
        try:

            b = RegistroContable(
                usuario = user,
                creador = request.user.username,
                fecha = request.POST['fecha'],
                estado = request.POST['tipo'],
                caja = caja,
                cuenta = request.POST['cuenta'],
                categoria = request.POST['categoria'],
                importe = float(request.POST['importe']),
                nota = request.POST['nota'],
                )

            try:
                b.adjunto = request.FILES['adjunto']
                b.save()

            except:
                b.save()

        except:
            pass
        try:
            if request.POST["borrar"]:
                sub_aux = RegistroContable.objects.get(id = request.POST["borrar"])
                sub_aux.delete()
        except:
            pass

        try:

            if request.POST['editar']:
                registro = RegistroContable.objects.get(id = request.POST['editar'])
                registro.fecha = request.POST['fecha']
                registro.caja = request.POST['caja']
                registro.cuenta = request.POST['cuenta']
                registro.categoria = request.POST['categoria']
                registro.importe = request.POST['importe']
                registro.nota = request.POST['nota']
                try:
                    registro.adjunto = request.FILES['adjunto']
                    registro.save()

                except:
                    registro.save()
        except:
            pass
    
    if request.user.username == user_caja:
        data = RegistroContable.objects.filter(usuario = user, caja = caja).order_by("-fecha")
    else:
        data = RegistroContable.objects.filter(usuario__identificacion = user_caja, caja = caja).order_by("-fecha")
    
    if estado == 0:
        data = data
    elif estado == 1:
        data = data.filter(estado = "INGRESOS")
    else:
        data = data.filter(estado = "GASTOS")
    
    if int(mes) != 0:
        data = data.filter(fecha__month = mes)

    if int(year) != 0:
        data = data.filter(fecha__year = year)

    context = {}
    context["data"] = data
    context["caja"] = caja
    context["user_caja"] = user_caja
    context["estado"] = estado
    context["mes"] = mes
    context["year"] = year
    context["list_year"] = list_year
    
    return render(request, 'chanchito/registro_contable_caja_detalle.html', context)

def registro_contable(request, date_i):

    hoy = datetime.date(int(date_i[0:4]), int(date_i[4:]), 1)

    user = datosusuario.objects.get(identificacion = request.user.username)

    registros_totales = RegistroContable.objects.filter(usuario = user)

    ##### Parte del control con arqueo

    cajas_disponibles = RegistroContable.objects.filter(usuario = user).values_list("caja", flat = True).distinct()
    fechas_cargadas = RegistroContable.objects.filter(usuario = user).values_list("fecha", flat = True).order_by("-fecha").distinct()
    cajas_arqueo = ArqueoChanchito.objects.filter(usuario = user).values_list("caja", flat = True)

    rango_maximo = 4

    for fecha in fechas_cargadas:
        rango_maximo -= 1
        if rango_maximo > 0:
            for caja in cajas_disponibles:
                if caja in cajas_arqueo:
                    try:
                        arqueo = Arqueo.objects.filter(fecha = fecha)[0]
                        data_frame = pd.read_excel(arqueo.arqueo)
                        print("Estoy con el dataframe")
                        data_caja = data_frame[data_frame['PROYECTO'] == ArqueoChanchito.objects.filter(usuario = user, caja = caja)[0].arqueo]['EFECTIVO']
                        print(np.array(data_caja))
                    except:
                        pass
        else:
            break

    if request.method == 'POST':

        try:
 
            if request.POST['fecha_m'] == "1":
                if hoy.month != 12:
                    new_date_i = str(hoy.year)+str(hoy.month + 1)
                if hoy.month == 12:
                    new_date_i = str(hoy.year + 1)+str(1)

                return redirect('Registro Contable', date_i = new_date_i)

            if request.POST['fecha_m'] == "0":
                if hoy.month != 1:
                    new_date_i = str(hoy.year)+str(hoy.month - 1)
                if hoy.month == 1:
                    new_date_i = str(hoy.year - 1)+str(12)

                return redirect('Registro Contable', date_i = new_date_i)

        except:
            pass
        try:
            
            b = RegistroContable(
                usuario = user,
                creador = request.user.username,
                fecha = request.POST['fecha'],
                estado = request.POST['tipo'],
                caja = "Personal",
                cuenta = request.POST['cuenta'],
                categoria = request.POST['categoria'],
                importe = float(request.POST['importe']),
                nota = request.POST['nota'],
                )

            try:
                b.adjunto = request.FILES['adjunto']
                b.save()

            except:
                b.save()

        except:
            pass

        try:

            if request.POST['editar']:
                registro = RegistroContable.objects.get(id = request.POST['editar'])
                registro.fecha = request.POST['fecha']
                registro.cuenta = request.POST['cuenta']
                registro.categortia = request.POST['categoria']
                registro.importe = request.POST['importe']
                registro.nota = request.POST['nota']
                try:
                    registro.adjunto = request.FILES['adjunto']
                    registro.save()

                except:
                    registro.save()
        except:
            pass
        try:
            if request.POST['eliminar']:
                registro = RegistroContable.objects.get(id = request.POST['eliminar'])
                registro.delete()
        except:
            pass

    ##### Esquema diario

    fecha_inicial = date(hoy.year, hoy.month, 1)

    if hoy.month == 12:

        fecha_final = date(hoy.year + 1, 1 , 1)

    else:

        fecha_final = date(hoy.year, hoy.month + 1, 1)

    fechas = RegistroContable.objects.filter(usuario = user, fecha__range=[fecha_inicial, fecha_final]).values_list("fecha", flat=True).order_by("-fecha").distinct()

    datos = []

    for f in fechas:

        ingresos_f = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha = f).values_list("importe", flat=True)))
        gastos_f = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha = f).values_list("importe", flat=True)))

        data = RegistroContable.objects.filter(usuario = user, fecha = f)

        datos.append([f, data, ingresos_f, gastos_f])

    ### Cuadros generales

    ##### Idea de la caja

    total_cajas = []

    cajas_cargadas =  RegistroContable.objects.filter(usuario = user).values_list("caja", flat=True).distinct()
    for c in cajas_cargadas:
        ingresos = sum(RegistroContable.objects.filter(usuario = user, caja = c, estado = "INGRESOS").values_list("importe", flat=True))
        gastos = sum(RegistroContable.objects.filter(usuario = user, caja = c, estado = "GASTOS").values_list("importe", flat=True))
        balance = ingresos - gastos
        total_cajas.append((c, ingresos, gastos, balance))

    cat_ingresos = RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha__range=[fecha_inicial, fecha_final]).values_list("categoria", flat=True).distinct()

    pie_ingresos = []

    for ci in cat_ingresos:
        aux = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha__range=[fecha_inicial, fecha_final], categoria = ci).values_list("importe", flat=True)))
        list = RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha__range=[fecha_inicial, fecha_final], categoria = ci)
        pie_ingresos.append([ci, aux, list])

    pie_ingresos = sorted(pie_ingresos, key=lambda X : -X[1])

    aux_color = 63
    aux_color_2 = 0
    for pie in pie_ingresos: 
        if aux_color_2 > (120/2):
            color = (aux_color, 63, 186-aux_color_2)
        else:
            color = (aux_color, 63, 186)
        pie.append(color)
        aux_color += (186 - 63)/len(pie_ingresos)
        aux_color_2 += 120/len(pie_ingresos)

    cat_gastos = RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha__range=[fecha_inicial, fecha_final]).values_list("categoria", flat=True).distinct()

    pie_gastos = []

    aux_color = 0

    for cg in cat_gastos:
        aux = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha__range=[fecha_inicial, fecha_final], categoria = cg).values_list("importe", flat=True)))        
        list = RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha__range=[fecha_inicial, fecha_final], categoria = cg)
        pie_gastos.append([cg, aux, list])

    pie_gastos = sorted(pie_gastos, key=lambda X : -X[1])

    aux_color = 33
    aux_color_2 = 0 
    for pie in pie_gastos: 
        if aux_color_2 > (170/2):
            color = ((214-aux_color_2), aux_color, 42)
        else:
            color = (214, aux_color, 42)
        pie.append(color)
        aux_color += (214 - 33)/len(pie_gastos)
        aux_color_2 += 170/len(pie_gastos)


    list_cat_gasto = RegistroContable.objects.filter(usuario = user, estado = "GASTOS").values_list("categoria", flat=True).distinct()
    list_cat_ing = RegistroContable.objects.filter(usuario = user, estado = "INGRESOS").values_list("categoria", flat=True).distinct()


    ## Esquema mensual

    data_month = []

    try:

        fecha_1 = RegistroContable.objects.all().order_by("fecha")[0].fecha
        fecha_f = RegistroContable.objects.all().order_by("-fecha")[0].fecha
        fecha_f_auxiliar = datetime.date(fecha_f.year, fecha_f.month, 1)
        fechas = []

        fecha_auxiliar = datetime.date(fecha_1.year, fecha_1.month, 1)

        if fecha_1.month == 12:

            fecha_auxiliar_2 = datetime.date(fecha_1.year + 1, 1, 1)

        else:

            fecha_auxiliar_2 = datetime.date(fecha_1.year, fecha_1.month + 1, 1)

        while fecha_auxiliar <= fecha_f_auxiliar:

            mes = fecha_auxiliar
            ingresos = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha__range=[fecha_auxiliar, fecha_auxiliar_2]).values_list("importe", flat=True)))
            gastos = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha__range=[fecha_auxiliar, fecha_auxiliar_2]).values_list("importe", flat=True)))
            balance = ingresos - gastos
            data_month.append((mes, ingresos, gastos, balance))

            if fecha_auxiliar.month == 12:

                fecha_auxiliar = datetime.date(fecha_auxiliar.year + 1, 1, 1)

            else:

                fecha_auxiliar = datetime.date(fecha_auxiliar.year, fecha_auxiliar.month + 1, 1)

            if fecha_auxiliar_2.month == 12:

                fecha_auxiliar_2 = datetime.date(fecha_auxiliar_2.year + 1, 1, 1)

            else:

                fecha_auxiliar_2 = datetime.date(fecha_auxiliar_2.year, fecha_auxiliar_2.month + 1, 1)

    except:
        pass


    ##### Generales

    ingresos = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha__range=[fecha_inicial, fecha_final]).values_list("importe", flat=True)))

    gastos = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha__range=[fecha_inicial, fecha_final]).values_list("importe", flat=True)))
  
    balance = ingresos - gastos

    ingresos_t = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "INGRESOS").values_list("importe", flat=True)))

    gastos_t = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "GASTOS").values_list("importe", flat=True)))
  
    balance_t = ingresos_t - gastos_t

    ##### Semana

    semana = hoy.weekday()

    semana_1 = hoy - datetime.timedelta(days=hoy.weekday())

    semana_2 = semana_1

    datos_week = []

    for i in range(5):
        semana_2 = semana_2 + datetime.timedelta(days=7)
        ingresos_week = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "INGRESOS", fecha__range=[semana_1, semana_2]).values_list("importe", flat=True)))
        gastos_week = sum(np.array(RegistroContable.objects.filter(usuario = user, estado = "GASTOS", fecha__range=[semana_1, semana_2]).values_list("importe", flat=True)))
        balance_week = ingresos - gastos
        datos_week.append((semana_1, semana_2, ingresos_week, gastos_week, balance_week))
        semana_1 = semana_1 + datetime.timedelta(days=7)    


    return render(request, "chanchito/registro_contable.html", {'total_cajas':total_cajas, 'registros_totales':registros_totales,'datos_week':datos_week, 'data_month':data_month, 'list_cat_ing':list_cat_ing, 'list_cat_gasto':list_cat_gasto, 'pie_gastos':pie_gastos, 'pie_ingresos':pie_ingresos, 'hoy':hoy, 'datos':datos, "ingresos":ingresos, "gastos":gastos, "balance":balance, "ingresos_t":ingresos_t, "gastos_t":gastos_t, "balance_t":balance_t})

def editar_registro_contable(request):

    users_registro = RegistroContable.objects.filter(creador = request.user.username).values_list("usuario__identificacion", flat=True).distinct()

    data = []

    for i in users_registro:
        user = datosusuario.objects.get(identificacion = i)
        data_user = RegistroContable.objects.filter(creador = request.user.username, usuario = user)
        total_cajas = []
        cajas_cargadas =  RegistroContable.objects.filter(creador = request.user.username, usuario = user).values_list("caja", flat=True).distinct()
        for c in cajas_cargadas:
            ingresos = sum(RegistroContable.objects.filter(creador = request.user.username, usuario = user, caja = c, estado = "INGRESOS").values_list("importe", flat=True))
            gastos = sum(RegistroContable.objects.filter(creador = request.user.username, usuario = user, caja = c, estado = "GASTOS").values_list("importe", flat=True))
            balance = ingresos - gastos
            total_cajas.append((c, ingresos, gastos, balance))
        data.append((user, data_user, total_cajas))

    return render(request, "chanchito/registro_contable_editar.html", {"data":data})

class DescargarRegistroContable(TemplateView):

    def get(self, request, *args, **kwargs):

        # --> Iniciamos el Workbook
        wb = Workbook()

        # --> Primeros calculos

        data = RegistroContable.objects.all()

        ws = wb.active
        ws.title = "Resumen"
        ws["A1"] = "Registros contables"
        ws["A1"].font = Font(bold = True)
        ws["A2"] = "Para usar este archivo para modificar, elimina las primeras 3 filas y en acción agrega 'EDITAR' , 'NUEVO', 'BORRAR' "

        ws["A4"] = "Usuario"
        ws["B4"] = "Creador"
        ws["C4"] = "Fecha"
        ws["D4"] = "Tipo"
        ws["E4"] = "Caja"
        ws["F4"] = "Cuenta"
        ws["G4"] = "Categoria"
        ws["H4"] = "Importe"
        ws["I4"] = "Nota"
        ws["J4"] = "Acción"
        ws["K4"] = "Id"


        ws["A4"].font = Font(bold = True, color= "E8F8F8")
        ws["A4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["B4"].font = Font(bold = True, color= "E8F8F8")
        ws["B4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["C4"].font = Font(bold = True, color= "E8F8F8")
        ws["C4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["D4"].font = Font(bold = True, color= "E8F8F8")
        ws["D4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["E4"].font = Font(bold = True, color= "E8F8F8")
        ws["E4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["F4"].font = Font(bold = True, color= "E8F8F8")
        ws["F4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["G4"].font = Font(bold = True, color= "E8F8F8")
        ws["G4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["H4"].font = Font(bold = True, color= "E8F8F8")
        ws["H4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["I4"].font = Font(bold = True, color= "E8F8F8")
        ws["I4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["J4"].font = Font(bold = True, color= "E8F8F8")
        ws["J4"].fill =  PatternFill("solid", fgColor= "2C9E9D")
        ws["K4"].font = Font(bold = True, color= "E8F8F8")
        ws["K4"].fill =  PatternFill("solid", fgColor= "2C9E9D")



        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 17
        ws.column_dimensions['D'].width = 17
        ws.column_dimensions['E'].width = 17
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 7

        cont = 5

        for d in data:

            if request.user.username in d.usuario.identificacion or request.user.username in d.creador:

                ws = wb.active

                ws["A"+str(cont)] = d.usuario.identificacion
                ws["B"+str(cont)] = d.creador
                ws["C"+str(cont)] = d.fecha
                ws["D"+str(cont)] = d.estado
                ws["E"+str(cont)] = d.caja
                ws["F"+str(cont)] = d.cuenta
                ws["G"+str(cont)] = d.categoria
                if d.estado == "INGRESOS":
                    ws["H"+str(cont)] = d.importe
                else:
                    ws["H"+str(cont)] = - d.importe
                ws["I"+str(cont)] = d.nota
                ws["K"+str(cont)] = d.id

                ws["H"+str(cont)].number_format = '"$"" "#,##0.00_-'
                

                cont += 1


        #Establecer el nombre del archivo
        nombre_archivo = "RegistroContable.-{}.xls".format(str(d.usuario.identificacion))
        
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo).replace(',', '_')
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
