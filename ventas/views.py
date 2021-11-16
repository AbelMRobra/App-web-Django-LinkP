# Modelos de la BBDD usados
from django.contrib.auth.decorators import permission_required
from .models import EstudioMercado, PricingResumen, FeaturesProjects, FeaturesUni, DosierDeVenta, Clientescontacto
from proyectos.models import Unidades, Proyectos
from finanzas.models import Almacenero
from rrhh.models import datosusuario
from ventas.models import Pricing, ArchivosAreaVentas, VentasRealizadas, ArchivoFechaEntrega, ArchivoVariacionHormigon, ReclamosPostventa, \
                            ImgEnlacesProyecto
from presupuestos.models import Constantes, Desde, Registrodeconstantes
from crm.models import Consulta, Tipologia

# Librerias matematicas
import numpy as np
import numpy_financial as npf
from django.shortcuts import render
from django.views.generic import View
from django.conf import settings
from numpy.lib import twodim_base

from django.shortcuts import redirect
from django.template.loader import get_template
import datetime
import string
import requests
from datetime import date
import operator
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side 
from django.views.generic.base import TemplateView 
from django.http import HttpResponse 
from xhtml2pdf import pisa
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from .functions import calculo_cotizacion
from .functions_unidades import calculo_m2_unidad, cliente_crm, plan_financiacion_cotizador, info_para_cotizador
from .wabot import WABot
from .funciones.f_pricing import *
from .funciones.f_atributos import *
import smtplib


def comercial_principal(request):

    context = {}

    context["proyectos"] = Unidades.objects.all().values_list("proyecto__nombre", flat=True).distinct()

    if request.method == 'POST':

        proyecto = Proyectos.objects.get(nombre = request.POST["proyecto"])

        return redirect( 'Pricing', id_proyecto = proyecto.id )


    return render(request, 'comercial_principal.html', context)

def apparchivoscomercial(request):
    return render(request, 'apparchivoscomercial.html')

def encuestapostventa(request):

    if request.method == 'POST':
        datos_elegidos = request.POST.items()
        for dato in datos_elegidos:

            if dato[0] == "fecha":
                b = ArchivosAreaVentas(
                    fecha = request.POST['fecha'],
                    encuesta_postventa = request.FILES['adjunto']
                )
          
                b.save()
            if dato[0] == "delete":
                archivo = ArchivosAreaVentas.objects.get(id = int(request.POST['delete']))
                archivo.encuesta_postventa = None
                archivo.save()

    data = ArchivosAreaVentas.objects.filter(encuesta_postventa__isnull = False).order_by("-fecha")

    datos = {"data":data}

    return render(request, 'encuestapostventa.html', {"datos":datos})

def resumenprecio(request):


    busqueda = 1
    datos_pricing = PricingResumen.objects.all()
    datos = 0
    fecha = 0

    fechas = []

    datos_presupuesto = []

    #--> Parte para calcular precio presupuesto

    datos_desde = Desde.objects.all()

    constantes = Constantes.objects.all()

    usd_blue = Constantes.objects.get(nombre = "USD_BLUE")

    for i in datos_desde:

        costo = i.presupuesto.valor

        #Aqui calculo el precio min y sugerido

        costo = (costo/(1 + i.parametros.tasa_des_p))*(1 + i.parametros.soft)
        
        costo = costo*(1 + i.parametros.imprevitso)

        porc_terreno = i.parametros.terreno/i.parametros.proyecto.m2*100
        porc_link = i.parametros.link/i.parametros.proyecto.m2*100

        aumento_tem = i.parametros.tem_iibb*i.parametros.por_temiibb*(1+i.parametros.ganancia)

        aumento_comer = i.parametros.comer*(1+(porc_terreno + porc_link)/100)*(1+i.parametros.ganancia)
        

        costo = costo/(1-aumento_tem- aumento_comer)
        
        m2 = (i.parametros.proyecto.m2 - i.parametros.terreno - i.parametros.link)

        valor_costo = costo/m2

        #Aqui coloco la tasa de descuento


        fecha_entrega =  datetime.datetime.strptime(str(i.presupuesto.proyecto.fecha_f), '%Y-%m-%d')
        ahora = datetime.datetime.utcnow()
        fecha_inicial = ahora + datetime.timedelta(days = (365*2))

        if fecha_entrega > fecha_inicial:
            y = fecha_entrega.year - fecha_inicial.year
            n = fecha_entrega.month - fecha_inicial.month

            meses = y*12 + n

            valor_costo = -np.pv(fv=valor_costo, rate=i.parametros.tasa_des, nper=meses, pmt=0)


        #Calculo el valor final
        
        valor_final = valor_costo*(1 + i.parametros.ganancia)


        # Valorizo en dolares el precio de costo y sugerido

        valor_costo_usd = 0

        valor_final_usd = 0

        for c in constantes:

            if str(c.nombre) == 'USD_BLUE':

                valor_costo_usd = valor_costo/c.valor

                valor_final_usd = valor_final/c.valor

        i.valor_costo = valor_costo
        i.valor_costo_usd = valor_costo_usd
        i.valor_final = valor_final
        i.valor_final_usd = valor_final_usd

        i.save()

        #--> Parte para calcular precio promedio contado

        datos_unidades = Unidades.objects.filter(proyecto = i.parametros.proyecto, estado = "DISPONIBLE")
        unidades_totales = len(Unidades.objects.filter(proyecto = i.parametros.proyecto))
        unidades_disponibles = len(datos_unidades)

        if unidades_totales == 0:
            porcentaje_vendido = 0

        else:
            porcentaje_vendido = (1 - (unidades_disponibles/unidades_totales))*100

        m2_totales = 0

        sumatoria_contado = 0
    
        for dato in datos_unidades:

            if dato.sup_equiv > 0:

                m2 = dato.sup_equiv

            else:

                m2 = dato.sup_propia + dato.sup_balcon + dato.sup_comun + dato.sup_patio

            try:

                m2_panel = dato.sup_propia + dato.sup_balcon + dato.sup_comun + dato.sup_patio

                venta = VentasRealizadas.objects.get(unidad = dato.id)

                venta.m2 = m2_panel

                venta.asignacion = dato.asig

                venta.save()
            
            except:

                basura = 1
            
            try:

                param_uni = Pricing.objects.get(unidad = dato)
                
                desde = dato.proyecto.desde

                if dato.tipo == "COCHERA":
                    desde = dato.proyecto.desde*dato.proyecto.descuento_cochera

                if param_uni.frente == "SI":
                    desde = desde*dato.proyecto.recargo_frente

                if param_uni.piso_intermedio == "SI":
                    desde =desde*dato.proyecto.recargo_piso_intermedio

                if param_uni.cocina_separada == "SI":
                    desde = desde*dato.proyecto.recargo_cocina_separada

                if param_uni.local == "SI":
                    desde = desde*dato.proyecto.recargo_local

                if param_uni.menor_45_m2 == "SI":
                    desde = desde*dato.proyecto.recargo_menor_45

                if param_uni.menor_50_m2 == "SI":
                    desde = desde*dato.proyecto.recargo_menor_50

                if param_uni.otros == "SI":
                    desde = desde*dato.proyecto.recargo_otros 

                #Aqui calculamos el contado/financiado
                
                contado = desde*m2 

                sumatoria_contado = sumatoria_contado + contado
                m2_totales = m2_totales + m2

            except:

                basura = 1


        if m2_totales == 0:

            precio_promedio_contado = 0

        else:

            precio_promedio_contado = sumatoria_contado/m2_totales

        precio_promedio_contado_plus = precio_promedio_contado*1.05

        if valor_final == 0:

            var = 0

        else:

            var = ((precio_promedio_contado/i.valor_final)-1)*100

        fecha_pricing = PricingResumen.objects.order_by("-fecha")

        datos_presupuesto.append((i, precio_promedio_contado, precio_promedio_contado_plus, var, porcentaje_vendido, fecha_pricing))

    

    for dato in datos_pricing:
        
        fechas.append((dato.fecha, str(dato.fecha)))

    fechas = list(set(fechas))

    fechas.sort( reverse=True)

    if request.method == 'POST':

        #Trae los datos elegidos
        datos_elegidos = request.POST.items()

        for dato in datos_elegidos:
            if dato[0] == "fecha":
                datos = PricingResumen.objects.filter(fecha = dato[1])
                busqueda = 0
                fecha = dato[1]


    datos = {"fechas":fechas,
    "busqueda":busqueda,
    "datos":datos,
    "fecha":fecha,
    "datos_presupuesto":datos_presupuesto}

    return render(request, 'resumenprecio.html', {"datos":datos})

def panelunidades(request):

    datos = Unidades.objects.all().order_by("orden")

    proyectos = []

    datos_unidades = 0

    mensaje = 0

    otros_datos = 0


    for dato in datos:
        proyectos.append(dato.proyecto)

    proyectos = list(set(proyectos))

    if request.method == 'POST':

        #Trae los datos elegidos
        datos_elegidos = request.POST.items()

        list_proyectos = []
        aisgnacion = []
        disponibilidad = []

        for dato in datos_elegidos:

            if "Asig" in str(dato[0]):
                aisgnacion.append(dato[1])

            elif "Disp" in str(dato[0]):
                disponibilidad.append(dato[1])

            elif str(dato[0]) == "csrfmiddlewaretoken":
                basura = 1

            else:
                list_proyectos.append(dato[0])

        if len(list_proyectos)==0 or len(aisgnacion)==0 or len(disponibilidad)==0:
            mensaje = 1

        else:
            mensaje = 2
            otros_datos = []
            datos_tabla_unidad = []
            m2_totales = 0
            monto_total = 0
            cocheras = 0

            for proy in list_proyectos:
                for asig in aisgnacion:
                    for disp in disponibilidad:
                        
                        datos_unidades = Unidades.objects.filter(proyecto__nombre = proy, asig = asig, estado=disp)
                        
                        for dato in datos_unidades:
                            if dato.sup_equiv > 0:

                                m2 = round(dato.sup_equiv, 2)

                            else:

                                m2 = round((dato.sup_propia + dato.sup_balcon + dato.sup_comun + dato.sup_patio), 2)
                            
                            contado = m2*dato.proyecto.desde

                            features_unidad = FeaturesUni.objects.filter(unidad = dato)

                            for f2 in features_unidad:

                                contado = contado*f2.feature.inc

                            monto_total = monto_total + contado

                            desde = round((contado/m2), 4)

                            datos_tabla_unidad.append((dato, m2, contado, dato.id))
                            
                            m2_totales = m2_totales + m2
                            
                            if dato.tipo == "COCHERA":
                                cocheras += 1
                            

            cantidad = len(datos_tabla_unidad)

            departamentos = cantidad - cocheras

            otros_datos.append((m2_totales, cantidad, departamentos, cocheras, monto_total))

            datos_unidades = datos_tabla_unidad

            datos_unidades.sort(key=lambda datos_unidades: datos_unidades[3], reverse=False)


    datos = {"proyectos":proyectos, "datos":datos, "mensaje":mensaje, "datos_unidades":datos_unidades, "otros_datos":otros_datos}

    return render(request, 'panelunidades.html', {"datos":datos})

def variacionh(request):

    datos = ArchivoVariacionHormigon.objects.order_by("-fecha")

    busqueda = 0

    if request.method == 'POST':

        try:

            b = ArchivoVariacionHormigon(

                archivo = request.FILES['adjunto'],
            )

            b.save()

        except:
            
            busqueda = ArchivoVariacionHormigon.objects.get(id = request.POST['fecha']) 

    datos_hormigon = Registrodeconstantes.objects.filter(constante__nombre = "Hº VIVIENDA").order_by('fecha')

    year = datos_hormigon[0].fecha.year

    year_now = datetime.date.today().year

    datos_h = []

    valor_anterior = 0

    valor_inicial = 0

    while year != (year_now + 1):

        datos_year = []

        month = 1

        variacion_anual = 0

        for i in range(12):
            
            dia = datetime.date(year, month, 1)

            valor = Registrodeconstantes.objects.filter(constante__nombre = "Hº VIVIENDA", fecha = dia)

            # -> Este es la parte del flujo en caso de haber registros

            if len(valor) != 0:

                # -> Esta es la parte de la variación anual

                if valor_inicial != 0 and month !=12:

                    variacion_anual = (valor[0].valor/valor_inicial-1)*100

                if valor_inicial == 0:

                    valor_inicial = valor[0].valor

                if valor_anterior == 0:

                    variacion = 0

                    datos_year.append((dia, valor[0].valor, variacion))

                    valor_anterior = valor[0].valor

                    if month == 12:
                        
                        valor_inicial = valor[0].valor

                else:

                    variacion = (valor[0].valor/valor_anterior-1)*100

                    datos_year.append((dia, valor[0].valor, variacion))

                    valor_anterior = valor[0].valor

                    if month == 12:
                        
                        valor_inicial = valor[0].valor

                

            # -> Este es la parte del flujo en caso de no haber registros

            else:
                
                datos_year.append((dia, 0, 0))

            if month == 12:

                datos_year.append(variacion_anual)

            month += 1

      
        datos_h.append(datos_year)

        datos_h.sort(key=lambda datos_h: datos_h, reverse=True)

        year += 1

        horm = Constantes.objects.get(nombre = "Hº VIVIENDA")

    try:

        year_now = datetime.date.today()

        if year_now.month != 1:

            fecha_aux = datetime.date(year_now.year, year_now.month - 1 , 1)

        else:

            fecha_aux = datetime.date(year_now.year -1 , 12 , 1)

        valor_aux = Registrodeconstantes.objects.get(constante__nombre = "Hº VIVIENDA", fecha = fecha_aux)

        var = (horm.valor/valor_aux.valor - 1)*100

    except:

        var = 0


    return render(request, 'variacionhormigon.html', {"var":var, "datos_h":datos_h, "datos":datos, "busqueda":busqueda, "horm":horm})

def editarasignacion(request, id_unidad):

    id_unidad = id_unidad

    datos = Unidades.objects.get(id = id_unidad)

    if request.method == 'POST':

        valor_elegido = request.POST.items()

        for valor in valor_elegido:
            
            if valor[0] == "asignacion":

                if valor[1] == "1":
                    datos.asig = "PROYECTO"

                    datos.save()

                if valor[1] == "2":
                    datos.asig = "TERRENO"

                    datos.save()

                if valor[1] == "3":
                    datos.asig = "HON. LINK"

                    datos.save()

                if valor[1] == "4":

                    datos.asig = "SOCIOS"

                    datos.save()


                return redirect ('Panel de unidades')

    return render(request, 'editarasig.html', {"datos":datos} )

def cotizador(request, id_unidad):


    context = {}

    datos = Unidades.objects.get(id = id_unidad)
    today = datetime.date.today()

    m2 = calculo_m2_unidad(datos)

    precio_contado = m2*datos.proyecto.desde

    features_unidad = FeaturesUni.objects.filter(unidad = datos)

    for f2 in features_unidad:

        precio_contado = precio_contado*f2.feature.inc

    desde = round((precio_contado/m2), 4)

    if request.method == 'POST':


        try:
            context["cliente"] = cliente_crm(request.POST['email'], nombre = request.POST['nombre'], apellido = request.POST['apellido'], telefono = request.POST['telefono'])

        except:

            context["cliente"] = cliente_crm(request.POST['email'])

        context["resultados"] = plan_financiacion_cotizador(request.POST["anticipo"], request.POST["cuotas_esp"], request.POST["aporte"], request.POST["cuotas_p"], request.POST['observacion'], request.POST['descuento'], precio_contado, datos)

        context["info_coti_email"] = info_para_cotizador(context["resultados"])

    
    context["imagenes_carru"] = ImgEnlacesProyecto.objects.filter(proyecto = datos.proyecto)
    context["tiempo_restante"] = (datos.proyecto.fecha_f.year - today.year)*12 + (datos.proyecto.fecha_f.month - today.month)
    context["datos"] = datos
    context["precio_contado"] = precio_contado
    context["m2"] = m2

    return render(request, 'cotizador.html', context)

class PdfCotiza(View):

    def link_callback(self, uri, rel):
            """
            Convert HTML URIs to absolute system paths so xhtml2pdf can access those
            resources
            """
            sUrl = settings.STATIC_URL        # Typically /static/
            sRoot = settings.STATIC_ROOT      # Typically /home/userX/project_static/
            mUrl = settings.MEDIA_URL         # Typically /media/
            mRoot = settings.MEDIA_ROOT       # Typically /home/userX/project_static/media/

            if uri.startswith(mUrl):
                path = os.path.join(mRoot, uri.replace(mUrl, ""))
            elif uri.startswith(sUrl):
                path = os.path.join(sRoot, uri.replace(sUrl, ""))
            else:
                return uri

            # make sure that file exists
            if not os.path.isfile(path):
                raise Exception(
                        'media URI must start with %s or %s' % (sUrl, mUrl)
                )
            return path

    def get(self, request, id_unidad, id_cliente, info_coti, *args, **kwargs):

        #Creamos la información
        response_servidor = {"messages": "Perri"}
        cliente = Clientescontacto.objects.get(id = id_cliente)
        unidad = Unidades.objects.get(id = id_unidad)
        features_unidad = FeaturesUni.objects.filter(unidad = unidad)
        valor_hormigon = Constantes.objects.get(id = 7)

        today = datetime.date.today()

        # Saludo de bienvenida
        hora_actual = datetime.datetime.now()
        if hora_actual.hour >= 20:
            mensaje_email = "¡Buenas noches {}!".format(cliente.nombre)
        elif hora_actual.hour >= 13:
            mensaje_email = "¡Buenas tardes {}!".format(cliente.nombre)
        else:
            mensaje_email = "¡Buen dia {}!".format(cliente.nombre)

        data = calculo_cotizacion(unidad, features_unidad, info_coti, valor_hormigon)
        # Aqui llamamos y armamos el PDF
      
        template = get_template('cotizadorpdf.html')
        contexto = {'cliente':cliente, 
        'unidad':unidad, 
        'data':data,
        'today':today, 
        'logo':'{}{}'.format(settings.STATIC_URL, 'img/link.png'),
        'cabecera':'{}{}'.format(settings.STATIC_URL, 'img/fondo.png'),
        'fondo':'{}{}'.format(settings.STATIC_URL, 'img/fondo.jpg')}
        html = template.render(contexto)
        response = HttpResponse(content_type = "application/pdf")
        
        #response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'
        
        pisaStatus = pisa.CreatePDF(html, dest=response, link_callback=self.link_callback)
        
        if pisaStatus.err:
            
            return HttpResponse("Hay un error")

        # Establecemos conexion con el servidor smtp de gmail
        mailServer = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
        mailServer.ehlo()
        mailServer.starttls()
        mailServer.ehlo()
        mailServer.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

        # Construimos el mensaje simple
        
        mensaje = MIMEMultipart()
        mensaje.attach(MIMEText("""
        
{}

Enviamos la cotización de hoy!, cualquier consulta no dudes en comunicarte con nosotros

Ingresa a www.linkinversiones.com.ar para mas información

LINK Desarrollos inmobilairios
Vos confias porque nosotros cumplimos

Por favor no responder este email

        
        """.format(mensaje_email), 'plain'))
        mensaje['From']=settings.EMAIL_HOST_USER
        mensaje['To']=cliente.email
        mensaje['Subject']="LINK - Tu cotización {}".format(cliente.nombre)

        # Esta es la parte para adjuntar (prueba)
        mRoot = settings.MEDIA_ROOT

        if unidad.plano_venta:
            plano_adjunto = open(mRoot + "/{}".format(unidad.plano_venta.name), 'rb')
            adjunto_MIME = MIMEBase('application', "octet-stream")
            adjunto_MIME.set_payload(plano_adjunto.read())
            encoders.encode_base64(adjunto_MIME)
            adjunto_MIME.add_header('Content-Disposition', 'attachment; filename="Plano de la unidad.pdf"')
            mensaje.attach(adjunto_MIME)


        adjunto_MIME = MIMEBase('application', "octet-stream")
        adjunto_MIME.set_payload(response.content)
        encoders.encode_base64(adjunto_MIME)
        adjunto_MIME.add_header('Content-Disposition', 'attachment; filename="Tu_cotizacion.pdf"')
        mensaje.attach(adjunto_MIME)
        
        with open(mRoot + "/cotizacion{}{}.pdf".format(cliente.nombre, today).replace(" ", ""), 'wb') as f:
            f.write(response.content)
        
        name_coti_adjunta = "cotizacion{}{}.pdf".format(str(cliente.nombre).replace(" ", ""), today)

        # Envio del mensaje
        mailServer.sendmail(settings.EMAIL_HOST_USER,
                        cliente.email,
                        mensaje.as_string())

        # Aviso por wp
        try:
            jefe_ventas = datosusuario.objects.get(cargo = "JEFE DE VENTAS")
            send= "Hola!, {} ha enviado una cotización a {} de la siguiente unidad: {}{} - {}".format(request.user.username, cliente.nombre, unidad.piso_unidad, unidad.nombre_unidad, unidad.proyecto)
            bot_wp = WABot(response_servidor)
            bot_wp.send_message_user(str(jefe_ventas.Telefono), send)
            
        except:
            pass
        try:
            send= "Hola {}!, enviamos tu cotización al siguiente email {}, cualquier error o consulta no dudes en consultar! - Equipo Link".format(cliente.nombre, cliente.email)
            bot_wp = WABot(response_servidor)
            bot_wp.send_message_user(str(cliente.telefono), send)
        except:
            pass

        # Creo la consulta

        usuario = datosusuario.objects.get(identificacion = request.user.username)

        try:
            tipologia = Tipologia.objects.get(nombre = unidad.tipologia)
        except:
            tipologia = Tipologia(
                nombre = unidad.tipologia
            )
            tipologia.save()

        #try:
        new_consulta = Consulta(
            fecha = today,
            proyecto = unidad.proyecto,
            cliente = cliente,
            medio_contacto = 'RECOMENDACION',
            usuario = usuario,
            adjunto_propuesta = (name_coti_adjunta),
        )

        new_consulta.save()
        new_consulta.tipologia2.add(tipologia)
        new_consulta.save()

        #except:
            #pass
        
        return redirect('modificarcliente', id = cliente.id)

class DescargaPricing(TemplateView):

    def get(self, request, id_proyecto, *args, **kwargs):
        
        wb = Workbook()

        id_proyecto = id_proyecto

        # -----------> Toda la parte de calculo de información

        # Traemos la información necesaria

        proyecto = Proyectos.objects.get(id = id_proyecto)
        datos = Unidades.objects.filter(proyecto = proyecto)

        # Calculamos/establecemos algunos parametros

        fecha_entrega =  datetime.datetime.strptime(str(proyecto.fecha_f), '%Y-%m-%d')
        ahora = datetime.datetime.utcnow()
        y = fecha_entrega.year - ahora.year
        n = fecha_entrega.month - ahora.month
        meses = y*12 + n
        anticipo = 0.4

        # Variables a calcular en el proceso

        financiado = 0
        financiado_m2 = 0
        fin_ant = 0
        valor_cuotas = 0
        mensaje = 2
        otros_datos = []
        datos_unidad = []
        m2_totales = 0
        m2_totales_disp = 0
        cocheras = 0
        ingreso_ventas = 0
        iibb = 0
        comision = 0
        unidades_socios = 0
        sumatoria_contado = 0
        sumatoria_financiado = 0

        # Proceso 1: Calculo de datos
        for dato in datos:

            # Calculamos los m2 equivalente y total

            m2_equivalente = round(dato.sup_equiv, 2)
            m2_total = round((dato.sup_propia + dato.sup_balcon + dato.sup_comun + dato.sup_patio), 2)

            # Formula del m2 para calculos

            if dato.sup_equiv > 0:
                m2 = round(dato.sup_equiv, 2)
            else:
                m2 = round((dato.sup_propia + dato.sup_balcon + dato.sup_comun + dato.sup_patio), 2)

            # Calculamos el precio contado y financiado de ser posible, sino establecemos que no esta definido

            
            contado = m2*dato.proyecto.desde
            aumento = 1
            desde = dato.proyecto.desde
            param_uni = FeaturesUni.objects.filter(unidad = dato)
            features_unidad = FeaturesUni.objects.filter(unidad = dato)

            for f2 in features_unidad:

                contado = contado*f2.feature.inc
                aumento = 1*f2.feature.inc
            

            #Aqui calculamos el contado/financiado
         
            values = [0]
            for m in range((meses)):
                values.append(1)
            anticipo = 0.4
            valor_auxiliar = npf.npv(rate=(dato.proyecto.tasa_f/100), values=values)
            incremento = (meses/(1-anticipo)/(((anticipo/(1-anticipo))*meses)+valor_auxiliar))
            financiado = contado*incremento
            financiado_m2 = financiado/m2                
            fin_ant = financiado*anticipo
            valor_cuotas = (financiado - fin_ant)/meses


            # Aqui establecemos los datos de la venta

            venta = 0
            try:    
                venta= VentasRealizadas.objects.filter(unidad = dato.id).exclude(estado = "BAJA")
                contador = 0
                for v in venta:
                    contador += 1
                if contador == 0:
                    venta = 0        
            except:
                venta = 0

            #Aqui vamos armando los m2 totales , m2 de proyecto, cantidad de cocheras

            m2_totales = m2_totales + m2
            if dato.estado == "DISPONIBLE":
                try:
                    if dato.asig == "PROYECTO":
                        sumatoria_contado = sumatoria_contado + contado
                        sumatoria_financiado = sumatoria_financiado + financiado
                        m2_totales_disp = m2_totales_disp + m2
                except:
                    pass           
            if dato.tipo == "COCHERA":
                cocheras += 1
                            
            #Aqui sumamos los datos

            datos_unidad.append((dato, m2_equivalente, m2_total, aumento,  contado, desde, financiado, financiado_m2, venta, param_uni))

        # Ordenamos las unidades

        datos_unidad = sorted(datos_unidad, key=lambda unidad : unidad[0].orden)


        #Aqui calculo promedio contado y promedio financiado y otros datos que usaremos en el resumen

        cantidad = len(datos_unidad)
        departamentos = cantidad - cocheras
        promedio_contado = sumatoria_contado/m2_totales_disp
        promedio_financiado = sumatoria_financiado/m2_totales_disp

        # Proceso 2: Creado del Excel

        ws = wb.active
        ws.title = "ADVERTENCIA"

        ws.merge_cells("A2:K2")
        ws["A2"] = "UNA ADVERTENCIA ANTES DE AVANZAR"

        ws["A2"].alignment = Alignment(horizontal = "left")
        ws["A2"].font = Font(bold = True, color= "23346D", size = 20)

        ws.merge_cells("A5:K25")
        ws["A5"] = """
        La información que contiene este documento se considera de caracter CONFIDENCIAL.
        
         Esto quiere decir debes garantizar su protección y no debe ser divulgada sin el consentimiento de Link Inversiones S.R.L.
         
         Algunas recomendaciones:

         1 - Habla con tu responsable de área antes de pasar este documento
         2 - Si usas este documento fuera de las computadoras de la empresa, borra el archivo y vacia la papelera
         
         Gracias por tu atención

         Saludos!
         """
        ws["A5"].alignment = Alignment(horizontal = "left", vertical = "center", wrap_text=True)
        ws["A5"].font = Font(bold = True)

        ws = wb.create_sheet("My sheet")
        ws.title = "RESUMEN"

        ws["A2"] = "PROYECTO"
        ws["A3"] = "FECHA DE ENTREGA"
        ws["A4"] = "CANTIDAD DE UNIDADES"
        ws["A5"] = "M2 TOTALES"
        ws["A6"] = "PRECIO DE REPOSICIÓN CONTADO"
        ws["A7"] = "PRECIO DE REPOSICIÓN FINANCIADO"
        ws["A8"] = "PRECIO DESDE"

        ws["A2"].font = Font(bold = True, color= "FDFFFF")
        ws["A2"].fill =  PatternFill("solid", fgColor= "23346D")
        ws["A3"].font = Font(bold = True, color= "FDFFFF")
        ws["A3"].fill =  PatternFill("solid", fgColor= "23346D")
        ws["A4"].font = Font(bold = True, color= "FDFFFF")
        ws["A4"].fill =  PatternFill("solid", fgColor= "23346D")
        ws["A5"].font = Font(bold = True, color= "FDFFFF")
        ws["A5"].fill =  PatternFill("solid", fgColor= "23346D")
        ws["A6"].font = Font(bold = True, color= "FDFFFF")
        ws["A6"].fill =  PatternFill("solid", fgColor= "23346D")
        ws["A7"].font = Font(bold = True, color= "FDFFFF")
        ws["A7"].fill =  PatternFill("solid", fgColor= "23346D")
        ws["A8"].font = Font(bold = True, color= "FDFFFF")
        ws["A8"].fill =  PatternFill("solid", fgColor= "23346D")

        ws["B2"] = proyecto.nombre
        ws["B2"].font = Font(bold = True)
        ws["B2"].alignment = Alignment(horizontal = "center")
        ws["B3"] = fecha_entrega
        ws["B4"] = cantidad
        ws["B5"] = m2_totales
        ws["B6"] = promedio_contado
        ws["B7"] = promedio_financiado
        ws["B8"] = dato.proyecto.desde
 
        ws["B5"].number_format = '#,##0.00_-"M2"'
        ws["B6"].number_format = '"$ "#,##0.00_-'
        ws["B7"].number_format = '"$ "#,##0.00_-'
        ws["B8"].number_format = '"$ "#,##0.00_-'



        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        
        cont = 1
        
        for d in datos_unidad:

            if cont == 1:
                ws = wb.create_sheet("My sheet")
                ws.title = "DETALLE"
                ws["A"+str(cont)] = "PISO"
                ws["B"+str(cont)] = "N"
                ws["C"+str(cont)] = "TIPO"
                ws["D"+str(cont)] = "TIPOLOGIA"
                ws["E"+str(cont)] = "M2 EQUIV"
                ws["F"+str(cont)] = "M2"
                ws["G"+str(cont)] = "ESTADO"
                ws["H"+str(cont)] = "CONTADO"
                ws["I"+str(cont)] = "CONTADO M2"
                ws["J"+str(cont)] = "FINANCIADO"
                ws["K"+str(cont)] = "FINANCIADO M2"
                ws["L"+str(cont)] = "ASIGNACIÓN"
                ws["M"+str(cont)] = "VENDIDO A"
                ws["N"+str(cont)] = "%V"

                ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["I"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["J"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["K"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["L"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["M"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["N"+str(cont)].alignment = Alignment(horizontal = "center")
                

                ws["A"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["A"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["B"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["B"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["C"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["C"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["D"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["D"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["E"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["E"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["F"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["F"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["G"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["G"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["H"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["H"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["I"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["I"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["J"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["J"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["K"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["K"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["L"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["L"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["M"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["M"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["N"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["N"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")

                features_project = FeaturesProjects.objects.filter(proyecto = proyecto)

                alphabet_string = string.ascii_uppercase
                alphabet_list = list(alphabet_string)
                cont_alph = 14
                
                for f in features_project:

                    ws[str(alphabet_list[cont_alph])+str(cont)] = f.nombre
                    ws.column_dimensions[str(alphabet_list[cont_alph])].width = 15
                    ws[str(alphabet_list[cont_alph])+str(cont)].alignment = Alignment(horizontal = "center")
                    ws[str(alphabet_list[cont_alph])+str(cont)].font = Font(bold = True, color= "FDFFFF")
                    ws[str(alphabet_list[cont_alph])+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                    cont_alph += 1
                
 
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 8
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 10
                ws.column_dimensions['F'].width = 10
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 15
                ws.column_dimensions['I'].width = 15
                ws.column_dimensions['J'].width = 15
                ws.column_dimensions['K'].width = 15
                ws.column_dimensions['L'].width = 15
                ws.column_dimensions['M'].width = 20
                ws.column_dimensions['N'].width = 10
                



                # Aqui empiezan los datos

                ws["A"+str(cont+1)] = d[0].piso_unidad
                ws["B"+str(cont+1)] = d[0].nombre_unidad
                ws["C"+str(cont+1)] = d[0].tipo
                ws["D"+str(cont+1)] = d[0].tipologia
                ws["E"+str(cont+1)] = d[1]
                ws["F"+str(cont+1)] = d[2]
                ws["G"+str(cont+1)] = d[0].estado
                ws["H"+str(cont+1)] = d[4]
                ws["I"+str(cont+1)] = d[5]
                ws["J"+str(cont+1)] = d[6]
                ws["K"+str(cont+1)] = d[7]
                ws["L"+str(cont+1)] = d[0].asig

                if d[8] == 0:

                    ws["M"+str(cont+1)] = "SIN COMPRADOR"

                else:
                    ws["M"+str(cont+1)] = d[8][0].comprador

                ws["N"+str(cont+1)] = d[3]


                cont_alph = 14
                
                for f in features_project:

                    aux = len(FeaturesUni.objects.filter(feature = f, unidad = d[0]))

                    if aux:

                        ws[str(alphabet_list[cont_alph])+str(cont+1)] = "SI"
                        ws[str(alphabet_list[cont_alph])+str(cont+1)].alignment = Alignment(horizontal = "center")

                    else:
                        ws[str(alphabet_list[cont_alph])+str(cont+1)] = "NO"
                        ws[str(alphabet_list[cont_alph])+str(cont+1)].alignment = Alignment(horizontal = "center")

                    cont_alph += 1

                


                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].font = Font(bold = True)
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].number_format = '#,##0.00_-"M2"'
                ws["F"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont+1)].number_format = '#,##0.00_-"M2"'
                ws["G"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["I"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["I"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["J"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["K"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["L"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["M"+str(cont+1)].alignment = Alignment(horizontal = "left")
                ws["N"+str(cont+1)].alignment = Alignment(horizontal = "center")  

                cont += 1

            else:
                ws = wb["DETALLE"]

                ws["A"+str(cont+1)] = d[0].piso_unidad
                ws["B"+str(cont+1)] = d[0].nombre_unidad
                ws["C"+str(cont+1)] = d[0].tipo
                ws["D"+str(cont+1)] = d[0].tipologia
                ws["E"+str(cont+1)] = d[1]
                ws["F"+str(cont+1)] = d[2]
                ws["G"+str(cont+1)] = d[0].estado
                ws["H"+str(cont+1)] = d[4]
                ws["I"+str(cont+1)] = d[5]
                ws["J"+str(cont+1)] = d[6]
                ws["K"+str(cont+1)] = d[7]
                ws["L"+str(cont+1)] = d[0].asig

                if d[8] == 0:

                    ws["M"+str(cont+1)] = "SIN COMPRADOR"

                else:
                    ws["M"+str(cont+1)] = d[8][0].comprador
                    
                ws["N"+str(cont+1)] = d[3]

                cont_alph = 14
                
                for f in features_project:

                    aux = len(FeaturesUni.objects.filter(feature = f, unidad = d[0]))

                    if aux:

                        ws[str(alphabet_list[cont_alph])+str(cont+1)] = "SI"
                        ws[str(alphabet_list[cont_alph])+str(cont+1)].alignment = Alignment(horizontal = "center")

                    else:
                        ws[str(alphabet_list[cont_alph])+str(cont+1)] = "NO"
                        ws[str(alphabet_list[cont_alph])+str(cont+1)].alignment = Alignment(horizontal = "center")

                    cont_alph += 1

                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].font = Font(bold = True)
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].number_format = '#,##0.00_-"M2"'
                ws["F"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont+1)].number_format = '#,##0.00_-"M2"'
                ws["G"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["I"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["I"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["J"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["K"+str(cont+1)].number_format = '"$ "#,##0.00_-'
                ws["L"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["M"+str(cont+1)].alignment = Alignment(horizontal = "left")
                ws["N"+str(cont+1)].alignment = Alignment(horizontal = "center")

                cont += 1

        #Establecer el nombre del archivo
        nombre_archivo = "PRICING {}.xls".format(proyecto.nombre)
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


