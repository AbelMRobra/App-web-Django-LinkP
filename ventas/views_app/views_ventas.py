import datetime
from django.shortcuts import render, redirect
from ventas.models import VentasRealizadas
from proyectos.models import Unidades, Proyectos
from presupuestos.funciones import f_desde
from ..funciones.f_pricing import *
from users.funciones import f_generales
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill 
from django.views.generic.base import TemplateView 
from django.http import HttpResponse 

def ventas_detalles(request, id_venta):

    context = {}
    datos = VentasRealizadas.objects.get(id = id_venta)
    
    context['datos'] = datos

    return render(request, 'ventas/ventas_detalles.html', context)

def ventas_principal(request):

    context = {}

    context['datos'] = VentasRealizadas.objects.order_by("-fecha")

    return render(request, 'ventas/ventas_principal.html', context)

def ventas_eliminar(request, id_venta):

    context = {}
    datos = VentasRealizadas.objects.get(id = id_venta)
    context['datos'] = datos


    if request.method == 'POST':

        unidad = Unidades.objects.get(id = datos.unidad.id)

        unidad.estado = "DISPONIBLE"

        unidad.save()

        datos.delete()

        return redirect( 'Cargar Venta' )

    return render(request, 'ventas/ventas_eliminar.html', context)

def ventas_editar(request, id_venta):

    context = {}

    datos = VentasRealizadas.objects.get(id = id_venta)
    

    if request.method == 'POST':

        try:

            datos.comprador = request.POST['comprador']
            datos.precio_venta = float(request.POST['precio_venta'])
            datos.precio_venta_hormigon = float(request.POST['precio_venta_H'])
            datos.precio_contado = float(request.POST["precio_contado"])
            datos.anticipo = float(request.POST['anticipo'])
            datos.cuotas_pend = request.POST['cuotas']
            datos.tipo_venta = request.POST['tipo_venta']
            datos.fecha = request.POST['fecha']
            datos.observaciones = request.POST['observaciones']
            datos.save()

            context['mensaje']  = [1, "Unidad editada correctamente!"]

            datos = VentasRealizadas.objects.get(id = id_venta)

        except:

            context['mensaje']  = [0, "Error inesperado al tratar de editar"]

    context['datos'] = datos

    return render(request, 'ventas/ventas_editar.html', context)

def ventas_agregar(request):

    context = {}

    context['datos'] = Unidades.objects.all()

    if request.method == 'POST':

        comprador = request.POST['comprador']
        precio_venta = request.POST['precio_venta']
        precio_venta_hormigon = request.POST['precio_venta_H']
        anticipo = request.POST['anticipo']
        cuotas_pend = request.POST['cuotas']
        tipo_venta = request.POST['tipo_venta']
        unidad = Unidades.objects.get(id = int(request.POST['unidad']))
        fecha = request.POST['fecha']
        proyecto = Proyectos.objects.get(id = unidad.proyecto.id)
        observaciones = request.POST['observaciones']

        operaciones = VentasRealizadas.objects.filter(unidad = unidad).exclude(estado = "BAJA")

        if len(operaciones) > 1:

            context['mensaje']  = [0, "Esta unidad se encuentra asignada"]

        else:

            # try:
                m2 = unidades_calculo_m2(unidad.id)
                precio_pricing = unidades_calculo_precio_final(unidad.id)
                try:
                    precio_desde = unidades_calculo_precio_desde(unidad.id)*f_desde.presupuestos_precio_desde(unidad.proyecto.id)

                except:
                    precio_desde = 0

                if not "COCHERA" in unidad.tipo:
                    tipo_unidad = "DTO"
                else:
                    tipo_unidad = "COCHERA"


                nueva_venta = VentasRealizadas(

                    comprador = comprador,
                    fecha = fecha,
                    tipo_venta = tipo_venta,
                    unidad = unidad,
                    tipo_unidad = tipo_unidad,
                    proyecto = proyecto,
                    m2 = m2,
                    asignacion = unidad.asig,
                    precio_venta = precio_venta,
                    precio_venta_hormigon = precio_venta_hormigon,
                    precio_contado = request.POST["precio_contado"],
                    precio_pricing = precio_pricing[1],
                    precio_desde = precio_desde,
                    anticipo = anticipo,
                    cuotas_pend = cuotas_pend,
                    observaciones = observaciones,

                )

                nueva_venta.save()

                unidad.estado = "SEÑADA"
                unidad.save()

                context['mensaje'] = [1, "Unidad cargada correctamente"]

            # except:

            #     context['mensaje'] = [0, "Error inesperado al tratar de cargar"]


    return render(request, 'ventas/ventas_agregar.html', context)

def ventas_cargarplano(request,**kwargs):
    if request.method=='POST':
        
        id_proyecto=kwargs['id']
        unidad_id=request.POST.get('unidad')
        plano=request.FILES.get('plano')
        
        unidad=Unidades.objects.get(pk=int(unidad_id))
        if unidad:

            unidad.plano_venta=plano

            unidad.save()

            categoria = "pricing"
            accion = f"Cargo un plano en {unidad.proyecto}"

            f_generales.generales_registro_actividad(request.user.username, categoria, accion)


            return redirect('Pricing',id_proyecto)
        else:
            mensaje='No se pudo guardar el pdf'
            return redirect('Pricing',id_proyecto)

class ExcelRegistroVentas(TemplateView):

    def get(self, request, *args, **kwargs):
        
        wb = Workbook()

        #Aqui coloco la formula para calcular

        datos = VentasRealizadas.objects.order_by("fecha")

        ws = wb.active
        ws.title = "ADVERTENCIA"

        ws.column_dimensions['A'].width = 50
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A2:K2")
        ws["A2"] = "UNA ADVERTENCIA ANTES DE AVANZAR"

        ws["A2"].alignment = Alignment(horizontal = "left")
        ws["A2"].font = Font(bold = True, color= "23346D", size = 20)

        ws.merge_cells("A5:K25")
        ws["A5"] = """
        La información que contiene este documento se considera de caracter CONFIDENCIAL.

         Esto quiere decir debes garantizar su protección y no debe ser divulgada sin el consentimiento de Link Inversiones S.R.L.
         
         Algunas recomendaciones:

         1 - Habla con tu responsable de área antes de pasar este documento
         2 - Si usas este documento fuera de las computadoras de la empresa, borra el archivo y vacia la papelera
         
         Gracias por tu atención

         Saludos!
         """
        ws["A5"].alignment = Alignment(horizontal = "left", vertical = "center", wrap_text=True)
        ws["A5"].font = Font(bold = True)
        
        cont = 1
        
        for d in datos:

            if cont == 1:
                ws = wb.create_sheet("My sheet")
                ws.title = "REGISTRO"
                ws["A"+str(cont)] = "FECHA"
                ws["B"+str(cont)] = "PROYECTO"
                ws["C"+str(cont)] = "COMPRADOR"
                ws["D"+str(cont)] = "PISO"
                ws["E"+str(cont)] = "NOM"
                ws["F"+str(cont)] = "TIPO"
                ws["G"+str(cont)] = "TIPOLOGIA"
                ws["H"+str(cont)] = "SUPERFICIE"
                ws["I"+str(cont)] = "ASIGNACIÓN"
                ws["J"+str(cont)] = "PRECIO DE VENTA"
                ws["K"+str(cont)] = "ANOTACIONES"


                ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["I"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["J"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["K"+str(cont)].alignment = Alignment(horizontal = "center")


                ws["A"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["A"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["B"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["B"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["C"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["C"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["D"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["D"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["E"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["E"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["F"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["F"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["G"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["G"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["H"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["H"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["I"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["I"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["J"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["J"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")
                ws["K"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                ws["K"+str(cont)].fill =  PatternFill("solid", fgColor= "23346D")


                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 22
                ws.column_dimensions['D'].width = 6.86
                ws.column_dimensions['E'].width = 5
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 10
                ws.column_dimensions['H'].width = 10.29
                ws.column_dimensions['I'].width = 11.86
                ws.column_dimensions['J'].width = 16
                ws.column_dimensions['K'].width = 40

                ws["A"+str(cont+1)] = d.fecha
                ws["B"+str(cont+1)] = d.proyecto.nombre
                ws["C"+str(cont+1)] = d.comprador
                ws["D"+str(cont+1)] = d.unidad.piso_unidad
                ws["E"+str(cont+1)] = d.unidad.nombre_unidad
                ws["F"+str(cont+1)] = d.unidad.tipo
                ws["G"+str(cont+1)] = d.unidad.tipologia
                ws["H"+str(cont+1)] = d.m2
                ws["I"+str(cont+1)] = d.asignacion
                ws["J"+str(cont+1)] = d.precio_venta
                ws["K"+str(cont+1)] = d.observaciones


                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].font = Font(bold = True)
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].number_format = '#,##0.00_-'
                ws["I"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["J"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["K"+str(cont+1)].alignment = Alignment(horizontal = "left")
  

                cont += 1

            else:
                ws = wb["REGISTRO"]

                ws["A"+str(cont+1)] = d.fecha
                ws["B"+str(cont+1)] = d.proyecto.nombre
                ws["C"+str(cont+1)] = d.comprador
                ws["D"+str(cont+1)] = d.unidad.piso_unidad
                ws["E"+str(cont+1)] = d.unidad.nombre_unidad
                ws["F"+str(cont+1)] = d.unidad.tipo
                ws["G"+str(cont+1)] = d.unidad.tipologia
                ws["H"+str(cont+1)] = d.m2
                ws["I"+str(cont+1)] = d.asignacion
                ws["J"+str(cont+1)] = d.precio_venta
                ws["K"+str(cont+1)] = d.observaciones


                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont+1)].font = Font(bold = True)
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont+1)].number_format = '#,##0.00_-'
                ws["I"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["J"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["K"+str(cont+1)].alignment = Alignment(horizontal = "left")

                cont += 1

        #Establecer el nombre del archivo
        nombre_archivo = "RegistroVentas.xls"
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
