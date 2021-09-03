from rest_framework.generics import ListAPIView
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views.generic.base import TemplateView  
from django.conf import settings
from proyectos.models import Proyectos
from computos.models import Computos
from finanzas.models import Almacenero
from compras.models import Compras
from ventas.models import PricingResumen, VentasRealizadas
from registro.models import RegistroValorProyecto, RegistroConstantes
from rrhh.models import datosusuario
from .models import Articulos, Constantes, DatosProyectos, Prametros, Desde, Analisis, CompoAnalisis, Modelopresupuesto, Capitulos, Presupuestos, Registrodeconstantes, PorcentajeCapitulo, PresupuestosAlmacenados
import sqlite3
import pandas as pd
import numpy as np
import json
import datetime
import csv
import requests
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from .serializers import ArtSerializer
from .functions import auditor_presupuesto,auditor_presupuesto_p
from .wabot import WABot

def registroconstante(request):

    datos_constante = Constantes.objects.all()

    datos = []


    for dato in datos_constante:

        registros = []

        try:
            registro_constante = Registrodeconstantes.objects.filter(constante = dato).order_by("fecha")

            if len(registro_constante)>0:

                valor_referencia = registro_constante[0].valor

                for registro in registro_constante:

                    registros.append((registro, registro.valor/valor_referencia))

                datos.append((dato, registros))

        except:

            mensaje = "No hay registros de esta constanre"

    # Aqui armo para un grafico comparativo

    hormigon = Registrodeconstantes.objects.filter(constante__nombre = "Hº VIVIENDA").order_by("fecha")

    hormigon_list = []

    fecha = Registrodeconstantes.objects.values_list('fecha').filter(constante__nombre = "Hº VIVIENDA").order_by("fecha")

    contador = 0

    for h in hormigon:

        if contador == 0:

            hormigon_list.append(0)

            contador = h.valor

        else:

            hormigon_list.append(((h.valor/contador)-1)*100)


    usd = []
    usd_blue = []
    uva = []
    cac = []

    contador == 0

    valor_usd = 0
    valor_blue = 0
    valor_uva = 0
    valor_cac = 0


    for f in fecha:

        if contador == 0:

            usd.append(0)
            usd_blue.append(0)
            uva.append(0)
            cac.append(0)

            contador = 1

        else:

            try:

                
               
                if len(Registrodeconstantes.objects.filter(constante__nombre = "USD", fecha = f[0])) >0:

                    if valor_usd == 0:

                        valor_usd = Registrodeconstantes.objects.filter(constante__nombre = "USD", fecha = f[0])[0].valor

                    usd.append((((Registrodeconstantes.objects.filter(constante__nombre = "USD", fecha = f[0])[0].valor)/valor_usd) -1)*100 )
                
                else:

                    usd.append("")
            except:

                usd.append(valor_usd)
            

                
            try:

                            
                if len(Registrodeconstantes.objects.filter(constante__nombre = "USD_BLUE", fecha = f[0])) >0:

                    if valor_blue == 0:

                        valor_blue = Registrodeconstantes.objects.filter(constante__nombre = "USD_BLUE", fecha = f[0])[0].valor

                    usd_blue.append((((Registrodeconstantes.objects.filter(constante__nombre = "USD_BLUE", fecha = f[0])[0].valor)/valor_blue) -1)*100 )
                
                else:

                    usd_blue.append("")
            
            except:

                usd_blue.append(valor_blue)
            
            try:
                
                if len(Registrodeconstantes.objects.filter(constante__nombre = "UVA", fecha = f[0])) >0:

                    if valor_uva == 0:

                        valor_uva = Registrodeconstantes.objects.filter(constante__nombre = "UVA", fecha = f[0])[0].valor

                    uva.append((((Registrodeconstantes.objects.filter(constante__nombre = "UVA", fecha = f[0])[0].valor)/valor_uva) -1)*100 )
                
                else:

                    uva.append("")
            
            except:

                uva.append(valor_uva)


            try:
                
                if len(Registrodeconstantes.objects.filter(constante__nombre = "CAC_GENERAL", fecha = f[0])) >0:

                    if valor_cac == 0:

                        valor_cac = Registrodeconstantes.objects.filter(constante__nombre = "CAC_GENERAL", fecha = f[0])[0].valor

                    cac.append((((Registrodeconstantes.objects.filter(constante__nombre = "CAC_GENERAL", fecha = f[0])[0].valor)/valor_cac) -1)*100 )
                
                else:

                    cac.append("")
            except:

                cac.append(valor_cac)

    return render(request, 'constantes/historico.html', {'datos':datos, 'fecha':fecha, 'hormigon':hormigon_list, 'usd':usd, 'usd_blue':usd_blue, 'uva':uva, 'cac':cac})

def panel_presupuestos(request):

    proyectos_inicial = Proyectos.objects.order_by("nombre")

    proyectos = []

    for proyecto in proyectos_inicial:

        try:
            prueba = Modelopresupuesto.objects.filter(proyecto = proyecto)

            if len(prueba) > 0:

                proyectos.append(proyecto)

        except:

            var = "No tiene cargado nada"

        if request.method=='POST':
            id_proyecto=request.POST.get('proyecto')

            return redirect('presupuesto_proyecto',id_proyecto)
    return render(request, 'presupuestos/principal_presupuestos.html', {'proyectos':proyectos})


def presupuestostotal(request,id):
    
    response_servidor = {"messages": "Perri"}
    numero_prueba = "5493813540261-1599137567@g.us"
    # No tengo idea para que sirve

    presupuestador = 0

    variacion = 0

    variacion_year = 0  
    
    variacion_year_2 = 0  

    variacion_anuales = 0 


    datos = 0

    registro = 0

    proyecto = Proyectos.objects.get(pk=id)

    try:
        if int(request.POST['action']) == 2:

            data_aux = PresupuestosAlmacenados.objects.filter(proyecto = proyecto, nombre = "vigente")

            variable = PresupuestosAlmacenados(
                proyecto = proyecto,
                nombre = str("{}".format(datetime.date.today())),
                archivo = data_aux[0].archivo,
            )

            variable.save()

        if int(request.POST['action']) == 1:
            archivo_vigente = PresupuestosAlmacenados.objects.get(proyecto = proyecto, nombre = "vigente")
            archivo_vigente.nombre = str("{}".format(datetime.date.today()))
            archivo_vigente.save()
    except:
        pass


    # Comprobamos si existe el fichero csv en el almacen

    if len(PresupuestosAlmacenados.objects.filter(proyecto = proyecto, nombre = "vigente")) == 0:
        today = datetime.date.today()
        wb = Workbook()
        ws = wb.active
        ws.title = "Almacen"
        ws["A1"] = "Capitulo"
        ws["B1"] = "Modelo"
        ws["C1"] = "Analisis"
        ws["D1"] = "Cantidad An"
        ws["E1"] = "Articulo"
        ws["F1"] = "Cantidad Ar"
        ws["G1"] = "Precio"
        ws["H1"] = "Cantidad Art Totales"
        ws["I1"] = "Monto"

        contador = 2

        capitulo = Capitulos.objects.all()
        compo = CompoAnalisis.objects.all()
        computo = Computos.objects.all()
        for c in capitulo:
            modelo = Modelopresupuesto.objects.filter(proyecto = proyecto, capitulo = c ).order_by("orden")
            for d in modelo:
                cantidad = d.cantidad
                if d.cantidad == None:
                    if "SOLO MANO DE OBRA" in str(d.analisis): 
                        cantidad = 0
                        for h in computo:
                            if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                                cantidad = cantidad + h.valor_vacio                      
                        for e in compo:
                            if e.analisis == d.analisis:
                                ws["A{}".format(contador)] = c.nombre
                                ws["B{}".format(contador)] = d.id
                                ws["C{}".format(contador)] = d.analisis.codigo
                                ws["D{}".format(contador)] = cantidad
                                ws["E{}".format(contador)] = e.articulo.codigo
                                ws["F{}".format(contador)] = e.cantidad
                                ws["G{}".format(contador)] = e.articulo.valor
                                ws["H{}".format(contador)] = e.cantidad * cantidad
                                ws["I{}".format(contador)] = e.cantidad * e.articulo.valor * cantidad
                                contador += 1

                    else:

                        cantidad = 0

                        for h in computo:

                            if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                                
                                cantidad = cantidad + h.valor_lleno

                        for e in compo:

                            if e.analisis == d.analisis:

                                ws["A{}".format(contador)] = c.nombre
                                ws["B{}".format(contador)] = d.id
                                ws["C{}".format(contador)] = d.analisis.codigo
                                ws["D{}".format(contador)] = cantidad
                                ws["E{}".format(contador)] = e.articulo.codigo
                                ws["F{}".format(contador)] = e.cantidad
                                ws["G{}".format(contador)] = e.articulo.valor
                                ws["H{}".format(contador)] = e.cantidad * cantidad
                                ws["I{}".format(contador)] = e.cantidad * e.articulo.valor * cantidad
                                contador += 1
                        
                else:

                    for e in compo:

                        if e.analisis == d.analisis:

                            ws["A{}".format(contador)] = c.nombre
                            ws["B{}".format(contador)] = d.id
                            ws["C{}".format(contador)] = d.analisis.codigo
                            ws["D{}".format(contador)] = cantidad
                            ws["E{}".format(contador)] = e.articulo.codigo
                            ws["F{}".format(contador)] = e.cantidad
                            ws["G{}".format(contador)] = e.articulo.valor
                            ws["H{}".format(contador)] = e.cantidad * cantidad
                            ws["I{}".format(contador)] = e.cantidad * e.articulo.valor * cantidad
                            contador += 1

        #Establecer el nombre del archivo
        nombre_archivo = "{}.{}Almacen.xls".format(str(proyecto.nombre).replace(" ", ""),str(today))
        nombre_archivo
        mUrl = settings.MEDIA_URL
        mRoot = settings.MEDIA_ROOT
        wb.save(mRoot + "/{}".format(nombre_archivo))
        
        variable = PresupuestosAlmacenados(
            proyecto = proyecto,
            nombre = "vigente",
            archivo = (nombre_archivo),
        )

        variable.save()

        # BOT la actualización

        try:

            archivo = PresupuestosAlmacenados.objects.filter(proyecto = proyecto, nombre = "vigente")[0].archivo
            df = pd.read_excel(archivo)
            repo_nuevo = sum(np.array(df['Monto'].values))

            anterior_archivo = PresupuestosAlmacenados.objects.filter(proyecto = proyecto).order_by("-id").exclude(nombre = "vigente")[0].archivo
            df = pd.read_excel(anterior_archivo)
            repo_anterior = sum(np.array(df['Monto'].values))

            var = round((repo_nuevo/repo_anterior-1)*100, 2)

            if var != 0:

                send = "{} ha actualizado {}. Variación: {}%".format(request.user.first_name, proyecto.nombre, var)
                id = "-455382561"
                token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"
                url = "https://api.telegram.org/bot" + token + "/sendMessage"
                params = {
                    'chat_id' : id,
                    'text' : send
                }
                requests.post(url, params=params)
                bot_wp = WABot(response_servidor)
                bot_wp.send_message(numero_prueba, send)

            else:
                send = "{} guardo una copia de {}".format(request.user.first_name, proyecto.nombre)
                id = "-455382561"
                token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"
                url = "https://api.telegram.org/bot" + token + "/sendMessage"
                params = {
                    'chat_id' : id,
                    'text' : send
                }
                requests.post(url, params=params)
                bot_wp = WABot(response_servidor)
                bot_wp.send_message(numero_prueba, send)

            if proyecto.presupuesto == "BASE" and var != 0:

                send = "{}, has actualizado el proyecto BASE, empieza el proceso de actualización de proyectos extrapolados".format(request.user.first_name)

                id = "-455382561"

                token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"

                url = "https://api.telegram.org/bot" + token + "/sendMessage"

                params = {
                    'chat_id' : id,
                    'text' : send
                }

                requests.post(url, params=params)
                bot_wp = WABot(response_servidor)
                bot_wp.send_message(numero_prueba, send)

                proyectos_extrapolados = Proyectos.objects.filter(presupuesto = "EXTRAPOLADO")

                send_1 = "Proyectos actualizados: "
                send_2 = "Proyectos sin actualizar: "

                for p in proyectos_extrapolados:

                    try:
                    
                    
                        aux_var = Presupuestos.objects.get(proyecto = p)
                        aux_var.valor = aux_var.valor * (1+(var/100))
                        aux_var.saldo = aux_var.saldo * (1+(var/100))
                        aux_var.saldo_mat = aux_var.saldo_mat * (1+(var/100))
                        aux_var.saldo_mo =  aux_var.saldo_mo * (1+(var/100))
                        aux_var.save()

                        almacenero = Almacenero.objects.get(proyecto = p)

                        iva_compras = (aux_var.imprevisto + aux_var.saldo_mat + aux_var.saldo_mo + aux_var.credito + aux_var.fdr)*0.07875

                        almacenero.pendiente_iva_ventas = iva_compras

                        almacenero.save()

                        send_1 += "{} con {}% - ".format(p, var)

                        

                    except:

                        send_2 += "{} - ".format(p)


                params = {
                    'chat_id' : id,
                    'text' : send_1
                }

                requests.post(url, params=params)
                bot_wp = WABot(response_servidor)
                bot_wp.send_message(numero_prueba, send_1)

                params = {
                    'chat_id' : id,
                    'text' : send_2
                }

                requests.post(url, params=params)
                bot_wp = WABot(response_servidor)
                bot_wp.send_message(numero_prueba, send_2)

                send = "Proceso de actualización de proyectos extrapolados completo. Tambien actualice el IVA en el almacenero. Disculpen los mensajes"

                params = {
                    'chat_id' : id,
                    'text' : send
                }

                requests.post(url, params=params)
                bot_wp = WABot(response_servidor)
                bot_wp.send_message(numero_prueba, send)


        except:
            pass

    try:

        presup_info = Presupuestos.objects.get(proyecto = proyecto)
        presupuestador = datosusuario.objects.get(identificacion=presup_info.presupuestador)

    except:

        presupuestador = 0

    # Completamos los datos del tablero

    archivo = PresupuestosAlmacenados.objects.filter(proyecto = proyecto, nombre = "vigente")[0].archivo

    df = pd.read_excel(archivo)

    valor_reposicion = sum(np.array(df['Monto'].values))


    # saldo

    list_articulos = df['Articulo'].unique()

    valor_saldo = 0
    valor_proyecto_materiales = 0
    valor_proyecto_mo = 0

    for art in list_articulos:
        cantidad_solicitada = sum(np.array(df[df['Articulo'] == art]['Cantidad Art Totales'].values))
        valor = Articulos.objects.get(codigo = art).valor
        comprados = sum(np.array(Compras.objects.filter(proyecto = proyecto, articulo__codigo = art).values_list("cantidad", flat = True)))
        saldo_art = cantidad_solicitada*valor - valor*comprados
        if saldo_art > 0:
            valor_saldo = valor_saldo + saldo_art

            if str(art)[0] == "3":
                valor_proyecto_materiales += saldo_art
            else:
                valor_proyecto_mo += saldo_art

    # Guardamos los saldos en el almacenero

    try:

        Saldo_act = Presupuestos.objects.get(proyecto = proyecto)
        Saldo_act.saldo = valor_saldo
        Saldo_act.saldo_mat = valor_proyecto_materiales
        Saldo_act.saldo_mo = valor_proyecto_mo

        Saldo_act.save()

        almacenero = Almacenero.objects.get(proyecto = proyecto)

        iva_compras = (Saldo_act.imprevisto + Saldo_act.saldo_mat + Saldo_act.saldo_mo + Saldo_act.credito + Saldo_act.fdr)*0.07875

        almacenero.pendiente_iva_ventas = iva_compras

        almacenero.save()

    except:
        pass

    # Crea el avance

    avance = 0

    if valor_reposicion != 0:
        avance = (1 - (valor_saldo/valor_reposicion))*100
        pendiente = 100 - avance

    # Guarda los datos

    datos = []

    datos.append((proyecto, valor_reposicion, valor_saldo, avance, pendiente))

    # Crea el historico

    valor_proyecto = RegistroValorProyecto.objects.filter(proyecto = proyecto)

    registro = []

    for valor in valor_proyecto:

        registro.append((valor.fecha, valor.precio_proyecto/1000000))

    try:

        Presup_act = Presupuestos.objects.get(proyecto = proyecto)
        Presup_act.valor = valor_reposicion
        Presup_act.save()

        # Trato de establecer el precio de Link-P

        valor_linkp = valor_reposicion
        parametros = Prametros.objects.get(proyecto = proyecto)
        valor_linkp = (valor_linkp/(1 + parametros.tasa_des_p))*(1 + parametros.soft)       
        valor_linkp = valor_linkp*(1 + parametros.imprevitso)
        porc_terreno = parametros.terreno/parametros.proyecto.m2*100
        porc_link = parametros.link/parametros.proyecto.m2*100
        aumento_tem = parametros.tem_iibb*parametros.por_temiibb*(1+parametros.ganancia)
        aumento_comer = parametros.comer*(1+(porc_terreno + porc_link)/100)*(1+parametros.ganancia)           
        valor_linkp = valor_linkp/(1-aumento_tem- aumento_comer)           
        m2 = (parametros.proyecto.m2 - parametros.terreno - parametros.link)
        valor_costo = valor_linkp/m2
        proyecto.precio_linkp = valor_costo
        proyecto.save()

    except:
        pass

    registro = registro[-60:]


    try:

        variacion = (((valor_reposicion/1000000)/registro[-30][1]) -1)*100

        today = datetime.date.today()

        date = datetime.date(today.year, 1, 1)

        date_2 = datetime.date((today.year - 1), 1, 1)

        try:

            dato = RegistroValorProyecto.objects.filter(fecha = date, proyecto = proyectos)

            valor = (((valor_reposicion)/dato[0].precio_proyecto) -1)*100

            variacion_year = [date, valor]


        except:

            variacion_year = 0

            variacion_anuales = 0

        try:
            dato_1 = RegistroValorProyecto.objects.filter(fecha = date, proyecto = proyectos)
            dato_2 = RegistroValorProyecto.objects.filter(fecha = date_2, proyecto = proyectos)

            valor = ((dato_1[0].precio_proyecto/dato_2[0].precio_proyecto) -1)*100

            variacion_year_2 = [date_2, valor]

        except:

            variacion_year_2 = 0

    except:

        variacion = 0

    proyectos = 0

    return render(request, 'presupuestos/presupuesto_proyecto.html', {"datos":datos, "proyectos":proyectos, "valor":registro, "presupuestador":presupuestador, "variacion":variacion, "variacion_year":variacion_year, "variacion_year_2":variacion_year_2, "variacion_anuales":variacion_anuales})

def presupuestorepcompleto(request, id_proyecto):

    proyecto = Proyectos.objects.get(id = id_proyecto)
    capitulo = Capitulos.objects.all()
    compo = CompoAnalisis.objects.all()
    computo = Computos.objects.all()

    archivo = PresupuestosAlmacenados.objects.filter(proyecto = proyecto, nombre = "vigente")[0].archivo
    df = pd.read_excel(archivo)

    crudo = []

    valor_proyecto = sum(np.array(df['Monto'].values))

    # En ese bucle revisamos los capitulos

    for c in capitulo:

        # Aqui armo el listado de analisis del capitulo

        listado_analisis = []

        valor_capitulo = 0

        # En este bucle revisamos el modelo

        modelo = Modelopresupuesto.objects.filter(proyecto = proyecto, capitulo = c ).order_by("orden")

        for d in modelo:

            valor_analisis = sum(np.array(df[df['Modelo'] == d.id]['Cantidad Ar'].values) * np.array(df[df['Modelo'] == d.id]['Precio'].values))

            
            if len(df[df['Modelo'] == d.id]) > 0:
                
                cantidad = df[df['Modelo'] == d.id]['Cantidad An'].values[0]

                valor_mod = sum(np.array(df[df['Modelo'] == d.id]['Monto'].values))

                listado_analisis.append((d, valor_analisis, cantidad, valor_mod))  

                valor_capitulo = valor_capitulo + valor_analisis*cantidad
        
        crudo.append((c, valor_capitulo, 0.0, listado_analisis))

    datos =[]

    for i in crudo:
        i = list(i)
        i[2] = i[1]/valor_proyecto*100

        try:

            dato = PorcentajeCapitulo.objects.get(proyecto = proyecto, capitulo = i[0])

            dato.porcentaje = i[2]

            dato.save()

        except:

            b = PorcentajeCapitulo(
                proyecto = proyecto,
                capitulo = i[0],
                porcentaje = i[2]
                )

            b.save()

        i = tuple(i)
        datos.append(i)

    valor_proyecto_completo = valor_proyecto

    datos = {"datos":datos, "proyecto":proyecto, "valor_proyecto":valor_proyecto,"valor_proyecto_completo":valor_proyecto_completo}
    
    return render(request, 'presupuestos/presuprepabierto.html', {"datos":datos, "id_proyecto": id_proyecto})

def saldocapitulo(request, id_proyecto):

    #Armamos los datos para ver el presupuesto por capitulo

    datos = PresupuestoPorCapitulo(id_proyecto)

    datos_viejos = datos

    datos_presupuesto = []

    for componentes in datos_viejos:

        valor_capitulo = 0

        for articulos in componentes[2]:

            valor_capitulo = valor_capitulo + articulos[0].valor*articulos[1]
        
        datos_presupuesto.append((componentes[0], componentes[1], valor_capitulo ))


    #Armamos el saldo de cada capitulo

    saldo = Saldoporcapitulo(id_proyecto)

    datos_viejos = saldo

    datos_saldo = []

    valor_saldo = 0

    for componentes in datos_viejos:

        saldo_capitulo = 0

        for articulos in componentes[2]:

            if articulos[1] > 0:

                saldo_capitulo = saldo_capitulo + articulos[0].valor*articulos[1]
        
        datos_saldo.append((componentes[0], componentes[1], saldo_capitulo ))

        valor_saldo = valor_saldo + saldo_capitulo

    # Combinamos ambos

    datos = []

    for p in datos_presupuesto:
        for s in datos_saldo:
            if p[0] == s[0]:

                avance = 0
                inc = 0

                if p[2] != 0:

                    avance = (1 - s[2]/p[2])*100

                    if valor_saldo != 0:

                        inc = (s[2]/valor_saldo)*100  
                    
                    else:
                        inc = 100

                datos.append((p[0], p[1], p[2], s[2], avance, inc))

    
    proyecto = Proyectos.objects.get(id = id_proyecto)

    datos = {"proyecto":proyecto, "datos":datos, "saldo":valor_saldo}
                
    return render(request, 'presupuestos/saldocapitulo.html', {"datos":datos, "id_proyecto":id_proyecto})

def SaldoCapArticulos(request, id_proyecto, id_capitulo):

    #Armamos el saldo de cada capitulo

    saldo = Saldoporcapitulo(id_proyecto)

    datos_viejos = saldo

    datos_saldo = []
    capitulo = []

    for componentes in datos_viejos:
        if int(componentes[1].id) == int(id_capitulo):
            
            datos_saldo.append(componentes[2])
            capitulo.append(componentes[1])

    articulos = []

    for articulo in datos_saldo[0]:
        articulos.append(articulo[0])

    articulos = list(set(articulos))

    articulos_cant = []

    for articulo in articulos:

        cantidad = 0

        for art_can in datos_saldo[0]:

            if articulo == art_can[0] and art_can[1]>0:
                cantidad = cantidad + art_can[1]
        articulos_cant.append((articulo, cantidad))
    
    saldo_cap = 0

    datos_viejos = articulos_cant
    datos_saldo = []

    for dato in datos_viejos:
        saldo_cap = saldo_cap + dato[0].valor*dato[1]
        datos_saldo.append((dato[0], dato[1], float(dato[0].valor*dato[1])))

    datos_viejos = datos_saldo
    datos_saldo = []

    for dato in datos_viejos:
        if saldo_cap != 0:
            inc = float(dato[2])/float(saldo_cap)*100
        else:
            inc = 0
        datos_saldo.append((dato[0], dato[1], dato[2], inc))


    if len(datos_saldo) == 0:
        datos_saldo = 0

    else:

        datos_saldo.sort(key=lambda tup: tup[3], reverse=True)

    proyecto = Proyectos.objects.get(id = id_proyecto)

    datos = {"proyecto":proyecto,
     "datos_saldo":datos_saldo, "capitulo":capitulo,
     "saldo":saldo_cap}

    return render(request, 'presupuestos/saldoartcapitulo.html', {"datos":datos, "id_proyecto": id_proyecto})

def debugsa(request, id_proyecto):

    datos = debugsaldo(id_proyecto)

    return render(request, 'presupuestos/debugsaldo.html', {"datos":datos})

def creditos(request, id_proyecto):

    proyecto = Proyectos.objects.get(id = id_proyecto)

    datos = Creditocapitulo(id_proyecto)

    valor_saldo = 0

    for dato in datos:
        valor_saldo = valor_saldo + dato[4]

    #Guardamos el valor del credito en la base de presupuestos

    try:

        Cred_act = Presupuestos.objects.get(proyecto = proyecto)

        Cred_act.credito = valor_saldo

        Cred_act.save()

    except:
        pass

     #Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        datos_viejos = datos

        datos = []   

        for i in palabra_buscar:

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                buscador = (str(i[0]))

                if palabra.lower() in buscador.lower():

                    datos.append(i)


    #Aqui termina el filtro

    datos = {"datos":datos,
    "proyecto":proyecto,
    "valor_saldo":valor_saldo}

  
    return render(request, 'presupuestos/creditos.html', {"datos":datos})

def fdr(request, id_proyecto):

    proyecto = Proyectos.objects.get(id = id_proyecto)

    datos = Fondosdereparo(id_proyecto)

    valor_fdr = 0

    for dato in datos:
        valor_fdr = valor_fdr + dato[1]

    try:

        Fdr_act = Presupuestos.objects.get(proyecto = proyecto)

        Fdr_act.fdr = -valor_fdr

        Fdr_act.save()

    except:
        pass

     #Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        datos_viejos = datos

        datos = []   

        for i in palabra_buscar:

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                buscador = (str(i[0]))

                if palabra.lower() in buscador.lower():

                    datos.append(i)


    #Aqui termina el filtro

    datos = {"datos":datos,
    "proyecto":proyecto,
    "valor_fdr":valor_fdr}

  
    return render(request, 'presupuestos/fdr.html', {"datos":datos, "id_proyecto": id_proyecto})

def anticiposf(request, id_proyecto):

    proyecto = Proyectos.objects.get(id = id_proyecto)

    datos = AnticiposFinan(id_proyecto)

    valor_ant = 0

    for dato in datos:
        valor_ant = valor_ant + dato[1]

    try:

        Ant_act = Presupuestos.objects.get(proyecto = proyecto)

        Ant_act.anticipos = -valor_ant

        Ant_act.save()

    except:
        pass

     #Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        datos_viejos = datos

        datos = []   

        for i in palabra_buscar:

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                buscador = (str(i[0]))

                if palabra.lower() in buscador.lower():

                    datos.append(i)


    #Aqui termina el filtro

    datos = {"datos":datos,
    "proyecto":proyecto,
    "valor_fdr":valor_ant}

  
    return render(request, 'presupuestos/anticiposf.html', {"datos":datos, "id_proyecto": id_proyecto})

def explosion(request, id_proyecto):

    proyecto = Proyectos.objects.get(id = id_proyecto)
    modelo = Modelopresupuesto.objects.filter(proyecto = proyecto)

    #Version 2 de explosión
    
    crudo_analisis = []

    for i in modelo:

        if i.cantidad != None:

            crudo_analisis.append((i.analisis, i.cantidad))

        else:

            if "SOLO MANO DE OBRA" in str(i.analisis.nombre):

                computos = Computos.objects.values_list('valor_vacio').filter(tipologia = i.vinculacion, proyecto = proyecto)

                cantidad = sum(np.array(computos))

                crudo_analisis.append((i.analisis, cantidad))

            else:

                computos = Computos.objects.values_list('valor_lleno').filter(tipologia = i.vinculacion, proyecto = proyecto)

                cantidad = sum(np.array(computos))

                crudo_analisis.append((i.analisis, cantidad))

    crudo_articulos = []


    for c in crudo_analisis:

        analisis = CompoAnalisis.objects.filter(analisis = c[0])

        cantidad = 0

        for d in analisis:

            cantidad = d.cantidad*c[1]

            crudo_articulos.append((d.articulo, cantidad))

    datos = []

    for t in crudo_articulos:
        datos.append(t[0])

    datos = list(set(datos))

    datos_viejos = datos
    datos = []

    for i in datos_viejos:
        cantidad = 0
        for c in crudo_articulos:
            if i == c[0]:
                cantidad = cantidad + c[1]
        datos.append((i, cantidad))

    datos_viejos = datos
    datos = []

    for i in datos_viejos:

        comprado = sum(np.array(Compras.objects.values_list('cantidad').filter(proyecto = proyecto, articulo = i[0])))
        
        cantidad_saldo = i[1] - comprado

        saldo = cantidad_saldo * i[0].valor
        
        datos.append((i[0], i[1], comprado, cantidad_saldo, saldo ))

          #Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        datos_viejos = datos

        datos = []   

        for i in palabra_buscar:

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                buscador = (str(i[0]))

                if palabra.lower() in buscador.lower():

                    datos.append(i)

    #Aqui termina el filtro

    datos = {"datos":datos,
    "proyecto":proyecto}

    return render(request, 'presupuestos/explosion.html', {"datos":datos,"id_proyecto":id_proyecto} )

def presupuestosanalisis(request, id_proyecto, id_capitulo):

    proyecto = Proyectos.objects.get(id = id_proyecto)
    capitulo = Capitulos.objects.get(id = id_capitulo)
    compo = CompoAnalisis.objects.all()
    computo = Computos.objects.all()
    modelo = Modelopresupuesto.objects.filter(proyecto = proyecto)

    crudo = []

    valor_capitulo = 0

    for d in modelo:

        if d.capitulo == capitulo and d.proyecto == proyecto:

            if d.cantidad == None:

                if "SOLO MANO DE OBRA" in str(d.analisis):

                    valor_analisis = 0

                    for e in compo:

                        if e.analisis == d.analisis:

                            valor_analisis = valor_analisis + e.articulo.valor*e.cantidad

                    cantidad = 0

                    for h in computo:
                        if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                            cantidad = cantidad + h.valor_vacio  

                    total_parcial = valor_analisis*cantidad
                    crudo.append((d.analisis, valor_analisis, cantidad, total_parcial, 0.0)) 

                    valor_capitulo = valor_capitulo + valor_analisis*cantidad

                else:
                    valor_analisis = 0

                    for e in compo:

                        if e.analisis == d.analisis:

                            valor_analisis = valor_analisis + e.articulo.valor*e.cantidad

                    cantidad = 0

                    for h in computo:
                        if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                            cantidad = cantidad + h.valor_lleno 

                    total_parcial = valor_analisis*cantidad
                    crudo.append((d.analisis, valor_analisis, cantidad, total_parcial, 0.0, d.vinculacion)) 

                    valor_capitulo = valor_capitulo + valor_analisis*cantidad
            else:

                valor_analisis = 0

                for e in compo:

                    if e.analisis == d.analisis:

                        valor_analisis = valor_analisis + e.articulo.valor*e.cantidad
                
                total_parcial = valor_analisis*float(d.cantidad)
                
                crudo.append((d.analisis, valor_analisis, d.cantidad, total_parcial, 0.0, d.vinculacion))
                
                valor_capitulo = valor_capitulo + valor_analisis*float(d.cantidad)
    datos =[]

    for i in crudo:
        i = list(i)
        i[4] = i[3]/valor_capitulo*100
        i = tuple(i)
        datos.append(i)

    datos.sort(key=lambda tup: tup[4], reverse=True)

    datos = {"datos":datos, "proyecto":proyecto, "capitulo":capitulo}

    return render(request, 'presupuestos/presupuestoanalisis.html', {"datos":datos, "id_proyecto":id_proyecto })

def presupuestoscapitulo(request, id_proyecto):

    proyecto = Proyectos.objects.get(id = id_proyecto)
    capitulo = Capitulos.objects.all()
    compo = CompoAnalisis.objects.all()
    computo = Computos.objects.all()
    modelo = Modelopresupuesto.objects.filter(proyecto = proyecto)

    if len(modelo) == 0:
        datos = 0

        datos = {"datos":datos, "proyecto":proyecto}
    
    else:

        crudo = []

        valor_proyecto = 0

        for c in capitulo:

            valor_capitulo = 0

            for d in modelo:

                if d.capitulo == c and d.proyecto == proyecto:

                    if d.cantidad == None:

                        if "SOLO MANO DE OBRA" in str(d.analisis):

                            valor_analisis = 0

                            for e in compo:

                                if e.analisis == d.analisis:

                                    valor_analisis = valor_analisis + e.articulo.valor*e.cantidad/1000000

                            cantidad = 0

                            for h in computo:
                                if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                                    cantidad = cantidad + h.valor_vacio   

                            valor_capitulo = valor_capitulo + valor_analisis*cantidad

                        else:
                            valor_analisis = 0

                            for e in compo:

                                if e.analisis == d.analisis:

                                    valor_analisis = valor_analisis + e.articulo.valor*e.cantidad/1000000

                            cantidad = 0

                            for h in computo:
                                if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                                    cantidad = cantidad + h.valor_lleno  

                            valor_capitulo = valor_capitulo + valor_analisis*cantidad
        
                    else:

                        valor_analisis = 0

                        for e in compo:

                            if e.analisis == d.analisis:

                                valor_analisis = valor_analisis + e.articulo.valor*e.cantidad/1000000

                        valor_capitulo = valor_capitulo + valor_analisis*float(d.cantidad)

            valor_proyecto = valor_proyecto + valor_capitulo
            
            crudo.append((c, valor_capitulo, 0.0))

        datos =[]

        for i in crudo:
            i = list(i)
            i[2] = i[1]/valor_proyecto*100
            i = tuple(i)
            datos.append(i)

        valor_proyecto_completo = valor_proyecto*1000000

        datos = {"datos":datos, "proyecto":proyecto, "valor_proyecto":valor_proyecto,"valor_proyecto_completo":valor_proyecto_completo}

    
    return render(request, 'presupuestos/presupuestocapitulo.html', {"datos":datos, "id_proyecto" : id_proyecto})

def ver_analisis(request, id_analisis):

    analisis = Analisis.objects.get(codigo = id_analisis)
    composi = CompoAnalisis.objects.filter(analisis = analisis)
    lista_compo = []

    for i in composi:
        if i.analisis == analisis:
            lista_compo.append(i)

    total = 0

    for c in lista_compo:
        total = total + c.articulo.valor*c.cantidad

    lista_final = []

    for d in lista_compo:
        total_renglon = d.cantidad*d.articulo.valor
        inc = (total_renglon/total)*100
        lista_final.append((d, total_renglon, inc))

    datos = {"analisis":analisis, "lista_final":lista_final, "total":total}

    
    return render(request, 'analisis/veranalisis.html', {"datos":datos})

def analisis_list(request):

    analisis = Analisis.objects.all()
    datos = []

    for i in analisis:

        valor = 0

        composicion = CompoAnalisis.objects.filter(analisis = i)

        for c in composicion:

            valor = valor +c.articulo.valor*c.cantidad

        datos.append((i, valor))


        #Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        datos_viejos = datos

        datos = []   

        for i in palabra_buscar:

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                lista_palabra = palabra.split()

                buscar = (str(i[0].nombre)+str(i[0].codigo)+str(i[1]))

                contador = 0

                for palabra in lista_palabra:

                    contador2 = 0

                    if palabra.lower() in buscar.lower():
  
                        contador += 1

                if contador == len(lista_palabra):

                    datos.append(i)


    #Aqui termina el filtro

    return render(request, 'analisis/listaanalisis.html', {"datos":datos})

def panelanalisis(request):

    analisis = Analisis.objects.all()
    
    datos = []

    for i in analisis:

        valor = 0

        composicion = CompoAnalisis.objects.filter(analisis = i)

        for c in composicion:

            if i == c.analisis:
                valor = valor +c.articulo.valor*c.cantidad

        datos.append((i, valor))

            #Aqui empieza el filtro

    if request.method == 'POST':

        palabra_buscar = request.POST.items()

        datos_viejos = datos

        datos = []   

        for i in palabra_buscar:

            if i[0] == "palabra":
        
                palabra_buscar = i[1]

        if str(palabra_buscar) == "":

            datos = datos_viejos

        else:
        
            for i in datos_viejos:

                palabra =(str(palabra_buscar))

                buscador = (str(i[0].nombre)+str(i[0].codigo))

                if palabra.lower() in buscador.lower():

                    datos.append(i)


    #Aqui termina el filtro

    return render(request, 'analisis/panelanalisis.html', {"datos":datos})
    
def crearanalisis(request):

    articulos = Articulos.objects.all()

    mensaje = ""

    datos = {'articulos':articulos, "mensaje":mensaje}

    if request.method == 'POST':

        datos_p = request.POST.items()

        datos_post = []

        for t in datos_p:

            datos_post.append(t)

        tupla = datos_post[1]

        codigo_tupla = tupla[1]

        if len(Analisis.objects.filter(codigo = codigo_tupla)) > 0:

            mensaje = "Este codigo ya se encuentra en la base"

            datos = {'articulos':articulos, "mensaje":mensaje}

        else:

            resto = []

            datos_p = request.POST.items()

            for i in datos_p:

                if i[0] == "codigo":
                    
                    codigo_analisis = i[1]

                elif i[0] == "nombre":
                    
                    nombre = i[1]

                elif i[0] == "unidad":
                    
                    unidad = i[1]
                
                else:

                    resto.append(i)        
                
            id_num = 1

            while Analisis.objects.filter(id = id_num):
                id_num = id_num + 1

           
            b = Analisis(

                id = id_num,
                codigo = codigo_analisis,
                nombre = nombre,
                unidad = unidad,
                )

            b.save()

            valor = 1

            for t in resto:

                if t[0] != "csrfmiddlewaretoken" and valor == 1:

                    valor = 2
    
                    nombre_articulo = t[1]

                elif t[0] != "csrfmiddlewaretoken" and valor == 2:

                    valor = 1

                    cantidad = t[1]

                    datos_compo = CompoAnalisis.objects.all()
                    
                    id_compo = []

                    for c in datos_compo:

                        id_compo.append(c.id)
                        
                    id_num_compo = 1

                    while id_num_compo in id_compo:
                        id_num_compo = id_num_compo + 1
            
                    b = CompoAnalisis(
                        id = id_num_compo,
                        articulo = Articulos.objects.get(nombre=nombre_articulo),
                        analisis = Analisis.objects.get(codigo=codigo_analisis),
                        cantidad = cantidad,
                    )

                    b.save()


            return redirect('Lista de analisis')

    return render(request, 'analisis/crearanalisis.html', {'datos':datos})

def modificaranalisis(request, id_analisis):

    analisis = Analisis.objects.get(codigo = id_analisis)
    articulos = Articulos.objects.all()
    compo = CompoAnalisis.objects.all()

    datos = []

    for i in compo:

        if i.analisis == analisis:

            datos.append((i.articulo, i.cantidad, i.id))
    

    datos = {"analisis":analisis,
    "articulos":articulos,
    "datos":datos}

    return render(request, 'analisis/modificaranalisis.html', {"datos":datos})

def parametros(request):

    proyectos = Proyectos.objects.all()

    datos = []

    for proyecto in proyectos:

        try:

            parametros = Prametros.objects.get(proyecto = proyecto)
            tasa_pl = parametros.tasa_des_p*100
            soft = parametros.soft*100
            imp = parametros.imprevitso*100
            comer = parametros.comer*100
            tem = parametros.tem_iibb*100
            ganan = parametros.ganancia*100
            porc_terreno = parametros.terreno/proyecto.m2*100
            porc_link = parametros.link/proyecto.m2*100
            tasa_des = parametros.tasa_des*100
            presupuesto = Presupuestos.objects.get(proyecto = proyecto)
            costo_m2 = (presupuesto.valor/(1+(tasa_pl/100)))/proyecto.m2
            costo_soft_m2 = costo_m2*(1+(soft/100))
            costo_imp = costo_soft_m2*(1+(imp/100))
            costo_terreno = (costo_imp*proyecto.m2)/(proyecto.m2-parametros.terreno)
            costo_hon = (costo_imp*proyecto.m2)/(proyecto.m2-parametros.terreno-parametros.link)
            costo_comer = costo_hon/(1 - (parametros.comer*(1+(porc_terreno + porc_link)/100)*(1+parametros.ganancia)))
            costo_tem = costo_hon/(1 - (parametros.comer*(1+(porc_terreno + porc_link)/100)*(1+parametros.ganancia)) - (parametros.tem_iibb*parametros.por_temiibb*(1+parametros.ganancia)))
            
            #Aqui se incorpora la tasa de descuento

            fecha_entrega =  datetime.datetime.strptime(str(proyecto.fecha_f), '%Y-%m-%d')
            ahora = datetime.datetime.utcnow()
            fecha_inicial = ahora + datetime.timedelta(days = (365*2))

            if fecha_entrega > fecha_inicial:
                y = fecha_entrega.year - fecha_inicial.year
                n = fecha_entrega.month - fecha_inicial.month
                meses = y*12 + n

                costo_desc = -np.pv(fv=costo_tem, rate=parametros.tasa_des, nper=meses, pmt=0)

            else:
                costo_desc = costo_tem

                meses = 0


            
            ganancia = costo_desc * (1+(ganan/100))

            datos.append((parametros, porc_terreno, porc_link, tasa_pl, soft, imp, comer, tem, ganan, costo_m2, costo_soft_m2, costo_imp, costo_terreno, costo_hon, costo_comer, costo_tem, ganancia, tasa_des, costo_desc, meses ))

        except: 
            basura = 1

    return render(request, 'desde/parametros.html', {'datos': datos})

def desde(request):

    if request.method=='POST':

        proyecto_elegido = Prametros.objects.get(id = request.POST['id'])
        proyecto_elegido.soft = request.POST['soft']
        proyecto_elegido.iva = request.POST['iva']
        proyecto_elegido.imprevitso = request.POST['imprevisto']
        proyecto_elegido.terreno = request.POST['terreno']
        proyecto_elegido.link = request.POST['link']
        proyecto_elegido.comer = request.POST['comer']
        proyecto_elegido.por_comer = request.POST['porc_comer']
        proyecto_elegido.tem_iibb = request.POST['temiibb']
        proyecto_elegido.por_temiibb = request.POST['porc_temiibb']
        proyecto_elegido.ganancia = request.POST['ganancia']
        proyecto_elegido.depto = request.POST['depto']
        proyecto_elegido.save()



    parametros_all = Prametros.objects.all()

    datos = []

    for parametros in parametros_all:

        try:

            presupuesto = Presupuestos.objects.get(proyecto = parametros.proyecto)
                       
            # Primero calculamos el costo por m2
            costo = (presupuesto.valor)
            costo_m2 = costo
            # Calculo del costo con imprevisto
            costo_imp = costo*(1 + parametros.imprevitso)
            # Calculo del costo IVA
            costo_iva = costo_imp*(1 + parametros.iva)
            # Calculo del costo SOFT
            costo_soft = costo_iva+(costo * parametros.soft)
            # Calculo del terreno
            costo_terreno = 0
            porc_terreno =  parametros.terreno/parametros.proyecto.m2
            # Calculo del honorario
            costo_honorario = 0
            porc_hon =  parametros.link/parametros.proyecto.m2
            # Calculo del TEM
            costo_tem = 0
            aumento_tem =  parametros.tem_iibb*parametros.por_temiibb
            # Calculo del COMER
            costo_comer = 0
            aumento_comer =  parametros.comer
            # Aumento por honorarios + TEM + COMER
            costo_completo = costo_soft/(1-((aumento_tem + aumento_comer)*(1 - porc_terreno - porc_hon))*(1 + parametros.ganancia)/(1 - porc_terreno - porc_hon))
            costo_completo = costo_completo/(1 - porc_terreno - porc_hon)
            # Valor con ganancia
            valor_ganancia = costo_completo*(1 + parametros.ganancia)
            # Recalculamos
            costo_terreno = costo_completo*porc_terreno + costo_soft
            costo_honorario = costo_completo*porc_hon + costo_terreno
            costo_comer = valor_ganancia * aumento_comer * (1 - porc_terreno - porc_hon) + costo_honorario
            costo_tem = costo_completo
            costo_depto = costo_completo*parametros.depto/parametros.proyecto.m2
            
            porc_terreno = porc_terreno*100
            porc_hon = porc_hon*100

            datos_costo_m2 = [0, costo_m2, costo_imp, costo_iva, costo_soft, costo_terreno, costo_honorario, costo_comer, costo_tem, valor_ganancia]
            datos_costo_m2 = np.array(datos_costo_m2)/parametros.proyecto.m2
            datos_porcentaje = [porc_terreno, porc_hon]
            datos_parametros = [parametros.imprevitso*100, parametros.iva*100, parametros.soft*100, aumento_comer*100, aumento_tem*100, parametros.ganancia*100, parametros.terreno, parametros.link]

            datos.append((datos_costo_m2, datos_porcentaje, datos_parametros, parametros.proyecto, costo_depto, parametros))
        
        except: 

            valor_proyecto = 0
            m2_proyecto = 0

            proyectos_all = Proyectos.objects.filter(nombre__icontains = parametros.proyecto_no_est)

            for proyecto in proyectos_all:
                if "INFRA" not in proyecto.nombre:
                    m2_proyecto += proyecto.m2
                if len(Presupuestos.objects.filter(proyecto = proyecto)) > 0:
                    valor_proyecto += Presupuestos.objects.filter(proyecto = proyecto)[0].valor
                proyecto.nombre = parametros.proyecto_no_est
                
            
            # Primero calculamos el costo por m2
            costo = valor_proyecto
            costo_m2 = costo
            # Calculo del costo con imprevisto
            costo_imp = costo*(1 + parametros.imprevitso)
            # Calculo del costo IVA
            costo_iva = costo_imp*(1 + parametros.iva)
            # Calculo del costo SOFT
            costo_soft = costo_iva+(costo * parametros.soft)
            # Calculo del terreno
            costo_terreno = 0
            porc_terreno =  parametros.terreno/m2_proyecto
            # Calculo del honorario
            costo_honorario = 0
            porc_hon =  parametros.link/m2_proyecto
            # Calculo del TEM
            costo_tem = 0
            aumento_tem =  parametros.tem_iibb*parametros.por_temiibb
            # Calculo del COMER
            costo_comer = 0
            aumento_comer =  parametros.comer
            # Aumento por honorarios + TEM + COMER
            costo_completo = costo_soft/(1-((aumento_tem + aumento_comer)*(1 - porc_terreno - porc_hon))*(1 + parametros.ganancia)/(1 - porc_terreno - porc_hon))
            costo_completo = costo_completo/(1 - porc_terreno - porc_hon)
            # Valor con ganancia
            valor_ganancia = costo_completo*(1 + parametros.ganancia)
            # Recalculamos
            costo_terreno = (valor_ganancia * porc_terreno - costo_completo*porc_terreno) + costo_soft
            costo_honorario = (valor_ganancia * porc_hon - costo_completo*porc_hon)  + costo_terreno
            costo_comer = costo_completo - valor_ganancia * aumento_comer * (1 - porc_terreno - porc_hon)
            costo_tem = costo_completo
            
            porc_terreno = porc_terreno*100
            porc_hon = porc_hon*100

            datos_costo_m2 = [0, costo_m2, costo_imp, costo_iva, costo_soft, costo_terreno, costo_honorario, costo_comer, costo_tem, valor_ganancia]
            datos_costo_m2 = np.array(datos_costo_m2)/m2_proyecto
            datos_porcentaje = [porc_terreno, porc_hon]
            datos_parametros = [parametros.imprevitso*100, parametros.iva*100, parametros.soft*100, aumento_comer*100, aumento_tem*100, parametros.ganancia*100, parametros.terreno, parametros.link]
            costo_depto = costo_completo*parametros.depto/m2_proyecto
 
            

            datos.append((datos_costo_m2, datos_porcentaje, datos_parametros, proyecto, costo_depto, parametros))

    return render(request, 'desde/desde.html', {'datos':datos})

def proyectos(request):

    datos = DatosProyectos.objects.all()

    return render(request, 'datos/projects.html', {'datos':datos})

def InformeArea(request):

    proyectos = Proyectos.objects.all()
    capitulos = Capitulos.objects.all()
    proy_presup = []
    contador = 0
    proyecto_300 = 0
    m2_300 = 0

    valor_proyecto_300 = 0
    vr_M2_300 = 0
    valor_proyecto_materiales_300 = 0
    valor_proyecto_mo_300 = 0
    total_creditos_300 = 0
    total_fdr_300 = 0
    total_ant_300 = 0
    imprevisto_300 = 0
    saldo_total_300 = 0

    for proyecto in proyectos:

        if "#300" in proyecto.nombre and "infra" not in proyecto.nombre:

            m2_300 = m2_300 + proyecto.m2

    for proyecto in proyectos:

        if "300" in proyecto.nombre:

            try:
                proyecto_300 = proyecto
                datos_presup = Presupuestos.objects.get(proyecto = proyecto)
                valor_proyecto_300 = valor_proyecto_300 + datos_presup.valor
                vr_M2_300 = vr_M2_300 + valor_proyecto_300/m2_300
                valor_proyecto_materiales_300 = valor_proyecto_materiales_300 + datos_presup.saldo_mat
                valor_proyecto_mo_300 = valor_proyecto_mo_300 + datos_presup.saldo_mo
                total_creditos_300  = total_creditos_300 + datos_presup.credito
                total_fdr_300 = total_fdr_300 + datos_presup.fdr
                total_ant_300  =  total_ant_300 + datos_presup.anticipos
                imprevisto_300 = imprevisto_300 + datos_presup.imprevisto
                saldo_total_300 = saldo_total_300 + valor_proyecto_materiales_300 + valor_proyecto_mo_300 + total_creditos_300 + total_fdr_300 + total_ant_300 + imprevisto_300

            except:
                 basura = 1

        else:

            try:

                datos_presup = Presupuestos.objects.get(proyecto = proyecto)
                valor_proyecto = datos_presup.valor
                vr_M2 = valor_proyecto/proyecto.m2
                valor_proyecto_materiales = datos_presup.saldo_mat
                valor_proyecto_mo = datos_presup.saldo_mo
                total_creditos = datos_presup.credito
                total_fdr = datos_presup.fdr
                total_ant = datos_presup.anticipos
                imprevisto = datos_presup.imprevisto

                saldo_total = valor_proyecto_materiales + valor_proyecto_mo + total_creditos + total_fdr + total_ant + imprevisto

                proy_presup.append((proyecto, valor_proyecto, vr_M2, valor_proyecto_materiales, valor_proyecto_mo, total_creditos, saldo_total, total_fdr, total_ant, imprevisto))
            except:

                basura = 1

            try:
                # Trato de establecer el precio de Link-P

                valor_linkp = Presupuestos.objects.get(proyecto = proyecto).valor
                parametros = Prametros.objects.get(proyecto = proyecto)
                valor_linkp = (valor_linkp/(1 + parametros.tasa_des_p))*(1 + parametros.soft)       
                valor_linkp = valor_linkp*(1 + parametros.imprevitso)
                porc_terreno = parametros.terreno/parametros.proyecto.m2*100
                porc_link = parametros.link/parametros.proyecto.m2*100
                aumento_tem = parametros.tem_iibb*parametros.por_temiibb*(1+parametros.ganancia)
                aumento_comer = parametros.comer*(1+(porc_terreno + porc_link)/100)*(1+parametros.ganancia)           
                valor_linkp = valor_linkp/(1-aumento_tem- aumento_comer)           
                m2 = (parametros.proyecto.m2 - parametros.terreno - parametros.link)
                valor_costo = valor_linkp/m2
                proyecto.precio_linkp = valor_costo
                proyecto.save()
            except:
                pass
    
    proy_presup.append((proyecto_300, valor_proyecto_300, vr_M2_300, valor_proyecto_materiales_300, valor_proyecto_mo_300, total_creditos_300, saldo_total_300, total_fdr_300, total_ant_300, imprevisto_300))

    cant_proy_act = len(proy_presup)

    datos = {"cantidad":cant_proy_act,   
    "datos":proy_presup}


    #Calculos para tablero de avance de presupuesto

    barras = []

    datos_barras = Presupuestos.objects.order_by("-saldo")

    for db in datos_barras:

        if db.valor != 0:

            avance = (100 - db.saldo/db.valor*100)

            barras.append((db, int(avance)))

    barras = sorted(barras,reverse=True, key=lambda tup: tup[1])


    #Aqui calculamos el radar

    proyectos_radar = PorcentajeCapitulo.objects.values_list('proyecto')

    proyecto_radar = list(set(proyectos_radar))

    datos_radar = []

    for proyect in proyecto_radar:

        datos_radar.append(PorcentajeCapitulo.objects.filter(proyecto = proyect).order_by("capitulo"))

    return render(request, 'presupuestos/informearea.html', {"datos":datos, "datos_barras":barras, 'capitulos':capitulos, 'datos_radar':datos_radar})

def AnticiposFinan(id_proyecto):
    proyecto = Proyectos.objects.get(id = id_proyecto)
    articulo = Articulos.objects.get(codigo = 9998005250)
    datos = Compras.objects.filter(proyecto = proyecto, articulo = articulo)

    datos_viejos = datos
    proveedores = []
    for dato in datos:
        proveedores.append(dato.proveedor)

    proveedores = list(set(proveedores))

    datos = []

    for proveedor in proveedores:
        monto_fdr = 0
        for dato in datos_viejos:
            if dato.proveedor == proveedor:
                monto_fdr = monto_fdr + articulo.valor*dato.cantidad
        datos.append((proveedor, monto_fdr))

    return datos

def Fondosdereparo(id_proyecto):
    proyecto = Proyectos.objects.get(id = id_proyecto)
    articulo = Articulos.objects.get(codigo = 9998005201)
    datos = Compras.objects.filter(proyecto = proyecto, articulo = articulo)

    datos_viejos = datos
    proveedores = []
    for dato in datos:
        proveedores.append(dato.proveedor)

    proveedores = list(set(proveedores))

    datos = []

    for proveedor in proveedores:
        monto_fdr = 0
        for dato in datos_viejos:
            if dato.proveedor == proveedor:
                monto_fdr = monto_fdr + articulo.valor*dato.cantidad
        datos.append((proveedor, monto_fdr))

    return datos

def PresupuestoPorCapitulo(id_proyecto):

    #Modelos que seran necesarios recorrer completos

    proyecto = Proyectos.objects.get(id = id_proyecto)
    capitulo = Capitulos.objects.all()    

    #La lista datos tiene que tener 37 Arrays por cada capitulo

    datos = []
    
    # Vamos a recorrer todos los capitulos y armar una array

    numero_capitulo = 1
    
    for cap in capitulo:

        capitulo = [] 

        modelo = Modelopresupuesto.objects.filter(proyecto = proyecto, capitulo = cap)

        for mod in modelo:

                if mod.cantidad == None:

                    if "SOLO MANO DE OBRA" in str(mod.analisis):

                        computo = Computos.objects.filter(proyecto = proyecto, tipologia = mod.vinculacion)

                        cantidad_computo = 0

                        for comp in computo:

                            cantidad_computo = cantidad_computo + comp.valor_vacio

                        articulos_analisis = CompoAnalisis.objects.filter(analisis = mod.analisis)

                        for compo in articulos_analisis:

                            articulo_cantidad = (compo.articulo, compo.cantidad*cantidad_computo)

                            capitulo.append(articulo_cantidad)


                    else:

                        computo = Computos.objects.filter(proyecto = proyecto, tipologia = mod.vinculacion)

                        cantidad_computo = 0

                        for comp in computo:

                            cantidad_computo = cantidad_computo + comp.valor_lleno

                        articulos_analisis = CompoAnalisis.objects.filter(analisis = mod.analisis)

                        for compo in articulos_analisis:

                            articulo_cantidad = (compo.articulo, compo.cantidad*cantidad_computo)

                            capitulo.append(articulo_cantidad)

    
                else:

                    articulos_analisis = CompoAnalisis.objects.filter(analisis = mod.analisis)

                    for compo in articulos_analisis:

                        articulo_cantidad = (compo.articulo, compo.cantidad*mod.cantidad )

                        capitulo.append(articulo_cantidad)

        datos.append((numero_capitulo, cap, capitulo))

        numero_capitulo += 1


    #Devuelve el numero del capitulo, el nombre y una lista de todos los insumos y la cantidad de cada uno             

    return datos

def Saldoporcapitulo(id_proyecto):

    #Traemos las compras y el presupuesto

    proyecto = Proyectos.objects.get(id = id_proyecto)
    compras = Compras.objects.filter(proyecto = proyecto)
    presupuesto_capitulo = PresupuestoPorCapitulo(id_proyecto)

    #Ordenamos cada capitulo con una lista donde no se repitan los articulos

    datos_viejos = presupuesto_capitulo
    presupuesto_capitulo = []

    contador = 0

    for i in range(37):

        dato = datos_viejos[contador]

        nuevo_art_cant = []

        lista_art_cap = []

        for art_cant in dato[2]:

            lista_art_cap.append(art_cant[0])

        lista_art_cap = list(set(lista_art_cap))
        
        for articulo in lista_art_cap:

            cantidad = 0

            for articulo2 in dato[2]:

                if articulo == articulo2[0]:

                    cantidad = cantidad + articulo2[1]

            nuevo_art_cant.append((articulo, cantidad))

        presupuesto_capitulo.append((dato[0], dato[1], nuevo_art_cant))    
        
        contador += 1


    #Ordenamos la compra para que sea una sola lista

    articulos_comprados = []

    for compra in compras:

        articulos_comprados.append(compra.articulo)

    articulos_comprados = list(set(articulos_comprados))

    #Armamos el stock con todas las compras realizadas de este proyecto

    stock_articulos = []

    for articulo in articulos_comprados:

        cantidad = sum(np.array(Compras.objects.values_list('cantidad').filter(proyecto = proyecto, articulo = articulo)))

        stock_articulos.append((articulo, cantidad))

    #Armamos el saldo --> Hay un error ya que al descartar menores a 0, olvidamos que restan consumo

    saldo_capitulo = []

    for capitulo_presupuesto in presupuesto_capitulo:

        articulos_saldo = []

        for articulos_presupuesto in capitulo_presupuesto[2]:

            if articulos_presupuesto[0] in articulos_comprados and articulos_presupuesto[1]>=0:

                contador = 0

                for articulos_stock in stock_articulos:

                    #Si encontramos el articulo del capitulo en el stock, activamos una de las 3 posibilidades

                    if articulos_stock[0] == articulos_presupuesto[0]:

                        articulos_stock = list(articulos_stock)

                        if articulos_stock[1] > articulos_presupuesto[1]:

                            articulos_stock[1] = float(articulos_stock[1]) - float(articulos_presupuesto[1])                           

                            stock_articulos[contador] = list(stock_articulos[contador])
                            stock_articulos[contador][1] = articulos_stock[1]

                            articulos_stock = tuple(articulos_stock)
                            stock_articulos[contador] = tuple(stock_articulos[contador])

                        elif articulos_stock[1] == articulos_presupuesto[1]:

                            articulos_stock[1] = 0
                            stock_articulos[contador] = list(stock_articulos[contador])
                            stock_articulos[contador][1] = articulos_stock[1]

                            articulos_stock = tuple(articulos_stock)
                            stock_articulos[contador] = tuple(stock_articulos[contador])

                        elif articulos_stock[1] < articulos_presupuesto[1]:

                            cantidad_saldo = float(articulos_presupuesto[1]) - float(articulos_stock[1])

                            articulos_stock[1] = 0

                            stock_articulos[contador] = list(stock_articulos[contador])
                            stock_articulos[contador][1] = articulos_stock[1]

                            articulos_saldo.append((articulos_presupuesto[0], cantidad_saldo))

                            articulos_stock = tuple(articulos_stock)
                            stock_articulos[contador] = tuple(stock_articulos[contador])
                    contador += 1
            else:
                articulos_saldo.append(articulos_presupuesto)

        #Modificado con el saldo
                
        saldo_capitulo.append((capitulo_presupuesto[0], capitulo_presupuesto[1], articulos_saldo))


    return saldo_capitulo

def Creditocapitulo(id_proyecto):

    proyecto = Proyectos.objects.get(id = id_proyecto)
    modelo = Modelopresupuesto.objects.filter(proyecto = proyecto)

    #Con el siguiente conjunto de formulas creamos la explosión de insumos
    
    crudo_analisis = []

    for i in modelo:

        if i.cantidad != None:

            crudo_analisis.append((i.analisis, i.cantidad))

        else:

            if "SOLO MANO DE OBRA" in str(i.analisis.nombre):

                cantidad = sum(np.array(Computos.objects.filter(tipologia = i.vinculacion, proyecto = proyecto).values_list("valor_vacio", flat = True)))

                crudo_analisis.append((i.analisis, cantidad))

            else:
                cantidad = sum(np.array(Computos.objects.filter(tipologia = i.vinculacion, proyecto = proyecto).values_list("valor_lleno", flat = True)))
                
                crudo_analisis.append((i.analisis, cantidad))

    crudo_articulos = []

    for c in crudo_analisis:

        analisis = CompoAnalisis.objects.filter(analisis = c[0])

        for d in analisis:

            cantidad = d.cantidad*c[1]

            crudo_articulos.append((d.articulo, cantidad))

    datos = []

    for t in crudo_articulos:
        datos.append(t[0])

    datos = list(set(datos))

    datos_viejos = datos
    datos = []

    for i in datos_viejos:
        cantidad = 0
        for c in crudo_articulos:
            if i == c[0]:
                cantidad = cantidad + c[1]
        datos.append((i, cantidad))

    compras = Compras.objects.filter(proyecto = proyecto)

    # Este auxiliar arma una cadena de texto de todos los articulos necesarios

    comprado_aux = ""

    for dato in datos:
        comprado_aux = comprado_aux + str(dato[0])

    datos_viejos = datos
    
    datos = []

    for i in datos_viejos:
        comprado = 0
        for c in compras:
            if c.proyecto == proyecto and c.articulo == i[0]:
                comprado = comprado + c.cantidad
        
        cantidad_saldo = i[1] - comprado

        saldo = cantidad_saldo * i[0].valor

        if saldo < 0:
        
            datos.append((i[0], i[1], comprado, cantidad_saldo, saldo ))

    # Esta parte arma los articulos que no estan en el presupuesto, compara el nombre si esta adentro de la cadena auxiliar 

    for compra in compras:
        if str(compra.articulo.nombre) not in comprado_aux and compra.proyecto == proyecto and str(compra.articulo.nombre)!="FONDO DE REPARO ACT. UOCRA" and str(compra.articulo.nombre)!="ANTICIPO FINANCIERO ACT. UOCRA" :
            saldo = compra.articulo.valor*compra.cantidad
            datos.append((compra.articulo, 0, compra.cantidad, -compra.cantidad, -saldo))

    return datos

def presupuesto_auditor(request):

    proyecto = 0
    fecha_desde = 0
    fecha_hasta = 0
    proyectos=PresupuestosAlmacenados.objects.values('proyecto__nombre','proyecto__id').distinct()
    data_resultante=0
    data_resultante_p=0
    mensaje='Aun no se han filtrado datos'

    if request.method=='POST':

        
        response=request.POST
        datos_filtro={}
        
        for dato in response:
            datos_filtro[dato]=response[dato]

                
        proyecto = Proyectos.objects.get(id =datos_filtro['proyecto'])
        fecha_desde=datos_filtro['fecha_desde']
        fecha_hasta=datos_filtro['fecha_hasta']

        try:
            response_servidor = {"messages": "Perri"}
            bot_wp = WABot(response_servidor)
            presupuestador = Presupuestos.objects.get(proyecto = proyecto).presupuestador
            presupuestador = datosusuario.objects.get(identificacion = presupuestador)
            send = "{}: El usuario {} esta utilizando Auditor para analizar {}, desde {} hasta {}.".format(presupuestador.nombre, request.user.first_name, proyecto, fecha_desde, fecha_hasta)
            bot_wp.send_message_user(presupuestador.Telefono, send)
        except:
            pass

        listado_dias = PresupuestosAlmacenados.objects.filter(proyecto = proyecto).exclude(nombre = "vigente").values_list("nombre", flat = True).distinct()
    
        data_resultante,mensaje= auditor_presupuesto(proyecto,fecha_desde, fecha_hasta)
      
        data_resultante_p = auditor_presupuesto_p(proyecto,fecha_desde, fecha_hasta)
   
            
    return render(request, "presupuestos/presupuesto_auditor.html",{'fecha_hasta':fecha_hasta, 'fecha_desde':fecha_desde, 'proyecto':proyecto, 'data_resultante_p':data_resultante_p, 'data_resultante':data_resultante,'proyectos':proyectos,'mensaje':mensaje})

class ReporteExplosion(TemplateView):

    def get(self, request, id_proyecto, *args, **kwargs):
        wb = Workbook()

        proyecto = Proyectos.objects.get(id = id_proyecto)
        modelo = Modelopresupuesto.objects.filter(proyecto = proyecto)

        #Con el siguiente conjunto de formulas creamos la explosión de insumos
        
        crudo_analisis = []

        for i in modelo:

            if i.cantidad != None:

                crudo_analisis.append((i.analisis, i.cantidad))

            else:

                if "SOLO MANO DE OBRA" in str(i.analisis.nombre):

                    computos = Computos.objects.filter(tipologia = i.vinculacion, proyecto = proyecto)

                    cantidad = 0

                    for r in computos:
                        cantidad = cantidad + r.valor_vacio

                    crudo_analisis.append((i.analisis, cantidad))

                else:

                    computos = Computos.objects.filter(tipologia = i.vinculacion, proyecto = proyecto)

                    cantidad = 0

                    for r in computos:
                        cantidad = cantidad + r.valor_lleno

                    crudo_analisis.append((i.analisis, cantidad))

        crudo_articulos = []


        for c in crudo_analisis:

            analisis = CompoAnalisis.objects.filter(analisis = c[0])

            for d in analisis:

                cantidad = d.cantidad*c[1]

                crudo_articulos.append((d.articulo, cantidad))

        datos = []

        for t in crudo_articulos:
            datos.append(t[0])

        datos = list(set(datos))

        datos_viejos = datos
        datos = []

        for i in datos_viejos:
            cantidad = 0
            for c in crudo_articulos:
                if i == c[0]:
                    cantidad = cantidad + c[1]
            datos.append((i, cantidad))


        compras = Compras.objects.filter(proyecto = proyecto)

        comprado_aux = ""

        for dato in datos:
            comprado_aux = comprado_aux + str(dato[0])

        datos_viejos = datos
        
        datos = []

        for i in datos_viejos:

            comprado = 0
            for c in compras:

                if c.proyecto == proyecto and c.articulo == i[0]:

                    comprado = comprado + c.cantidad
            
            cantidad_saldo = i[1] - comprado

            saldo = cantidad_saldo * i[0].valor
            
            datos.append((i[0], i[1], comprado, cantidad_saldo, saldo ))

        #Esta parte arma los articulos que no estan en el presupuesto

        mat_no_presup = []

        for compra in compras:
            if str(compra.articulo.nombre) not in comprado_aux and compra.proyecto == proyecto:
                mat_no_presup.append((compra.articulo.nombre, compra.articulo.valor, compra.cantidad))

        cont = 1
        for d in datos:

            if cont == 1:
                ws = wb.active
                ws.title = "Explosion"
                ws["A"+str(cont)] = "CODIGO"
                ws["B"+str(cont)] = "ARTICULO"
                ws["C"+str(cont)] = "UNIDAD"
                ws["D"+str(cont)] = "VALOR"
                ws["E"+str(cont)] = "CANT. PRESPUESTO"
                ws["F"+str(cont)] = "PRESUPUESTO"
                ws["G"+str(cont)] = "COMPRADO"
                ws["H"+str(cont)] = "PENDIENTE"
                ws["I"+str(cont)] = "SALDO PENDIENTE"

                ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["F"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["G"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["H"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["I"+str(cont)].alignment = Alignment(horizontal = "center")

                ws["A"+str(cont)].font = Font(bold = True)
                ws["B"+str(cont)].font = Font(bold = True)
                ws["C"+str(cont)].font = Font(bold = True)
                ws["D"+str(cont)].font = Font(bold = True)
                ws["E"+str(cont)].font = Font(bold = True)
                ws["F"+str(cont)].font = Font(bold = True)
                ws["G"+str(cont)].font = Font(bold = True)
                ws["H"+str(cont)].font = Font(bold = True)
                ws["I"+str(cont)].font = Font(bold = True)

                ws.column_dimensions['A'].width = 11.29
                ws.column_dimensions['B'].width = 58.57
                ws.column_dimensions['C'].width = 8.57
                ws.column_dimensions['D'].width = 12.14
                ws.column_dimensions['E'].width = 18.57
                ws.column_dimensions['F'].width = 17.57
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 11.86
                ws.column_dimensions['I'].width = 17.57

                ws["A"+str(cont+1)] = d[0].codigo
                ws["B"+str(cont+1)] = d[0].nombre
                ws["C"+str(cont+1)] = d[0].unidad
                ws["D"+str(cont+1)] = d[0].valor
                ws["E"+str(cont+1)] = d[1]
                ws["F"+str(cont+1)] = "=D"+str(cont)+"*E"+str(cont)
                ws["G"+str(cont+1)] = d[2]
                ws["H"+str(cont+1)] = d[3]
                ws["I"+str(cont+1)] = d[4]

                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].number_format = '#,##0.00_-'
                ws["F"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["G"+str(cont+1)].number_format = '#,##0.00_-'
                ws["H"+str(cont+1)].number_format = '#,##0.00_-'
                ws["I"+str(cont+1)].font = Font(bold = True)
                ws["I"+str(cont+1)].number_format = '"$"#,##0.00_-'

                cont += 1

            else: 
                ws = wb.active
                ws["A"+str(cont+1)] = d[0].codigo
                ws["B"+str(cont+1)] = d[0].nombre
                ws["C"+str(cont+1)] = d[0].unidad
                ws["D"+str(cont+1)] = d[0].valor
                ws["E"+str(cont+1)] = d[1]
                ws["F"+str(cont+1)] = "=D"+str(cont+1)+"*E"+str(cont+1)
                ws["G"+str(cont+1)] = d[2]
                ws["H"+str(cont+1)] = d[3]
                ws["I"+str(cont+1)] = d[4]

                ws["A"+str(cont+1)].font = Font(bold = True)
                ws["A"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["D"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["C"+str(cont+1)].alignment = Alignment(horizontal = "center")
                ws["E"+str(cont+1)].number_format = '#,##0.00_-'
                ws["F"+str(cont+1)].number_format = '"$"#,##0.00_-'
                ws["G"+str(cont+1)].number_format = '#,##0.00_-'
                ws["H"+str(cont+1)].number_format = '#,##0.00_-'
                ws["I"+str(cont+1)].font = Font(bold = True)
                ws["I"+str(cont+1)].number_format = '"$"#,##0.00_-'

                cont += 1
        cont = 1
        for m in mat_no_presup:

            if cont == 1:
                ws = wb.create_sheet('Art-no-pre')
                ws["A"+str(cont)] = "ARTICULO"
                ws["B"+str(cont)] = "VALOR"
                ws["C"+str(cont)] = "CANTIDAD"

                ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                ws["C"+str(cont)].alignment = Alignment(horizontal = "center")

                ws["A"+str(cont)].font = Font(bold = True)
                ws["B"+str(cont)].font = Font(bold = True)
                ws["C"+str(cont)].font = Font(bold = True)

                ws.column_dimensions['A'].width = 58.57
                ws.column_dimensions['B'].width = 13
                ws.column_dimensions['C'].width = 13

                cont += 1

            else:
                
                ws = wb['Art-no-pre']
                ws["A"+str(cont)] = m[0]
                ws["B"+str(cont)] = m[1]
                ws["C"+str(cont)] = m[2]

                cont += 1

        #Establecer el nombre del archivo
        nombre_archivo = "Explosion-{0}.xls".format(str(proyecto.nombre))
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteExplosionCap(TemplateView):

    def get(self, request, id_proyecto, *args, **kwargs):
        wb = Workbook()

        #Aqui coloco la formula para calcular

        contador_cap = 0

        for i in range(37):

            saldo = Saldoporcapitulo(id_proyecto)

            datos_viejos = saldo

            datos_saldo = []
            capitulo = []

            contador_cap += 1

            for componentes in datos_viejos:

                if int(componentes[1].id) == int(contador_cap):
                    
                    datos_saldo.append(componentes[2])
                    capitulo.append(componentes[1])

            articulos = []

            for articulo in datos_saldo[0]:
                articulos.append(articulo[0])

            articulos = list(set(articulos))

            articulos_cant = []

            for articulo in articulos:

                cantidad = 0

                for art_can in datos_saldo[0]:

                    if articulo == art_can[0] and art_can[1]>0:
                        cantidad = cantidad + art_can[1]
                articulos_cant.append((articulo, cantidad))
            
            saldo_cap = 0

            datos_viejos = articulos_cant
            datos_saldo = []

            for dato in datos_viejos:
                saldo_cap = saldo_cap + dato[0].valor*dato[1]
                datos_saldo.append((dato[0], dato[1], float(dato[0].valor*dato[1])))

            datos_viejos = datos_saldo
            datos_saldo = []

            for dato in datos_viejos:
                if saldo_cap != 0:
                    inc = float(dato[2])/float(saldo_cap)*100
                else:
                    inc = 0
                datos_saldo.append((dato[0], dato[1], dato[2], inc))


            if len(datos_saldo) == 0:
                datos_saldo = 0

            else:

                datos_saldo.sort(key=lambda tup: tup[3], reverse=True)

                proyecto = Proyectos.objects.get(id = id_proyecto)

                ws = wb.active
                ws.title = "ADVERTENCIA"

                ws.merge_cells("B2:K2")
                ws["B2"] = "LEER ATENTAMENTE ANTES DE USAR ESTE DOCUMENTO"

                ws["B2"].alignment = Alignment(horizontal = "center")
                ws["B2"].font = Font(bold = True, color= "CF433F", size = 20)

                ws.merge_cells("B5:K25")
                ws["B5"] = "Este documento contiene informción --> PRIVADA <-- del área de presupuestos, \n la misma es solo para uso interno de LINK INVERSIONES y no debe ser compartida sin previa autorización. Compartir este archivo puede ser considerado como divulgar información confidencial. Si usted esta utilizando este archivo en una computadora que no pertenezca a la empresa, al finalizar --> ELIMINE <-- el archivo. Gracias --AR"
                ws["B5"].alignment = Alignment(horizontal = "center", vertical = "center", wrap_text=True)
                ws["B5"].font = Font(bold = True)
                cont = 1
                for d in datos_saldo:

                    if cont == 1:
                        ws = wb.create_sheet("My sheet")
                        ws.title = "CAP{0}".format(str(contador_cap))
                        ws["A"+str(cont)] = "CODIGO"
                        ws["B"+str(cont)] = "ARTICULO"
                        ws["C"+str(cont)] = "UNIDAD"
                        ws["D"+str(cont)] = "VALOR"
                        ws["E"+str(cont)] = "PENDIENTE"
                        ws["F"+str(cont)] = "SALDO PENDIENTE"
                        ws["G"+str(cont)] = "INC"


                        ws["A"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["B"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["C"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["D"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["E"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["F"+str(cont)].alignment = Alignment(horizontal = "center")
                        ws["G"+str(cont)].alignment = Alignment(horizontal = "center")


                        ws["A"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["A"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["B"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["B"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["C"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["C"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["D"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["D"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["E"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["E"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["F"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["F"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")
                        ws["G"+str(cont)].font = Font(bold = True, color= "FDFFFF")
                        ws["G"+str(cont)].fill =  PatternFill("solid", fgColor= "159ABB")


                        ws.column_dimensions['A'].width = 11.29
                        ws.column_dimensions['B'].width = 58.57
                        ws.column_dimensions['C'].width = 8.57
                        ws.column_dimensions['D'].width = 12.14
                        ws.column_dimensions['E'].width = 18.57
                        ws.column_dimensions['F'].width = 17.57
                        ws.column_dimensions['G'].width = 12

                        ws["A"+str(cont+1)] = d[0].codigo
                        ws["B"+str(cont+1)] = d[0].nombre
                        ws["C"+str(cont+1)] = d[0].unidad
                        ws["D"+str(cont+1)] = d[0].valor
                        ws["E"+str(cont+1)] = d[1]
                        ws["F"+str(cont+1)] = d[2]
                        ws["G"+str(cont+1)] = d[3]


                        ws["A"+str(cont+1)].font = Font(bold = True)
                        ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                        ws["C"+str(cont+1)].number_format = '"$"#,##0.00_-'
                        ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                        ws["E"+str(cont+1)].number_format = '#,##0.00_-'
                        ws["F"+str(cont+1)].number_format = '"$"#,##0.00_-'
                        ws["G"+str(cont+1)].number_format = '#,##0.00_-"%"'

                        cont += 1

                    else:
                        ws = wb["CAP{0}".format(str(contador_cap))]
                        ws["A"+str(cont+1)] = d[0].codigo
                        ws["B"+str(cont+1)] = d[0].nombre
                        ws["C"+str(cont+1)] = d[0].unidad
                        ws["D"+str(cont+1)] = d[0].valor
                        ws["E"+str(cont+1)] = d[1]
                        ws["F"+str(cont+1)] = d[2]
                        ws["G"+str(cont+1)] = d[3]


                        ws["A"+str(cont+1)].font = Font(bold = True)
                        ws["B"+str(cont+1)].alignment = Alignment(horizontal = "center")
                        ws["C"+str(cont+1)].number_format = '"$"#,##0.00_-'
                        ws["D"+str(cont+1)].alignment = Alignment(horizontal = "center")
                        ws["E"+str(cont+1)].number_format = '#,##0.00_-'
                        ws["F"+str(cont+1)].number_format = '"$"#,##0.00_-'
                        ws["G"+str(cont+1)].number_format = '#,##0.00_-"%"'
        

                        cont += 1

        #Establecer el nombre del archivo
        nombre_archivo = "ExplosionCap-{0}.xls".format(str(proyecto.nombre))
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

def debugsaldo(id_proyecto):

    #Traemos las compras y el presupuesto
    proyecto = Proyectos.objects.get(id = id_proyecto)
    compras = Compras.objects.filter(proyecto = proyecto)
    presupuesto_capitulo = PresupuestoPorCapitulo(id_proyecto)

    #Ordenamos cada capitulo con una lista donde no se repitan los articulos

    datos_viejos = presupuesto_capitulo
    presupuesto_capitulo = []

    contador = 0

    mensaje = []

    for i in range(37):

        dato = datos_viejos[contador]

        lista_art_cap = []

        for art_cant in dato[2]:

            lista_art_cap.append(art_cant[0])

        lista_art_cap = list(set(lista_art_cap))
        
        for articulo in lista_art_cap:

            cantidad = 0

            for articulo2 in dato[2]:

                if articulo == articulo2[0]:
                    cantidad = cantidad + articulo2[1]

            if cantidad<0:
            

                mensaje.append((articulo.nombre, dato[1]))  
        

        contador += 1

        return mensaje
    
# ---> Aqui empiezan los servicios

class ArticulosListApiView(ListAPIView):
    serializer_class = ArtSerializer

    def get_queryset(self):
        kword = self.request.query_params.get('kword')

        if kword == None:

            return Articulos.objects.all()[0:40]

        else:

            return Articulos.objects.filter(nombre__icontains = kword)[0:40]|Articulos.objects.filter(codigo__icontains = kword)[0:40]


def registro_contable(request):

    return render(request, 'registro_contable.html', f )







