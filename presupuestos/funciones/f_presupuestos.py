from openpyxl import Workbook
from presupuestos.models import Capitulos,CompoAnalisis,Modelopresupuesto, Presupuestos,PresupuestosAlmacenados
from django.conf import settings
from presupuestos.wabot import WABot
import datetime as dt
import pandas as pd
import numpy as np
import statistics
import json
from presupuestos.models import Capitulos, PresupuestosAlmacenados, Analisis, Articulos
from compras.models import Compras
from computos.models import Computos
from proyectos.models import Proyectos

def presupuestos_datos_bot():

    diccionario_datos = {}
    diccionario_datos['Telegram_grupo_presupuesto_id'] = '-455382561'
    diccionario_datos['Telegram_grupo_presupuesto_token']= '1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA'

    return diccionario_datos

def presupuesto_recalcular_presupuesto(proyecto):

    archivo_vigente = PresupuestosAlmacenados.objects.filter(proyecto = proyecto, nombre = "vigente")[0].archivo
    df = pd.read_excel(archivo_vigente)
    datos_saldo = presupuesto_calculo_saldo(proyecto)
    presupuesto = Presupuestos.objects.get(proyecto = proyecto)
    valor_actual = presupuesto.valor
    presupuesto.valor = sum(np.array(df['Monto'].values))
    presupuesto.saldo = datos_saldo['saldo_total']
    presupuesto.saldo_mat = datos_saldo['saldo_mat']
    presupuesto.saldo_mo = datos_saldo['saldo_mo']
    presupuesto.save()

    if presupuesto.proyecto.presupuesto == "BASE" and valor_actual != 0:
        presupuestos_actualizar = Presupuestos.objects.filter(proyecto_base = proyecto)

        for presup_actualizar in presupuestos_actualizar:
            presup_actualizar.valor *= (presupuesto.valor/valor_actual)
            presup_actualizar.saldo *= (presupuesto.valor/valor_actual)
            presup_actualizar.saldo_mat *= (presupuesto.valor/valor_actual)
            presup_actualizar.saldo_mo *= (presupuesto.valor/valor_actual)
            presup_actualizar.save()

    return "Calculate"

def presupuestos_revision_registros(proyecto):

    presupuestos_alm = PresupuestosAlmacenados.objects.filter(proyecto = proyecto, nombre = "vigente")[1:]

    for presupuesto in presupuestos_alm:

        presupuesto.nombre = str(dt.date.today())
        presupuesto.save()

def presupuesto_datos_proyecto(proyecto):
    presupuesto = Presupuestos.objects.get(proyecto = proyecto)

    datos = {
        'valor_reposicion': presupuesto.valor,
        'saldo_total': presupuesto.saldo,
        'saldo_material': presupuesto.saldo_mat,
        'saldo_mo': presupuesto.saldo_mo,
        'imprevisto': presupuesto.imprevisto,
    }

    return datos

def presupuesto_calculo_saldo(proyecto):

    presupuesto = Presupuestos.objects.get(proyecto = proyecto)
    presupuesto_detalle = json.loads(presupuesto.balance_details)
    valor_saldo_proyecto_mo = 0
    valor_saldo_proyecto_materiales = 0

    for detalle in presupuesto_detalle:
        key = list(detalle.keys())[0]
        valor_saldo_proyecto_mo += detalle[key]['saldo_mo']
        valor_saldo_proyecto_materiales += detalle[key]['saldo_mat']
    
    datos = {
        "saldo_mo": valor_saldo_proyecto_mo,
        "saldo_mat": valor_saldo_proyecto_materiales,
        "saldo_total": valor_saldo_proyecto_mo + valor_saldo_proyecto_materiales,
    }

    return datos

def auditor_presupuesto(proyecto,fecha_desde, fecha_hasta):
    #esta linea trae 29 objetos
    mensaje = ""
    data_almacenada_desde_objects = PresupuestosAlmacenados.objects.filter(nombre = fecha_desde,proyecto=proyecto)
    
    #esta linea da error porque el queryset se encuentra vacio
    data_almacenada_hasta_objects = PresupuestosAlmacenados.objects.filter(nombre = fecha_hasta,proyecto=proyecto)
    
    if data_almacenada_desde_objects.exists() and data_almacenada_hasta_objects.exists():

        try:
        
            data_almacenada_desde=data_almacenada_desde_objects[0]
            data_almacenada_hasta=data_almacenada_hasta_objects.latest("id")

            df_desde = pd.read_excel(data_almacenada_desde.archivo)
            df_hasta = pd.read_excel(data_almacenada_hasta.archivo)

            
            # Estudio de diferencia de cantidades

            valor_proyecto_desde=sum(list(df_desde['Monto'].values))
            valor_proyecto_hasta=sum(list(df_hasta['Monto'].values))

            

            df_desde_Q = df_desde.drop(['Precio', 'Monto'], axis = 1) #se eliminan
            df_hasta_Q = df_hasta.drop(['Precio', 'Monto'], axis = 1) #se eliminan

            
            df_dif_Q = df_hasta_Q.merge(df_desde_Q, how='outer', indicator='union')
            df_dif_Q.drop(df_dif_Q[df_dif_Q['union'] == "both"].index, inplace=True) #se unen
            
            df_dif_Q_desde = df_dif_Q[df_dif_Q['union'] == "left_only"]
            df_dif_Q_hasta = df_dif_Q[df_dif_Q['union'] == "right_only"]

            
            lista_capitulos_afectado = list(set(df_dif_Q["Capitulo"].values)) #set elimina los nombres de ls capitulos repetidos

            data_procesada = []

            cuantif_q = 0
            errores = 0
            
            data_capitulo = []

            # Definimos un diccionario con los precios en M2

            articulos_M2={}

            for art in df_hasta.index:
                
                articulos_M2[df_hasta['Articulo'][art]]=df_hasta['Precio'][art]

            for art in df_desde.index:
                if df_desde['Articulo'][art] not in articulos_M2:
                    articulos_M2[df_desde['Articulo'][art]]=df_desde['Precio'][art]


            cont_cap = 0
            for capitulo in lista_capitulos_afectado:
                # Primeros los articulos eliminados
                
                cuantif_q_cap = 0
                errores_cap = 0

                df_dif_Q_desde_aux = df_dif_Q_desde[df_dif_Q_desde['Capitulo'] == capitulo]

                articulos_eliminados = []

                for i in df_dif_Q_desde_aux.index:
                    try:
                        
                        analisis = Analisis.objects.get(codigo = df_dif_Q_desde_aux['Analisis'][i])
                    except:
                        analisis = "¿¿"+str(df_dif_Q_desde_aux['Analisis'][i])+"??"
                        errores += 1
                        errores_cap += 1
                    try:
                        articulo = Articulos.objects.get(codigo = df_dif_Q_desde_aux['Articulo'][i])
                    except:
                        articulo = "¿¿"+str(df_dif_Q_desde_aux['Articulo'][i])+"??"

                    cantidad = df_dif_Q_desde_aux['Cantidad Art Totales'][i]
                    try:
                        modif = cantidad*articulos_M2[articulo.codigo]
                        cuantif_q += modif
                        cuantif_q_cap += modif
                    except:
                        modif = "??"
                        errores += 1
                        errores_cap += 1

                    articulos_eliminados.append((analisis, articulo, cantidad, -modif))

                df_dif_Q_hasta_aux = df_dif_Q_hasta[df_dif_Q_hasta['Capitulo'] == capitulo]

                articulos_agregados = []

                for i in df_dif_Q_hasta_aux.index:
                    try:
                        analisis = Analisis.objects.get(codigo = df_dif_Q_hasta_aux['Analisis'][i])
                    except:
                        analisis = "¿¿"+str(df_dif_Q_hasta_aux['Analisis'][i])+"??"
                        errores += 1
                        errores_cap += 1
                    try:
                        articulo = Articulos.objects.get(codigo = df_dif_Q_hasta_aux['Articulo'][i])
                    except:
                        articulo = "¿¿"+str(df_dif_Q_hasta_aux['Articulo'][i])+"??"
                        

                    cantidad = df_dif_Q_hasta_aux['Cantidad Art Totales'][i]
                    try:
                        modif = cantidad*articulos_M2[articulo.codigo]
                        cuantif_q -= modif
                        cuantif_q_cap -= modif
                    except:
                        modif = "??"
                        errores += 1
                        errores_cap += 1

                    articulos_agregados.append((analisis, articulo, cantidad, modif))
                
                cont_cap += 1

                data_capitulo.append((capitulo, articulos_agregados, articulos_eliminados, cuantif_q_cap, errores_cap, cont_cap))
            diferencia_precio = valor_proyecto_hasta - valor_proyecto_desde
            data_procesada.append((data_capitulo, cuantif_q, errores, valor_proyecto_desde , valor_proyecto_hasta, diferencia_precio ))
        except:
            data_procesada=None
            mensaje='Hubo un error en el proceso'
    else:
        data_procesada=None
        mensaje='No se encontraron registros en esa fecha para ese proyecto, prueba con: {}'.format(list(PresupuestosAlmacenados.objects.filter(proyecto = proyecto).exclude(nombre="vigente").values_list("nombre", flat=True).distinct()))      

    return data_procesada,mensaje

def auditor_presupuesto_p(proyecto, fecha_desde, fecha_hasta):

    cont=0
    data_almacenada_desde_objects = PresupuestosAlmacenados.objects.filter(nombre = fecha_desde,proyecto=proyecto)
    
    #esta linea da error porque el queryset se encuentra vacio
    data_almacenada_hasta_objects = PresupuestosAlmacenados.objects.filter(nombre = fecha_hasta,proyecto=proyecto)
    
    if data_almacenada_desde_objects.exists() and data_almacenada_hasta_objects.exists():
        
        
        data_almacenada_desde=data_almacenada_desde_objects[0]
        data_almacenada_hasta=data_almacenada_hasta_objects.latest("id")

        df_desde = pd.read_excel(data_almacenada_desde.archivo)

        valor_desde=sum(list(df_desde['Monto'].values))
        df_hasta = pd.read_excel(data_almacenada_hasta.archivo)

        articulos_M2={}
        dif_articulos={}

        for art in df_hasta.index:
            
            #de la columna 'articulos' toma los registros con el codigo de cada fila
            articulos_M2[df_hasta['Articulo'][art]]=df_hasta['Precio'][art]

        df_desde_m2=df_desde
        
        cont_dif_art = 0
        list_resultante_p = []
        for i in df_desde_m2.index:
            try:

                # Diferencia absoluta de la variación de precios

                dif_articulos[df_desde_m2['Articulo'][i]]= [abs((articulos_M2[df_desde_m2['Articulo'][i]]/df_desde_m2['Precio'][i]-1)*100), cont_dif_art]


                # Ponderacion de diferencia de precios por articulo
                articulo = df_desde_m2['Articulo'][i]
                var = round(abs((articulos_M2[df_desde_m2['Articulo'][i]]/df_desde_m2['Precio'][i]-1)*100), 2)
                dif_precio_articulos = round(articulos_M2[df_desde_m2['Articulo'][i]] - df_desde_m2['Precio'][i], 2)
                variacion_por_articulo = round(dif_precio_articulos*float(df_desde_m2['Cantidad Ar'][i]), 2)
                capitulo = df_desde_m2['Capitulo'][i]
                
                if variacion_por_articulo != 0:

                    list_resultante_p.append((articulo, var, dif_precio_articulos, variacion_por_articulo, capitulo))

                df_desde_m2['Precio'][i]=articulos_M2[df_desde_m2['Articulo'][i]]

                
                df_desde_m2['Monto'][i]=articulos_M2[df_desde_m2['Articulo'][i]]*df_desde_m2['Cantidad Art Totales'][i]
            
                cont_dif_art += 1
            except:
                cont=cont+1

        valor_desde_m2=sum(list(df_desde_m2['Monto'].values))
        dif_P=valor_desde_m2-valor_desde

        data_resultante_p = [valor_desde_m2, dif_P, cont, dif_articulos]

        list_resultante_p = sorted(list_resultante_p, key = lambda x: abs(x[3]), reverse=True)

    else:
        data_resultante_p = 0
        list_resultante_p = 0
   
    return [data_resultante_p, list_resultante_p]

def presupuesto_generar_xls_proyecto(proyecto):

    destruir_analisis_ajuste(proyecto.id)
    
    computo = Computos.objects.all()
    capitulo = Capitulos.objects.all()
    compo = CompoAnalisis.objects.all()
    
    today = dt.date.today()
    wb = Workbook()
    ws = wb.active
    ws.title = "Almacen"
    ws["A1"] = "Capitulo"
    ws["B1"] = "Modelo"
    ws["C1"] = "Analisis"
    ws["D1"] = "Cantidad An"
    ws["E1"] = "Articulo"
    ws["F1"] = "Cantidad Ar"
    ws["G1"] = "Precio"
    ws["H1"] = "Cantidad Art Totales"
    ws["I1"] = "Monto"
    ws["J1"] = "Constante"
    ws["K1"] = "Tipo"

    contador = 2

    for c in capitulo:
        modelo = Modelopresupuesto.objects.filter(proyecto = proyecto, capitulo = c ).order_by("orden")
        for d in modelo:
            cantidad = d.cantidad

            for e in compo:

                if e.analisis == d.analisis:

                    ws["A{}".format(contador)] = c.nombre
                    ws["B{}".format(contador)] = d.id
                    ws["C{}".format(contador)] = d.analisis.codigo
                    ws["D{}".format(contador)] = cantidad
                    ws["E{}".format(contador)] = e.articulo.codigo
                    ws["F{}".format(contador)] = e.cantidad
                    ws["G{}".format(contador)] = e.articulo.valor
                    ws["H{}".format(contador)] = e.cantidad * cantidad
                    ws["I{}".format(contador)] = e.cantidad * e.articulo.valor * cantidad
                    if e.articulo.constante:
                            ws["J{}".format(contador)] = e.articulo.constante.nombre
                    else:
                        ws["J{}".format(contador)] = "SIN ASIGNAR"

                    if str(e.articulo.codigo)[0] == "3":
                        ws["K{}".format(contador)] = "MATERIALES"

                    else:
                        ws["K{}".format(contador)] = "MANO DE OBRA/SUB"

                    contador += 1

    #Establecer el nombre del archivo
    nombre_archivo = "{}.{}Almacen.xls".format(str(proyecto.nombre).replace(" ", ""),str(today))
    nombre_archivo
    mRoot = settings.MEDIA_ROOT
    wb.save(mRoot + "/{}".format(nombre_archivo))
    
    nuevo_presupuesto = PresupuestosAlmacenados(
        proyecto = proyecto,
        nombre = "vigente",
        archivo = (nombre_archivo),
    )

    nuevo_presupuesto.save()

    return "Save"

def destruir_analisis_ajuste(id_proyecto):
    proyecto = Proyectos.objects.get(id = id_proyecto)
    analisis = Analisis.objects.filter(nombre__icontains = 'AUTO-AJUSTE-' + str(proyecto.nombre))

    for analisis_in in analisis:
        analisis_in.delete()

def presupuestos_saldo_capitulo(id_proyecto):

    '''
    1 - Destruyo todos los analisis de AUTO-AJUSTE (Esto lo hago antes de generar el Excel)
    2 - Traemos datos generales (Compras, Articulos, Capitulos, Presupuestos, Proyecto) y el EXCEL que usaremos como base para calcular
    3 - Armamos un OBJETO de la necesidad de cada capitulo, tendra este formato:
        CAPITULO: {
            'id': ..,
            'valor_capitulo': ..,
            'inc': ..,
            'saldo': ..,
            'data': [
                'CODIGO ART': {
                    'articulo', ..,
                    'cantidad': ..,
                    'contante': ..,
                    'comprado': .., 
                    'precio': ..
                }
            ],
        }
    4 - Armamos el STOCK de las compras
    1 - Recorro todos los capitulos buscando un listado de todos los articulos comprados especificamente para ese capitulo
    2 - Creo una clave para ese articulo, para especificar su detalle de consumo
    
    '''

    ##########################################################################################################
    proyecto = Proyectos.objects.get(id = id_proyecto)
    presupuesto = Presupuestos.objects.get(proyecto = proyecto)
    archivo = PresupuestosAlmacenados.objects.get(proyecto = proyecto, nombre = "vigente").archivo
    df = pd.read_excel(archivo)
    compras = Compras.objects.filter(proyecto = proyecto)
    articulos = Articulos.objects.all()
    capitulos = Capitulos.objects.all()

    # Primero hacemos el JSON con la dara sin tener en cuenta las compras

    articulo_capitulo = []

    for capitulo in capitulos:
        if float(sum(df[df['Capitulo'] == capitulo.nombre]['Monto'])) > 0:
            inc = float(sum(df[df['Capitulo'] == capitulo.nombre]['Monto'])/sum(df['Monto'])*100)
        else:
            inc = 0

        saldo = float(sum(df[df['Capitulo'] == capitulo.nombre]['Monto']))
        mask_1 = df['Capitulo'] == capitulo.nombre 
        mask_2 = df['Tipo'] == 'MATERIALES'
        saldo_mat = float(sum(df[mask_1&mask_2]['Monto']))

        saldo_mo = saldo - saldo_mat

        info_saldo_capitulo = {
            capitulo.nombre: {
                'id': capitulo.id,
                'valor_capitulo': float(sum(df[df['Capitulo'] == capitulo.nombre]['Monto'])),
                'inc': inc,
                'saldo': saldo,
                'saldo_mat': saldo_mat,
                'saldo_mo': saldo_mo,
                'data': []
                }  
        }

        listado_articulos_capitulo = df[df['Capitulo'] == capitulo.nombre]['Articulo'].unique()
        for articulo in listado_articulos_capitulo:
            cantidad_articulo = float(sum(df[df['Capitulo'] == capitulo.nombre][df['Articulo'] == articulo]['Cantidad Art Totales']))
            precio_articulo = float(statistics.mean(df[df['Articulo'] == articulo]['Precio']))
            articulo_select = articulos.get(codigo = articulo)
            dict_articulo = { 
                str(articulo): {
                    'articulo': articulo_select.nombre, 
                    'cantidad': cantidad_articulo,
                    'contante': str(articulo_select.constante),
                    'comprado': 0, 
                    'precio': precio_articulo
                }
            }
            info_saldo_capitulo[capitulo.nombre]['data'].append(dict_articulo)
        
        articulo_capitulo.append(info_saldo_capitulo)

    # Detalle del consumo
    consumption_details = []
    articulos_comprados = compras.values_list("articulo", flat=True).distinct()
    stock_articulos = []

    for articulo in articulos_comprados:
        cantidad = sum(np.array(compras.filter(articulo = articulo, capitulo = None).values_list('cantidad', flat=True)))
        stock_articulos.append([articulo, cantidad])

    for stock in stock_articulos:

        dicc_stock = {
            stock[0]: {
                "compras": [],
                "detalle": []
                }
            }
        
        compras_articulo = compras.filter(articulo = stock[0], capitulo = None)
        total_comprado = 0
        
        for compra in compras_articulo:
            dicc_stock[stock[0]]["compras"].append(f"Compra {compra.documento}, al proveedor {compra.proveedor.name}, cantidad {round(compra.cantidad, 2)}")
            total_comprado += compra.cantidad

        total_comprado_des = total_comprado
        dicc_stock[stock[0]]["compras"].append(f"** Total comprado al cierre {round(total_comprado, 2)}")
        total_asignado = 0

        for capitulo in articulo_capitulo: ## Aqui recorro todos los capitulos
        
            total_asignado_capitulo = 0

            for key in capitulo.keys(): ## Aqui solo acceso al nombre del cap
                qy_capitulo = Capitulos.objects.get(nombre = key)
                compras_espeficias = compras.filter(capitulo__nombre = key, articulo = stock[0])
                cantidad_especifica = 0
                
                if len(compras_espeficias) > 0:
                    dicc_stock[stock[0]]["compras"].append(f"-> Compras especficias del capitulo {key}")

                    for c_esp in compras_espeficias:
                        dicc_stock[stock[0]]["compras"].append(f"-- Compra {c_esp.documento}, al proveedor {c_esp.proveedor.name}, cantidad {round(c_esp.cantidad, 2)}")
                        cantidad_especifica += c_esp.cantidad

                for i in range(len(capitulo[key]['data'])): ## Aqui recorrdo todos los articulos de ese capitulo
                    
                    if str(stock[0]) in list(capitulo[key]['data'][i].keys()): ## Pregunto si el articuo se necesita en el capitulo
                        necesidad_a_cubirir = float(capitulo[key]['data'][i][str(stock[0])]['cantidad']) - float(capitulo[key]['data'][i][str(stock[0])]['comprado'])
                        
                        if necesidad_a_cubirir > 0:
                        
                            if cantidad_especifica:
                                if cantidad_especifica >= necesidad_a_cubirir:
                                    total_asignado_capitulo += necesidad_a_cubirir
                                    capitulo[key]['data'][i][str(stock[0])]['comprado'] = necesidad_a_cubirir
                                    capitulo[key]['saldo'] = float(capitulo[key]['saldo'] - (necesidad_a_cubirir*capitulo[key]['data'][i][str(stock[0])]['precio']))

                                    if str(stock[0])[0] == "3":
                                        capitulo[key]['saldo_mat'] -= (necesidad_a_cubirir*capitulo[key]['data'][i][str(stock[0])]['precio'])

                                    else:
                                        capitulo[key]['saldo_mo'] -= (necesidad_a_cubirir*capitulo[key]['data'][i][str(stock[0])]['precio'])

                                    necesidad_a_cubirir = 0

                                else:
                                    total_asignado_capitulo += cantidad_especifica
                                    capitulo[key]['data'][i][str(stock[0])]['comprado'] = necesidad_a_cubirir
                                    capitulo[key]['saldo'] = float(capitulo[key]['saldo'] - (cantidad_especifica*capitulo[key]['data'][i][str(stock[0])]['precio']))

                                    if str(stock[0])[0] == "3":
                                        capitulo[key]['saldo_mat'] -= (cantidad_especifica*capitulo[key]['data'][i][str(stock[0])]['precio'])

                                    else:
                                        capitulo[key]['saldo_mo'] -= (cantidad_especifica*capitulo[key]['data'][i][str(stock[0])]['precio'])

                                    necesidad_a_cubirir -= cantidad_especifica
                                    cantidad_especifica = 0

                            if total_comprado_des >= necesidad_a_cubirir:
                                total_asignado_capitulo += necesidad_a_cubirir
                                capitulo[key]['data'][i][str(stock[0])]['comprado'] += necesidad_a_cubirir
                                capitulo[key]['saldo'] = float(capitulo[key]['saldo'] - (necesidad_a_cubirir*capitulo[key]['data'][i][str(stock[0])]['precio']))

                                if str(stock[0])[0] == "3":
                                    capitulo[key]['saldo_mat'] -= (necesidad_a_cubirir*capitulo[key]['data'][i][str(stock[0])]['precio'])

                                else:
                                    capitulo[key]['saldo_mo'] -= (necesidad_a_cubirir*capitulo[key]['data'][i][str(stock[0])]['precio'])

                                total_comprado_des = total_comprado_des - necesidad_a_cubirir
                                necesidad_a_cubirir = 0


                            else:
                                total_asignado_capitulo += total_comprado_des                                
                                capitulo[key]['data'][i][str(stock[0])]['comprado'] += total_comprado_des
                                capitulo[key]['saldo'] = float(capitulo[key]['saldo'] - (total_comprado_des*capitulo[key]['data'][i][str(stock[0])]['precio']))

                                if str(stock[0])[0] == "3":
                                    capitulo[key]['saldo_mat'] -= (total_comprado_des*capitulo[key]['data'][i][str(stock[0])]['precio'])

                                else:
                                    capitulo[key]['saldo_mo'] -= (total_comprado_des*capitulo[key]['data'][i][str(stock[0])]['precio'])

                                necesidad_a_cubirir -= total_comprado_des
                                total_comprado_des = 0

                    dicc_stock[stock[0]]["detalle"].append(f"*** Capitulo {key}, se asigno {round(total_asignado_capitulo, 2)}")

                if cantidad_especifica > 0:
                    nombre = f'AUTO-AJUSTE-{proyecto.nombre}-{key}'
                    analisis_ajuste = Analisis.objects.filter(nombre = nombre)

                    if len(analisis_ajuste):
                        analisis_ajuste = analisis_ajuste.first()

                    else:
                        now = dt.datetime.now()
                        codigo = "80"+str(qy_capitulo.id)+str(now.year)+str(now.month)+str(now.day)+str(now.hour)+str(now.minute)+str(now.second)
                        unidad = 'GL'
                        analisis_ajuste = Analisis.objects.create(id = 99, codigo=int(codigo), nombre = nombre, unidad = unidad)

                    articulo = Articulos.objects.get(codigo = stock[0])
                    compo_analisis = CompoAnalisis.objects.create(articulo = articulo, analisis = analisis_ajuste, cantidad = cantidad_especifica)

                    modelo_filter = Modelopresupuesto.objects.filter(analisis = analisis_ajuste, capitulo = qy_capitulo, proyecto = proyecto, cantidad = 1, orden = 99)
                    if not len(modelo_filter):
                        modelo = Modelopresupuesto.objects.create(analisis = analisis_ajuste, capitulo = qy_capitulo, proyecto = proyecto, cantidad = 1, orden = 99,
                        comentario = "Analisis automatico")

            total_asignado += total_asignado_capitulo

        dicc_stock[stock[0]]["detalle"].append(f"Total asignado {round(total_asignado, 2)}")
        dicc_stock[stock[0]]["detalle"].append(f"Total sobrante/pendiente de comprar {round((total_comprado - total_asignado), 2)}")
        consumption_details.append(dicc_stock)

    presupuesto.balance_details = json.dumps(articulo_capitulo)
    presupuesto.consumption_details = json.dumps(consumption_details)
    presupuesto.save()




    # # Ordenamos cada capitulo con una lista donde no se repitan los articulos

    # presupuesto_capitulo = []

    # contador = 0

    # for i in range(37):

    #     dato = datos_viejos[contador]

    #     nuevo_art_cant = []

    #     lista_art_cap = []

    #     for art_cant in dato[2]:

    #         lista_art_cap.append(art_cant[0])

    #     lista_art_cap = list(set(lista_art_cap))
        
    #     for articulo in lista_art_cap:

    #         cantidad = 0

    #         for articulo2 in dato[2]:

    #             if articulo == articulo2[0]:

    #                 cantidad = cantidad + articulo2[1]

    #         nuevo_art_cant.append((articulo, cantidad))

    #     presupuesto_capitulo.append((dato[0], dato[1], nuevo_art_cant))    
        
    #     contador += 1


    # # Ordenamos la compra para que sea una sola lista

    # articulos_comprados = []

    # for compra in compras:

    #     articulos_comprados.append(compra.articulo)

    # articulos_comprados = list(set(articulos_comprados))

    # #Armamos el stock con todas las compras realizadas de este proyecto

    # stock_articulos = []

    # for articulo in articulos_comprados:

    #     cantidad = sum(np.array(Compras.objects.values_list('cantidad').filter(proyecto = proyecto, articulo = articulo)))

    #     stock_articulos.append((articulo, cantidad))

    # #Armamos el saldo --> Hay un error ya que al descartar menores a 0, olvidamos que restan consumo

    # saldo_capitulo = []

    # for capitulo_presupuesto in presupuesto_capitulo:

    #     articulos_saldo = []

    #     for articulos_presupuesto in capitulo_presupuesto[2]:

    #         if articulos_presupuesto[0] in articulos_comprados and articulos_presupuesto[1]>=0:

    #             contador = 0

    #             for articulos_stock in stock_articulos:

    #                 #Si encontramos el articulo del capitulo en el stock, activamos una de las 3 posibilidades

    #                 if articulos_stock[0] == articulos_presupuesto[0]:

    #                     articulos_stock = list(articulos_stock)

    #                     if articulos_stock[1] > articulos_presupuesto[1]:

    #                         articulos_stock[1] = float(articulos_stock[1]) - float(articulos_presupuesto[1])                           

    #                         stock_articulos[contador] = list(stock_articulos[contador])
    #                         stock_articulos[contador][1] = articulos_stock[1]

    #                         articulos_stock = tuple(articulos_stock)
    #                         stock_articulos[contador] = tuple(stock_articulos[contador])

    #                     elif articulos_stock[1] == articulos_presupuesto[1]:

    #                         articulos_stock[1] = 0
    #                         stock_articulos[contador] = list(stock_articulos[contador])
    #                         stock_articulos[contador][1] = articulos_stock[1]

    #                         articulos_stock = tuple(articulos_stock)
    #                         stock_articulos[contador] = tuple(stock_articulos[contador])

    #                     elif articulos_stock[1] < articulos_presupuesto[1]:

    #                         cantidad_saldo = float(articulos_presupuesto[1]) - float(articulos_stock[1])

    #                         articulos_stock[1] = 0

    #                         stock_articulos[contador] = list(stock_articulos[contador])
    #                         stock_articulos[contador][1] = articulos_stock[1]

    #                         articulos_saldo.append((articulos_presupuesto[0], cantidad_saldo))

    #                         articulos_stock = tuple(articulos_stock)
    #                         stock_articulos[contador] = tuple(stock_articulos[contador])
    #                 contador += 1
    #         else:
    #             articulos_saldo.append(articulos_presupuesto)

    #     #Modificado con el saldo
                
    #     saldo_capitulo.append((capitulo_presupuesto[0], capitulo_presupuesto[1], articulos_saldo))


    # return saldo_capitulo



