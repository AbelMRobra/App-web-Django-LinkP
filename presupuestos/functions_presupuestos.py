from openpyxl import Workbook
from presupuestos.models import Capitulos,CompoAnalisis,Modelopresupuesto,PresupuestosAlmacenados
from django.conf import settings
import requests
from .wabot import WABot

import datetime as dt

def generar_excel(computo,proyecto):
    capitulo = Capitulos.objects.all()
    compo = CompoAnalisis.objects.all()
    
    
    today = dt.date.today()
    wb = Workbook()
    ws = wb.active
    ws.title = "Almacen"
    ws["A1"] = "Capitulo"
    ws["B1"] = "Modelo"
    ws["C1"] = "Analisis"
    ws["D1"] = "Cantidad An"
    ws["E1"] = "Articulo"
    ws["F1"] = "Cantidad Ar"
    ws["G1"] = "Precio"
    ws["H1"] = "Cantidad Art Totales"
    ws["I1"] = "Monto"

    contador = 2

  
    for c in capitulo:
        modelo = Modelopresupuesto.objects.filter(proyecto = proyecto, capitulo = c ).order_by("orden")
        for d in modelo:
            cantidad = d.cantidad
            if d.cantidad == None:
                if "SOLO MANO DE OBRA" in str(d.analisis): 
                    cantidad = 0
                    for h in computo:
                        if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                            cantidad = cantidad + h.valor_vacio                      
                    for e in compo:
                        if e.analisis == d.analisis:
                            ws["A{}".format(contador)] = c.nombre
                            ws["B{}".format(contador)] = d.id
                            ws["C{}".format(contador)] = d.analisis.codigo
                            ws["D{}".format(contador)] = cantidad
                            ws["E{}".format(contador)] = e.articulo.codigo
                            ws["F{}".format(contador)] = e.cantidad
                            ws["G{}".format(contador)] = e.articulo.valor
                            ws["H{}".format(contador)] = e.cantidad * cantidad
                            ws["I{}".format(contador)] = e.cantidad * e.articulo.valor * cantidad
                            contador += 1

                else:

                    cantidad = 0

                    for h in computo:

                        if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                            
                            cantidad = cantidad + h.valor_lleno

                    for e in compo:

                        if e.analisis == d.analisis:

                            ws["A{}".format(contador)] = c.nombre
                            ws["B{}".format(contador)] = d.id
                            ws["C{}".format(contador)] = d.analisis.codigo
                            ws["D{}".format(contador)] = cantidad
                            ws["E{}".format(contador)] = e.articulo.codigo
                            ws["F{}".format(contador)] = e.cantidad
                            ws["G{}".format(contador)] = e.articulo.valor
                            ws["H{}".format(contador)] = e.cantidad * cantidad
                            ws["I{}".format(contador)] = e.cantidad * e.articulo.valor * cantidad
                            contador += 1
                    
            else:

                for e in compo:

                    if e.analisis == d.analisis:

                        ws["A{}".format(contador)] = c.nombre
                        ws["B{}".format(contador)] = d.id
                        ws["C{}".format(contador)] = d.analisis.codigo
                        ws["D{}".format(contador)] = cantidad
                        ws["E{}".format(contador)] = e.articulo.codigo
                        ws["F{}".format(contador)] = e.cantidad
                        ws["G{}".format(contador)] = e.articulo.valor
                        ws["H{}".format(contador)] = e.cantidad * cantidad
                        ws["I{}".format(contador)] = e.cantidad * e.articulo.valor * cantidad
                        contador += 1

    #Establecer el nombre del archivo
    nombre_archivo = "{}.{}Almacen.xls".format(str(proyecto.nombre).replace(" ", ""),str(today))
    nombre_archivo
    mRoot = settings.MEDIA_ROOT
    wb.save(mRoot + "/{}".format(nombre_archivo))
    
    nuevo_presupuesto = PresupuestosAlmacenados(
        proyecto = proyecto,
        nombre = "vigente",
        archivo = (nombre_archivo),
    )

    nuevo_presupuesto.save()


def notificar_botwp(mensaje,response_servidor,numero_prueba):
    send = mensaje
    id = "-455382561"
    params = {
        'chat_id' : id,
        'text' : send
    }
    
    token = "1880193427:AAH-Ej5ColiocfDZrDxUpvsJi5QHWsASRxA"
    url = "https://api.telegram.org/bot" + token + "/sendMessage"
 
    requests.post(url, params=params)
    bot_wp = WABot(response_servidor)
    bot_wp.send_message(numero_prueba, send)

