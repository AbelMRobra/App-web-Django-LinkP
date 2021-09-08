from openpyxl import Workbook
from presupuestos.models import Capitulos,CompoAnalisis,Modelopresupuesto,PresupuestosAlmacenados
from django.conf import settings
import requests
from presupuestos.wabot import WABot
import datetime as dt
import pandas as pd
import requests
from presupuestos.models import Capitulos, PresupuestosAlmacenados, Analisis, Articulos



def bot_telegram(send, id, token):
    url = "https://api.telegram.org/bot" + token + "/sendMessage"

    params = {
        'chat_id' : id,
        'text' : send
    }

    requests.post(url, params=params)

def auditor_presupuesto(proyecto,fecha_desde, fecha_hasta):
    #esta linea trae 29 objetos
    mensaje = ""
    data_almacenada_desde_objects = PresupuestosAlmacenados.objects.filter(nombre = fecha_desde,proyecto=proyecto)
    
    #esta linea da error porque el queryset se encuentra vacio
    data_almacenada_hasta_objects = PresupuestosAlmacenados.objects.filter(nombre = fecha_hasta,proyecto=proyecto)
    
    if data_almacenada_desde_objects.exists() and data_almacenada_hasta_objects.exists():

        try:
        
            data_almacenada_desde=data_almacenada_desde_objects[0]
            data_almacenada_hasta=data_almacenada_hasta_objects[0]

            df_desde = pd.read_excel(data_almacenada_desde.archivo)
            df_hasta = pd.read_excel(data_almacenada_hasta.archivo)

            
            # Estudio de diferencia de cantidades

            valor_proyecto_desde=sum(list(df_desde['Monto'].values))
            valor_proyecto_hasta=sum(list(df_hasta['Monto'].values))

            

            df_desde_Q = df_desde.drop(['Precio', 'Monto'], axis = 1) #se eliminan
            df_hasta_Q = df_hasta.drop(['Precio', 'Monto'], axis = 1) #se eliminan

            
            df_dif_Q = df_hasta_Q.merge(df_desde_Q, how='outer', indicator='union')
            df_dif_Q.drop(df_dif_Q[df_dif_Q['union'] == "both"].index, inplace=True) #se unen
            
            df_dif_Q_desde = df_dif_Q[df_dif_Q['union'] == "left_only"]
            df_dif_Q_hasta = df_dif_Q[df_dif_Q['union'] == "right_only"]

            
            lista_capitulos_afectado = list(set(df_dif_Q["Capitulo"].values)) #set elimina los nombres de ls capitulos repetidos

            data_procesada = []

            cuantif_q = 0
            errores = 0
            
            data_capitulo = []

            # Definimos un diccionario con los precios en M2

            articulos_M2={}

            for art in df_hasta.index:
                
                articulos_M2[df_hasta['Articulo'][art]]=df_hasta['Precio'][art]

            for art in df_desde.index:
                if df_desde['Articulo'][art] not in articulos_M2:
                    articulos_M2[df_desde['Articulo'][art]]=df_desde['Precio'][art]


            cont_cap = 0
            for capitulo in lista_capitulos_afectado:
                # Primeros los articulos eliminados
                
                cuantif_q_cap = 0
                errores_cap = 0

                df_dif_Q_desde_aux = df_dif_Q_desde[df_dif_Q_desde['Capitulo'] == capitulo]

                articulos_eliminados = []

                for i in df_dif_Q_desde_aux.index:
                    try:
                        
                        analisis = Analisis.objects.get(codigo = df_dif_Q_desde_aux['Analisis'][i])
                    except:
                        analisis = "¿¿"+str(df_dif_Q_desde_aux['Analisis'][i])+"??"
                        errores += 1
                        errores_cap += 1
                    try:
                        articulo = Articulos.objects.get(codigo = df_dif_Q_desde_aux['Articulo'][i])
                    except:
                        articulo = "¿¿"+str(df_dif_Q_desde_aux['Articulo'][i])+"??"

                    cantidad = df_dif_Q_desde_aux['Cantidad Art Totales'][i]
                    try:
                        modif = cantidad*articulos_M2[articulo.codigo]
                        cuantif_q += modif
                        cuantif_q_cap += modif
                    except:
                        modif = "??"
                        errores += 1
                        errores_cap += 1

                    articulos_eliminados.append((analisis, articulo, cantidad, -modif))

                df_dif_Q_hasta_aux = df_dif_Q_hasta[df_dif_Q_hasta['Capitulo'] == capitulo]

                articulos_agregados = []

                for i in df_dif_Q_hasta_aux.index:
                    try:
                        analisis = Analisis.objects.get(codigo = df_dif_Q_hasta_aux['Analisis'][i])
                    except:
                        analisis = "¿¿"+str(df_dif_Q_hasta_aux['Analisis'][i])+"??"
                        errores += 1
                        errores_cap += 1
                    try:
                        articulo = Articulos.objects.get(codigo = df_dif_Q_hasta_aux['Articulo'][i])
                    except:
                        articulo = "¿¿"+str(df_dif_Q_hasta_aux['Articulo'][i])+"??"
                        

                    cantidad = df_dif_Q_hasta_aux['Cantidad Art Totales'][i]
                    try:
                        modif = cantidad*articulos_M2[articulo.codigo]
                        cuantif_q -= modif
                        cuantif_q_cap -= modif
                    except:
                        modif = "??"
                        errores += 1
                        errores_cap += 1

                    articulos_agregados.append((analisis, articulo, cantidad, modif))
                
                cont_cap += 1

                data_capitulo.append((capitulo, articulos_agregados, articulos_eliminados, cuantif_q_cap, errores_cap, cont_cap))
            diferencia_precio = valor_proyecto_hasta - valor_proyecto_desde
            data_procesada.append((data_capitulo, cuantif_q, errores, valor_proyecto_desde , valor_proyecto_hasta, diferencia_precio ))
        except:
            data_procesada=None
            mensaje='Hubo un error en el proceso'
    else:
        data_procesada=None
        mensaje='No se encontraron registros en esa fecha para ese proyecto, prueba con: {}'.format(list(PresupuestosAlmacenados.objects.filter(proyecto = proyecto).exclude(nombre="vigente").values_list("nombre", flat=True).distinct()))      

    return data_procesada,mensaje

def auditor_presupuesto_p(proyecto, fecha_desde, fecha_hasta):

    cont=0
    data_almacenada_desde_objects = PresupuestosAlmacenados.objects.filter(nombre = fecha_desde,proyecto=proyecto)
    
    #esta linea da error porque el queryset se encuentra vacio
    data_almacenada_hasta_objects = PresupuestosAlmacenados.objects.filter(nombre = fecha_hasta,proyecto=proyecto)
    
    if data_almacenada_desde_objects.exists() and data_almacenada_hasta_objects.exists():
        
        
        data_almacenada_desde=data_almacenada_desde_objects[0]
        data_almacenada_hasta=data_almacenada_hasta_objects[0]

        df_desde = pd.read_excel(data_almacenada_desde.archivo)

        valor_desde=sum(list(df_desde['Monto'].values))
        df_hasta = pd.read_excel(data_almacenada_hasta.archivo)

        articulos_M2={}
        dif_articulos={}

        for art in df_hasta.index:
            
            #de la columna 'articulos' toma los registros con el codigo de cada fila
            articulos_M2[df_hasta['Articulo'][art]]=df_hasta['Precio'][art]

        df_desde_m2=df_desde
        
        cont_dif_art = 0
        for i in df_desde_m2.index:
            try:

                dif_articulos[df_desde_m2['Articulo'][i]]= [abs((articulos_M2[df_desde_m2['Articulo'][i]]/df_desde_m2['Precio'][i]-1)*100), cont_dif_art]


                df_desde_m2['Precio'][i]=articulos_M2[df_desde_m2['Articulo'][i]]

                
                df_desde_m2['Monto'][i]=articulos_M2[df_desde_m2['Articulo'][i]]*df_desde_m2['Cantidad Art Totales'][i]
            
                cont_dif_art += 1
            except:
                cont=cont+1

        valor_desde_m2=sum(list(df_desde_m2['Monto'].values))
        dif_P=valor_desde_m2-valor_desde

        data_resultante_p = [valor_desde_m2, dif_P, cont, dif_articulos]

    else:
        data_resultante_p = 0
   
    return data_resultante_p


def generar_excel(computo,proyecto):
    capitulo = Capitulos.objects.all()
    compo = CompoAnalisis.objects.all()
    
    
    today = dt.date.today()
    wb = Workbook()
    ws = wb.active
    ws.title = "Almacen"
    ws["A1"] = "Capitulo"
    ws["B1"] = "Modelo"
    ws["C1"] = "Analisis"
    ws["D1"] = "Cantidad An"
    ws["E1"] = "Articulo"
    ws["F1"] = "Cantidad Ar"
    ws["G1"] = "Precio"
    ws["H1"] = "Cantidad Art Totales"
    ws["I1"] = "Monto"

    contador = 2

  
    for c in capitulo:
        modelo = Modelopresupuesto.objects.filter(proyecto = proyecto, capitulo = c ).order_by("orden")
        for d in modelo:
            cantidad = d.cantidad
            if d.cantidad == None:
                if "SOLO MANO DE OBRA" in str(d.analisis): 
                    cantidad = 0
                    for h in computo:
                        if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                            cantidad = cantidad + h.valor_vacio                      
                    for e in compo:
                        if e.analisis == d.analisis:
                            ws["A{}".format(contador)] = c.nombre
                            ws["B{}".format(contador)] = d.id
                            ws["C{}".format(contador)] = d.analisis.codigo
                            ws["D{}".format(contador)] = cantidad
                            ws["E{}".format(contador)] = e.articulo.codigo
                            ws["F{}".format(contador)] = e.cantidad
                            ws["G{}".format(contador)] = e.articulo.valor
                            ws["H{}".format(contador)] = e.cantidad * cantidad
                            ws["I{}".format(contador)] = e.cantidad * e.articulo.valor * cantidad
                            contador += 1

                else:

                    cantidad = 0

                    for h in computo:

                        if h.proyecto == proyecto and h.tipologia == d.vinculacion:
                            
                            cantidad = cantidad + h.valor_lleno

                    for e in compo:

                        if e.analisis == d.analisis:

                            ws["A{}".format(contador)] = c.nombre
                            ws["B{}".format(contador)] = d.id
                            ws["C{}".format(contador)] = d.analisis.codigo
                            ws["D{}".format(contador)] = cantidad
                            ws["E{}".format(contador)] = e.articulo.codigo
                            ws["F{}".format(contador)] = e.cantidad
                            ws["G{}".format(contador)] = e.articulo.valor
                            ws["H{}".format(contador)] = e.cantidad * cantidad
                            ws["I{}".format(contador)] = e.cantidad * e.articulo.valor * cantidad
                            contador += 1
                    
            else:

                for e in compo:

                    if e.analisis == d.analisis:

                        ws["A{}".format(contador)] = c.nombre
                        ws["B{}".format(contador)] = d.id
                        ws["C{}".format(contador)] = d.analisis.codigo
                        ws["D{}".format(contador)] = cantidad
                        ws["E{}".format(contador)] = e.articulo.codigo
                        ws["F{}".format(contador)] = e.cantidad
                        ws["G{}".format(contador)] = e.articulo.valor
                        ws["H{}".format(contador)] = e.cantidad * cantidad
                        ws["I{}".format(contador)] = e.cantidad * e.articulo.valor * cantidad
                        contador += 1

    #Establecer el nombre del archivo
    nombre_archivo = "{}.{}Almacen.xls".format(str(proyecto.nombre).replace(" ", ""),str(today))
    nombre_archivo
    mRoot = settings.MEDIA_ROOT
    wb.save(mRoot + "/{}".format(nombre_archivo))
    
    nuevo_presupuesto = PresupuestosAlmacenados(
        proyecto = proyecto,
        nombre = "vigente",
        archivo = (nombre_archivo),
    )

    nuevo_presupuesto.save()




