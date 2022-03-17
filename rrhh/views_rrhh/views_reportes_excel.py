import numpy as np
import string
import datetime
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from django.views.generic.base import TemplateView 
from django.http import HttpResponse
from ..models import MonedaLink, EntregaMoneda, CanjeMonedas, datosusuario


class ExcelReporteLinkcoins(TemplateView):

    def get(self, request, *args, **kwargs):
        query_monedas = MonedaLink.objects.all()
        query_entregas = EntregaMoneda.objects.all()
        query_canje = CanjeMonedas.objects.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Linkcoins"

        ws["A2"] = "Usuario"
        ws["B2"] = "Genero"
        ws["C2"] = "Genero bajo LINK"
        ws["D2"] = "Entrego"
        ws["E2"] = "Entrego bajo LINK"
        ws["F2"] = "Recibio"
        ws["G2"] = "Canjeo"

        ws["A2"].alignment = Alignment(horizontal = "center", vertical="center")
        ws["A2"].font = Font(bold = True, color="EAE3F2")
        ws["A2"].fill = PatternFill("solid", fgColor= "625E66")

        ws["B2"].alignment = Alignment(horizontal = "center", vertical="center")
        ws["B2"].font = Font(bold = True, color="EAE3F2")
        ws["B2"].fill = PatternFill("solid", fgColor= "625E66")

        ws["D2"].alignment = Alignment(horizontal = "center", vertical="center")
        ws["D2"].font = Font(bold = True, color="EAE3F2")
        ws["D2"].fill = PatternFill("solid", fgColor= "625E66")

        ws["E2"].alignment = Alignment(horizontal = "center", vertical="center")
        ws["E2"].font = Font(bold = True, color="EAE3F2")
        ws["E2"].fill = PatternFill("solid", fgColor= "625E66")

        ws["F2"].alignment = Alignment(horizontal = "center", vertical="center")
        ws["F2"].font = Font(bold = True, color="EAE3F2")
        ws["F2"].fill = PatternFill("solid", fgColor= "625E66")

        ws["G2"].alignment = Alignment(horizontal = "center", vertical="center")
        ws["G2"].font = Font(bold = True, color="EAE3F2")
        ws["G2"].fill = PatternFill("solid", fgColor= "625E66")

        ws["C2"].alignment = Alignment(horizontal = "center", vertical="center")
        ws["C2"].font = Font(bold = True, color="EAE3F2")
        ws["C2"].fill = PatternFill("solid", fgColor= "625E66")

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 22

        contador = 3
        usuarios = query_monedas.values_list("usuario_portador__id", flat=True).order_by("usuario_portador__identificacion").distinct()

        for usuario in usuarios:
            user = datosusuario.objects.get(id = usuario)

            ws["A"+str(contador)] = str(user.identificacion) + ": " + str(user.nombre)
            ws["B"+str(contador)] = query_monedas.filter(usuario_portador=user).count()
            ws["C"+str(contador)] = query_monedas.filter(usuario_portador=user, nombre__icontains='LINK').count()
            ws["D"+str(contador)] = query_entregas.filter(moneda__usuario_portador=user).count()
            ws["E"+str(contador)] = query_entregas.filter(moneda__usuario_portador=user, moneda__nombre__icontains='LINK').count()
            ws["F"+str(contador)] = query_entregas.filter(usuario_recibe=user).count()
            ws["G"+str(contador)] = sum(query_canje.filter(usuario=user).values_list('monedas', flat=True))

            contador += 1

        #Establecer el nombre del archivo
        nombre_archivo = "ReporteLinkcoins.xls"
        
        #Definir tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo).replace(',', '_')
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
